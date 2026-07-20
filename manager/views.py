from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import JsonResponse, HttpResponse, HttpResponseNotAllowed, FileResponse, Http404
from django.contrib import messages
from django.db import transaction
from django.views.decorators.http import require_http_methods, require_POST, require_GET
from django.db.models import Sum, Avg, Q, Count, Max, F, Prefetch, OuterRef, Subquery, Value, IntegerField
from django.db.models.functions import Coalesce
from django.utils import timezone
from decimal import Decimal
from datetime import timedelta
from django.core.paginator import Paginator
import uuid
import json
from . import models
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.utils import PlotlyJSONEncoder

from django.utils.text import slugify
from django.core.mail import send_mail
from django.conf import settings as django_settings
from django.views.decorators.csrf import csrf_protect, ensure_csrf_cookie
from django.core.cache import cache
from django.utils.translation import gettext as _

# Marketplace imports for unified cart
from marketplace.models import (
    Product, Course, SupermarketItem, MarketplaceCartItem,
    MarketplaceOrder, MarketplaceOrderItem, CourseProgress, CourseLesson, CourseSection,
    PostDeliveryReview,
)
from marketplace.utils import (
    build_attribute_groups,
    normalize_selected_attributes,
    validate_selected_attributes,
)
from marketplace.pricing_rules import validate_quantity
from book_Project.payment_config import build_payment_options
import hashlib
import logging

logger = logging.getLogger(__name__)


def _annotate_book_delivered(qs):
    """Sum OrderItem quantities for delivered book orders (payment completed)."""
    sub = models.OrderItem.objects.filter(
        book_id=OuterRef('pk'),
        order__status='delivered',
        order__payment_status='completed',
    ).values('book_id').annotate(total=Sum('quantity')).values('total')[:1]
    return qs.annotate(
        sold_delivered=Coalesce(Subquery(sub, output_field=IntegerField()), Value(0))
    )


def _annotate_product_delivered(qs):
    sub = MarketplaceOrderItem.objects.filter(
        item_type='product',
        item_id=OuterRef('pk'),
        order__status='delivered',
        order__payment_status='completed',
    ).values('item_id').annotate(total=Sum('quantity')).values('total')[:1]
    return qs.annotate(
        sold_delivered=Coalesce(Subquery(sub, output_field=IntegerField()), Value(0))
    )


def _annotate_course_delivered(qs):
    sub = MarketplaceOrderItem.objects.filter(
        item_type='course',
        item_id=OuterRef('pk'),
        order__status='delivered',
        order__payment_status='completed',
    ).values('item_id').annotate(total=Sum('quantity')).values('total')[:1]
    return qs.annotate(
        sold_delivered=Coalesce(Subquery(sub, output_field=IntegerField()), Value(0))
    )


def _annotate_supermarket_delivered(qs):
    sub = MarketplaceOrderItem.objects.filter(
        item_type='supermarket',
        item_id=OuterRef('pk'),
        order__status='delivered',
        order__payment_status='completed',
    ).values('item_id').annotate(total=Sum('quantity')).values('total')[:1]
    return qs.annotate(
        sold_delivered=Coalesce(Subquery(sub, output_field=IntegerField()), Value(0))
    )


# ---------------------------------------------------------------------------
# Simple IP-based brute-force guard (no extra packages needed).
# Limits: 5 failed attempts per 5 minutes per IP.
# ---------------------------------------------------------------------------
_LOGIN_MAX_ATTEMPTS = 5
_LOGIN_WINDOW_SECONDS = 300  # 5 minutes


def _get_client_ip(request):
    """Return the real client IP, honouring X-Forwarded-For behind a proxy."""
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        return x_forwarded_for.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR", "0.0.0.0")


def _is_rate_limited(ip: str) -> bool:
    """Return True when the IP has exceeded the login failure threshold."""
    key = f"login_fail:{ip}"
    failures = cache.get(key, 0)
    return failures >= _LOGIN_MAX_ATTEMPTS


def _record_login_failure(ip: str) -> int:
    """Increment failure counter; returns current count."""
    key = f"login_fail:{ip}"
    try:
        return cache.incr(key)
    except ValueError:
        cache.set(key, 1, _LOGIN_WINDOW_SECONDS)
        return 1


def _reset_login_failures(ip: str) -> None:
    cache.delete(f"login_fail:{ip}")


# ---------------------------------------------------------------------------
# Generalized keyed rate limiter — same cache-based approach as the login
# guard above, reused for any endpoint that needs a per-IP request cap
# (signup, password reset, checkout, payment polling) without duplicating
# the counter logic per endpoint.
# ---------------------------------------------------------------------------
def _is_rate_limited_key(key: str, max_attempts: int) -> bool:
    """True when `key`'s counter has already reached/exceeded max_attempts."""
    return cache.get(key, 0) >= max_attempts


def _record_attempt_key(key: str, window_seconds: int) -> int:
    """Increment (or start) `key`'s counter; returns the new count."""
    try:
        return cache.incr(key)
    except ValueError:
        cache.set(key, 1, window_seconds)
        return 1


# 方法实现（数据库操作和页面跳转）
# ====================   默认跳转  ===========================
def index(request):
    return redirect("/manager/login")


# ====================   管理员登录  ===========================
def manager_login(request):
    """Admin login with PBKDF2 password check and IP-based rate limiting."""
    if request.method == "GET":
        return render(request, "admin/admin.html")

    if request.method == "POST":
        ip = _get_client_ip(request)

        # --- Rate-limit check ---
        if _is_rate_limited(ip):
            return render(
                request,
                "admin/admin.html",
                {"error": "登录尝试次数过多，请5分钟后再试。"},
                status=429,
            )

        number = request.POST.get("number", "").strip()
        password = request.POST.get("password", "")

        if not number or not password:
            _record_login_failure(ip)
            return render(request, "admin/admin.html", {"error": "账号和密码不能为空"})

        try:
            manager = models.Manager.objects.get(number=number)
        except models.Manager.DoesNotExist:
            _record_login_failure(ip)
            return render(request, "admin/admin.html", {"error": "账号或密码错误"})

        if not manager.check_password(password):
            _record_login_failure(ip)
            return render(request, "admin/admin.html", {"error": "账号或密码错误"})

        # Successful login
        _reset_login_failures(ip)
        request.session.cycle_key()          # prevent session fixation
        request.session["name"] = manager.name
        request.session["manager_id"] = manager.id
        request.session["is_admin"] = manager.is_admin
        return redirect("/manager/dashboard/")

    return redirect("/manager/login")


# ====================   管理员登出  ===========================
def manager_logout(request):
    """Manager logout — clears session and redirects to login page."""
    request.session.flush()   # flush() deletes AND regenerates the key
    return redirect("/manager/login")


# Keep the old logout function for backward compatibility
def logout(request):
    """Legacy logout function — delegates to manager_logout."""
    return manager_logout(request)


# ====================   一、出版社模块  ===========================
# 01添加出版社
def add_publisher(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    # 直接访问（get请求），跳转界面
    if request.method == "GET":
        return render(request, 'publisher/add_publisher.html', {"name": request.session["name"]})
    # 提交表单请求（POST）,处理数据库,跳转到列表页面
    if request.method == "POST":
        # 1.获取请求参数
        publisher_name = request.POST.get("publisher_name")
        publisher_address = request.POST.get("publisher_address")
        # 2.将数据保存到数据库中（insert）
        # publisher_name/publisher_address are django-modeltranslation
        # fields — passing them as .objects.create() kwargs silently drops
        # them (the library skips populating the per-language column while
        # _mt_init is set during __init__), so assign as plain attributes.
        new_publisher = models.Publisher()
        new_publisher.publisher_name = publisher_name
        new_publisher.publisher_address = publisher_address
        new_publisher.save()
        messages.success(request, f'出版社 "{new_publisher.publisher_name}" 添加成功！')
        # 3.重添加成功，返回出版社列表
        return redirect("/manager/publisher_list")


# 02查询所有出版社信息
def publisher_list(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    # 查询数据库中的所以信息（select * from）
    publisher_list = models.Publisher.objects.all()
    # 跳转到publisher_list页面，传入publisher_list数据
    return render(request, "publisher/publisher_list.html",
                  {"publisher_obj_list": publisher_list, "name": request.session["name"]})


# 03修改出版社信息
def edit_publisher(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    # 1获取到的是表单提交过来的内容（POST），获取对应的值
    if request.method == "POST":
        id = request.POST.get('id')
        publisher_name = request.POST.get("publisher_name")
        publisher_address = request.POST.get("publisher_address")
        # 2根据id去数据库查找对象（where id = id）
        publisher_obj = models.Publisher.objects.get(id=id)
        # 3修改
        publisher_obj.publisher_name = publisher_name
        publisher_obj.publisher_address = publisher_address
        # 04更新数据库（update）
        publisher_obj.save()
        messages.success(request, f'出版社 "{publisher_obj.publisher_name}" 更新成功！')
        # 4重定向到出版社列表
        return redirect('/manager/publisher_list/')
    # get请求跳转界面（获取原始数据）
    else:
        # 1获取id
        id = request.GET.get('id')
        # 2去数据库中查找相应的数据
        publisher_obj = models.Publisher.objects.get(id=id)
        publisher_obj_list = models.Publisher.objects.all()
        # 3返回页面
        return render(request, "publisher/edit_publisher.html",
                      {"publisher_obj": publisher_obj, "publisher_obj_list": publisher_obj_list,
                       "name": request.session["name"]})


# 04删除出版社信息
def delete_publisher(request):
    if "name" not in request.session:
        return redirect("/manager/login")
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    pub_id = request.POST.get('id')
    try:
        pub_qs = models.Publisher.objects.filter(id=pub_id)
        pub_name = pub_qs.first().publisher_name if pub_qs.exists() else None
        deleted, _ = pub_qs.delete()
        if deleted:
            messages.success(request, f'出版社 "{pub_name}" 已删除')
        else:
            messages.error(request, '出版社不存在')
    except Exception as exc:
        logger.error("delete_publisher failed id=%s: %s", pub_id, exc)
        messages.error(request, '删除出版社时发生错误')
    return redirect('/manager/publisher_list')


# ============================  二、图书模块操作   ===============================
# 01获取所有图书信息
def book_list(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    # Admin panel only manages books belonging to the Duno360 Official Store —
    # vendor-created books are managed exclusively under Vendors (vendor_books).
    book_obj_list = models.Book.objects.filter(
        vendorbook__vendor__is_official=True
    ).select_related('publisher', 'category').distinct()
    # 2将数据渲染到页面上
    return render(request, 'book/book_list.html', {'book_obj_list': book_obj_list, "name": request.session["name"]})


# 02添加图书
def add_book(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")

    # Handle inline creation of new publisher
    if request.method == 'POST' and request.POST.get('action') == 'create_publisher':
        pub_name = request.POST.get('publisher_name', '').strip()
        pub_address = request.POST.get('publisher_address', '').strip()
        if not pub_name:
            return JsonResponse({'success': False, 'message': '出版社名称不能为空'})
        if models.Publisher.objects.filter(publisher_name__iexact=pub_name).exists():
            return JsonResponse({'success': False, 'message': f'出版社 "{pub_name}" 已存在，请使用不同的名称。'})
        # publisher_name/publisher_address are django-modeltranslation
        # fields — .objects.create() kwargs silently drop them, so
        # assign as plain attributes instead.
        pub = models.Publisher()
        pub.publisher_name = pub_name
        pub.publisher_address = pub_address
        pub.save()
        return JsonResponse({'success': True, 'id': pub.id, 'name': pub.publisher_name})

    # Handle inline creation of new author
    if request.method == 'POST' and request.POST.get('action') == 'create_author':
        author_name = request.POST.get('author_name', '').strip()
        if not author_name:
            return JsonResponse({'success': False, 'message': '作者名称不能为空'})
        if models.Author.objects.filter(name__iexact=author_name).exists():
            return JsonResponse({'success': False, 'message': f'作者 "{author_name}" 已存在，请使用不同的名称。'})
        author = models.Author()
        author.name = author_name
        author.save()
        return JsonResponse({'success': True, 'id': author.id, 'name': author.name})

    if request.method == 'POST':
        # 1获取表单提交过来的内容
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '')
        price = request.POST.get('price')
        inventory = request.POST.get('inventory')
        publisher_id = request.POST.get('publisher_id')
        category_id = request.POST.get('category_id') or None
        author_ids = request.POST.getlist('author_ids')
        cover_image = request.FILES.get('cover_image')

        # Book titles must be unique across the official store's own
        # catalog (vendor-created books are checked separately, scoped to
        # their own vendor — see marketplace/views.py's _reject_duplicate_title).
        if name and models.Book.objects.filter(name__iexact=name, vendorbook__vendor__is_official=True).exists():
            messages.error(request, f'图书 "{name}" 已存在，请使用不同的书名。')
            return redirect('/manager/add_book/')
        book_file = request.FILES.get('book_file')
        download_link = request.POST.get('download_link', '').strip()
        delivery_days_min, delivery_days_max = _parse_delivery_days_override(request.POST)

        # 2保存到数据库（insert）— sale_num is never set manually: it starts
        # at 0 and is only ever incremented/decremented automatically by
        # manager/inventory_service.py as orders are confirmed or returned.
        book = models.Book(
            price=price,
            inventory=inventory,
            sale_num=0,
            publisher_id=publisher_id,
            category_id=category_id,
            delivery_days_min=delivery_days_min,
            delivery_days_max=delivery_days_max,
        )
        # name/description are django-modeltranslation fields —
        # .objects.create() kwargs silently drop them, so assign as plain
        # attributes instead. Saved immediately so book.id exists below
        # (generate_cover_image needs a real PK).
        book.name = name
        book.description = description
        book.save()

        # Handle image upload or auto-generate cover
        if cover_image:
            book.cover_image = cover_image
        else:
            # Auto-generate a stylish cover
            try:
                from manager.cover_generator import generate_cover_image
                book.cover_image = generate_cover_image(book.name, book.id)
            except Exception:
                pass
        
        # Handle book file upload
        if book_file:
            book.book_file = book_file
        
        # Handle download link
        if download_link:
            book.download_link = download_link
            
        book.save()

        # Associate authors with the book
        if author_ids:
            authors = models.Author.objects.filter(id__in=author_ids)
            for author in authors:
                author.book.add(book)

        from manager.official_store import get_official_vendor
        official = get_official_vendor(create=True)
        if official:
            models.VendorBook.objects.get_or_create(
                vendor=official,
                book=book,
                defaults={'is_active': True},
            )

        messages.success(request, f'图书 "{book.name}" 添加成功！')
        # 3重定向到图书列表页面
        return redirect('/manager/book_list/')
    else:
        # 1获取所有的出版社（点击添加图书按钮时，得到所有出版社信息供用户选择）
        publisher_obj_list = models.Publisher.objects.all()
        author_obj_list = models.Author.objects.all()
        category_obj_list = models.BookCategory.objects.filter(is_active=True)
        # 2返回html页面（在页面中遍历出版社对象列表）
        return render(request, 'book/add_book.html', {
            'publisher_obj_list': publisher_obj_list,
            'author_obj_list': author_obj_list,
            'category_obj_list': category_obj_list,
            'name': request.session['name'],
        })


# 03修改图书信息
def edit_book(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    # 点击修改图书（获取要修改图书的原本信息）
    if request.method == 'GET':
        id = request.GET.get('id')
        book_obj = get_object_or_404(
            models.Book.objects.select_related('category'),
            id=id, vendorbook__vendor__is_official=True,
        )
        publisher_obj_list = models.Publisher.objects.all()
        author_obj_list = models.Author.objects.all()
        book_author_ids = list(book_obj.author_set.values_list('id', flat=True))
        category_obj_list = models.BookCategory.objects.filter(is_active=True)
        book_obj_list = models.Book.objects.filter(vendorbook__vendor__is_official=True).distinct()
        return render(request, "book/edit_book.html",
                      {"book_obj": book_obj, "book_obj_list": book_obj_list, "publisher_obj_list": publisher_obj_list,
                       "author_obj_list": author_obj_list, "book_author_ids": book_author_ids,
                       "category_obj_list": category_obj_list, "name": request.session["name"]})
    # 修改图书信息（POST表单）
    else:
        id = request.POST.get('id')
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        inventory = request.POST.get('inventory')
        price = request.POST.get('price')
        publisher_id = request.POST.get('publisher_id')
        category_id = request.POST.get('category_id') or None
        author_ids = request.POST.getlist('author_ids')
        cover_image = request.FILES.get('cover_image')
        book_file = request.FILES.get('book_file')
        download_link = request.POST.get('download_link', '').strip()

        # 获取要更新的图书对象（仅限官方直营图书，卖家图书通过卖家后台管理）
        book = get_object_or_404(models.Book, id=id, vendorbook__vendor__is_official=True)

        # 数据库中更新图书信息 — sale_num is never edited manually here: it's
        # only ever changed automatically by manager/inventory_service.py
        # as orders are confirmed or returned, so it's simply left untouched.
        book.name = name
        book.description = description
        book.inventory = inventory
        book.price = price
        book.publisher_id = publisher_id
        book.category_id = category_id
        book.delivery_days_min, book.delivery_days_max = _parse_delivery_days_override(request.POST)

        # Handle image upload
        if cover_image:
            book.cover_image = cover_image

        # Handle book file upload
        if book_file:
            book.book_file = book_file

        # Handle download link
        if download_link:
            book.download_link = download_link
        elif 'clear_download_link' in request.POST:
            book.download_link = None

        book.save()

        if author_ids:
            book.author_set.set(models.Author.objects.filter(id__in=author_ids))

        messages.success(request, f'图书 "{book.name}" 更新成功！')
        return redirect("/manager/book_list/")


# 04删除图书
def delete_book(request):
    if "name" not in request.session:
        return redirect("/manager/login")
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    book_id = request.POST.get('id')
    try:
        # Admin panel may only delete Duno360 Official Store books.
        book_qs = models.Book.objects.filter(id=book_id, vendorbook__vendor__is_official=True)
        book_name = book_qs.first().name if book_qs.exists() else None
        deleted, _ = book_qs.delete()
        if deleted:
            messages.success(request, f'图书 "{book_name}" 已删除')
        else:
            messages.error(request, '图书不存在或无法删除')
    except Exception as exc:
        logger.error("delete_book failed id=%s: %s", book_id, exc)
        messages.error(request, '删除图书时发生错误')
    return redirect('/manager/book_list/')


# ================================  三、作者操作模块  =============================
# 01作者列表
def author_list(request):
    if "name" not in request.session:
        return redirect("/manager/login")
    # Use prefetch_related to avoid N+1 queries when accessing author.book.all()
    author_obj_list = models.Author.objects.prefetch_related('book').all()
    ret_list = [
        {'author_obj': author, 'book_list': author.book.all()}
        for author in author_obj_list
    ]
    return render(request, 'author/author_list.html', {'ret_list': ret_list, "name": request.session["name"]})


# 02添加作家
def add_author(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    # 添加作家页面
    if request.method == 'GET':
        # 1获取所有的图书
        book_obj_list = models.Book.objects.all()
        # 2返回页面
        return render(request, 'author/add_author.html',
                      {'book_obj_list': book_obj_list, "name": request.session["name"]})
    # 数据库处理添加
    else:
        # 1.获取表单提交过来的数据
        name = (request.POST.get('name') or '').strip()
        book_ids = request.POST.getlist('books')
        max_len = models.Author._meta.get_field('name').max_length
        if not name:
            messages.error(request, '作者姓名不能为空')
            return redirect('/manager/add_author/')
        if len(name) > max_len:
            messages.error(request, f'作者姓名不能超过 {max_len} 个字符')
            return redirect('/manager/add_author/')
        # 2 保存数据库
        # name is a django-modeltranslation field — .objects.create() kwargs
        # silently drop it, so assign as a plain attribute instead.
        author_obj = models.Author()
        author_obj.name = name
        author_obj.save()
        author_obj.book.set(book_ids)  # 设置关系
        messages.success(request, f'作者 "{author_obj.name}" 添加成功！')
        # 3 重定向到列表页面
        return redirect('/manager/author_list/')


# 03修改作者信息
def edit_author(request):
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    
    # 跳转修改界面
    if request.method == 'GET':
        id = request.GET.get('id')
        try:
            author_obj = models.Author.objects.get(id=id)
            book_obj_list = models.Book.objects.all()
            # Get currently associated books
            current_books = author_obj.book.all()
            return render(request, 'author/edit_author.html',
                          {
                              'author_obj': author_obj, 
                              'book_obj_list': book_obj_list,
                              'current_books': current_books,
                              "name": request.session["name"]
                          })
        except models.Author.DoesNotExist:
            return redirect('/manager/author_list/')
    
    # 提交修改表单
    else:
        id = request.POST.get('id')
        name = (request.POST.get('name') or '').strip()
        book_ids = request.POST.getlist('books')  # Get list of selected book IDs

        if not name:
            messages.error(request, '作者姓名不能为空')
            return redirect(f'/manager/edit_author/?id={id}')
        max_len = models.Author._meta.get_field('name').max_length
        if len(name) > max_len:
            messages.error(request, f'作者姓名不能超过 {max_len} 个字符')
            return redirect(f'/manager/edit_author/?id={id}')

        try:
            # 找到作者对象
            author_obj = models.Author.objects.filter(id=id).first()
            if author_obj:
                author_obj.name = name
                # Clear existing relationships and set new ones
                author_obj.book.set(book_ids)  # This handles the many-to-many relationship
                author_obj.save()
                messages.success(request, f'作者 "{author_obj.name}" 更新成功！')
            else:
                messages.error(request, '作者不存在')
            return redirect('/manager/author_list/')
        except Exception as e:
            # Handle any errors gracefully
            messages.error(request, '更新作者时发生错误')
            return redirect(f'/manager/edit_author/?id={id}')

# 04 删除作者
def delete_author(request):
    if "name" not in request.session:
        return redirect("/manager/login")
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    author_id = request.POST.get('id')
    try:
        author_qs = models.Author.objects.filter(id=author_id)
        author_name = author_qs.first().name if author_qs.exists() else None
        deleted, _ = author_qs.delete()
        if deleted:
            messages.success(request, f'作者 "{author_name}" 已删除')
        else:
            messages.error(request, '作者不存在')
    except Exception as exc:
        logger.error("delete_author failed id=%s: %s", author_id, exc)
        messages.error(request, '删除作者时发生错误')
    return redirect('/manager/author_list/')


# ====================   PUBLIC USER INTERFACE  ===========================


def _build_trending_feed(max_items=24):
    """Trending carousel: top sellers (30d), new listings (30d), most wishlisted."""
    since = timezone.now() - timedelta(days=30)
    cache_key = f'home:trending:v1:{max_items}:{since.date().isoformat()}'
    cached = cache.get(cache_key)
    if cached:
        return cached

    tag_rank = {'hot': 3, 'fav': 2, 'new': 1}
    items = {}

    def upsert(item_type, pk, name, price, image_url, detail_url, tag, score):
        key = (item_type, int(pk))
        entry = {
            'item_type': item_type,
            'name': name,
            'price': str(price) if price is not None else '0',
            'image_url': image_url or '/static/img/default_cover.png',
            'detail_url': detail_url,
            'tag': tag,
            'rank_score': int(score or 0),
        }
        existing = items.get(key)
        if not existing:
            items[key] = entry
            return
        if tag_rank.get(tag, 0) > tag_rank.get(existing['tag'], 0):
            existing['tag'] = tag
        existing['rank_score'] = max(existing['rank_score'], int(score or 0))

    paid_book_statuses = ['paid', 'confirmed', 'processing', 'shipped', 'delivered']
    paid_mkt_statuses = ['paid', 'processing', 'shipped', 'delivered']
    per_bucket = max(8, max_items // 3)

    try:
        book_sales = models.OrderItem.objects.filter(
            order__created_at__gte=since,
            order__payment_status='completed',
            order__status__in=paid_book_statuses,
            book__is_active=True,
        ).values('book_id').annotate(sold=Sum('quantity')).order_by('-sold')[:per_bucket]
        book_ids = [row['book_id'] for row in book_sales]
        books = {b.id: b for b in models.Book.objects.filter(id__in=book_ids)}
        for row in book_sales:
            b = books.get(row['book_id'])
            if b:
                upsert('book', b.id, b.name, b.price, b.get_cover_url(), f'/manager/public/books/{b.id}/', 'hot', row['sold'])

        marketplace_sales = MarketplaceOrderItem.objects.filter(
            order__created_at__gte=since,
            order__payment_status='completed',
            order__status__in=paid_mkt_statuses,
        ).values('item_type', 'item_id').annotate(sold=Sum('quantity')).order_by('-sold')[:per_bucket * 3]
        product_ids = [row['item_id'] for row in marketplace_sales if row['item_type'] == 'product']
        course_ids = [row['item_id'] for row in marketplace_sales if row['item_type'] == 'course']
        supermarket_ids = [row['item_id'] for row in marketplace_sales if row['item_type'] == 'supermarket']
        products = {p.id: p for p in Product.objects.filter(id__in=product_ids, is_active=True)}
        courses = {c.id: c for c in Course.objects.filter(id__in=course_ids, is_active=True)}
        supermarket_items = {s.id: s for s in SupermarketItem.objects.filter(id__in=supermarket_ids, is_active=True)}
        for row in marketplace_sales:
            if row['item_type'] == 'product':
                item = products.get(row['item_id'])
                if item:
                    upsert('product', item.id, item.name, item.price, item.get_image_url(), f'/marketplace/products/{item.slug}/', 'hot', row['sold'])
            elif row['item_type'] == 'course':
                item = courses.get(row['item_id'])
                if item:
                    upsert('course', item.id, item.title, item.price, item.get_image_url(), f'/marketplace/courses/{item.slug}/', 'hot', row['sold'])
            elif row['item_type'] == 'supermarket':
                item = supermarket_items.get(row['item_id'])
                if item:
                    upsert('supermarket', item.id, item.name, item.price, item.get_image_url(), f'/marketplace/supermarket/{item.slug}/', 'hot', row['sold'])

        recent_book_ids = models.VendorBook.objects.filter(
            created_at__gte=since, is_active=True, book__is_active=True,
        ).order_by('-created_at').values_list('book_id', flat=True)[:per_bucket]
        for b in models.Book.objects.filter(id__in=list(recent_book_ids)):
            upsert('book', b.id, b.name, b.price, b.get_cover_url(), f'/manager/public/books/{b.id}/', 'new', 1)
        for p in Product.objects.filter(is_active=True, created_at__gte=since).order_by('-created_at')[:per_bucket]:
            upsert('product', p.id, p.name, p.price, p.get_image_url(), f'/marketplace/products/{p.slug}/', 'new', 1)
        for c in Course.objects.filter(is_active=True, created_at__gte=since).order_by('-created_at')[:per_bucket]:
            upsert('course', c.id, c.title, c.price, c.get_image_url(), f'/marketplace/courses/{c.slug}/', 'new', 1)
        for s in SupermarketItem.objects.filter(is_active=True, created_at__gte=since).order_by('-created_at')[:per_bucket]:
            upsert('supermarket', s.id, s.name, s.price, s.get_image_url(), f'/marketplace/supermarket/{s.slug}/', 'new', 1)

        book_favs = models.Wishlist.objects.filter(
            item_type='book', book__is_active=True,
        ).values('book_id').annotate(fav_count=Count('id')).order_by('-fav_count')[:per_bucket]
        fav_book_ids = [row['book_id'] for row in book_favs if row['book_id']]
        fav_books = {b.id: b for b in models.Book.objects.filter(id__in=fav_book_ids)}
        for row in book_favs:
            b = fav_books.get(row['book_id'])
            if b:
                upsert('book', b.id, b.name, b.price, b.get_cover_url(), f'/manager/public/books/{b.id}/', 'fav', row['fav_count'])

        mkt_favs = models.Wishlist.objects.exclude(item_type='book').filter(
            item_id__isnull=False,
        ).values('item_type', 'item_id').annotate(fav_count=Count('id')).order_by('-fav_count')[:per_bucket * 3]
        fav_product_ids = [row['item_id'] for row in mkt_favs if row['item_type'] == 'product']
        fav_course_ids = [row['item_id'] for row in mkt_favs if row['item_type'] == 'course']
        fav_super_ids = [row['item_id'] for row in mkt_favs if row['item_type'] == 'supermarket']
        fav_products = {p.id: p for p in Product.objects.filter(id__in=fav_product_ids, is_active=True)}
        fav_courses = {c.id: c for c in Course.objects.filter(id__in=fav_course_ids, is_active=True)}
        fav_super = {s.id: s for s in SupermarketItem.objects.filter(id__in=fav_super_ids, is_active=True)}
        for row in mkt_favs:
            if row['item_type'] == 'product':
                item = fav_products.get(row['item_id'])
                if item:
                    upsert('product', item.id, item.name, item.price, item.get_image_url(), f'/marketplace/products/{item.slug}/', 'fav', row['fav_count'])
            elif row['item_type'] == 'course':
                item = fav_courses.get(row['item_id'])
                if item:
                    upsert('course', item.id, item.title, item.price, item.get_image_url(), f'/marketplace/courses/{item.slug}/', 'fav', row['fav_count'])
            elif row['item_type'] == 'supermarket':
                item = fav_super.get(row['item_id'])
                if item:
                    upsert('supermarket', item.id, item.name, item.price, item.get_image_url(), f'/marketplace/supermarket/{item.slug}/', 'fav', row['fav_count'])
    except Exception as exc:
        logger.warning(f'Failed to build trending feed: {exc}')

    trending = list(items.values())
    trending.sort(key=lambda x: (tag_rank.get(x['tag'], 0), x['rank_score']), reverse=True)
    trending = trending[:max_items]
    cache.set(cache_key, trending, 300)
    return trending


def public_home(request):
    """Public homepage with platform statistics and featured content."""
    book_count = models.Book.objects.filter(is_active=True).count()
    author_count = models.Author.objects.count()
    publisher_count = models.Publisher.objects.count()
    user_count = models.SiteUser.objects.filter(is_active=True).count()

    featured_products = []
    featured_courses = []
    flash_sales = []
    flash_sale_end = None
    search_categories = []
    try:
        from marketplace.models import Category, Product, Course, SupermarketItem, FlashSale
        from django.utils import timezone as tz
        product_count = Product.objects.filter(is_active=True).count()
        course_count = Course.objects.filter(is_active=True).count()
        supermarket_product_count = SupermarketItem.objects.filter(is_active=True).count()
        vendor_count = models.Vendor.objects.filter(is_active=True).count()
        total_orders_count = models.Order.objects.count()
        featured_products = list(
            _annotate_product_delivered(
                Product.objects.filter(is_active=True).select_related('category')
            ).order_by('-sold_delivered', '-sales_count')[:6]
        )
        featured_courses = list(
            _annotate_course_delivered(Course.objects.filter(is_active=True)).order_by(
                '-sold_delivered', '-enrollment_count'
            )[:6]
        )
        featured_supermarket = list(
            SupermarketItem.objects.filter(is_active=True).select_related('category').order_by('-sales_count', '-id')[:6]
        )
        recent_products = list(
            Product.objects.filter(is_active=True).select_related('category').order_by('-id')[:4]
        )
        recent_courses = list(
            Course.objects.filter(is_active=True).order_by('-id')[:4]
        )
        recent_supermarket = list(
            SupermarketItem.objects.filter(is_active=True).select_related('category').order_by('-id')[:4]
        )
        now = tz.now()
        flash_sales = list(FlashSale.objects.filter(
            is_active=True, start_time__lte=now, end_time__gte=now
        ).select_related('product', 'course', 'supermarket_item').order_by('end_time')[:10])
        if flash_sales:
            flash_sale_end = flash_sales[0].end_time
        search_categories = list(
            Category.objects.filter(is_active=True, section='products', parent__isnull=True)
            .order_by('display_order', 'name')
            .values('name', 'slug')
        )
    except Exception:
        product_count = 0
        course_count = 0
        supermarket_product_count = 0
        vendor_count = 0
        total_orders_count = models.Order.objects.count()
        featured_supermarket = []
        recent_products = []
        recent_courses = []
        recent_supermarket = []

    featured_books = list(
        _annotate_book_delivered(
            models.Book.objects.filter(is_active=True).select_related('publisher', 'category')
        ).order_by('-sold_delivered', '-sale_num')[:6]
    )
    recent_books = list(
        _annotate_book_delivered(
            models.Book.objects.filter(is_active=True).select_related('publisher', 'category')
        ).order_by('-id')[:8]
    )
    book_categories = list(models.BookCategory.objects.filter(is_active=True, parent__isnull=True)[:12])
    latest_blogs = list(models.BlogPost.objects.filter(status='published').order_by('-created_at')[:3])
    trending_items = _build_trending_feed(24)

    # Deals for the once-per-session "welcome deals" popup — reuses the
    # flash_sales and trending_items querysets already built above (no new
    # DB queries), normalized into one shape the popup template can loop
    # over regardless of source. Flash sales first (real discounts), then
    # top-selling ("hot") trending items fill any remaining slots. Rendered
    # as a horizontally scrollable list, so more items is fine.
    popup_deals = []
    for sale in flash_sales[:6]:
        popup_deals.append({
            'name': sale.get_item_name(),
            'image_url': sale.get_item_image_url(),
            'detail_url': sale.get_item_url(),
            'price': sale.flash_price,
            'original_price': sale.get_original_price(),
            'discount_percent': sale.get_discount_percent(),
            'is_flash': True,
        })
    if len(popup_deals) < 10:
        for item in trending_items:
            if item.get('tag') != 'hot' or len(popup_deals) >= 10:
                continue
            popup_deals.append({
                'name': item['name'],
                'image_url': item['image_url'],
                'detail_url': item['detail_url'],
                'price': item['price'],
                'original_price': None,
                'discount_percent': 0,
                'is_flash': False,
            })

    total_catalog_items = book_count + product_count + course_count + supermarket_product_count + author_count + publisher_count + vendor_count + user_count

    # Personalized first page of the mobile home feed (pages 2+ come from the
    # /api/feed/ endpoint, which is personalized the same way). Falls back to
    # the popularity-ordered featured_* loops for anonymous / signal-less users.
    home_feed_items = []
    home_feed_personalized = False
    _uid = request.session.get('site_user_id')
    if _uid:
        try:
            from manager import recommendations as _reco
            if not request.session.session_key:
                request.session.save()
            _skey = request.session.session_key or ''
            _items, _hm, _has_sig = _reco.recommend(_uid, _skey, domain='mixed', page=1, per_page=12)
            if _has_sig and _items:
                home_feed_items = _items
                home_feed_personalized = True
        except Exception:
            pass

    ctx = {
        'book_count': book_count,
        'author_count': author_count,
        'publisher_count': publisher_count,
        'product_count': product_count,
        'course_count': course_count,
        'supermarket_product_count': supermarket_product_count,
        'total_catalog_items': total_catalog_items,
        'vendor_count': vendor_count,
        'user_count': user_count,
        'total_orders_count': total_orders_count,
        'featured_books': featured_books,
        'featured_products': featured_products,
        'featured_courses': featured_courses,
        'featured_supermarket': featured_supermarket,
        'recent_books': recent_books,
        'recent_products': recent_products,
        'recent_courses': recent_courses,
        'recent_supermarket': recent_supermarket,
        'book_categories': book_categories,
        'flash_sales': flash_sales,
        'flash_sale_end': flash_sale_end,
        'latest_blogs': latest_blogs,
        'trending_items': trending_items,
        'popup_deals': popup_deals,
        'search_categories': search_categories,
        'home_feed_items': home_feed_items,
        'home_feed_personalized': home_feed_personalized,
    }

    return render(request, 'public/home.html', ctx)


# ── Direct-message guards (shared by buyer & vendor send paths) ──────────
MESSAGE_MAX_LENGTH = 4000
_MSG_RATE_WINDOW_SECONDS = 60
_MSG_RATE_MAX_PER_WINDOW = 20
MESSAGE_RECALL_WINDOW_SECONDS = 5 * 60
MESSAGE_ATTACHMENT_MAX_BYTES = {
    'image': 10 * 1024 * 1024,
    'video': 50 * 1024 * 1024,
    'file': 20 * 1024 * 1024,
}
MESSAGE_ATTACHMENT_EXTENSIONS = {
    'image': {'.jpg', '.jpeg', '.png', '.gif', '.webp'},
    'video': {'.mp4', '.webm', '.mov', '.m4v'},
    'file': {'.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.txt', '.zip', '.rar'},
}


def _message_guard(request, message_text, allow_empty=False):
    """Validate a direct-message payload. Returns an error string or None.

    - caps content length (TextField is unbounded at the DB level)
    - simple per-session sliding-window rate limit against spam floods
    """
    if not message_text and not allow_empty:
        return '消息不能为空'
    if len(message_text) > MESSAGE_MAX_LENGTH:
        return f'消息过长（最多{MESSAGE_MAX_LENGTH}字符）'
    now_ts = timezone.now().timestamp()
    window = [t for t in request.session.get('msg_send_times', [])
              if now_ts - t < _MSG_RATE_WINDOW_SECONDS]
    if len(window) >= _MSG_RATE_MAX_PER_WINDOW:
        return '发送过于频繁，请稍后再试'
    window.append(now_ts)
    request.session['msg_send_times'] = window
    return None


def _validate_message_attachment(f):
    """Validate an uploaded chat attachment (image, video, or common document) within size limits.

    Returns (attachment_type, error_message). attachment_type is '' on error.
    """
    import os as _os
    ext = _os.path.splitext(f.name)[1].lower()
    for kind, exts in MESSAGE_ATTACHMENT_EXTENSIONS.items():
        if ext in exts:
            if f.size > MESSAGE_ATTACHMENT_MAX_BYTES[kind]:
                limit_mb = MESSAGE_ATTACHMENT_MAX_BYTES[kind] // (1024 * 1024)
                return '', f'File too large (max {limit_mb}MB)'
            return kind, None
    return '', 'Unsupported file type. Allowed: images, videos, PDF, Office documents, TXT, ZIP/RAR'


def _is_vendor_blocked(buyer_id, vendor_id):
    if not buyer_id or not vendor_id:
        return None
    return models.VendorBlock.objects.filter(buyer_id=buyer_id, vendor_id=vendor_id).first()


# ── WhatsApp-Business-style automatic replies (welcome / away / keywords) ──

def _send_auto_reply(conversation, vendor, content):
    """Send a canned message on the vendor's behalf, tagged is_auto_reply for the UI."""
    if not content:
        return None
    msg = models.DirectMessage.objects.create(
        conversation=conversation,
        sender_type='vendor',
        sender_name=vendor.company_name if vendor else '',
        content=content,
        is_auto_reply=True,
    )
    conversation.save(update_fields=['updated_at'])
    return msg


def _maybe_send_welcome_reply(conversation, vendor):
    """Fires once, right after a buyer's first-ever message in a new conversation."""
    if not vendor:
        return
    if conversation.direct_messages.filter(sender_type='buyer').count() != 1:
        return
    settings_obj = models.AutoReplySettings.objects.filter(vendor=vendor).first()
    if settings_obj and settings_obj.welcome_enabled and settings_obj.welcome_message.strip():
        _send_auto_reply(conversation, vendor, settings_obj.welcome_message.strip())


def _maybe_send_keyword_reply(conversation, vendor, message_text):
    """Scans the buyer's message for configured trigger keywords and auto-replies
    with the first match (case-insensitive substring match)."""
    if not vendor or not message_text:
        return
    text_lower = message_text.lower()
    rules = models.AutoReplyKeyword.objects.filter(vendor=vendor, is_active=True)
    for rule in rules:
        if any(kw in text_lower for kw in rule.keyword_list()):
            _send_auto_reply(conversation, vendor, rule.reply_message)
            return


def _maybe_send_away_reply(conversation, vendor, last_msg):
    """Opportunistically checked whenever the buyer polls a conversation: if the
    last visible message is from the buyer and has gone unanswered longer than
    the configured delay, auto-send the away message once per waiting message."""
    if not vendor or not last_msg or last_msg.sender_type != 'buyer':
        return
    settings_obj = models.AutoReplySettings.objects.filter(vendor=vendor).first()
    if not settings_obj or not settings_obj.away_enabled or not settings_obj.away_message.strip():
        return
    if conversation.away_reply_sent_at and conversation.away_reply_sent_at >= last_msg.created_at:
        return
    delay = timezone.timedelta(minutes=settings_obj.away_delay_minutes)
    if timezone.now() - last_msg.created_at < delay:
        return
    _send_auto_reply(conversation, vendor, settings_obj.away_message.strip())
    conversation.away_reply_sent_at = timezone.now()
    conversation.save(update_fields=['away_reply_sent_at'])


@ensure_csrf_cookie
def public_messages(request):
    """Messages page: buyer/vendor direct discussions plus support chatbot entries."""
    user_id = request.session.get('site_user_id')
    if not user_id:
        return redirect(f"/manager/public/user/login/?next={request.get_full_path()}")
    user = None
    user_messages = []
    direct_conversations = []
    if user_id:
        try:
            user = models.SiteUser.objects.get(pk=user_id)
            user_messages = list(models.ContactMessage.objects.filter(
                email=user.email
            ).order_by('-created_at')[:20])
            direct_conversations = list(models.Conversation.objects.filter(
                buyer=user
            ).select_related('vendor').prefetch_related('direct_messages')[:30])
        except models.SiteUser.DoesNotExist:
            return redirect(f"/manager/public/user/login/?next={request.get_full_path()}")
    # Also include messages by session key
    if not request.session.session_key:
        request.session.save()
    session_key = request.session.session_key or ''
    if session_key:
        session_msgs = list(models.ContactMessage.objects.filter(
            session_key=session_key
        ).order_by('-created_at')[:10])
        seen_ids = {m.id for m in user_messages}
        for m in session_msgs:
            if m.id not in seen_ids:
                user_messages.append(m)
        user_messages.sort(key=lambda m: m.created_at, reverse=True)
    unread_count = len([m for m in user_messages if not m.is_read])
    context = {
        'site_user': user,
        'user_messages': user_messages,
        'direct_conversations': direct_conversations,
        'unread_count': unread_count,
    }
    return render(request, 'public/messages.html', context)


def start_conversation(request):
    """Create/open a buyer-seller conversation from product pages."""
    if not request.session.get('site_user_id'):
        return redirect(f"/manager/public/user/login/?next={request.get_full_path()}")
    user = get_object_or_404(models.SiteUser, pk=request.session['site_user_id'])
    item_type = request.GET.get('item_type', 'support')
    item_id = request.GET.get('item_id')
    vendor = None
    subject = request.GET.get('subject', '')
    try:
        if request.GET.get('vendor_id'):
            vendor = models.Vendor.objects.filter(pk=request.GET.get('vendor_id'), is_active=True).first()
        elif item_type == 'book':
            book = models.Book.objects.filter(pk=item_id).first()
            if book:
                subject = subject or book.name
                vb = models.VendorBook.objects.filter(book=book, is_active=True).select_related('vendor').first()
                vendor = vb.vendor if vb else None
        elif item_type == 'product':
            from marketplace.models import Product
            item = Product.objects.select_related('vendor').filter(pk=item_id).first()
            vendor = item.vendor if item else None
            subject = subject or (item.name if item else '')
        elif item_type == 'course':
            from marketplace.models import Course
            item = Course.objects.select_related('vendor').filter(pk=item_id).first()
            vendor = item.vendor if item else None
            subject = subject or (item.title if item else '')
        elif item_type == 'supermarket':
            from marketplace.models import SupermarketItem
            item = SupermarketItem.objects.select_related('vendor').filter(pk=item_id).first()
            vendor = item.vendor if item else None
            subject = subject or (item.name if item else '')
    except Exception:
        vendor = None

    if not vendor:
        referer = request.META.get('HTTP_REFERER', '/manager/public/')
        separator = '&' if '?' in referer else '?'
        return redirect(f"{referer}{separator}open_chatbot=1")

    if _is_vendor_blocked(user.id, vendor.id):
        referer = request.META.get('HTTP_REFERER', '/manager/public/')
        separator = '&' if '?' in referer else '?'
        return redirect(f"{referer}{separator}vendor_blocked=1")

    conversation, created = models.Conversation.objects.get_or_create(
        buyer=user,
        vendor=vendor,
        conversation_type='buyer_seller',
        ref_item_type=item_type,
        ref_item_id=item_id or None,
        defaults={'subject': subject[:200] or 'Support'}
    )
    auto_link = '1' if (created or not conversation.direct_messages.exists()) else '0'
    return redirect(f"/manager/public/messages/?convo={conversation.id}&auto_link={auto_link}")


def _conversation_ref_item_payload(conversation):
    """Build a normalized product preview payload for chat UIs."""
    if not conversation or not conversation.ref_item_type or not conversation.ref_item_id:
        return None
    item_type = conversation.ref_item_type
    item_id = conversation.ref_item_id
    try:
        if item_type == 'book':
            book = models.Book.objects.filter(pk=item_id, is_active=True).first()
            if not book:
                return None
            return {
                'type': 'book',
                'name': book.name,
                'price': str(book.price),
                'image': book.get_cover_url(),
                'url': f'/manager/public/books/{book.id}/',
            }
        if item_type == 'product':
            item = Product.objects.filter(pk=item_id, is_active=True).first()
            if not item:
                return None
            return {
                'type': 'product',
                'name': item.name,
                'price': str(item.price),
                'image': item.get_image_url(),
                'url': f'/marketplace/products/{item.slug}/',
            }
        if item_type == 'course':
            item = Course.objects.filter(pk=item_id, is_active=True).first()
            if not item:
                return None
            return {
                'type': 'course',
                'name': item.title,
                'price': str(item.price),
                'image': item.get_image_url(),
                'url': f'/marketplace/courses/{item.slug}/',
            }
        if item_type == 'supermarket':
            item = SupermarketItem.objects.filter(pk=item_id, is_active=True).first()
            if not item:
                return None
            return {
                'type': 'supermarket',
                'name': item.name,
                'price': str(item.price),
                'image': item.get_image_url(),
                'url': f'/marketplace/supermarket/{item.slug}/',
            }
    except Exception:
        return None
    return None


def public_send_message(request):
    """AJAX: send a support or direct message from the messages page."""
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '无效请求'})
    user_id = request.session.get('site_user_id')
    message_text = request.POST.get('message', request.POST.get('content', '')).strip()
    attachment_file = request.FILES.get('attachment')
    guard_error = _message_guard(request, message_text, allow_empty=bool(attachment_file))
    if guard_error:
        return JsonResponse({'success': False, 'message': guard_error})
    conversation_id = request.POST.get('conversation_id')
    if conversation_id:
        if not user_id:
            return JsonResponse({'success': False, 'message': '请先登录'}, status=401)
        conversation = get_object_or_404(models.Conversation, pk=conversation_id, buyer_id=user_id)
        if conversation.vendor_id and _is_vendor_blocked(user_id, conversation.vendor_id):
            return JsonResponse({'success': False, 'message': 'You have blocked this vendor. Unblock to send messages.'})
        attachment_type = ''
        if attachment_file:
            attachment_type, attach_error = _validate_message_attachment(attachment_file)
            if attach_error:
                return JsonResponse({'success': False, 'message': attach_error})
        reply_to = None
        reply_to_id = request.POST.get('reply_to_id')
        if reply_to_id:
            reply_to = conversation.direct_messages.filter(pk=reply_to_id, is_recalled=False).first()
        sender_name = request.session.get('site_user_name', '')
        msg = models.DirectMessage.objects.create(
            conversation=conversation,
            sender_type='buyer',
            sender_name=sender_name,
            content=message_text,
            attachment=attachment_file or None,
            attachment_type=attachment_type,
            reply_to=reply_to,
        )
        conversation.save(update_fields=['updated_at'])
        if conversation.vendor_id:
            create_vendor_notification(
                conversation.vendor_id, 'new_message',
                f'New message from {sender_name or "Customer"}',
                (message_text[:100] if message_text else f'[{attachment_type or "attachment"}]'),
                icon='fas fa-comment-dots',
                color='#3b82f6',
                link=f'/manager/vendor/messages/?conversation={conversation.id}',
                related_id=conversation.id,
            )
            _maybe_send_welcome_reply(conversation, conversation.vendor)
            if message_text:
                _maybe_send_keyword_reply(conversation, conversation.vendor, message_text)
        return JsonResponse({'success': True, 'message': '消息已发送', 'id': msg.id})
    try:
        user = models.SiteUser.objects.get(pk=user_id) if user_id else None
        name = user.name if user else request.POST.get('name', 'Guest')
        email = user.email if user else request.POST.get('email', '')
        if not email:
            return JsonResponse({'success': False, 'message': '请先登录或提供邮箱'})
        if not request.session.session_key:
            request.session.save()
        msg = models.ContactMessage.objects.create(
            name=name,
            email=email,
            subject='来自消息页面',
            message=message_text,
            session_key=request.session.session_key or '',
        )
        return JsonResponse({'success': True, 'message': '消息已发送', 'id': msg.id,
                             'created_at': msg.created_at.strftime('%H:%M')})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


def api_buyer_unread_count(request):
    """API: total unread message count across all of the buyer's conversations,
    used to badge the navbar Messages button on every page."""
    user_id = request.session.get('site_user_id')
    if not user_id:
        return JsonResponse({'unread_count': 0})
    count = models.DirectMessage.objects.filter(
        conversation__buyer_id=user_id, conversation__buyer_hidden=False,
        is_read=False, is_recalled=False, deleted_for_buyer=False,
    ).exclude(sender_type='buyer').count()
    return JsonResponse({'unread_count': count})


def api_conversations(request):
    """API: list all conversations for the logged-in user."""
    user_id = request.session.get('site_user_id')
    if not user_id:
        return JsonResponse({'conversations': [], 'error': 'not_logged_in'})
    convos = models.Conversation.objects.filter(
        buyer_id=user_id, buyer_hidden=False,
    ).select_related('vendor').prefetch_related('direct_messages').order_by('-updated_at')
    blocked_vendor_ids = set(
        models.VendorBlock.objects.filter(buyer_id=user_id).values_list('vendor_id', flat=True)
    )
    result = []
    for c in convos:
        visible_msgs = c.direct_messages.filter(is_recalled=False, deleted_for_buyer=False)
        last_msg = visible_msgs.order_by('-created_at').first()
        if c.vendor_id:
            _maybe_send_away_reply(c, c.vendor, last_msg)
        unread = visible_msgs.filter(is_read=False).exclude(sender_type='buyer').count()
        vendor_name = c.vendor.company_name if c.vendor else (c.subject or 'Support')
        vendor_avatar = vendor_name[0].upper() if vendor_name else 'S'
        ref = _conversation_ref_item_payload(c)
        last_preview = last_msg.content[:80] if (last_msg and last_msg.content) else (
            '📷 Photo' if (last_msg and last_msg.attachment_type == 'image') else
            '🎥 Video' if (last_msg and last_msg.attachment_type == 'video') else
            '📎 File' if (last_msg and last_msg.attachment_type == 'file') else ''
        )
        result.append({
            'id': c.id,
            'type': c.conversation_type,
            'vendor_name': vendor_name,
            'vendor_avatar': vendor_avatar,
            'vendor_id': c.vendor_id,
            'subject': c.subject or '',
            'last_message': last_preview,
            'last_sender_type': last_msg.sender_type if last_msg else '',
            'last_created_at': last_msg.created_at.isoformat() if last_msg else '',
            'unread_count': unread,
            'unread': unread,  # keep backward compat
            'is_closed': c.is_closed,
            'is_blocked': c.vendor_id in blocked_vendor_ids,
            'ref_item': ref,
            # flatten ref_item fields for quick access in UI
            'ref_item_name': ref.get('name', '') if ref else '',
            'ref_item_price': ref.get('price', '') if ref else '',
            'ref_item_image': ref.get('image', '') if ref else '',
            'ref_item_url': ref.get('url', '') if ref else '',
        })
    return JsonResponse({'conversations': result})


def _message_brief(m):
    """Short preview of a message, used for quote/reply snippets."""
    if not m or m.is_recalled:
        return None
    if m.content:
        text = m.content[:80]
    elif m.attachment_type == 'image':
        text = '📷 Photo'
    elif m.attachment_type == 'video':
        text = '🎥 Video'
    else:
        text = ''
    return {'id': m.id, 'sender_type': m.sender_type, 'sender_name': m.sender_name, 'content': text}


def api_conversation_messages(request, conversation_id):
    """API: get messages for a specific conversation."""
    user_id = request.session.get('site_user_id')
    if not user_id:
        return JsonResponse({'messages': [], 'error': 'not_logged_in'})
    convo = get_object_or_404(models.Conversation, pk=conversation_id, buyer_id=user_id)
    if convo.vendor_id:
        last_visible = convo.direct_messages.filter(is_recalled=False, deleted_for_buyer=False).order_by('-created_at').first()
        _maybe_send_away_reply(convo, convo.vendor, last_visible)
    # Mark vendor messages as read (whole thread)
    convo.direct_messages.filter(is_read=False).exclude(sender_type='buyer').update(is_read=True)
    block = _is_vendor_blocked(user_id, convo.vendor_id) if convo.vendor_id else None
    all_msgs = convo.direct_messages.exclude(deleted_for_buyer=True).select_related('reply_to').order_by('created_at')
    if block:
        # Hide vendor messages sent after the buyer blocked them.
        all_msgs = all_msgs.exclude(sender_type='vendor', created_at__gt=block.created_at)
    total = all_msgs.count()
    # Pagination: page 1 = newest window; higher pages = older windows (no overlap)
    try:
        page = max(1, int(request.GET.get('page', 1)))
    except (TypeError, ValueError):
        page = 1
    page_size = 50
    end_idx = total - (page - 1) * page_size
    if end_idx <= 0:
        msgs = []
        has_more = False
    else:
        start_idx = max(0, end_idx - page_size)
        msgs = list(all_msgs[start_idx:end_idx])
        has_more = start_idx > 0
    now = timezone.now()
    result = []
    for m in msgs:
        can_recall = (
            m.sender_type == 'buyer' and not m.is_recalled
            and (now - m.created_at).total_seconds() < MESSAGE_RECALL_WINDOW_SECONDS
        )
        result.append({
            'id': m.id,
            'sender_type': m.sender_type,
            'sender_name': m.sender_name,
            'content': '' if m.is_recalled else m.content,
            'attachment_url': (m.attachment.url if (m.attachment and not m.is_recalled) else ''),
            'attachment_type': ('' if m.is_recalled else m.attachment_type),
            'reply_to': _message_brief(m.reply_to) if m.reply_to_id else None,
            'is_recalled': m.is_recalled,
            'can_recall': can_recall,
            'can_delete_for_me': m.sender_type == 'buyer' and not m.is_recalled,
            'created_at': m.created_at.isoformat(),
            'is_read': m.is_read,
            'is_auto_reply': m.is_auto_reply,
        })
    vendor_name = convo.vendor.company_name if convo.vendor else (convo.subject or 'Support')
    ref = _conversation_ref_item_payload(convo)
    return JsonResponse({
        'messages': result,
        'has_more': has_more,
        'conversation': {
            'id': convo.id,
            'vendor_name': vendor_name,
            'vendor_id': convo.vendor_id,
            'subject': convo.subject or '',
            'type': convo.conversation_type,
            'is_closed': convo.is_closed,
            'is_blocked': bool(block),
            'ref_item': ref,
            'ref_item_name': ref.get('name', '') if ref else '',
            'ref_item_price': ref.get('price', '') if ref else '',
            'ref_item_image': ref.get('image', '') if ref else '',
            'ref_item_url': ref.get('url', '') if ref else '',
        }
    })


def api_conversation_delete(request, conversation_id):
    """API: buyer removes a conversation from their own list (vendor keeps it)."""
    user_id = request.session.get('site_user_id')
    if not user_id:
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    convo = get_object_or_404(models.Conversation, pk=conversation_id, buyer_id=user_id)
    convo.buyer_hidden = True
    convo.save(update_fields=['buyer_hidden'])
    return JsonResponse({'success': True})


def api_conversation_block_vendor(request, conversation_id):
    """API: buyer blacklists the vendor behind a conversation."""
    user_id = request.session.get('site_user_id')
    if not user_id:
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    convo = get_object_or_404(models.Conversation, pk=conversation_id, buyer_id=user_id)
    if not convo.vendor_id:
        return JsonResponse({'success': False, 'message': 'No vendor to block'})
    models.VendorBlock.objects.get_or_create(buyer_id=user_id, vendor_id=convo.vendor_id)
    return JsonResponse({'success': True})


def api_conversation_unblock_vendor(request, conversation_id):
    """API: buyer removes a vendor from their blacklist."""
    user_id = request.session.get('site_user_id')
    if not user_id:
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    convo = get_object_or_404(models.Conversation, pk=conversation_id, buyer_id=user_id)
    if convo.vendor_id:
        models.VendorBlock.objects.filter(buyer_id=user_id, vendor_id=convo.vendor_id).delete()
    return JsonResponse({'success': True})


def api_message_recall(request, message_id):
    """API: buyer recalls (unsends) their own message within the recall window."""
    user_id = request.session.get('site_user_id')
    if not user_id:
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    msg = get_object_or_404(
        models.DirectMessage, pk=message_id, conversation__buyer_id=user_id, sender_type='buyer',
    )
    if msg.is_recalled:
        return JsonResponse({'success': True})
    age = (timezone.now() - msg.created_at).total_seconds()
    if age >= MESSAGE_RECALL_WINDOW_SECONDS:
        return JsonResponse({'success': False, 'message': 'Recall window has expired (5 minutes)'})
    msg.is_recalled = True
    msg.save(update_fields=['is_recalled'])
    return JsonResponse({'success': True})


def api_message_delete_for_me(request, message_id):
    """API: buyer hides their own sent message from their own view only."""
    user_id = request.session.get('site_user_id')
    if not user_id:
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    msg = get_object_or_404(
        models.DirectMessage, pk=message_id, conversation__buyer_id=user_id, sender_type='buyer',
    )
    msg.deleted_for_buyer = True
    msg.save(update_fields=['deleted_for_buyer'])
    return JsonResponse({'success': True})


def api_mark_conversation_read(request, conversation_id):
    """API: mark all messages in a conversation as read."""
    user_id = request.session.get('site_user_id')
    if not user_id:
        return JsonResponse({'success': False, 'error': 'not_logged_in'})
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    convo = get_object_or_404(models.Conversation, pk=conversation_id, buyer_id=user_id)
    convo.direct_messages.filter(is_read=False).exclude(sender_type='buyer').update(is_read=True)
    return JsonResponse({'success': True})


def api_vendor_conversations(request):
    """API: list conversations for the logged-in vendor."""
    vendor_id = request.session.get('vendor_id')
    if not vendor_id:
        return JsonResponse({'conversations': [], 'error': 'not_logged_in'})
    convos = models.Conversation.objects.filter(
        vendor_id=vendor_id, vendor_hidden=False,
    ).select_related('buyer').prefetch_related('direct_messages').order_by('-updated_at')
    result = []
    for c in convos:
        visible_msgs = c.direct_messages.filter(is_recalled=False, deleted_for_vendor=False)
        last_msg = visible_msgs.order_by('-created_at').first()
        unread = visible_msgs.filter(is_read=False).exclude(sender_type='vendor').count()
        buyer_name = c.buyer.name if c.buyer else 'Unknown'
        last_preview = last_msg.content[:80] if (last_msg and last_msg.content) else (
            '📷 Photo' if (last_msg and last_msg.attachment_type == 'image') else
            '🎥 Video' if (last_msg and last_msg.attachment_type == 'video') else
            '📎 File' if (last_msg and last_msg.attachment_type == 'file') else ''
        )
        result.append({
            'id': c.id,
            'buyer_name': buyer_name,
            'buyer_avatar': buyer_name[0].upper() if buyer_name else 'U',
            'buyer_id': c.buyer_id,
            'subject': c.subject,
            'last_message': last_preview,
            'last_created_at': last_msg.created_at.isoformat() if last_msg else '',
            'unread': unread,
            'is_closed': c.is_closed,
        })
    return JsonResponse({'conversations': result})


def api_vendor_conversation_messages(request, conversation_id):
    """API: get messages for a vendor conversation."""
    vendor_id = request.session.get('vendor_id')
    if not vendor_id:
        return JsonResponse({'messages': [], 'error': 'not_logged_in'})
    convo = get_object_or_404(models.Conversation, pk=conversation_id, vendor_id=vendor_id)
    msgs = convo.direct_messages.exclude(deleted_for_vendor=True).select_related('reply_to').order_by('created_at')
    msgs.filter(is_read=False).exclude(sender_type='vendor').update(is_read=True)
    now = timezone.now()
    result = []
    for m in msgs:
        can_recall = (
            m.sender_type == 'vendor' and not m.is_recalled
            and (now - m.created_at).total_seconds() < MESSAGE_RECALL_WINDOW_SECONDS
        )
        result.append({
            'id': m.id,
            'sender_type': m.sender_type,
            'sender_name': m.sender_name,
            'content': '' if m.is_recalled else m.content,
            'attachment_url': (m.attachment.url if (m.attachment and not m.is_recalled) else ''),
            'attachment_type': ('' if m.is_recalled else m.attachment_type),
            'attachment_name': ('' if (m.is_recalled or not m.attachment) else m.attachment.name.rsplit('/', 1)[-1]),
            'reply_to': _message_brief(m.reply_to) if m.reply_to_id else None,
            'is_recalled': m.is_recalled,
            'can_recall': can_recall,
            'can_delete_for_me': m.sender_type == 'vendor' and not m.is_recalled,
            'created_at': m.created_at.isoformat(),
            'is_read': m.is_read,
            'is_auto_reply': m.is_auto_reply,
        })
    buyer_name = convo.buyer.name if convo.buyer else 'Unknown'
    return JsonResponse({
        'messages': result,
        'conversation': {
            'id': convo.id,
            'buyer_name': buyer_name,
            'buyer_id': convo.buyer_id,
            'subject': convo.subject,
            'type': convo.conversation_type,
            'is_closed': convo.is_closed,
        }
    })


def api_vendor_conversation_delete(request, conversation_id):
    """API: vendor removes a conversation from their own list (buyer keeps it)."""
    vendor_id = request.session.get('vendor_id')
    if not vendor_id:
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    convo = get_object_or_404(models.Conversation, pk=conversation_id, vendor_id=vendor_id)
    convo.vendor_hidden = True
    convo.save(update_fields=['vendor_hidden'])
    return JsonResponse({'success': True})


def api_vendor_mark_conversation_read(request, conversation_id):
    """API: vendor marks all buyer messages in a conversation as read."""
    vendor_id = request.session.get('vendor_id')
    if not vendor_id:
        return JsonResponse({'success': False, 'error': 'not_logged_in'})
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    convo = get_object_or_404(models.Conversation, pk=conversation_id, vendor_id=vendor_id)
    convo.direct_messages.filter(is_read=False).exclude(sender_type='vendor').update(is_read=True)
    return JsonResponse({'success': True})


def api_vendor_message_recall(request, message_id):
    """API: vendor recalls (unsends) their own message within the recall window."""
    vendor_id = request.session.get('vendor_id')
    if not vendor_id:
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    msg = get_object_or_404(
        models.DirectMessage, pk=message_id, conversation__vendor_id=vendor_id, sender_type='vendor',
    )
    if msg.is_recalled:
        return JsonResponse({'success': True})
    age = (timezone.now() - msg.created_at).total_seconds()
    if age >= MESSAGE_RECALL_WINDOW_SECONDS:
        return JsonResponse({'success': False, 'message': 'Recall window has expired (5 minutes)'})
    msg.is_recalled = True
    msg.save(update_fields=['is_recalled'])
    return JsonResponse({'success': True})


def api_vendor_message_delete_for_me(request, message_id):
    """API: vendor hides their own sent message from their own view only."""
    vendor_id = request.session.get('vendor_id')
    if not vendor_id:
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    msg = get_object_or_404(
        models.DirectMessage, pk=message_id, conversation__vendor_id=vendor_id, sender_type='vendor',
    )
    msg.deleted_for_vendor = True
    msg.save(update_fields=['deleted_for_vendor'])
    return JsonResponse({'success': True})


# ── Auto-reply settings & keyword rules (shared by vendor + admin/official store) ──

def _auto_reply_settings_payload(vendor):
    settings_obj, _ = models.AutoReplySettings.objects.get_or_create(vendor=vendor)
    return {
        'welcome_enabled': settings_obj.welcome_enabled,
        'welcome_message': settings_obj.welcome_message,
        'away_enabled': settings_obj.away_enabled,
        'away_message': settings_obj.away_message,
        'away_delay_minutes': settings_obj.away_delay_minutes,
    }


def _auto_reply_settings_get_or_update(request, vendor):
    if request.method == 'POST':
        settings_obj, _ = models.AutoReplySettings.objects.get_or_create(vendor=vendor)
        settings_obj.welcome_enabled = request.POST.get('welcome_enabled') == '1'
        settings_obj.welcome_message = (request.POST.get('welcome_message') or '').strip()[:2000]
        settings_obj.away_enabled = request.POST.get('away_enabled') == '1'
        settings_obj.away_message = (request.POST.get('away_message') or '').strip()[:2000]
        try:
            delay = int(request.POST.get('away_delay_minutes', 5))
        except (TypeError, ValueError):
            delay = 5
        settings_obj.away_delay_minutes = max(1, min(delay, 1440))
        settings_obj.save()
    return JsonResponse({'success': True, 'settings': _auto_reply_settings_payload(vendor)})


def _auto_reply_keyword_payload(k):
    return {'id': k.id, 'keywords': k.keywords, 'reply_message': k.reply_message, 'is_active': k.is_active}


def _auto_reply_keywords_list(vendor):
    rules = models.AutoReplyKeyword.objects.filter(vendor=vendor).order_by('-created_at')
    return JsonResponse({'keywords': [_auto_reply_keyword_payload(k) for k in rules]})


def _auto_reply_keyword_create(request, vendor):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    keywords = (request.POST.get('keywords') or '').strip()[:255]
    reply_message = (request.POST.get('reply_message') or '').strip()[:2000]
    if not keywords or not reply_message:
        return JsonResponse({'success': False, 'message': 'Keywords and reply message are required'})
    k = models.AutoReplyKeyword.objects.create(vendor=vendor, keywords=keywords, reply_message=reply_message)
    return JsonResponse({'success': True, 'keyword': _auto_reply_keyword_payload(k)})


def _auto_reply_keyword_update(request, vendor, keyword_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    k = get_object_or_404(models.AutoReplyKeyword, pk=keyword_id, vendor=vendor)
    if 'keywords' in request.POST:
        k.keywords = (request.POST.get('keywords') or '').strip()[:255]
    if 'reply_message' in request.POST:
        k.reply_message = (request.POST.get('reply_message') or '').strip()[:2000]
    if 'is_active' in request.POST:
        k.is_active = request.POST.get('is_active') == '1'
    k.save()
    return JsonResponse({'success': True, 'keyword': _auto_reply_keyword_payload(k)})


def _auto_reply_keyword_delete(vendor, keyword_id):
    k = get_object_or_404(models.AutoReplyKeyword, pk=keyword_id, vendor=vendor)
    k.delete()
    return JsonResponse({'success': True})


def api_vendor_auto_reply_settings(request):
    """API: vendor gets/updates their welcome + away auto-reply settings."""
    vendor_id = request.session.get('vendor_id')
    if not vendor_id:
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    vendor = get_object_or_404(models.Vendor, pk=vendor_id)
    return _auto_reply_settings_get_or_update(request, vendor)


def api_vendor_auto_reply_keywords(request):
    """API: vendor lists/creates keyword-triggered auto-reply rules."""
    vendor_id = request.session.get('vendor_id')
    if not vendor_id:
        return JsonResponse({'keywords': [], 'error': 'not_logged_in'})
    vendor = get_object_or_404(models.Vendor, pk=vendor_id)
    if request.method == 'POST':
        return _auto_reply_keyword_create(request, vendor)
    return _auto_reply_keywords_list(vendor)


def api_vendor_auto_reply_keyword_update(request, keyword_id):
    vendor_id = request.session.get('vendor_id')
    if not vendor_id:
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    vendor = get_object_or_404(models.Vendor, pk=vendor_id)
    return _auto_reply_keyword_update(request, vendor, keyword_id)


def api_vendor_auto_reply_keyword_delete(request, keyword_id):
    vendor_id = request.session.get('vendor_id')
    if not vendor_id:
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    vendor = get_object_or_404(models.Vendor, pk=vendor_id)
    return _auto_reply_keyword_delete(vendor, keyword_id)


def api_admin_store_auto_reply_settings(request):
    """API: admin gets/updates the official store's welcome + away auto-reply settings."""
    if not _admin_authed(request):
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    from manager.official_store import get_official_vendor
    vendor = get_official_vendor(create=True)
    return _auto_reply_settings_get_or_update(request, vendor)


def api_admin_store_auto_reply_keywords(request):
    """API: admin lists/creates keyword-triggered auto-reply rules for the official store."""
    if not _admin_authed(request):
        return JsonResponse({'keywords': [], 'error': 'not_logged_in'})
    from manager.official_store import get_official_vendor
    vendor = get_official_vendor(create=True)
    if request.method == 'POST':
        return _auto_reply_keyword_create(request, vendor)
    return _auto_reply_keywords_list(vendor)


def api_admin_store_auto_reply_keyword_update(request, keyword_id):
    if not _admin_authed(request):
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    from manager.official_store import get_official_vendor
    vendor = get_official_vendor(create=True)
    return _auto_reply_keyword_update(request, vendor, keyword_id)


def api_admin_store_auto_reply_keyword_delete(request, keyword_id):
    if not _admin_authed(request):
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    from manager.official_store import get_official_vendor
    vendor = get_official_vendor(create=True)
    return _auto_reply_keyword_delete(vendor, keyword_id)


def _vendor_listing_payload(item, item_type):
    """Normalize a vendor's own listing into a compact picker payload."""
    if item_type == 'book':
        return {
            'type': 'book', 'id': item.id, 'name': item.name,
            'price': str(item.price), 'image': item.get_cover_url(),
            'url': f'/manager/public/books/{item.id}/',
        }
    if item_type == 'product':
        return {
            'type': 'product', 'id': item.id, 'name': item.name,
            'price': str(item.price), 'image': item.get_image_url(),
            'url': f'/marketplace/products/{item.slug}/',
        }
    if item_type == 'course':
        return {
            'type': 'course', 'id': item.id, 'name': item.title,
            'price': str(item.price), 'image': item.get_image_url(),
            'url': f'/marketplace/courses/{item.slug}/',
        }
    if item_type == 'supermarket':
        return {
            'type': 'supermarket', 'id': item.id, 'name': item.name,
            'price': str(item.price), 'image': item.get_image_url(),
            'url': f'/marketplace/supermarket/{item.slug}/',
        }
    return None


def api_vendor_my_listings(request):
    """API: search the vendor's own catalog, for the 'send a listing link' picker."""
    vendor_id = request.session.get('vendor_id')
    if not vendor_id:
        return JsonResponse({'listings': [], 'error': 'not_logged_in'})
    q = (request.GET.get('q') or '').strip()
    results = []
    vendor_books = models.VendorBook.objects.filter(vendor_id=vendor_id, is_active=True).select_related('book')
    if q:
        vendor_books = vendor_books.filter(book__name__icontains=q)
    for vb in vendor_books[:15]:
        results.append(_vendor_listing_payload(vb.book, 'book'))
    from marketplace.models import Product, Course, SupermarketItem
    products = Product.objects.filter(vendor_id=vendor_id, is_active=True)
    if q:
        products = products.filter(name__icontains=q)
    for p in products[:15]:
        results.append(_vendor_listing_payload(p, 'product'))
    courses = Course.objects.filter(vendor_id=vendor_id, is_active=True)
    if q:
        courses = courses.filter(title__icontains=q)
    for c in courses[:15]:
        results.append(_vendor_listing_payload(c, 'course'))
    items = SupermarketItem.objects.filter(vendor_id=vendor_id, is_active=True)
    if q:
        items = items.filter(name__icontains=q)
    for it in items[:15]:
        results.append(_vendor_listing_payload(it, 'supermarket'))
    return JsonResponse({'listings': [r for r in results if r][:40]})


def vendor_create_conversation(request):
    """Vendor starts a new conversation with a buyer by email."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'})
    vendor_id = request.session.get('vendor_id')
    if not vendor_id:
        return JsonResponse({'success': False, 'message': 'Not logged in'}, status=401)
    vendor = get_object_or_404(models.Vendor, pk=vendor_id)
    buyer_email = request.POST.get('buyer_email', '').strip()
    subject = request.POST.get('subject', '').strip()[:200] or 'Message from vendor'
    message_text = request.POST.get('message', '').strip()
    if not buyer_email:
        return JsonResponse({'success': False, 'message': 'Email and message are required'})
    guard_error = _message_guard(request, message_text)
    if guard_error:
        return JsonResponse({'success': False, 'message': guard_error})
    buyer = models.SiteUser.objects.filter(email__iexact=buyer_email).first()
    if not buyer:
        return JsonResponse({'success': False, 'message': f'No customer found with email: {buyer_email}'})
    convo, created = models.Conversation.objects.get_or_create(
        buyer=buyer,
        vendor=vendor,
        defaults={
            'conversation_type': 'buyer_seller',
            'subject': subject,
        }
    )
    if not created and subject:
        convo.subject = subject
        convo.save(update_fields=['subject', 'updated_at'])
    models.DirectMessage.objects.create(
        conversation=convo,
        sender_type='vendor',
        sender_name=vendor.company_name,
        content=message_text,
    )
    convo.save(update_fields=['updated_at'])
    return JsonResponse({'success': True, 'conversation_id': convo.id})


def vendor_send_message(request):
    """Vendor sends a message in a conversation."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'})
    vendor_id = request.session.get('vendor_id')
    if not vendor_id:
        return JsonResponse({'success': False, 'message': 'Please log in first'}, status=401)
    conversation_id = request.POST.get('conversation_id')
    content = request.POST.get('content', '').strip()
    attachment_file = request.FILES.get('attachment')
    guard_error = _message_guard(request, content, allow_empty=bool(attachment_file))
    if guard_error:
        return JsonResponse({'success': False, 'message': guard_error})
    if not conversation_id:
        return JsonResponse({'success': False, 'message': 'No conversation specified'})
    try:
        convo = models.Conversation.objects.get(pk=conversation_id, vendor_id=vendor_id)
    except models.Conversation.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Conversation not found'}, status=404)
    try:
        vendor = models.Vendor.objects.get(pk=vendor_id)
    except models.Vendor.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Vendor not found'}, status=404)
    attachment_type = ''
    if attachment_file:
        attachment_type, attach_error = _validate_message_attachment(attachment_file)
        if attach_error:
            return JsonResponse({'success': False, 'message': attach_error})
    reply_to = None
    reply_to_id = request.POST.get('reply_to_id')
    if reply_to_id:
        reply_to = convo.direct_messages.filter(pk=reply_to_id, is_recalled=False).first()
    msg = models.DirectMessage.objects.create(
        conversation=convo,
        sender_type='vendor',
        sender_name=vendor.company_name,
        content=content,
        attachment=attachment_file or None,
        attachment_type=attachment_type,
        reply_to=reply_to,
    )
    convo.save(update_fields=['updated_at'])
    return JsonResponse({'success': True, 'id': msg.id})


@ensure_csrf_cookie
def vendor_messages_page(request):
    """Vendor messaging dashboard page."""
    vendor_id = request.session.get('vendor_id')
    if not vendor_id:
        return redirect('manager:vendor_login')
    vendor = get_object_or_404(models.Vendor, pk=vendor_id)
    return render(request, 'public/vendor_messages.html', {'vendor': vendor})


# ── Admin: messages addressed to the Duno360 Official Store ──────────────
# The official store is a platform-owned Vendor with no logged-in owner
# (see manager/official_store.py), so its inbox is only reachable by admin
# staff. These views mirror the vendor messaging endpoints above but are
# gated on admin session auth instead of a vendor session.

def _admin_authed(request):
    return 'name' in request.session


@ensure_csrf_cookie
def admin_store_messages_page(request):
    """Admin dashboard for conversations sent to the Duno360 Official Store."""
    if not _admin_authed(request):
        return redirect('/manager/login/')
    from manager.official_store import get_official_vendor
    vendor = get_official_vendor(create=True)
    return render(request, 'admin/store_messages.html', {
        'vendor': vendor,
        'name': request.session.get('name', ''),
    })


def api_admin_store_conversations(request):
    """API: list conversations addressed to the official store, for admin staff."""
    if not _admin_authed(request):
        return JsonResponse({'conversations': [], 'error': 'not_logged_in'})
    from manager.official_store import get_official_vendor
    vendor = get_official_vendor(create=True)
    convos = models.Conversation.objects.filter(
        vendor=vendor, vendor_hidden=False,
    ).select_related('buyer').prefetch_related('direct_messages').order_by('-updated_at')
    result = []
    for c in convos:
        visible_msgs = c.direct_messages.filter(is_recalled=False, deleted_for_vendor=False)
        last_msg = visible_msgs.order_by('-created_at').first()
        unread = visible_msgs.filter(is_read=False).exclude(sender_type='vendor').count()
        buyer_name = c.buyer.name if c.buyer else 'Unknown'
        last_preview = last_msg.content[:80] if (last_msg and last_msg.content) else (
            '📷 Photo' if (last_msg and last_msg.attachment_type == 'image') else
            '🎥 Video' if (last_msg and last_msg.attachment_type == 'video') else
            '📎 File' if (last_msg and last_msg.attachment_type == 'file') else ''
        )
        result.append({
            'id': c.id,
            'buyer_name': buyer_name,
            'buyer_avatar': buyer_name[0].upper() if buyer_name else 'U',
            'buyer_id': c.buyer_id,
            'buyer_email': c.buyer.email if c.buyer else '',
            'subject': c.subject,
            'last_message': last_preview,
            'last_created_at': last_msg.created_at.isoformat() if last_msg else '',
            'unread': unread,
            'is_closed': c.is_closed,
        })
    return JsonResponse({'conversations': result})


def api_admin_store_conversation_messages(request, conversation_id):
    """API: messages in a conversation addressed to the official store."""
    if not _admin_authed(request):
        return JsonResponse({'messages': [], 'error': 'not_logged_in'})
    from manager.official_store import get_official_vendor
    vendor = get_official_vendor(create=True)
    convo = get_object_or_404(models.Conversation, pk=conversation_id, vendor=vendor)
    msgs = convo.direct_messages.exclude(deleted_for_vendor=True).select_related('reply_to').order_by('created_at')
    msgs.filter(is_read=False).exclude(sender_type='vendor').update(is_read=True)
    now = timezone.now()
    result = []
    for m in msgs:
        can_recall = (
            m.sender_type == 'vendor' and not m.is_recalled
            and (now - m.created_at).total_seconds() < MESSAGE_RECALL_WINDOW_SECONDS
        )
        result.append({
            'id': m.id,
            'sender_type': m.sender_type,
            'sender_name': m.sender_name,
            'content': '' if m.is_recalled else m.content,
            'attachment_url': (m.attachment.url if (m.attachment and not m.is_recalled) else ''),
            'attachment_type': ('' if m.is_recalled else m.attachment_type),
            'attachment_name': ('' if (m.is_recalled or not m.attachment) else m.attachment.name.rsplit('/', 1)[-1]),
            'reply_to': _message_brief(m.reply_to) if m.reply_to_id else None,
            'is_recalled': m.is_recalled,
            'can_recall': can_recall,
            'can_delete_for_me': m.sender_type == 'vendor' and not m.is_recalled,
            'created_at': m.created_at.isoformat(),
            'is_read': m.is_read,
            'is_auto_reply': m.is_auto_reply,
        })
    buyer_name = convo.buyer.name if convo.buyer else 'Unknown'
    return JsonResponse({
        'messages': result,
        'conversation': {
            'id': convo.id,
            'vendor_id': vendor.id,
            'buyer_name': buyer_name,
            'buyer_id': convo.buyer_id,
            'buyer_email': convo.buyer.email if convo.buyer else '',
            'subject': convo.subject,
            'type': convo.conversation_type,
            'is_closed': convo.is_closed,
        }
    })


def api_admin_store_conversation_delete(request, conversation_id):
    """API: admin removes a conversation from the store's own list (buyer keeps it)."""
    if not _admin_authed(request):
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    from manager.official_store import get_official_vendor
    vendor = get_official_vendor(create=True)
    convo = get_object_or_404(models.Conversation, pk=conversation_id, vendor=vendor)
    convo.vendor_hidden = True
    convo.save(update_fields=['vendor_hidden'])
    return JsonResponse({'success': True})


def api_admin_store_mark_read(request, conversation_id):
    """API: admin marks all buyer messages in a store conversation as read."""
    if not _admin_authed(request):
        return JsonResponse({'success': False, 'error': 'not_logged_in'})
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    from manager.official_store import get_official_vendor
    vendor = get_official_vendor(create=True)
    convo = get_object_or_404(models.Conversation, pk=conversation_id, vendor=vendor)
    convo.direct_messages.filter(is_read=False).exclude(sender_type='vendor').update(is_read=True)
    return JsonResponse({'success': True})


def api_admin_store_message_recall(request, message_id):
    """API: admin recalls (unsends) a store reply within the recall window."""
    if not _admin_authed(request):
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    from manager.official_store import get_official_vendor
    vendor = get_official_vendor(create=True)
    msg = get_object_or_404(
        models.DirectMessage, pk=message_id, conversation__vendor=vendor, sender_type='vendor',
    )
    if msg.is_recalled:
        return JsonResponse({'success': True})
    age = (timezone.now() - msg.created_at).total_seconds()
    if age >= MESSAGE_RECALL_WINDOW_SECONDS:
        return JsonResponse({'success': False, 'message': 'Recall window has expired (5 minutes)'})
    msg.is_recalled = True
    msg.save(update_fields=['is_recalled'])
    return JsonResponse({'success': True})


def api_admin_store_message_delete_for_me(request, message_id):
    """API: admin hides a store reply from the admin panel's own view only."""
    if not _admin_authed(request):
        return JsonResponse({'success': False, 'error': 'not_logged_in'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'method_not_allowed'})
    from manager.official_store import get_official_vendor
    vendor = get_official_vendor(create=True)
    msg = get_object_or_404(
        models.DirectMessage, pk=message_id, conversation__vendor=vendor, sender_type='vendor',
    )
    msg.deleted_for_vendor = True
    msg.save(update_fields=['deleted_for_vendor'])
    return JsonResponse({'success': True})


def api_admin_store_listings(request):
    """API: search the official store's own catalog, for the 'send a listing link' picker."""
    if not _admin_authed(request):
        return JsonResponse({'listings': [], 'error': 'not_logged_in'})
    from manager.official_store import get_official_vendor
    vendor = get_official_vendor(create=True)
    q = (request.GET.get('q') or '').strip()
    results = []
    vendor_books = models.VendorBook.objects.filter(vendor=vendor, is_active=True).select_related('book')
    if q:
        vendor_books = vendor_books.filter(book__name__icontains=q)
    for vb in vendor_books[:15]:
        results.append(_vendor_listing_payload(vb.book, 'book'))
    from marketplace.models import Product, Course, SupermarketItem
    products = Product.objects.filter(vendor=vendor, is_active=True)
    if q:
        products = products.filter(name__icontains=q)
    for p in products[:15]:
        results.append(_vendor_listing_payload(p, 'product'))
    courses = Course.objects.filter(vendor=vendor, is_active=True)
    if q:
        courses = courses.filter(title__icontains=q)
    for c in courses[:15]:
        results.append(_vendor_listing_payload(c, 'course'))
    items = SupermarketItem.objects.filter(vendor=vendor, is_active=True)
    if q:
        items = items.filter(name__icontains=q)
    for it in items[:15]:
        results.append(_vendor_listing_payload(it, 'supermarket'))
    return JsonResponse({'listings': [r for r in results if r][:40]})


def admin_store_send_message(request):
    """Admin staff replies to a buyer on behalf of the official store."""
    if not _admin_authed(request):
        return JsonResponse({'success': False, 'message': 'Please log in first'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'})
    from manager.official_store import get_official_vendor
    vendor = get_official_vendor(create=True)
    conversation_id = request.POST.get('conversation_id')
    content = request.POST.get('content', '').strip()
    attachment_file = request.FILES.get('attachment')
    guard_error = _message_guard(request, content, allow_empty=bool(attachment_file))
    if guard_error:
        return JsonResponse({'success': False, 'message': guard_error})
    if not conversation_id:
        return JsonResponse({'success': False, 'message': 'No conversation specified'})
    convo = get_object_or_404(models.Conversation, pk=conversation_id, vendor=vendor)
    attachment_type = ''
    if attachment_file:
        attachment_type, attach_error = _validate_message_attachment(attachment_file)
        if attach_error:
            return JsonResponse({'success': False, 'message': attach_error})
    reply_to = None
    reply_to_id = request.POST.get('reply_to_id')
    if reply_to_id:
        reply_to = convo.direct_messages.filter(pk=reply_to_id, is_recalled=False).first()
    msg = models.DirectMessage.objects.create(
        conversation=convo,
        sender_type='vendor',
        sender_name=vendor.company_name,
        content=content,
        attachment=attachment_file or None,
        attachment_type=attachment_type,
        reply_to=reply_to,
    )
    convo.save(update_fields=['updated_at'])
    return JsonResponse({
        'success': True,
        'id': msg.id,
    })


def _build_user_wishlist_items(user):
    """Return normalized wishlist entries for profile/wishlist pages."""
    wishlist_entries = models.Wishlist.objects.filter(user=user).select_related('book').order_by('-created_at')
    wishlist_items = []
    for wish in wishlist_entries:
        item = wish.get_item()
        item_type = wish.item_type
        item_url = '#'
        if item_type == 'book' and wish.book_id:
            item_url = f'/manager/public/books/{wish.book_id}/'
        elif item and getattr(item, 'slug', None):
            if item_type == 'product':
                item_url = f'/marketplace/products/{item.slug}/'
            elif item_type == 'course':
                item_url = f'/marketplace/courses/{item.slug}/'
            elif item_type == 'supermarket':
                item_url = f'/marketplace/supermarket/{item.slug}/'
        wishlist_items.append({
            'type': item_type,
            'type_label': wish.get_item_type_display(),
            'item_id': wish.book_id if item_type == 'book' else wish.item_id,
            'name': wish.get_item_name(),
            'price': wish.get_item_price(),
            'image_url': wish.get_item_image_url(),
            'url': item_url,
            'created_at': wish.created_at,
        })
    return wishlist_items


def public_my_profile(request):
    """PWA 'My' tab — shows login prompt if not logged in, else full profile."""
    from django.db.models import Sum, Count
    user_id = request.session.get('site_user_id')
    if not user_id:
        return render(request, 'public/my_profile.html', {'site_user': None})
    try:
        user = models.SiteUser.objects.get(pk=user_id)
    except models.SiteUser.DoesNotExist:
        return render(request, 'public/my_profile.html', {'site_user': None})

    # Loyalty
    try:
        loyalty = models.LoyaltyPoints.objects.get(user=user)
    except Exception:
        loyalty = None

    # Orders
    orders = models.Order.objects.filter(customer_email=user.email).order_by('-created_at')[:5]

    # Followed shops
    followed_shops = models.UserFollowedShop.objects.filter(user=user).select_related('publisher')[:20]
    followed_vendors = models.UserFollowedVendor.objects.filter(user=user).select_related('vendor')[:20]

    # Wishlist: include books + marketplace items for the "Me" page.
    from django.db.models import Q
    wishlist_books = models.Book.objects.filter(
        wishlist__user=user
    ).select_related('publisher')[:20]
    wishlist_items = _build_user_wishlist_items(user)

    # Auto-follow official Duno360 store
    from manager.official_store import get_official_vendor
    admin_vendor = get_official_vendor(create=True)
    if admin_vendor and not models.UserFollowedVendor.objects.filter(user=user, vendor=admin_vendor).exists():
        models.UserFollowedVendor.objects.create(user=user, vendor=admin_vendor)
        followed_vendors = models.UserFollowedVendor.objects.filter(user=user).select_related('vendor')[:20]

    # Mixed feed: books + marketplace items
    feed_books = list(models.Book.objects.filter(is_active=True).select_related('publisher').order_by('-id')[:10])
    feed_items = []
    try:
        from marketplace.models import Product, Course
        feed_products = list(Product.objects.filter(is_active=True).order_by('-created_at')[:6])
        feed_courses = list(Course.objects.filter(is_active=True).order_by('-created_at')[:4])
        for p in feed_products:
            feed_items.append({
                'type': 'product', 'name': p.name, 'price': p.price,
                'image_url': p.get_image_url(), 'url': f'/marketplace/products/{p.slug}/',
            })
        for c in feed_courses:
            feed_items.append({
                'type': 'course', 'name': c.title, 'price': c.price,
                'image_url': c.get_image_url(), 'url': f'/marketplace/courses/{c.slug}/',
            })
    except Exception:
        pass

    # All publishers and vendors for discover
    all_publishers = models.Publisher.objects.all()[:20]
    all_vendors = models.Vendor.objects.filter(is_active=True)[:20]

    from . import views_review

    # Wallet
    wallet_balance = Decimal('0.00')
    try:
        wallet = models.UserWallet.objects.get(user=user)
        wallet_balance = wallet.balance
    except Exception:
        pass

    context = {
        'site_user': user,
        'loyalty': loyalty,
        'orders': orders,
        'followed_shops': followed_shops,
        'followed_vendors': followed_vendors,
        'following_count': len(followed_shops) + len(followed_vendors),
        'wishlist_books': wishlist_books,
        'wishlist_items': wishlist_items,
        'wishlist_count': len(wishlist_items),
        'feed_books': feed_books,
        'feed_items': feed_items,
        'all_publishers': all_publishers,
        'all_vendors': all_vendors,
        'pending_review_items': views_review.collect_pending_reviews_for_user(user),
        'wallet_balance': wallet_balance,
    }
    return render(request, 'public/my_profile.html', context)


@ensure_csrf_cookie
def public_wishlist(request):
    """Dedicated wishlist page for mobile/desktop."""
    user_id = request.session.get('site_user_id')
    if not user_id:
        return redirect('manager:user_login')
    user = get_object_or_404(models.SiteUser, pk=user_id)
    wishlist_items = _build_user_wishlist_items(user)
    return render(request, 'public/wishlist.html', {
        'site_user': user,
        'wishlist_items': wishlist_items,
        'wishlist_count': len(wishlist_items),
    })


def follow_publisher(request, publisher_id):
    """AJAX: toggle follow/unfollow a publisher."""
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '无效请求'})
    user_id = request.session.get('site_user_id')
    if not user_id:
        return JsonResponse({'success': False, 'message': '请先登录', 'login_required': True}, status=401)
    try:
        user = models.SiteUser.objects.get(pk=user_id)
        publisher = models.Publisher.objects.get(pk=publisher_id)
        follow_obj, created = models.UserFollowedShop.objects.get_or_create(
            user=user, publisher=publisher
        )
        if not created:
            follow_obj.delete()
            return JsonResponse({'success': True, 'following': False, 'message': f'已取消关注 {publisher.publisher_name}'})
        return JsonResponse({'success': True, 'following': True, 'message': f'已关注 {publisher.publisher_name}'})
    except models.Publisher.DoesNotExist:
        return JsonResponse({'success': False, 'message': '店铺不存在'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


def follow_vendor(request, vendor_id):
    """AJAX: toggle follow/unfollow a marketplace vendor."""
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '无效请求'})
    user_id = request.session.get('site_user_id')
    if not user_id:
        return JsonResponse({'success': False, 'message': '请先登录', 'login_required': True}, status=401)
    try:
        user = models.SiteUser.objects.get(pk=user_id)
        vendor = models.Vendor.objects.get(pk=vendor_id, is_active=True)
        follow_obj, created = models.UserFollowedVendor.objects.get_or_create(
            user=user, vendor=vendor
        )
        if not created:
            follow_obj.delete()
            return JsonResponse({'success': True, 'following': False, 'message': f'已取消关注 {vendor.company_name}'})
        create_vendor_notification(
            vendor.id,
            'new_follower',
            f'{user.name} follows your shop',
            f'{user.name} ({user.email}) started following {vendor.company_name}.',
            icon='fas fa-user-plus',
            color='#8b5cf6',
            link='/manager/vendor/dashboard/',
            related_id=follow_obj.id,
        )
        return JsonResponse({'success': True, 'following': True, 'message': f'已关注 {vendor.company_name}'})
    except models.Vendor.DoesNotExist:
        return JsonResponse({'success': False, 'message': '卖家不存在'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


def public_books(request):
    """Public book listing with search, category filter and pagination."""
    search_query = request.GET.get('search', request.GET.get('q', '')).strip()
    sort_by = request.GET.get('sort', 'name')
    if sort_by == 'new':
        sort_by = 'newest'
    category_slug = request.GET.get('category', '').strip()
    min_price = request.GET.get('min_price', '').strip()
    max_price = request.GET.get('max_price', '').strip()
    
    books = _annotate_book_delivered(
        models.Book.objects.filter(is_active=True).select_related('publisher', 'category')
    )
    
    if search_query:
        books = books.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(publisher__publisher_name__icontains=search_query) |
            Q(author__name__icontains=search_query)
        ).distinct()
    if category_slug:
        books = books.filter(category__slug=category_slug)
    if min_price:
        try:
            books = books.filter(price__gte=min_price)
        except Exception:
            pass
    if max_price:
        try:
            books = books.filter(price__lte=max_price)
        except Exception:
            pass
    
    if sort_by == 'price_low':
        books = books.order_by('price')
    elif sort_by == 'price_high':
        books = books.order_by('-price')
    elif sort_by in ('popular', '-sale_num'):
        books = books.order_by('-sold_delivered', '-sale_num')
    elif sort_by == 'newest':
        books = books.order_by('-id')
    else:
        books = books.order_by('name')
    
    paginator = Paginator(books, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    categories = models.BookCategory.objects.filter(is_active=True, parent__isnull=True)

    current_category_obj = None
    current_category_name = ''
    if category_slug:
        try:
            current_category_obj = models.BookCategory.objects.get(slug=category_slug, is_active=True)
            current_category_name = current_category_obj.name
        except models.BookCategory.DoesNotExist:
            pass

    response_format = request.GET.get('format')

    if response_format == 'suggest':
        suggest_q = request.GET.get('q', search_query).strip()
        suggestions = []
        if suggest_q:
            suggest_books = models.Book.objects.filter(
                is_active=True,
                name__icontains=suggest_q
            ).order_by('-sale_num')[:8]
            for sbook in suggest_books:
                suggestions.append({
                    'id': sbook.id,
                    'name': sbook.name,
                    'price': str(sbook.price),
                    'image': sbook.get_cover_url(),
                    'url': f'/manager/public/books/{sbook.id}/',
                    'publisher': sbook.publisher.publisher_name if sbook.publisher else '',
                })
        return JsonResponse({'results': suggestions, 'query': suggest_q})

    if response_format == 'json':
        data_books = []
        for book in page_obj:
            data_books.append({
                'id': book.id,
                'name': book.name,
                'price': str(book.price),
                'cover_url': book.get_cover_url(),
                'url': f'/manager/public/books/{book.id}/',
                'publisher_name': book.publisher.publisher_name if book.publisher else '',
                'inventory': getattr(book, 'inventory', 0),
                'sale_num': book.sold_delivered,
                'has_cover': bool(book.cover_image),
            })
        return JsonResponse({
            'books': data_books,
            'page': page_obj.number,
            'has_more': page_obj.has_next(),
            'total_count': paginator.count,
        })

    context = {
        'books': page_obj,
        'page_obj': page_obj,
        'categories': categories,
        'search_query': search_query,
        'query': search_query,
        'sort_by': sort_by,
        'current_sort': sort_by,
        'current_category': category_slug,
        'current_category_name': current_category_name,
        'current_category_obj': current_category_obj,
        'active_category': category_slug,
        'min_price': min_price,
        'max_price': max_price,
        'total_count': paginator.count,
    }
    return render(request, 'public/books.html', context)

def public_book_detail(request, book_id):
    """Public book detail view"""
    book = get_object_or_404(models.Book.objects.select_related('publisher', 'category'), id=book_id, is_active=True)
    authors = book.author_set.all()
    
    # Related books by same publisher
    related_books = models.Book.objects.filter(
        publisher=book.publisher
    ).exclude(id=book.id)[:4]

    # Find the vendor selling this book (if any)
    vendor_book = models.VendorBook.objects.filter(book=book, is_active=True).select_related('vendor').first()
    book_vendor = vendor_book.vendor if vendor_book else None
    if not book_vendor:
        from manager.official_store import get_official_vendor
        book_vendor = get_official_vendor(create=True)

    from marketplace.review_service import reviews_for_listing, review_summary

    is_following_vendor = False
    is_following_publisher = False
    vendor_follower_count = 0
    user_id = request.session.get('site_user_id')
    if book_vendor:
        vendor_follower_count = models.UserFollowedVendor.objects.filter(vendor=book_vendor).count()
        if user_id:
            is_following_vendor = models.UserFollowedVendor.objects.filter(
                user_id=user_id, vendor=book_vendor
            ).exists()
    if user_id:
        is_following_publisher = models.UserFollowedShop.objects.filter(
            user_id=user_id, publisher=book.publisher
        ).exists()
    publisher_follower_count = models.UserFollowedShop.objects.filter(publisher=book.publisher).count()

    from manager.fulfillment_service import get_delivery_estimate
    context = {
        'book': book,
        'authors': authors,
        'related_books': related_books,
        'book_vendor': book_vendor,
        'is_following_vendor': is_following_vendor,
        'is_following_publisher': is_following_publisher,
        'vendor_follower_count': vendor_follower_count,
        'publisher_follower_count': publisher_follower_count,
        'sold_delivered': book.get_units_sold_delivered(),
        'listing_reviews_preview': list(reviews_for_listing('book', book.pk)[:3]),
        'listing_review_summary': review_summary('book', book.pk),
        'listing_kind': 'book',
        'listing_id': book.pk,
        'delivery_estimate': get_delivery_estimate('book', book.pk),
    }
    return render(request, 'public/book_detail.html', context)

def public_authors(request):
    """Public authors listing"""
    search_query = request.GET.get('search', '')
    
    authors = models.Author.objects.prefetch_related('book').all()
    
    if search_query:
        authors = authors.filter(name__icontains=search_query)
    
    authors = authors.order_by('name')

    # Use aggregate instead of per-author Python loop to avoid N+1
    from django.db.models import Count
    total_authors = authors.count()
    total_books = models.Book.objects.filter(author__in=authors).distinct().count()
    
    paginator = Paginator(authors, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'authors': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'total_authors': total_authors,
        'total_books': total_books,
    }
    return render(request, 'public/authors.html', context)

def public_author_detail(request, author_id):
    """Public author detail view"""
    author = get_object_or_404(models.Author, id=author_id)
    # Fix: Use 'book' instead of 'book_set' for ManyToMany relationship
    books = author.book.select_related('publisher').all()
    
    # Calculate statistics
    total_sales = sum(book.sale_num for book in books)
    total_inventory = sum(book.inventory for book in books)
    avg_price = sum(book.price for book in books) / len(books) if books else 0
    
    context = {
        'author': author,
        'books': books,
        'total_sales': total_sales,
        'total_inventory': total_inventory,
        'avg_price': avg_price,
    }
    return render(request, 'public/author_detail.html', context)

def public_publishers(request):
    """Public publishers listing"""
    search_query = request.GET.get('search', '')
    
    publishers = models.Publisher.objects.all()
    
    if search_query:
        publishers = publishers.filter(publisher_name__icontains=search_query)
    
    publishers = publishers.order_by('publisher_name')

    # Use aggregate instead of per-publisher Python loop to avoid N+1
    from django.db.models import Count
    total_publishers = publishers.count()
    total_books = models.Book.objects.filter(publisher__in=publishers).count()
    
    paginator = Paginator(publishers, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'publishers': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'total_publishers': total_publishers,
        'total_books': total_books,
    }
    return render(request, 'public/publishers.html', context)

def public_publisher_detail(request, publisher_id):
    """Public publisher detail view"""
    publisher = get_object_or_404(models.Publisher, id=publisher_id)
    books = models.Book.objects.filter(publisher=publisher).order_by('name')

    # Calculate statistics
    total_sales = sum(book.sale_num for book in books)
    total_inventory = sum(book.inventory for book in books)
    avg_price = sum(book.price for book in books) / len(books) if books else 0
    total_revenue = sum(book.price * book.sale_num for book in books)

    # Count unique authors
    author_count = models.Author.objects.filter(book__publisher=publisher).distinct().count()

    is_following_publisher = False
    publisher_follower_count = models.UserFollowedShop.objects.filter(publisher=publisher).count()
    user_id = request.session.get('site_user_id')
    if user_id:
        is_following_publisher = models.UserFollowedShop.objects.filter(
            user_id=user_id, publisher=publisher
        ).exists()

    context = {
        'publisher': publisher,
        'books': books,
        'total_sales': total_sales,
        'total_inventory': total_inventory,
        'avg_price': avg_price,
        'total_revenue': total_revenue,
        'author_count': author_count,
        'is_following_publisher': is_following_publisher,
        'publisher_follower_count': publisher_follower_count,
    }
    return render(request, 'public/publisher_detail.html', context)

# ==================== E-COMMERCE FUNCTIONALITY ====================

def get_session_key(request):
    """Get or create session key for cart functionality"""
    if not request.session.session_key:
        request.session.create()
    return request.session.session_key


def _get_marketplace_item(item_type, item_id):
    """Helper to resolve a marketplace item from its type and ID."""
    if item_type == 'product':
        return Product.objects.filter(pk=item_id, is_active=True).first()
    elif item_type == 'course':
        return Course.objects.filter(pk=item_id, is_active=True).first()
    elif item_type == 'supermarket':
        return SupermarketItem.objects.filter(pk=item_id, is_active=True).first()
    return None


def _get_marketplace_attribute_context(item):
    if not item or not hasattr(item, 'attributes'):
        return {'groups': [], 'selectable_groups': [], 'specification_groups': []}
    return build_attribute_groups(item.attributes.all())


def _build_unified_cart(session_key):
    """Build a unified cart list merging books and marketplace items."""
    items = []

    # Book cart items
    book_items = models.CartItem.objects.filter(session_key=session_key).select_related('book', 'book__publisher')
    for ci in book_items:
        items.append({
            'id': ci.id,
            'item_type': 'book',
            'item_id': ci.book.id,
            'name': ci.book.name,
            'price': ci.book.price,
            'quantity': ci.quantity,
            'image_url': ci.book.get_cover_url() if hasattr(ci.book, 'get_cover_url') else (ci.book.cover_image.url if ci.book.cover_image else ''),
            'total_price': ci.get_total_price(),
            'publisher': ci.book.publisher.publisher_name if ci.book.publisher else '',
            'inventory': ci.book.inventory,
            'item_obj': ci.book,
            'cart_item': ci,
        })

    # Marketplace cart items
    mkt_items = MarketplaceCartItem.objects.filter(session_key=session_key).order_by('-created_at')
    for ci in mkt_items:
        item = ci.get_item()
        if item:
            items.append({
                'id': ci.id,
                'item_type': ci.item_type,
                'item_id': ci.item_id,
                'name': ci.get_item_name(),
                'price': ci.get_item_price(),
                'quantity': ci.quantity,
                'image_url': ci.get_item_image_url() or '/static/img/default_product.png',
                'total_price': ci.get_total_price(),
                'publisher': getattr(item, 'brand', '') or '',
                'inventory': getattr(item, 'stock', 999),
                'item_obj': item,
                'cart_item': ci,
                'selected_attributes': ci.selected_attributes or {},
                'selected_attribute_list': ci.get_selected_attributes_display(),
                'pricing_rule_log': ci.pricing_rule_log or {},
            })

    return items


@require_POST
def add_to_cart(request, book_id):
    """Add book to shopping cart"""
    try:
        book = get_object_or_404(models.Book, id=book_id)
        quantity = int(request.POST.get('quantity', 1))
        session_key = get_session_key(request)
        
        # Validate quantity
        if quantity > book.inventory:
            return JsonResponse({
                'success': False,
                'message': f'库存不足！当前库存：{book.inventory}本'
            })
        
        # Get or create cart item
        cart_item, created = models.CartItem.objects.get_or_create(
            session_key=session_key,
            book=book,
            defaults={'quantity': quantity}
        )
        
        if not created:
            new_quantity = cart_item.quantity + quantity
            if new_quantity > book.inventory:
                return JsonResponse({
                    'success': False,
                    'message': f'库存不足！当前库存：{book.inventory}本，购物车中已有：{cart_item.quantity}本'
                })
            cart_item.quantity = new_quantity
            cart_item.save()
        
        # Get total unified cart count
        cart_count = _get_unified_cart_count(session_key)
        
        return JsonResponse({
            'success': True,
            'message': f'已将《{book.name}》添加到购物车',
            'cart_count': cart_count,
            'item_quantity': cart_item.quantity
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': '添加到购物车失败，请重试'
        })


@require_POST
def add_marketplace_item_to_cart(request):
    """Add marketplace item (product/course/supermarket) to unified cart."""
    try:
        item_type = request.POST.get('item_type')
        item_id = int(request.POST.get('item_id'))
        quantity = int(request.POST.get('quantity', 1))
        session_key = get_session_key(request)

        if item_type not in ('product', 'course', 'supermarket'):
            return JsonResponse({'success': False, 'message': '无效的商品类型'})

        item = _get_marketplace_item(item_type, item_id)
        if not item:
            return JsonResponse({'success': False, 'message': '商品不存在或已下架'})

        # Validate stock
        if item_type in ('product', 'supermarket'):
            qty_check = validate_quantity(item, quantity)
            if not qty_check.is_valid:
                return JsonResponse({'success': False, 'message': qty_check.message, 'suggested_quantity': qty_check.suggested_quantity})
            if quantity > item.stock:
                return JsonResponse({'success': False, 'message': f'库存不足！当前库存：{item.stock}'})
        if item_type == 'course':
            quantity = 1

        selected_attributes = {}
        if item_type in ('product', 'supermarket'):
            validation = validate_selected_attributes(
                _get_marketplace_attribute_context(item),
                request.POST.get('selected_attributes', '{}')
            )
            if not validation['is_valid']:
                return JsonResponse({'success': False, 'message': validation['errors'][0]})
            selected_attributes = validation['cleaned']

        item_name = item.title if item_type == 'course' else item.name

        cart_item = MarketplaceCartItem.objects.filter(
            session_key=session_key,
            item_type=item_type,
            item_id=item_id,
            selected_attributes=selected_attributes,
        ).first()
        created = cart_item is None
        if created:
            cart_item = MarketplaceCartItem.objects.create(
                session_key=session_key,
                item_type=item_type,
                item_id=item_id,
                quantity=quantity,
                selected_attributes=selected_attributes,
            )

        if not created:
            if item_type == 'course':
                return JsonResponse({'success': False, 'message': '该课程已在购物车中'})
            new_qty = cart_item.quantity + quantity
            if item_type in ('product', 'supermarket'):
                qty_check = validate_quantity(item, new_qty)
                if not qty_check.is_valid:
                    return JsonResponse({'success': False, 'message': qty_check.message, 'suggested_quantity': qty_check.suggested_quantity})
                if new_qty > item.stock:
                    return JsonResponse({'success': False, 'message': f'库存不足！购物车已有{cart_item.quantity}件'})
            cart_item.quantity = new_qty
            cart_item.save()

        cart_count = _get_unified_cart_count(session_key)
        return JsonResponse({
            'success': True,
            'message': f'已将「{item_name}」添加到购物车',
            'cart_count': cart_count,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': '添加失败，请重试'})


def _get_unified_cart_count(session_key):
    """Get total count of items in both carts."""
    book_count = models.CartItem.objects.filter(session_key=session_key).count()
    mkt_count = MarketplaceCartItem.objects.filter(session_key=session_key).count()
    return book_count + mkt_count


@ensure_csrf_cookie
def view_cart(request):
    """Display unified shopping cart (books + marketplace)"""
    session_key = get_session_key(request)
    unified_items = _build_unified_cart(session_key)
    
    total_amount = sum(item['total_price'] for item in unified_items)
    total_items = sum(item['quantity'] for item in unified_items)
    
    context = {
        'cart_items': unified_items,
        'total_amount': total_amount,
        'total_items': total_items,
    }
    
    return render(request, 'public/cart.html', context)

@require_POST
def update_cart(request):
    """Update cart item quantities via AJAX - handles both books and marketplace items"""
    try:
        data = json.loads(request.body)
        item_id = data.get('item_id')
        item_type = data.get('item_type', 'book')
        quantity = int(data.get('quantity', 1))
        session_key = get_session_key(request)
        
        item_name = ''

        if item_type == 'book':
            cart_item = get_object_or_404(models.CartItem, id=item_id, session_key=session_key)
            if quantity > cart_item.book.inventory:
                return JsonResponse({
                    'success': False,
                    'message': f'库存不足！最大可购买：{cart_item.book.inventory}本'
                })
            item_name = cart_item.book.name
            if quantity <= 0:
                cart_item.delete()
            else:
                cart_item.quantity = quantity
                cart_item.save()
            item_total = float(cart_item.get_total_price()) if quantity > 0 else 0
        else:
            cart_item = MarketplaceCartItem.objects.filter(
                id=item_id, session_key=session_key
            ).first()
            if not cart_item:
                return JsonResponse({'success': False, 'message': '商品不在购物车中'})

            item = cart_item.get_item()
            if quantity > 0 and item_type in ('product', 'supermarket') and item:
                qty_check = validate_quantity(item, quantity)
                if not qty_check.is_valid:
                    return JsonResponse({'success': False, 'message': qty_check.message, 'suggested_quantity': qty_check.suggested_quantity})
                if quantity > item.stock:
                    return JsonResponse({'success': False, 'message': f'库存不足！最大可购买：{item.stock}'})
            
            item_name = cart_item.get_item_name()
            if quantity <= 0:
                cart_item.delete()
            else:
                cart_item.quantity = quantity
                cart_item.save()
            item_total = float(cart_item.get_total_price()) if quantity > 0 else 0
        
        # Recalculate unified totals
        unified_items = _build_unified_cart(session_key)
        total_amount = sum(i['total_price'] for i in unified_items)
        total_items = sum(i['quantity'] for i in unified_items)
        
        message = f'已更新「{item_name}」数量' if quantity > 0 else f'已移除「{item_name}」'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'item_total': item_total,
            'cart_total': float(total_amount),
            'total_items': total_items
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': '更新购物车失败'
        })

def remove_from_cart(request, item_id):
    """Remove item from cart - handles both books and marketplace items"""
    session_key = get_session_key(request)
    item_type = request.POST.get('item_type', request.GET.get('item_type', 'book'))

    item_name = ''
    if item_type == 'book':
        cart_item = get_object_or_404(models.CartItem, id=item_id, session_key=session_key)
        item_name = cart_item.book.name
        cart_item.delete()
    else:
        cart_item = MarketplaceCartItem.objects.filter(
            id=item_id, session_key=session_key
        ).first()
        if cart_item:
            item_name = cart_item.get_item_name()
            cart_item.delete()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.method == 'POST':
        unified_items = _build_unified_cart(session_key)
        total_amount = sum(i['total_price'] for i in unified_items)
        total_items = sum(i['quantity'] for i in unified_items)
        cart_count = len(unified_items)
        return JsonResponse({
            'success': True,
            'message': f'已移除「{item_name}」',
            'cart_total': float(total_amount),
            'total_items': total_items,
            'cart_count': cart_count,
        })

    messages.success(request, f'已移除「{item_name}」')
    return redirect('manager:view_cart')


def clear_cart(request):
    """Clear all items from both carts"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    session_key = get_session_key(request)
    models.CartItem.objects.filter(session_key=session_key).delete()
    MarketplaceCartItem.objects.filter(session_key=session_key).delete()
    return JsonResponse({'success': True, 'message': '购物车已清空'})

def get_cart_count(request):
    """Get unified cart item count"""
    session_key = get_session_key(request)
    cart_count = _get_unified_cart_count(session_key)
    return JsonResponse({'cart_count': cart_count})

@require_POST
def buy_now(request, book_id):
    """Direct purchase without cart"""
    try:
        book = get_object_or_404(models.Book, id=book_id)
        quantity = int(request.POST.get('quantity', 1))
        session_key = get_session_key(request)
        
        if quantity > book.inventory:
            messages.error(request, f'库存不足！当前库存：{book.inventory}本')
            return redirect('manager:public_book_detail', book_id=book_id)
        
        # Clear book cart items only
        models.CartItem.objects.filter(session_key=session_key).delete()
        MarketplaceCartItem.objects.filter(session_key=session_key).delete()
        
        models.CartItem.objects.create(
            session_key=session_key,
            book=book,
            quantity=quantity
        )
        
        return redirect('manager:checkout')
        
    except Exception as e:
        messages.error(request, '购买失败，请重试')
        return redirect('manager:public_book_detail', book_id=book_id)

# Optional checkout donation supporting children in need — flat amount,
# not user-adjustable, shown as a clearly optional toggle at checkout.
DONATION_AMOUNT = Decimal('500.00')


def donation_admin_note(amount):
    """Bilingual note auto-attached to admin_notes whenever a checkout
    included a donation, so whoever ends up looking at the order/payment
    (admin, finance, vendor) immediately sees that part of what was
    collected is a solidarity donation, not revenue for an item."""
    from manager.templatetags.currency_filters import to_fcfa
    amt = to_fcfa(amount)
    return (
        f"\U0001F49B Ce paiement inclut un don solidaire de {amt} "
        f"(soutien aux enfants dans le besoin) — à ne pas compter comme "
        f"chiffre d'affaires produit.\n"
        f"This payment includes a {amt} solidarity donation "
        f"(supporting children in need) — do not count as product revenue."
    )


def _parse_delivery_days_override(post_data, prefix=''):
    """Parse an optional per-item delivery-days override from a POST dict.
    Returns (None, None) unless both min and max are present, positive
    integers, min <= max, and within a sane 1-90 day range — an invalid or
    partial submission just falls back to no override (vendor default /
    platform default) rather than erroring the whole form."""
    min_raw = post_data.get(f'{prefix}delivery_days_min', '').strip()
    max_raw = post_data.get(f'{prefix}delivery_days_max', '').strip()
    if not min_raw or not max_raw:
        return None, None
    if not (min_raw.isdigit() and max_raw.isdigit()):
        return None, None
    d_min, d_max = int(min_raw), int(max_raw)
    if 0 < d_min <= d_max <= 90:
        return d_min, d_max
    return None, None


def checkout(request):
    """Unified checkout - handles books and marketplace items together"""
    session_key = get_session_key(request)
    unified_items = _build_unified_cart(session_key)
    
    if not unified_items:
        messages.warning(request, '购物车为空，请先添加商品')
        return redirect('manager:public_books')
    
    # Separate items by type
    book_items = [i for i in unified_items if i['item_type'] == 'book']
    marketplace_items = [i for i in unified_items if i['item_type'] != 'book']

    # Calculate totals - books use actual price, marketplace uses item price
    total_amount = sum(i['total_price'] for i in unified_items)
    total_items_count = sum(i['quantity'] for i in unified_items)

    # Delivery estimate: most conservative (widest) range across every
    # physical item in the cart. Courses are digital and skipped.
    from manager.fulfillment_service import get_delivery_estimate
    overall_delivery_estimate = None
    for i in unified_items:
        if i['item_type'] == 'course':
            continue
        est = get_delivery_estimate(i['item_type'], i['item_id'])
        if overall_delivery_estimate is None or est['days_max'] > overall_delivery_estimate['days_max']:
            overall_delivery_estimate = est
    
    if request.method == 'POST':
        ip = _get_client_ip(request)
        rl_key = f'checkout_fail:{ip}'
        if _is_rate_limited_key(rl_key, 20):
            messages.error(request, '请求过于频繁，请稍后再试。')
            return redirect('manager:checkout')
        _record_attempt_key(rl_key, 300)
        try:
            payment_confirmed = request.POST.get('payment_confirmed', 'no')
            kkiapay_transaction_id = request.POST.get('kkiapay_transaction_id', '').strip()

            if payment_confirmed == 'kkiapay_success':
                # Payment completed via KKiaPay widget — verify server-side FIRST
                if not kkiapay_transaction_id:
                    messages.error(request, 'Transaction ID manquant. Contactez le support.')
                    return redirect('manager:checkout')
                from manager.payments.kkiapay import is_transaction_successful
                verified, _tx = is_transaction_successful(kkiapay_transaction_id)
                if not verified:
                    messages.error(request, 'Vérification du paiement échouée. Contactez le support.')
                    return redirect('manager:checkout')
                initial_status = 'processing'
                payment_status = 'completed'
            elif payment_confirmed == 'yes':
                initial_status = 'processing'
                payment_status = 'pending'
            else:
                initial_status = 'payment_pending'
                payment_status = 'pending'
            
            customer_name = request.POST.get('customer_name')
            customer_email = request.POST.get('customer_email')
            customer_phone = request.POST.get('customer_phone')
            country = request.POST.get('country', 'China')
            city = request.POST.get('city', '').strip()
            shipping_address = request.POST.get('shipping_address', '').strip()
            payment_method = request.POST.get('payment_method')
            customer_notes = request.POST.get('customer_notes', '')

            from book_Project.checkout_cities import is_valid_checkout_city
            if not is_valid_checkout_city(country, city):
                messages.error(request, '请选择有效的城市。')
                return redirect('manager:checkout')

            available_methods = {
                option['method']
                for region_options in build_payment_options(country).values()
                for option in region_options
            }
            if payment_method not in available_methods:
                messages.error(request, _('当前国家暂不支持该支付方式，请重新选择。'))
                return redirect('manager:checkout')

            # Optional 500 FCFA donation supporting children in need — added
            # to whichever order actually gets charged (matches the
            # target_order selection below: book_order takes priority when
            # both book and marketplace items are in the same cart).
            donation = DONATION_AMOUNT if request.POST.get('donate') == 'yes' else Decimal('0.00')
            book_donation = donation if book_items else Decimal('0.00')
            mkt_donation = donation if not book_items else Decimal('0.00')

            book_order = None
            mkt_order = None

            # Create book order if there are book items
            if book_items:
                book_total = sum(i['total_price'] for i in book_items) + book_donation
                book_order = models.Order.objects.create(
                    customer_name=customer_name,
                    customer_email=customer_email,
                    customer_phone=customer_phone,
                    country=country,
                    city=city,
                    shipping_address=shipping_address,
                    payment_method=payment_method,
                    total_amount=book_total,
                    donation_amount=book_donation,
                    status=initial_status,
                    # Always created 'pending' — even for the already-verified
                    # kkiapay_success case — so _update_order_status below is
                    # the one place that ever promotes it to 'completed' and
                    # fires shipment creation / email / inventory exactly once.
                    payment_status='pending',
                    customer_notes=customer_notes,
                    admin_notes=donation_admin_note(book_donation) if book_donation else '',
                    payment_transaction_id=kkiapay_transaction_id or None,
                )
                for item in book_items:
                    models.OrderItem.objects.create(
                        order=book_order,
                        book=item['item_obj'],
                        quantity=item['quantity'],
                        unit_price=item['price'],
                        total_price=item['total_price']
                    )
                # Stock/sales are deducted only once payment is actually
                # confirmed (see below and manager.payments.views), never
                # here at order creation — an abandoned/failed payment must
                # never permanently reduce stock.

                # Clear book cart
                models.CartItem.objects.filter(session_key=session_key).delete()

                create_notification(
                    'new_order',
                    f'新订单 {book_order.order_number}',
                    f'{book_order.customer_name} 下了一个新订单，共 ¥{book_order.total_amount}',
                    icon='fas fa-shopping-bag',
                    color='#10b981',
                    link=f'/manager/order_detail/{book_order.id}/',
                    related_id=book_order.id,
                )

                if payment_status == 'completed':
                    # KKiaPay was already verified synchronously above — run
                    # it through the same pipeline a webhook would (creates
                    # shipments, sends confirmation email, deducts stock)
                    # instead of leaving those permanently skipped.
                    from manager.payments.views import _update_order_status
                    _update_order_status(book_order, 'SUCCESSFUL', transaction_id=kkiapay_transaction_id or None)

            # Create marketplace order if there are marketplace items
            if marketplace_items:
                for item in marketplace_items:
                    obj = item['item_obj']
                    if item['item_type'] in ('product', 'supermarket'):
                        qty_check = validate_quantity(obj, item['quantity'])
                        if not qty_check.is_valid:
                            messages.error(request, qty_check.message)
                            return redirect('manager:checkout')
                        if item['quantity'] > obj.stock:
                            messages.error(request, f'库存不足！当前库存：{obj.stock}')
                            return redirect('manager:checkout')
                mkt_total = sum(i['total_price'] for i in marketplace_items) + mkt_donation
                mkt_order = MarketplaceOrder(
                    user_name=customer_name,
                    user_email=customer_email,
                    customer_phone=customer_phone,
                    country=country,
                    city=city,
                    payment_method=payment_method,
                    total_amount=mkt_total,
                    donation_amount=mkt_donation,
                    status=initial_status,
                    # See book_order above — always created 'pending';
                    # _update_order_status is the only place that promotes it.
                    payment_status='pending',
                    shipping_address=shipping_address or request.POST.get('shipping_address', ''),
                    notes=customer_notes,
                    customer_notes=customer_notes,
                    admin_notes=donation_admin_note(mkt_donation) if mkt_donation else '',
                    payment_transaction_id=kkiapay_transaction_id or None,
                )
                mkt_order.save()

                for item in marketplace_items:
                    MarketplaceOrderItem.objects.create(
                        order=mkt_order,
                        item_type=item['item_type'],
                        item_id=item['item_id'],
                        item_name=item['name'],
                        item_image=item['image_url'],
                        quantity=item['quantity'],
                        unit_price=item['price'],
                        selected_attributes=item.get('selected_attributes', {}),
                        pricing_rule_log=item.get('pricing_rule_log', {}),
                    )
                    # Stock/sales/enrollment are deducted only once payment is
                    # actually confirmed (see below), never here.

                # Clear marketplace cart
                MarketplaceCartItem.objects.filter(session_key=session_key).delete()

                if payment_status == 'completed':
                    from manager.payments.views import _update_order_status
                    _update_order_status(mkt_order, 'SUCCESSFUL', transaction_id=kkiapay_transaction_id or None)

            # Grant access via session before redirecting
            if book_order:
                _accessible = request.session.get('accessible_orders', [])
                _order_key = str(book_order.order_number)
                if _order_key not in _accessible:
                    _accessible.append(_order_key)
                request.session['accessible_orders'] = _accessible
                
            if mkt_order:
                _accessible = request.session.get('accessible_orders', [])
                _order_key = str(mkt_order.order_number)
                if _order_key not in _accessible:
                    _accessible.append(_order_key)
                request.session['accessible_orders'] = _accessible
                
            # If payment method is KKiaPay, redirect to the new dedicated payment page
            if payment_method == 'kkiapay':
                target_order = book_order if book_order else mkt_order
                return redirect('manager:kkiapay_pay', order_number=target_order.order_number)

            if payment_method == 'pawapay':
                target_order = book_order if book_order else mkt_order
                return redirect('manager:pawapay_pay', order_number=target_order.order_number)
            
            target_order_for_redirect = book_order if book_order else mkt_order
            return redirect('manager:order_confirmation', order_number=target_order_for_redirect.order_number)
            
        except Exception as e:
            messages.error(request, '订单创建失败，请重试')
    
    payment_methods_by_region = build_payment_options()

    total_amount_fcfa = round(float(total_amount))

    from book_Project.payment_config import get_kkiapay_country_codes
    from book_Project.checkout_cities import get_checkout_cities_by_country
    kkiapay_countries = get_kkiapay_country_codes()

    from manager.models import KkiapayCountry
    kkiapay_countries_data = KkiapayCountry.get_for_widget()

    context = {
        'cart_items': unified_items,
        'book_items': book_items,
        'marketplace_items': marketplace_items,
        'total_amount': total_amount,
        'total_amount_fcfa': total_amount_fcfa,
        'total_items': total_items_count,
        'payment_methods_by_region': payment_methods_by_region,
        'KKIAPAY_PUBLIC_KEY': django_settings.KKIAPAY_PUBLIC_KEY,
        'KKIAPAY_SANDBOX': django_settings.KKIAPAY_SANDBOX,
        'kkiapay_countries': kkiapay_countries,
        'kkiapay_countries_data': kkiapay_countries_data,
        'checkout_cities_by_country': get_checkout_cities_by_country(),
        'donation_amount': DONATION_AMOUNT,
        'delivery_estimate': overall_delivery_estimate,
    }

    response = render(request, 'public/checkout.html', context)
    # Prevent browser from caching the checkout page (avoids stale JS bugs)
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    return response


def kkiapay_pay(request, order_number):
    """
    Dedicated view for KKiaPay payment flow.
    The order is already created with 'payment_pending'.
    This view hosts the KKiaPay widget.
    """
    from django.conf import settings as django_settings
    
    accessible = request.session.get('accessible_orders', [])
    if str(order_number) not in accessible:
        messages.warning(request, 'Non autorisé ou session expirée.')
        return redirect('manager:public_home')
        
    order = None
    try:
        order = models.Order.objects.get(order_number=order_number)
    except models.Order.DoesNotExist:
        try:
            from marketplace.models import MarketplaceOrder
            order = MarketplaceOrder.objects.get(order_number=order_number)
        except MarketplaceOrder.DoesNotExist:
            messages.error(request, 'Commande non trouvée.')
            return redirect('manager:public_home')

    if order.payment_status == 'completed':
        messages.info(request, 'Commande déjà payée.')
        return redirect('manager:order_confirmation', order_number=order.order_number)

    # Pass the order parameters safely to Kkiapay
    context = {
        'order': order,
        'order_number': order.order_number,
        'total_amount_fcfa': int(order.total_amount),
        'customer_name': order.customer_name if hasattr(order, 'customer_name') else getattr(order, 'user_name', ''),
        'customer_email': order.customer_email if hasattr(order, 'customer_email') else getattr(order, 'user_email', ''),
        'customer_phone': order.customer_phone,
        'kkiapay_public_key': django_settings.KKIAPAY_PUBLIC_KEY,
        'kkiapay_sandbox': django_settings.KKIAPAY_SANDBOX,
        'callback_url': request.build_absolute_uri(f"/manager/public/kkiapay/success/{order_number}/")
    }
    
    return render(request, 'public/kkiapay_pay.html', context)


def pawapay_pay(request, order_number):
    """PawaPay checkout — creates a hosted Payment Page session and sends the
    customer to PawaPay's own checkout widget (operator selection, phone entry,
    PIN-prompt waiting screen — all rendered on pawapay.io), then brings them
    back to pawapay_return() once they finish there."""
    accessible = request.session.get('accessible_orders', [])
    if str(order_number) not in accessible:
        messages.warning(request, 'Non autorisé ou session expirée.')
        return redirect('manager:public_home')

    order = None
    try:
        order = models.Order.objects.get(order_number=order_number)
    except models.Order.DoesNotExist:
        try:
            from marketplace.models import MarketplaceOrder
            order = MarketplaceOrder.objects.get(order_number=order_number)
        except MarketplaceOrder.DoesNotExist:
            messages.error(request, 'Commande non trouvée.')
            return redirect('manager:public_home')

    if order.payment_status == 'completed':
        messages.info(request, 'Commande déjà payée.')
        return redirect('manager:order_confirmation', order_number=order.order_number)

    order_country = getattr(order, 'country', '') or ''
    customer_phone = order.customer_phone or ''
    customer_name = order.customer_name if hasattr(order, 'customer_name') else getattr(order, 'user_name', '')

    from manager.payments.pawapay import create_payment_page_session
    from django.utils import translation

    return_url = request.build_absolute_uri(
        reverse('manager:pawapay_return', args=[order_number])
    )
    lang = 'EN' if translation.get_language() == 'en' else 'FR'
    result = create_payment_page_session(
        amount=int(order.total_amount),
        order_number=order_number,
        return_url=return_url,
        phone_number=customer_phone,
        country=order_country,
        reason=f'DUNO 360 — {customer_name}'[:50] if customer_name else None,
        language=lang,
    )
    if not result.get('success'):
        messages.error(
            request,
            result.get('error') or "Impossible d'ouvrir la page de paiement PawaPay. Veuillez réessayer."
        )
        return redirect('manager:checkout')

    order.payment_transaction_id = result['deposit_id']
    order.payment_status = 'processing'
    order.save(update_fields=['payment_transaction_id', 'payment_status'])

    return redirect(result['redirect_url'])


def pawapay_return(request, order_number):
    """Landing view PawaPay's hosted Payment Page redirects back to once the
    customer finishes (success, failure, or cancel) — receives ?depositId=...
    and checks the real status server-side before showing the customer
    anything (never trust the redirect itself as proof of payment)."""
    order = None
    try:
        order = models.Order.objects.get(order_number=order_number)
    except models.Order.DoesNotExist:
        try:
            from marketplace.models import MarketplaceOrder
            order = MarketplaceOrder.objects.get(order_number=order_number)
        except MarketplaceOrder.DoesNotExist:
            messages.error(request, 'Commande non trouvée.')
            return redirect('manager:public_home')

    deposit_id = request.GET.get('depositId') or order.payment_transaction_id
    if not deposit_id:
        messages.error(request, 'Paiement introuvable.')
        return redirect('manager:pawapay_pay', order_number=order_number)

    if order.payment_status != 'completed':
        from manager.payments.pawapay import get_deposit_status_v2, normalize_pawapay_status
        from manager.payments.views import _update_order_status

        result = get_deposit_status_v2(deposit_id)
        internal = normalize_pawapay_status(result.get('status', 'PENDING'))
        if internal in ('SUCCESSFUL', 'FAILED'):
            _update_order_status(order, internal, transaction_id=deposit_id)
        elif result.get('status') == 'NOT_FOUND':
            # Customer left the Payment Page without completing anything.
            messages.warning(request, 'Paiement annulé ou non terminé.')
            return redirect('manager:pawapay_pay', order_number=order_number)

    if order.payment_status == 'completed':
        return redirect('manager:order_confirmation', order_number=order_number)
    if order.payment_status == 'failed':
        messages.error(request, 'Le paiement a échoué. Veuillez réessayer.')
        return redirect('manager:pawapay_pay', order_number=order_number)

    # Still pending/processing — PawaPay confirms most sandbox test flows
    # instantly, but leave room for the rare async case; send the customer to
    # the order confirmation page, which already polls payment status.
    return redirect('manager:order_confirmation', order_number=order_number)


def kkiapay_success_redirect(request, order_number):
    """
    Redirected here after KKiaPay widget completes via callback GET.
    We just show confirmation page. The async webhook or client JS 
    will/has hit the verify endpoint. But we can also proactively verify here if we want!
    """
    from manager.payments.kkiapay import is_transaction_successful
    transaction_id = request.GET.get('transaction_id')
    
    if transaction_id:
        try:
            order = models.Order.objects.get(order_number=order_number)
        except models.Order.DoesNotExist:
            from marketplace.models import MarketplaceOrder
            order = MarketplaceOrder.objects.filter(order_number=order_number).first()
            
        if order and order.payment_status != 'completed':
            success, tx = is_transaction_successful(transaction_id)
            if success:
                # Route through the shared pipeline instead of setting
                # fields directly — this used to silently skip shipment
                # creation, the confirmation email, and inventory deduction
                # whenever this GET landing page ran before the webhook did.
                from manager.payments.views import _update_order_status
                _update_order_status(order, 'SUCCESSFUL', transaction_id=transaction_id)
    
    return redirect('manager:order_confirmation', order_number=order_number)


def order_confirmation(request, order_number):
    """Order confirmation page - handles both book and marketplace orders"""
    # Ownership check: order_number must be in the session's accessible list
    # (granted at checkout or after email-verified track_order lookup).
    accessible = request.session.get('accessible_orders', [])
    if str(order_number) not in accessible:
        messages.warning(request, _('请通过订单查询验证您的身份后查看订单详情'))
        return redirect('manager:track_order')
    book_order = None
    book_order_items = []
    mkt_order = None
    mkt_order_items = []

    # Try book order first
    try:
        book_order = models.Order.objects.get(order_number=order_number)
        book_order_items = models.OrderItem.objects.filter(order=book_order).select_related('book')
    except models.Order.DoesNotExist:
        pass

    # Try marketplace order
    if not book_order:
        try:
            mkt_order = MarketplaceOrder.objects.get(order_number=order_number)
            mkt_order_items = mkt_order.items.all()
        except MarketplaceOrder.DoesNotExist:
            pass

    if not book_order and not mkt_order:
        messages.error(request, '订单不存在')
        return redirect('manager:public_home')

    resolved_order = book_order or mkt_order
    payment_time_remaining_seconds = 0
    if hasattr(resolved_order, 'get_payment_time_remaining'):
        payment_time_remaining = resolved_order.get_payment_time_remaining()
        if payment_time_remaining:
            # Order.get_payment_time_remaining() (book orders) returns an int of
            # seconds; MarketplaceOrder's returns a timedelta. Handle both rather
            # than assuming one shape (previously crashed every fresh book order).
            if hasattr(payment_time_remaining, 'total_seconds'):
                payment_time_remaining_seconds = max(int(payment_time_remaining.total_seconds()), 0)
            else:
                payment_time_remaining_seconds = max(int(payment_time_remaining), 0)

    order_source = 'book' if book_order else 'marketplace'
    shipments = list(
        models.Shipment.objects.filter(order_source=order_source, order_id=resolved_order.id)
        .select_related('vendor')
        .prefetch_related('return_requests')
    )

    context = {
        'order': resolved_order,
        'book_order': book_order,
        'book_order_items': book_order_items,
        'mkt_order': mkt_order,
        'mkt_order_items': mkt_order_items,
        'order_items': book_order_items,  # backward compat
        'payment_time_remaining_seconds': payment_time_remaining_seconds,
        'shipments': shipments,
        'order_source': order_source,
        'return_reasons': models.OrderReturnRequest.REASON_CHOICES,
    }

    return render(request, 'public/order_confirmation.html', context)


def _buyer_owns_shipment(request, shipment):
    """Ownership check mirrors order_confirmation()'s: order_number must be
    in the session's accessible_orders list — this platform doesn't require
    login for guest book checkout, so shipments are protected the same way
    orders already are, not via a hard user FK."""
    accessible = request.session.get('accessible_orders', [])
    return shipment.order_number in accessible


@require_POST
def confirm_delivery_receipt(request):
    """Buyer clicks 'Confirm receipt' — the ONLY normal path to a shipment
    being marked delivered (see fulfillment_service docstring for why it's
    never the seller). Starts the escrow release countdown for this
    shipment's items."""
    from manager import fulfillment_service as fs

    shipment_id = request.POST.get('shipment_id')
    shipment = get_object_or_404(models.Shipment, id=shipment_id)
    if not _buyer_owns_shipment(request, shipment):
        return JsonResponse({'success': False, 'message': str(_('Non autorisé.'))}, status=403)

    if not shipment.can_confirm_receipt:
        return JsonResponse({'success': False, 'message': str(_('Cette expédition ne peut pas être confirmée maintenant.'))}, status=400)

    ok = fs.confirm_delivery(shipment, confirmed_by='buyer')
    if not ok:
        return JsonResponse({'success': False, 'message': str(_('Action impossible.'))}, status=400)
    return JsonResponse({'success': True, 'message': str(_('Réception confirmée — merci !'))})


@require_POST
def submit_return_request(request):
    """Buyer opens a return/dispute on a delivered shipment."""
    from manager import fulfillment_service as fs

    shipment_id = request.POST.get('shipment_id')
    shipment = get_object_or_404(models.Shipment, id=shipment_id)
    if not _buyer_owns_shipment(request, shipment):
        return JsonResponse({'success': False, 'message': str(_('Non autorisé.'))}, status=403)

    reason = request.POST.get('reason', '').strip()
    description = request.POST.get('description', '').strip()
    valid_reasons = dict(models.OrderReturnRequest.REASON_CHOICES)
    if reason not in valid_reasons:
        return JsonResponse({'success': False, 'message': str(_('Motif invalide.'))}, status=400)

    buyer_name = getattr(shipment, '_buyer_name', '') or ''
    order = None
    if shipment.order_source == 'book':
        order = models.Order.objects.filter(id=shipment.order_id).first()
        buyer_email = order.customer_email if order else ''
        buyer_name = order.customer_name if order else ''
    else:
        order = MarketplaceOrder.objects.filter(id=shipment.order_id).first()
        buyer_email = order.user_email if order else ''
        buyer_name = order.user_name if order else ''

    req = fs.open_return_request(shipment, buyer_name, buyer_email, reason, description=description)
    if not req:
        return JsonResponse({'success': False, 'message': str(_('Impossible d\'ouvrir une demande de retour pour cette expédition (délai dépassé ou statut invalide).'))}, status=400)
    return JsonResponse({'success': True, 'message': str(_('Demande de retour envoyée.'))})


def track_order(request):
    """Order tracking page - Search by order number or email"""
    order = None
    orders = None
    mkt_order = None
    mkt_orders = None
    has_downloadable_books = False

    status_filter = (request.GET.get('status') or '').strip().lower()
    user_id = request.session.get('site_user_id')
    site_user = None
    all_orders = None
    all_mkt_orders = None
    if user_id and request.method == 'GET' and not request.GET.get('order_number'):
        try:
            site_user = models.SiteUser.objects.get(pk=user_id)
            all_orders = models.Order.objects.filter(
                customer_email=site_user.email
            ).order_by('-created_at')
            orders = all_orders
            try:
                from django.db.models import Q as _Q

                all_mkt_orders = MarketplaceOrder.objects.filter(
                    _Q(user_email=site_user.email) | _Q(user_id=site_user.id)
                ).order_by('-created_at')
                mkt_orders = all_mkt_orders
                if not all_mkt_orders.exists():
                    mkt_orders = None
            except Exception:
                mkt_orders = None
                all_mkt_orders = None

            if status_filter:
                from manager.order_status import filter_orders_by_bucket
                orders = filter_orders_by_bucket(all_orders, status_filter)
                mkt_orders = (
                    filter_orders_by_bucket(all_mkt_orders, status_filter)
                    if all_mkt_orders is not None else None
                )

            if orders.exists():
                _accessible = request.session.get('accessible_orders', [])
                for _o in orders:
                    if str(_o.order_number) not in _accessible:
                        _accessible.append(str(_o.order_number))
                request.session['accessible_orders'] = _accessible
            if not orders.exists():
                orders = None
            if mkt_orders is not None and not mkt_orders.exists():
                mkt_orders = None
        except models.SiteUser.DoesNotExist:
            pass

    if request.method == 'GET' and request.GET.get('order_number'):
        order_number = request.GET.get('order_number').strip()
        if order_number:
            try:
                order = models.Order.objects.get(order_number=order_number)
                order.apply_ttl_rules()
                _accessible = request.session.get('accessible_orders', [])
                if str(order.order_number) not in _accessible:
                    _accessible.append(str(order.order_number))
                request.session['accessible_orders'] = _accessible
                has_downloadable_books = any(
                    item.book.has_download for item in order.orderitem_set.all()
                )
            except models.Order.DoesNotExist:
                try:
                    mkt_order = MarketplaceOrder.objects.get(order_number=order_number)
                    mkt_order.apply_ttl_rules()
                except MarketplaceOrder.DoesNotExist:
                    messages.error(request, '订单号不存在')

    if request.method == 'POST':
        search_type = request.POST.get('search_type', 'order_number')
        
        if search_type == 'email':
            customer_email = (request.POST.get('customer_email') or '').strip().lower()
            order_proof = (request.POST.get('email_order_number') or '').strip()
            if customer_email and order_proof:
                verified_book_order = models.Order.objects.filter(
                    order_number__iexact=order_proof,
                    customer_email__iexact=customer_email,
                ).first()
                verified_mkt_order = None
                try:
                    verified_mkt_order = MarketplaceOrder.objects.filter(
                        order_number__iexact=order_proof,
                        user_email__iexact=customer_email,
                    ).first()
                except Exception:
                    verified_mkt_order = None

                if not verified_book_order and not verified_mkt_order:
                    messages.error(request, _('For your security, the email and order number do not match any order.'))
                    orders = None
                    mkt_orders = None
                else:
                    orders = models.Order.objects.filter(
                        customer_email__iexact=customer_email
                    ).order_by('-created_at')
                    try:
                        mkt_orders = MarketplaceOrder.objects.filter(
                            user_email__iexact=customer_email
                        ).order_by('-created_at')
                        if not mkt_orders.exists():
                            mkt_orders = None
                    except Exception:
                        mkt_orders = None
                    _accessible = request.session.get('accessible_orders', [])
                    for _o in (orders if orders.exists() else []):
                        if str(_o.order_number) not in _accessible:
                            _accessible.append(str(_o.order_number))
                    for _o in (mkt_orders or []):
                        if str(_o.order_number) not in _accessible:
                            _accessible.append(str(_o.order_number))
                    request.session['accessible_orders'] = _accessible
                    if not orders.exists():
                        orders = None
            else:
                messages.error(request, _('Enter both your email and one order number linked to that email.'))
                orders = None
                mkt_orders = None
        else:
            # Search by order number - return single order
            order_number = request.POST.get('order_number')
            if order_number:
                # Try book order first
                try:
                    order = models.Order.objects.get(order_number=order_number)
                    order.apply_ttl_rules()

                    # Grant session access for order found by order_number
                    _accessible = request.session.get('accessible_orders', [])
                    if str(order.order_number) not in _accessible:
                        _accessible.append(str(order.order_number))
                    request.session['accessible_orders'] = _accessible

                    # Check if any books have downloads available
                    has_downloadable_books = any(
                        item.book.has_download 
                        for item in order.orderitem_set.all()
                    )
                except models.Order.DoesNotExist:
                    # Try marketplace order
                    try:
                        mkt_order = MarketplaceOrder.objects.get(order_number=order_number)
                        mkt_order.apply_ttl_rules()
                    except MarketplaceOrder.DoesNotExist:
                        messages.error(request, '订单号不存在')
    
    # Desktop unified overview rows (better visibility for logged-in users).
    desktop_order_rows = []
    status_counts = {
        'all': 0,
        'pending': 0,
        'processing': 0,
        'shipped': 0,
        'delivered': 0,
        'cancelled': 0,
    }
    source_orders = all_orders if all_orders is not None else orders
    source_mkt_orders = all_mkt_orders if all_mkt_orders is not None else mkt_orders
    from manager.order_status import order_status_bucket, order_matches_bucket
    for o in (source_orders or []):
        bucket = order_status_bucket(o.status)
        if bucket in status_counts:
            status_counts[bucket] += 1
        status_counts['all'] += 1
        desktop_order_rows.append({
            'kind': 'book',
            'channel_label': '图书',
            'channel_icon': 'fas fa-book',
            'order_number': o.order_number,
            'created_at': o.created_at,
            'status': o.status,
            'status_display': o.get_status_display(),
            'payment_status': o.payment_status,
            'payment_status_display': o.get_payment_status_display(),
            'total_amount': o.total_amount,
            'item_count': o.orderitem_set.count(),
            'detail_url': f'/manager/order-confirmation/{o.order_number}/',
        })
    for o in (source_mkt_orders or []):
        bucket = order_status_bucket(o.status)
        if bucket in status_counts:
            status_counts[bucket] += 1
        status_counts['all'] += 1
        desktop_order_rows.append({
            'kind': 'marketplace',
            'channel_label': '市场',
            'channel_icon': 'fas fa-store',
            'order_number': o.order_number,
            'created_at': o.created_at,
            'status': o.status,
            'status_display': o.get_status_display(),
            'payment_status': o.payment_status,
            'payment_status_display': o.get_payment_status_display(),
            'total_amount': o.total_amount,
            'item_count': o.items.count(),
            'detail_url': f'/manager/order-confirmation/{o.order_number}/',
        })
    desktop_order_rows.sort(key=lambda x: x['created_at'], reverse=True)
    if status_filter:
        desktop_order_rows = [
            r for r in desktop_order_rows
            if order_matches_bucket(r['status'], status_filter)
        ]

    from . import views_review

    pending_review_items = (
        views_review.collect_pending_reviews_for_user(site_user) if site_user else []
    )

    context = {
        'order': order,
        'orders': orders,
        'mkt_order': mkt_order,
        'mkt_orders': mkt_orders,
        'has_downloadable_books': has_downloadable_books,
        'site_user': site_user,
        'desktop_order_rows': desktop_order_rows,
        'desktop_status_counts': status_counts,
        'status_filter': status_filter,
        'show_legacy_order_results': bool((orders or mkt_orders) and not desktop_order_rows),
        'pending_review_items': pending_review_items,
    }
    return render(request, 'public/track_order.html', context)


def download_book(request, order_id, book_id):
    """Download purchased ebook - supports files and external links"""
    import os

    # Verify order and book
    order = get_object_or_404(models.Order, id=order_id)
    book = get_object_or_404(models.Book, id=book_id)

    # Ownership check — order_number must be in session's accessible list
    accessible = request.session.get('accessible_orders', [])
    if str(order.order_number) not in accessible:
        messages.error(request, _('您没有权限下载此文件，请先通过订单查询验证您的身份'))
        return redirect('manager:track_order')
    
    # Check if order status allows download (shipped or delivered means payment is done)
    valid_statuses = ['paid', 'confirmed', 'processing', 'shipped', 'delivered']
    if order.status not in valid_statuses:
        messages.error(request, '订单尚未完成支付，无法下载')
        return redirect('manager:order_confirmation', order_number=order.order_number)
    
    # Additional check: if status is shipped/delivered, payment must be completed
    # (For backwards compatibility, we allow download if status is shipped regardless of payment_status)
    if order.status not in ['shipped', 'delivered'] and order.payment_status != 'completed':
        messages.error(request, '订单支付未完成，无法下载')
        return redirect('manager:order_confirmation', order_number=order.order_number)
    
    # Check if this book is in the order
    order_item = models.OrderItem.objects.filter(order=order, book=book).first()
    if not order_item:
        messages.error(request, '此订单不包含该图书')
        return redirect('manager:order_confirmation', order_number=order.order_number)
    
    # Check if book has download available
    if not book.has_download():
        messages.error(request, '该图书暂无下载文件')
        return redirect('manager:order_confirmation', order_number=order.order_number)
    
    # If book has a file, serve it directly
    if book.book_file:
        try:
            response = FileResponse(book.book_file.open('rb'))
            file_name = os.path.basename(book.book_file.name)
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            return response
        except Exception as e:
            messages.error(request, f'文件下载失败: {str(e)}')
            return redirect('manager:order_confirmation', order_number=order.order_number)
    
    # If book has external download link, redirect to it
    elif book.download_link:
        return redirect(book.download_link)
    
    # Fallback: Create a placeholder PDF with book information
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from io import BytesIO
        
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        
        # Add content to PDF
        p.setFont("Helvetica-Bold", 24)
        p.drawString(100, 750, f"{book.name}")
        
        p.setFont("Helvetica", 12)
        p.drawString(100, 720, f"Publisher: {book.publisher.publisher_name}")
        p.drawString(100, 700, f"Order Number: {order.order_number}")
        p.drawString(100, 680, f"Customer: {order.customer_name}")
        
        p.setFont("Helvetica-Oblique", 10)
        p.drawString(100, 650, "Thank you for your purchase!")
        p.drawString(100, 630, "This is a sample ebook file.")
        p.drawString(100, 610, "Please contact support to get the actual ebook file.")
        
        p.showPage()
        p.save()
        
        buffer.seek(0)
        
        response = FileResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{book.name}.pdf"'
        
        return response
        
    except ImportError:
        # If reportlab is not installed, return a simple text file
        content = f"""
        电子书信息
        ==========
        
        书名: {book.name}
        出版社: {book.publisher.publisher_name}
        订单号: {order.order_number}
        客户: {order.customer_name}
        
        感谢您的购买！
        
        注意: 这是一个示例文件。实际生产环境中，这里应该提供真实的电子书文件。
        """
        
        response = HttpResponse(content, content_type='text/plain; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{book.name}.txt"'
        return response


# ====================   API Views for Order Actions  ===========================
def api_cancel_order(request):
    """API endpoint to cancel an order"""
    from django.http import JsonResponse
    import json
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            order_number = data.get('order_number')
            
            order = models.Order.objects.get(order_number=order_number)
            
            # Only allow cancellation for pending or payment_pending orders
            if order.status in ['pending', 'payment_pending']:
                order.status = 'cancelled'
                order.save()
                return JsonResponse({'success': True, 'message': '订单已取消'})
            else:
                return JsonResponse({'success': False, 'message': '此订单状态无法取消'})
                
        except models.Order.DoesNotExist:
            return JsonResponse({'success': False, 'message': '订单不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': '无效请求'})


def api_confirm_payment(request):
    """API endpoint for the customer to self-report 'I've paid' on manual/QR
    payment methods with no real-time gateway webhook (WeChat Pay, Alipay,
    bank transfer) — status moves to 'processing' as a signal for ops to go
    verify manually; payment_status stays 'pending' until an admin actually
    confirms it, so this alone cannot mark an order paid or trigger escrow/
    shipment creation. It had no ownership check at all before this fix —
    anyone who could guess/enumerate an order_number could flip its status."""
    from django.http import JsonResponse
    import json

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            order_number = data.get('order_number')

            accessible = request.session.get('accessible_orders', [])
            if str(order_number) not in accessible:
                return JsonResponse({'success': False, 'message': '未授权'}, status=403)

            order = models.Order.objects.get(order_number=order_number)

            # Update order status to processing
            if order.status == 'payment_pending':
                order.status = 'processing'
                order.payment_status = 'pending'
                order.save()
                create_notification(
                    'order_paid',
                    f'订单 {order.order_number} 确认付款',
                    f'{order.customer_name} 确认了订单付款，金额 ¥{order.total_amount}',
                    icon='fas fa-credit-card',
                    color='#10b981',
                    link=f'/manager/order_detail/{order.id}/',
                    related_id=order.id,
                )
                return JsonResponse({'success': True, 'message': '支付确认成功'})
            else:
                return JsonResponse({'success': False, 'message': '订单状态无法确认支付'})
                
        except models.Order.DoesNotExist:
            return JsonResponse({'success': False, 'message': '订单不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': '无效请求'})


def api_home_feed(request):
    """Paginated feed API for infinite scroll on homepage.
    Returns mixed books + products, 12 per page."""
    try:
        page = max(1, int(request.GET.get('page', 1)))
    except ValueError:
        page = 1
    feed_type = request.GET.get('feed', 'mixed')
    PER_PAGE = 12
    offset = (page - 1) * PER_PAGE

    items = []
    feed_type = feed_type if feed_type in ('mixed', 'books', 'products', 'courses', 'supermarket') else 'mixed'

    # ── Personalized feed for logged-in shoppers ──────────────────────────
    # When the visitor has usable preference signals, rank the feed with the
    # recommendation engine instead of raw popularity. Falls through to the
    # popularity feed below for anonymous / signal-less visitors.
    user_id = request.session.get('site_user_id')
    if user_id:
        try:
            from manager import recommendations as _reco
            domain_map = {'mixed': 'mixed', 'books': 'books',
                          'products': 'marketplace', 'courses': 'marketplace',
                          'supermarket': 'marketplace'}
            if not request.session.session_key:
                request.session.save()
            skey = request.session.session_key or ''
            if feed_type in ('books', 'mixed'):
                domain = domain_map.get(feed_type, 'mixed')
            else:
                domain = feed_type  # single marketplace type handled by _domain_types? use per-type
            # For single marketplace types keep them isolated.
            if feed_type in ('products', 'courses', 'supermarket'):
                rec_items, has_more, has_sig = _reco.recommend(
                    user_id, skey, domain='marketplace', page=page, per_page=PER_PAGE)
                type_key = feed_type.rstrip('s') if feed_type != 'supermarket' else 'supermarket'
                rec_items = [it for it in rec_items if it['type'] == type_key] or rec_items
            else:
                rec_items, has_more, has_sig = _reco.recommend(
                    user_id, skey, domain=domain, page=page, per_page=PER_PAGE)
            if has_sig and rec_items:
                return JsonResponse({'items': rec_items, 'page': page,
                                     'has_more': has_more, 'personalized': True})
        except Exception:
            pass  # any issue → fall back to the popularity feed below

    def _slice_wrap(qs, count, limit):
        if not count or limit <= 0:
            return []
        safe_offset = offset % count
        rows = list(qs[safe_offset:safe_offset + limit])
        if len(rows) < limit:
            rows += list(qs[:limit - len(rows)])
        return rows

    sources = []
    # Books
    books_qs = models.Book.objects.filter(is_active=True).select_related('publisher').order_by('-sale_num')
    sources.append({
        'type': 'book',
        'label': '图书',
        'qs': books_qs,
        'count': books_qs.count(),
        'to_item': lambda b: {
            'type': 'book',
            'type_label': '图书',
            'name': b.name[:30],
            'price': str(b.price),
            'image': b.get_cover_url(),
            'url': f'/manager/public/books/{b.id}/',
        }
    })

    try:
        from marketplace.models import Product, Course, SupermarketItem
        p_qs = Product.objects.filter(is_active=True).order_by('-sales_count')
        c_qs = Course.objects.filter(is_active=True).order_by('-enrollment_count')
        s_qs = SupermarketItem.objects.filter(is_active=True).order_by('-sales_count')
        sources += [
            {
                'type': 'product',
                'label': '商品',
                'qs': p_qs,
                'count': p_qs.count(),
                'to_item': lambda p: {
                    'type': 'product',
                    'type_label': '商品',
                    'name': p.name[:30],
                    'price': str(p.price),
                    'image': p.get_image_url(),
                    'url': f'/marketplace/products/{p.slug}/',
                }
            },
            {
                'type': 'course',
                'label': '课程',
                'qs': c_qs,
                'count': c_qs.count(),
                'to_item': lambda c: {
                    'type': 'course',
                    'type_label': '课程',
                    'name': c.title[:30],
                    'price': str(c.price),
                    'image': c.get_image_url(),
                    'url': f'/marketplace/courses/{c.slug}/',
                }
            },
            {
                'type': 'supermarket',
                'label': '超市',
                'qs': s_qs,
                'count': s_qs.count(),
                'to_item': lambda s: {
                    'type': 'supermarket',
                    'type_label': '超市',
                    'name': s.name[:30],
                    'price': str(s.price),
                    'image': s.get_image_url(),
                    'url': f'/marketplace/supermarket/{s.slug}/',
                }
            },
        ]
    except Exception:
        pass

    if feed_type == 'mixed':
        active_sources = [s for s in sources if s['count'] > 0]
        if active_sources:
            each = max(1, PER_PAGE // len(active_sources))
            for src in active_sources:
                for row in _slice_wrap(src['qs'], src['count'], each):
                    items.append(src['to_item'](row))
            # Fill remainder from the largest source.
            remainder = PER_PAGE - len(items)
            if remainder > 0:
                largest = max(active_sources, key=lambda s: s['count'])
                for row in _slice_wrap(largest['qs'], largest['count'], remainder):
                    items.append(largest['to_item'](row))
    else:
        src = next((s for s in sources if s['type'] == feed_type.rstrip('s')), None)
        if src is None:
            src = next((s for s in sources if s['type'] == feed_type), None)
        if src:
            for row in _slice_wrap(src['qs'], src['count'], PER_PAGE):
                items.append(src['to_item'](row))

    return JsonResponse({'items': items, 'page': page, 'has_more': len(items) > 0})


def api_recommendations(request):
    """Personalized 'Recommended for you' rail for the marketplace and book
    catalog. Returns [] when the visitor has no usable preference signal so the
    caller can simply hide the rail (no awkward empty/duplicate section)."""
    user_id = request.session.get('site_user_id')
    domain = request.GET.get('domain', 'mixed')
    if domain not in ('mixed', 'books', 'marketplace'):
        domain = 'mixed'
    try:
        limit = min(24, max(4, int(request.GET.get('limit', 12))))
    except (TypeError, ValueError):
        limit = 12
    if not user_id:
        return JsonResponse({'items': [], 'personalized': False})
    try:
        from manager import recommendations as _reco
        if not request.session.session_key:
            request.session.save()
        skey = request.session.session_key or ''
        items, _has_more, has_sig = _reco.recommend(
            user_id, skey, domain=domain, page=1, per_page=limit, exclude_owned=True)
        if not has_sig:
            return JsonResponse({'items': [], 'personalized': False})
        # Only surface items that actually matched a preference (a genuine
        # recommendation), not popularity filler.
        rec = [it for it in items if it.get('recommended')][:limit]
        return JsonResponse({'items': rec, 'personalized': bool(rec)})
    except Exception:
        return JsonResponse({'items': [], 'personalized': False})


def api_unified_search(request):
    """Unified search across books, products, courses, and supermarket items."""
    query = request.GET.get('q', '').strip()
    if not query:
        return JsonResponse({'results': [], 'query': '', 'total': 0, 'suggestions': []})

    from marketplace.models import Product, Course, SupermarketItem, Category as MktCategory
    results = []
    MAX_PER_TYPE = 8

    books = models.Book.objects.filter(
        Q(name__icontains=query) | Q(description__icontains=query) |
        Q(publisher__publisher_name__icontains=query) |
        Q(category__name__icontains=query),
        is_active=True
    ).select_related('publisher', 'category')[:MAX_PER_TYPE]
    for b in books:
        results.append({
            'type': 'book', 'type_label': '图书',
            'name': b.name, 'price': str(b.price),
            'image': b.get_cover_url(),
            'url': f'/manager/public/books/{b.id}/',
            'desc': (b.description or '')[:80],
        })

    products = Product.objects.filter(
        Q(name__icontains=query) | Q(description__icontains=query) |
        Q(category__name__icontains=query),
        is_active=True
    ).select_related('category')[:MAX_PER_TYPE]
    for p in products:
        results.append({
            'type': 'product', 'type_label': '商品',
            'name': p.name, 'price': str(p.price),
            'image': p.get_image_url(),
            'url': f'/marketplace/products/{p.slug}/',
            'desc': (p.description or '')[:80],
        })

    courses = Course.objects.filter(
        Q(title__icontains=query) | Q(description__icontains=query),
        is_active=True
    )[:MAX_PER_TYPE]
    for c in courses:
        results.append({
            'type': 'course', 'type_label': '课程',
            'name': c.title, 'price': str(c.price),
            'image': c.get_image_url(),
            'url': f'/marketplace/courses/{c.slug}/',
            'desc': (c.description or '')[:80],
        })

    supermarket = SupermarketItem.objects.filter(
        Q(name__icontains=query) | Q(description__icontains=query) |
        Q(category__name__icontains=query),
        is_active=True
    ).select_related('category')[:MAX_PER_TYPE]
    for s in supermarket:
        results.append({
            'type': 'supermarket', 'type_label': '超市',
            'name': s.name, 'price': str(s.price),
            'image': s.get_image_url(),
            'url': f'/marketplace/supermarket/{s.slug}/',
            'desc': (s.description or '')[:80],
        })

    suggestions = []
    if not results:
        suggest_books = models.Book.objects.filter(is_active=True).order_by('-sale_num')[:4]
        suggest_products = Product.objects.filter(is_active=True).order_by('-sales_count')[:4]
        for b in suggest_books:
            suggestions.append({
                'type': 'book', 'type_label': '图书',
                'name': b.name, 'price': str(b.price),
                'image': b.get_cover_url(),
                'url': f'/manager/public/books/{b.id}/',
            })
        for p in suggest_products:
            suggestions.append({
                'type': 'product', 'type_label': '商品',
                'name': p.name, 'price': str(p.price),
                'image': p.get_image_url(),
                'url': f'/marketplace/products/{p.slug}/',
            })

    return JsonResponse({
        'results': results,
        'query': query,
        'total': len(results),
        'suggestions': suggestions,
    })


def public_search(request):
    """Full-page unified search results."""
    query = request.GET.get('q', '').strip()
    return render(request, 'public/search_results.html', {'query': query})


def api_spin_wheel(request):
    """Daily spin wheel — awards 10-100 random points once per day per user."""
    import random
    from datetime import date

    if not request.session.get('site_user_id'):
        return JsonResponse({'success': False, 'message': '请先登录', 'code': 'login_required'})

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '无效请求'})

    user_id = request.session['site_user_id']
    try:
        user = models.SiteUser.objects.get(pk=user_id)
        loyalty, created = models.LoyaltyPoints.objects.get_or_create(user=user)

        today = date.today()
        if loyalty.last_spin == today:
            return JsonResponse({'success': False, 'message': '今天已经转过了，明天再来！', 'code': 'already_spun'})

        pts = random.choice([10, 20, 30, 50, 80, 100])
        loyalty.points_balance += pts
        loyalty.lifetime_points += pts
        loyalty.last_spin = today
        loyalty.update_tier()
        loyalty.save()

        models.PointTransaction.objects.create(
            user=user, points=pts, reason='daily_spin',
            description=f'每日转盘获得 {pts} 积分'
        )
        return JsonResponse({
            'success': True,
            'points_won': pts,
            'new_balance': loyalty.points_balance,
            'tier': loyalty.get_tier_display(),
        })
    except models.SiteUser.DoesNotExist:
        return JsonResponse({'success': False, 'message': '用户不存在'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ====================   订单管理模块  ===========================
def order_list(request):
    """Admin order list view"""
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    payment_status_filter = request.GET.get('payment_status', '')
    search_query = request.GET.get('search', '')
    
    # Start with all orders
    orders = models.Order.objects.all().select_related().prefetch_related('orderitem_set__book')
    
    # Apply filters
    if status_filter:
        orders = orders.filter(status=status_filter)
    if payment_status_filter:
        orders = orders.filter(payment_status=payment_status_filter)
    if search_query:
        orders = orders.filter(
            Q(order_number__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(customer_email__icontains=search_query) |
            Q(customer_phone__icontains=search_query)
        )
      # Order by creation date (newest first)
    orders = orders.order_by('-created_at')
    
    # Get statistics
    # "待处理" means genuinely awaiting action — placed, paid or confirmed
    # but not yet shipped/delivered/cancelled/refunded. It used to check
    # only status='pending', a value that's essentially never set in
    # practice (orders actually land in 'payment_pending' first), so this
    # stat silently showed 0 even when orders needed attention.
    # "已完成" is aligned with the same definition the dashboard uses
    # (status='delivered') rather than payment_status='completed', which
    # conflated "we got paid" with "the order is done" — a paid order can
    # still be sitting unshipped.
    total_orders = models.Order.objects.count()
    pending_orders = models.Order.objects.filter(
        status__in=['pending', 'payment_pending', 'paid', 'confirmed']
    ).count()
    completed_orders = models.Order.objects.filter(status='delivered').count()
    total_revenue = models.Order.objects.filter(
        payment_status__in=['completed', 'pending']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    context = {
        'orders': orders,
        'name': request.session["name"],
        'status_choices': models.ORDER_STATUS_CHOICES,
        'payment_status_choices': models.PAYMENT_STATUS_CHOICES,
        'current_status_filter': status_filter,
        'current_payment_status_filter': payment_status_filter,
        'current_search': search_query,
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'completed_orders': completed_orders,
        'total_revenue': total_revenue,
    }
    
    return render(request, 'order/order_list.html', context)


def order_detail(request, order_id):
    """Admin order detail view"""
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    
    order = get_object_or_404(models.Order, id=order_id)
    order_items = models.OrderItem.objects.filter(order=order).select_related('book')
    
    context = {
        'order': order,
        'order_items': order_items,
        'name': request.session["name"],
        'status_choices': models.ORDER_STATUS_CHOICES,
        'payment_status_choices': models.PAYMENT_STATUS_CHOICES,
    }
    
    return render(request, 'order/order_detail.html', context)


@require_POST
def update_order_status(request):
    """Update order status via AJAX"""
    # 登录判断
    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '请先登录'})
    
    try:
        order_id = request.POST.get('order_id')
        new_status = request.POST.get('status')
        admin_notes = request.POST.get('admin_notes', '')
        
        order = get_object_or_404(models.Order, id=order_id)
        old_status = order.status

        # This endpoint only flips the status label — it has no inventory
        # logic. Cancelling/refunding an order that already had stock
        # deducted (inventory_applied=True, meaning payment was confirmed)
        # needs restore_inventory_for_shipment via the Returns & Shipments
        # queue instead, or the deducted stock is silently never restored.
        # Orders that were cancelled before payment ever completed
        # (inventory_applied=False) have nothing to restore, so those are
        # still fine to cancel directly here.
        if new_status in ('cancelled', 'refunded') and order.inventory_applied and old_status not in ('cancelled', 'refunded'):
            return JsonResponse({
                'success': False,
                'message': '该订单库存已扣减（已付款），请通过"退货与物流"处理退款以正确恢复库存，而不是直接修改订单状态。',
            })

        order.status = new_status
        if admin_notes:
            order.admin_notes = admin_notes
        order.save()

        # dict(...).get(...) rather than direct indexing — a handful of
        # legacy orders carry a status value that predates the current
        # ORDER_STATUS_CHOICES list (e.g. an old 'paid' payment_status), and
        # direct indexing raised KeyError there, which the broad except
        # below turned into a false "update failed" even though order.save()
        # above had already succeeded.
        status_labels = dict(models.ORDER_STATUS_CHOICES)
        return JsonResponse({
            'success': True,
            'message': f'订单状态已从 "{status_labels.get(old_status, old_status)}" 更新为 "{status_labels.get(new_status, new_status)}"',
            'new_status': new_status,
            'new_status_display': status_labels.get(new_status, new_status),
            'new_status_color': order.get_status_color()
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': '更新失败，请重试'})


@require_POST
def update_payment_status(request):
    """Update payment status via AJAX"""
    # 登录判断
    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '请先登录'})
    
    try:
        order_id = request.POST.get('order_id')
        new_payment_status = request.POST.get('payment_status')
        transaction_id = request.POST.get('transaction_id', '')
        
        order = get_object_or_404(models.Order, id=order_id)
        old_payment_status = order.payment_status

        if new_payment_status == 'completed':
            # Route through the shared pipeline (shipment creation,
            # confirmation email, inventory deduction) instead of just
            # flipping the field — this used to silently skip all of that
            # whenever an admin manually marked an order paid (e.g. cash on
            # delivery, bank transfer confirmed by hand).
            from manager.payments.views import _update_order_status
            _update_order_status(order, 'SUCCESSFUL', transaction_id=transaction_id or None)
        elif new_payment_status == 'refunded' and order.inventory_applied and old_payment_status != 'refunded':
            # Same gap as update_order_status above: flipping this field
            # directly has no inventory logic, so marking a paid order
            # "refunded" here would never give the stock back. Route through
            # Returns & Shipments instead, which calls
            # restore_inventory_for_shipment for the real reversal.
            return JsonResponse({
                'success': False,
                'message': '该订单库存已扣减，请通过"退货与物流"处理退款以正确恢复库存，而不是直接修改支付状态。',
            })
        else:
            order.payment_status = new_payment_status
            if transaction_id:
                order.payment_transaction_id = transaction_id
            order.save()

        # dict(...).get(...) rather than direct indexing — see the same note
        # in update_order_status: legacy orders can carry a status value
        # that predates the current choices lists, and direct indexing
        # raised KeyError there, turning a successful save into a false
        # "update failed" response.
        payment_labels = dict(models.PAYMENT_STATUS_CHOICES)
        order_labels = dict(models.ORDER_STATUS_CHOICES)
        return JsonResponse({
            'success': True,
            'message': f'支付状态已从 "{payment_labels.get(old_payment_status, old_payment_status)}" 更新为 "{payment_labels.get(new_payment_status, new_payment_status)}"',
            'new_payment_status': new_payment_status,
            'new_payment_status_display': payment_labels.get(new_payment_status, new_payment_status),
            'new_payment_status_color': order.get_payment_status_color(),
            'order_status': order.status,
            'order_status_display': order_labels.get(order.status, order.status),
            'order_status_color': order.get_status_color()
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': '更新失败，请重试'})

@require_http_methods(["GET"])
def export_orders(request):
    """Export orders to CSV or Excel format"""
    # 登录判断
    if "name" not in request.session:
        return redirect("/manager/login")
    
    export_format = request.GET.get('format', 'csv')  # csv or excel
    status_filter = request.GET.get('status', '')
    payment_status_filter = request.GET.get('payment_status', '')
    search_filter = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Build queryset with filters
    orders = models.Order.objects.all().order_by('-created_at')
    
    if status_filter:
        orders = orders.filter(status=status_filter)
    
    if payment_status_filter:
        orders = orders.filter(payment_status=status_filter)
    
    if search_filter:
        orders = orders.filter(
            Q(order_number__icontains=search_filter) |
            Q(customer_name__icontains=search_filter) |
            Q(customer_email__icontains=search_filter) |
            Q(customer_phone__icontains=search_filter)
        )
    
    if date_from:
        try:
            from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__gte=from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__lte=to_date)
        except ValueError:
            pass
    
    if export_format == 'excel':
        return export_orders_excel(orders)
    else:
        return export_orders_csv(orders)


def export_orders_csv(orders):
    """Export orders to CSV format"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="orders_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # Add BOM for UTF-8 to ensure proper encoding in Excel
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow([
        '订单号', '客户姓名', '客户邮箱', '客户电话', '收货地址', '城市', '省份',
        '支付方式', '订单状态', '支付状态', '总金额', '商品数量', '创建时间', '客户备注', '管理员备注'
    ])
    
    # Write data
    for order in orders:
        writer.writerow([
            order.order_number,
            order.customer_name,
            order.customer_email,
            order.customer_phone,
            order.shipping_address,
            order.shipping_city,
            order.shipping_state,
            order.get_payment_method_display(),
            order.get_status_display(),
            order.get_payment_status_display(),
            f'¥{order.total_amount}',
            order.orderitem_set.count(),
            order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            order.customer_notes or '',
            order.admin_notes or ''
        ])
    
    return response


def export_orders_excel(orders):
    """Export orders to Excel format"""
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "订单导出"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        '订单号', '客户姓名', '客户邮箱', '客户电话', '收货地址', '城市', '省份',
        '支付方式', '订单状态', '支付状态', '总金额', '商品数量', '创建时间', '客户备注', '管理员备注'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row_num, order in enumerate(orders, 2):
        data = [
            order.order_number,
            order.customer_name,
            order.customer_email,
            order.customer_phone,
            order.shipping_address,
            order.shipping_city,
            order.shipping_state,
            order.get_payment_method_display(),
            order.get_status_display(),
            order.get_payment_status_display(),
            float(order.total_amount),
            order.orderitem_set.count(),
            order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            order.customer_notes or '',
            order.admin_notes or ''
        ]
        
        for col_num, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="orders_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response


@require_http_methods(["GET"])
def order_analytics(request):
    """Get order analytics data"""
    if not request.session.get('is_login'):
        return JsonResponse({'error': '请先登录'}, status=401)
    
    # Date range for analytics (default: last 30 days)
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    # Get date range from request
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if date_from:
        try:
            start_date = timezone.datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if date_to:
        try:
            end_date = timezone.datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Analytics queries
    orders_in_range = models.Order.objects.filter(
        created_at__date__range=[start_date, end_date]
    )
    
    # Daily order counts
    daily_orders = []
    current_date = start_date
    while current_date <= end_date:
        count = orders_in_range.filter(created_at__date=current_date).count()
        revenue = orders_in_range.filter(
            created_at__date=current_date,
            payment_status='completed'
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        daily_orders.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'orders': count,
            'revenue': float(revenue)
        })
        current_date += timedelta(days=1)
    
    # Status distribution
    status_counts = {}
    for status, label in models.ORDER_STATUS_CHOICES:
        count = orders_in_range.filter(status=status).count()
        status_counts[label] = count
    
    # Payment status distribution
    payment_status_counts = {}
    for status, label in models.PAYMENT_STATUS_CHOICES:
        count = orders_in_range.filter(payment_status=status).count()
        payment_status_counts[label] = count
    
    # Top customers
    top_customers = orders_in_range.values('customer_name', 'customer_email').annotate(
        order_count=models.Count('id'),
        total_spent=Sum('total_amount')
    ).order_by('-total_spent')[:10]
    
    return JsonResponse({
        'success': True,
        'data': {
            'date_range': {
                'start': start_date.strftime('%Y-%m-%d'),
                'end': end_date.strftime('%Y-%m-%d')
            },
            'daily_orders': daily_orders,
            'status_distribution': status_counts,
            'payment_status_distribution': payment_status_counts,
            'top_customers': list(top_customers),
            'summary': {
                'total_orders': orders_in_range.count(),
                'total_revenue': float(orders_in_range.filter(payment_status='completed').aggregate(
                    total=Sum('total_amount'))['total'] or 0),
                'average_order_value': float(orders_in_range.aggregate(
                    avg=Avg('total_amount'))['avg'] or 0),
                'completion_rate': round(
                    (orders_in_range.filter(status='delivered').count() / 
                     max(orders_in_range.count(), 1)) * 100, 2
                )
            }
        }
    })

# ====================   MANAGER DASHBOARD  ===========================
def _dashboard_context(request):
    """
    Shared context builder for both dashboards:
    - manager_dashboard: unified (books + marketplace), the post-login landing page.
    - book_dashboard: books-only, nested under the Books sub-panel (mirrors
      Marketplace's own dashboard being the first item in its sub-panel nav).
    Returns None if the caller should redirect to login.
    """
    if "name" not in request.session:
        return None
    
    from datetime import datetime, timedelta
    from django.utils import timezone
    from django.db.models import Sum, Count, Q
    import json
    
    # Get current date ranges
    now = timezone.now()
    current_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month = (current_month - timedelta(days=1)).replace(day=1)
    last_7_days = now.date() - timedelta(days=7)
    
    # ==== BASIC STATISTICS ====
    total_books = models.Book.objects.count()
    total_publishers = models.Publisher.objects.count()
    total_authors = models.Author.objects.count()
    
    # Books added this month - Real dynamic calculation
    new_books_this_month = models.Book.objects.filter(
        created_at__gte=current_month
    ).count() if hasattr(models.Book, 'created_at') else 0
    
    # ==== ORDER STATISTICS ====
    try:
        total_orders = models.Order.objects.count()
        orders_this_month = models.Order.objects.filter(created_at__gte=current_month).count()

        # Full status breakdown for the order-status chart — grouped by
        # actual status present in the data rather than a hardcoded subset
        # of statuses, so every order is represented no matter what status
        # it's in (an earlier version only counted payment_pending/
        # processing/shipped/delivered/cancelled, silently dropping any
        # order sitting in pending, paid, confirmed or refunded).
        _status_rows = list(
            models.Order.objects.values('status').annotate(c=Count('id')).order_by('-c')
        )
        _status_display = dict(models.ORDER_STATUS_CHOICES)
        order_status_chart = {
            'labels': [_status_display.get(r['status'], r['status']) for r in _status_rows],
            'data': [r['c'] for r in _status_rows],
        }
        if not order_status_chart['labels']:
            order_status_chart = {'labels': ['暂无数据'], 'data': [1]}

        # Revenue calculations
        total_revenue = models.Order.objects.filter(
            payment_status__in=['completed', 'pending']
        ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        revenue_this_month = models.Order.objects.filter(
            created_at__gte=current_month,
            payment_status__in=['completed', 'pending']
        ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
    except Exception as e:
        total_orders = 0
        orders_this_month = 0
        total_revenue = 0
        revenue_this_month = 0
        order_status_chart = {'labels': ['暂无数据'], 'data': [1]}

    # ==== INVENTORY STATISTICS ====
    low_inventory_books = models.Book.objects.filter(inventory__lt=10).count()
    total_inventory = models.Book.objects.aggregate(Sum('inventory'))['inventory__sum'] or 0
    total_sales = models.Book.objects.aggregate(Sum('sale_num'))['sale_num__sum'] or 0
    
    # Top selling books
    top_books = models.Book.objects.order_by('-sale_num')[:5]
    
    # Recent orders (if available)
    try:
        recent_orders = models.Order.objects.order_by('-created_at')[:5]
    except:
        recent_orders = []
    
    # ==== CHART DATA ==== 
    # Daily sales for the last 7 days
    daily_sales = []
    for i in range(6, -1, -1):  # Reverse order to show oldest to newest
        date = (now.date() - timedelta(days=i))
        try:
            day_orders = models.Order.objects.filter(
                created_at__date=date
            ).count()
            day_revenue = models.Order.objects.filter(
                created_at__date=date,
                payment_status__in=['completed', 'pending']
            ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        except:
            day_orders = 0
            day_revenue = 0
        
        daily_sales.append({
            'date': date.strftime('%m-%d'),
            'orders': day_orders,
            'revenue': float(day_revenue)
        })
    
    # Publisher distribution
    publisher_stats = []
    for publisher in models.Publisher.objects.all()[:5]:
        book_count = models.Book.objects.filter(publisher=publisher).count()
        publisher_stats.append({
            'name': publisher.publisher_name,
            'books': book_count
        })
    
    # Add default data if no publishers
    if not publisher_stats:
        publisher_stats = [{'name': '暂无数据', 'books': 1}]

    # ==== ADDITIONAL CHART DATA ====
    # Top 5 books: sales vs inventory comparison
    top_books_comparison = []
    for book in models.Book.objects.order_by('-sale_num')[:5]:
        top_books_comparison.append({
            'name': book.name[:12],
            'sales': book.sale_num,
            'inventory': book.inventory,
        })
    if not top_books_comparison:
        top_books_comparison = [{'name': '暂无数据', 'sales': 0, 'inventory': 0}]

    # Price range distribution
    price_ranges = {
        '0-20': models.Book.objects.filter(price__lte=20).count(),
        '20-50': models.Book.objects.filter(price__gt=20, price__lte=50).count(),
        '50-100': models.Book.objects.filter(price__gt=50, price__lte=100).count(),
        '100-200': models.Book.objects.filter(price__gt=100, price__lte=200).count(),
        '200+': models.Book.objects.filter(price__gt=200).count(),
    }
    price_distribution = [{'range': k, 'count': v} for k, v in price_ranges.items()]

    # Monthly order trend (last 6 months)
    monthly_orders = []
    for i in range(5, -1, -1):
        month_start = (now.replace(day=1) - timedelta(days=i * 30)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if i > 0:
            next_month = (month_start + timedelta(days=32)).replace(day=1)
        else:
            next_month = now
        try:
            month_count = models.Order.objects.filter(
                created_at__gte=month_start, created_at__lt=next_month
            ).count()
            month_rev = float(models.Order.objects.filter(
                created_at__gte=month_start, created_at__lt=next_month,
                payment_status__in=['completed', 'pending']
            ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0)
        except Exception:
            month_count = 0
            month_rev = 0
        monthly_orders.append({
            'month': month_start.strftime('%Y-%m'),
            'orders': month_count,
            'revenue': month_rev,
        })

    # ==== MARKETPLACE STATS ====
    # Unified dashboard (task: merge Book Dashboard + Marketplace Dashboard) —
    # pulls in the same counts/chart payload the standalone marketplace admin
    # dashboard uses (marketplace/views.py:admin_dashboard), so both domains
    # render with identical chart types side by side instead of duplicating
    # the chart-building logic here.
    total_products = 0
    total_courses = 0
    total_supermarket = 0
    mkt_order_count = 0
    mkt_revenue = 0
    mkt_low_stock = 0
    mkt_recent_orders = []
    mkt_chart_payload = {}
    try:
        from marketplace.models import Product, Course, SupermarketItem, MarketplaceOrder as MktOrder
        from marketplace.views import _marketplace_admin_chart_payload
        total_products = Product.objects.filter(is_active=True).count()
        total_courses = Course.objects.filter(is_active=True).count()
        total_supermarket = SupermarketItem.objects.filter(is_active=True).count()
        mkt_low_stock = Product.objects.filter(is_active=True, stock__lt=10).count()
        _mkt_qs = MktOrder.objects.all()
        mkt_order_count = _mkt_qs.count()
        mkt_revenue = float(_mkt_qs.filter(payment_status='completed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0)
        mkt_recent_orders = _mkt_qs.order_by('-created_at')[:5]
        mkt_chart_payload = _marketplace_admin_chart_payload(total_products, total_courses, total_supermarket)
    except Exception:
        pass

    # Combined KPIs — top row of the unified dashboard (books + marketplace).
    combined_catalog_count = total_books + total_products + total_courses + total_supermarket
    combined_order_count = total_orders + mkt_order_count
    combined_revenue = float(total_revenue) + mkt_revenue
    combined_low_stock = low_inventory_books + mkt_low_stock

    context = {
        'name': request.session["name"],
        'current_date': f'{now.year}年{now.month:02d}月{now.day:02d}日',
        
        # Basic stats
        'total_books': total_books,
        'total_publishers': total_publishers,
        'total_authors': total_authors,
        'new_books_this_month': new_books_this_month,

        # Order stats — the full per-status breakdown lives in
        # order_status_chart below (covers every status, not just the 5
        # historically hardcoded here), so no unused per-bucket counts here.
        'total_orders': total_orders,
        'orders_this_month': orders_this_month,
        'total_revenue': total_revenue,
        'revenue_this_month': revenue_this_month,
        
        # Inventory stats
        'low_inventory_books': low_inventory_books,
        'total_inventory': total_inventory,
        'total_sales': total_sales,
        
        # Lists
        'top_books': top_books,
        'recent_orders': recent_orders,
        'recent_activities': [],  # Default empty list
        
        # Marketplace stats
        'total_products': total_products,
        'total_courses': total_courses,
        'total_supermarket': total_supermarket,
        'mkt_order_count': mkt_order_count,
        'mkt_revenue': mkt_revenue,
        'mkt_recent_orders': mkt_recent_orders,

        # Combined KPIs (unified dashboard top row)
        'combined_catalog_count': combined_catalog_count,
        'combined_order_count': combined_order_count,
        'combined_revenue': combined_revenue,
        'combined_low_stock': combined_low_stock,

        # Chart data (as JSON)
        'daily_sales_json': json.dumps(daily_sales),
        'order_status_json': json.dumps(order_status_chart),
        'publisher_stats_json': json.dumps(publisher_stats),
        'mkt_daily_sales_json': mkt_chart_payload.get('mkt_daily_sales_json', '[]'),
        'mkt_order_status_json': mkt_chart_payload.get('mkt_order_status_json', '{}'),
        'mkt_catalog_json': mkt_chart_payload.get('mkt_catalog_json', '{}'),
        'top_books_comparison_json': json.dumps(top_books_comparison),
        'price_distribution_json': json.dumps(price_distribution),
        'monthly_orders_json': json.dumps(monthly_orders),
    }

    return context


def manager_dashboard(request):
    """Unified dashboard (books + marketplace) — the post-login landing page."""
    context = _dashboard_context(request)
    if context is None:
        return redirect("/manager/login")
    return render(request, 'manager/dashboard.html', context)


def book_dashboard(request):
    """Books-only dashboard, nested under the Books sub-panel — same data as
    manager_dashboard, rendered without the marketplace-domain widgets."""
    context = _dashboard_context(request)
    if context is None:
        return redirect("/manager/login")
    return render(request, 'manager/book_dashboard.html', context)


@require_http_methods(["GET"])
def dashboard_analytics_api(request):
    """API endpoint for dashboard analytics data"""
    if "name" not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    # Return analytics data for AJAX requests
    return JsonResponse({
        'success': True,
        'data': {
            'total_books': models.Book.objects.count(),
            'total_orders': models.Order.objects.count() if hasattr(models, 'Order') else 0,
            'low_inventory': models.Book.objects.filter(inventory__lt=10).count(),
        }
    })

@require_POST
def delete_order(request, order_id):
    """Delete an order via AJAX"""
    # 登录判断
    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '请先登录'})
    
    try:
        order = get_object_or_404(models.Order, id=order_id)
        
        # Check if order can be deleted (business rules)
        if order.status == 'delivered':
            return JsonResponse({
                'success': False, 
                'message': '已送达的订单不能删除'
            })
        
        if order.payment_status == 'completed':
            return JsonResponse({
                'success': False, 
                'message': '已完成支付的订单不能删除，请先处理退款'
            })
        
        # Store order info for response
        order_number = order.order_number
        customer_name = order.customer_name
        
        # Delete the order (cascade will delete related OrderItems)
        order.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'订单 {order_number} (客户: {customer_name}) 已成功删除'
        })
        
    except models.Order.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'message': '订单不存在'
        })
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'删除失败：{str(e)}'
        })


# ====================   Public Static Pages  ===========================
def _build_info_page_context(page_meta, extra=None):
    """Build template context for modern info/legal pages."""
    from django.urls import reverse

    ctx = {
        'page_title': page_meta['title'],
        'page_subtitle': page_meta.get('subtitle', ''),
        'page_icon': page_meta.get('icon', 'fa-circle-info'),
        'page_gradient': page_meta.get('gradient', ('#667eea', '#764ba2')),
        'page_sections': page_meta.get('sections', []),
        'page_breadcrumb': page_meta.get('breadcrumb', []),
        'content_template': page_meta.get('content_template'),
        'page_highlights': page_meta.get('highlights', []),
    }
    cta = page_meta.get('cta')
    if cta:
        url = cta.get('url')
        if not url and cta.get('url_name'):
            url = reverse(cta['url_name'], kwargs=cta.get('url_kwargs') or None)
        ctx['page_cta'] = {
            'title': cta['title'],
            'text': cta['text'],
            'button': cta['button'],
            'url': url,
            'icon': cta.get('icon', 'fa-arrow-right'),
        }
    if extra:
        ctx.update(extra)
    return ctx


def public_legal_privacy(request):
    """Politique de confidentialité + CGU (document légal regroupé)."""
    from django.urls import reverse
    from django.utils.translation import gettext_lazy as _
    from manager.info_pages import LEGAL_PAGE

    meta = dict(LEGAL_PAGE)
    meta['breadcrumb'] = [
        {'label': _('首页'), 'url': reverse('manager:public_home')},
        {'label': LEGAL_PAGE['title'], 'url': ''},
    ]
    return render(request, 'public/pages/legal.html', _build_info_page_context(meta))


def public_legal_terms(request):
    """Redirect to CGU section on the legal page."""
    from django.shortcuts import redirect
    from django.urls import reverse
    return redirect(reverse('manager:legal_privacy') + '#conditions-utilisation')


def public_info_page(request, slug):
    """Dynamic info pages (refund, shipping, partner, premium, support)."""
    from django.http import Http404
    from django.urls import reverse
    from django.utils.translation import gettext_lazy as _
    from manager.info_pages import get_info_page

    page = get_info_page(slug)
    if not page:
        raise Http404
    meta = dict(page)
    meta['breadcrumb'] = [
        {'label': _('首页'), 'url': reverse('manager:public_home')},
        {'label': page['title'], 'url': ''},
    ]
    ctx = _build_info_page_context(meta)
    return render(request, 'public/pages/info_page_content.html', ctx)


def public_site_map(request):
    """Interactive site map."""
    from django.urls import reverse
    from django.utils.translation import gettext_lazy as _
    from manager.info_pages import get_sitemap_sections

    sections = []
    total_links = 0
    for group in get_sitemap_sections():
        links = []
        for link in group['links']:
            url_name = link['url_name']
            kwargs = link.get('url_kwargs')
            links.append({
                'label': link['label'],
                'url': reverse(url_name, kwargs=kwargs) if kwargs else reverse(url_name),
            })
        total_links += len(links)
        sections.append({
            'slug': group.get('slug', group['title']),
            'title': group['title'],
            'icon': group['icon'],
            'color': group.get('color', '#667eea'),
            'links': links,
            'link_count': len(links),
        })
    return render(request, 'public/pages/sitemap.html', {
        'sitemap_sections': sections,
        'total_links': total_links,
        'section_count': len(sections),
        'page_breadcrumb': [
            {'label': _('首页'), 'url': reverse('manager:public_home')},
            {'label': _('网站地图'), 'url': ''},
        ],
    })


def public_about(request):
    """About Us page — content from PAGES DUNO 360."""
    from django.urls import reverse
    from django.utils.translation import gettext_lazy as _
    from manager.info_pages import ABOUT_PAGE

    meta = dict(ABOUT_PAGE)
    meta['breadcrumb'] = [
        {'label': _('首页'), 'url': reverse('manager:public_home')},
        {'label': ABOUT_PAGE['title'], 'url': ''},
    ]
    meta['cta'] = {
        'title': _('加入 DUNO 360'),
        'text': _('创建账户或开设您的店铺。'),
        'button': _('创建账户'),
        'url_name': 'manager:user_register',
        'icon': 'fa-user-plus',
    }
    ctx = _build_info_page_context(meta)
    ctx['book_count'] = models.Book.objects.count()
    ctx['author_count'] = models.Author.objects.count()
    ctx['publisher_count'] = models.Publisher.objects.count()
    return render(request, 'public/pages/info_page_content.html', ctx)


def public_services(request):
    """Services page"""
    return render(request, 'public/services.html')


@csrf_protect
def public_contact(request):
    """Contact Us page with email sending via configured EmailAccount(s)"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()

        if not name or not email or not message:
            return JsonResponse({'success': False, 'error': 'Please fill in all required fields.'}, status=400)

        # Save to database first (never lose a message)
        contact_msg = models.ContactMessage.objects.create(
            name=name, email=email, subject=subject, message=message
        )

        # Create admin notification for contact form message
        create_notification(
            'contact_message',
            f'新联系消息 - {name}',
            f'{subject or "无主题"}: {message[:80]}',
            icon='fas fa-envelope-open-text',
            color='#10b981',
            link=f'/manager/email/?folder=contact',
            related_id=contact_msg.id,
        )

        # Build notification email content
        email_subject = f'[Contact Form] {subject or "No Subject"} - from {name}'
        email_body = (
            f'New message from the contact form\n'
            f'{"-" * 40}\n\n'
            f'Name:    {name}\n'
            f'Email:   {email}\n'
            f'Subject: {subject or "N/A"}\n\n'
            f'Message:\n{message}\n\n'
            f'{"-" * 40}\n'
            f'Sent from DUNO 360 contact form\n'
        )

        # Notify platform inbox via Zoho SMTP (admin@duno360.com)
        sent_ok = False
        try:
            from django.core.mail import EmailMessage
            msg = EmailMessage(
                subject=email_subject,
                body=email_body,
                from_email=django_settings.DEFAULT_FROM_EMAIL,
                to=[django_settings.CONTACT_EMAIL],
                reply_to=[email] if email else None,
            )
            msg.send(fail_silently=False)
            sent_ok = True
        except Exception as e:
            logger.warning(f'Contact form email failed (msg #{contact_msg.id}): {e}')

        if sent_ok:
            contact_msg.email_sent = True
            contact_msg.save(update_fields=['email_sent'])

        return JsonResponse({'success': True})

    return render(request, 'public/contact.html')


# ====================   Public Blog  ===========================
def public_blog(request):
    """Blog listing page"""
    search = request.GET.get('search', request.GET.get('q', '')).strip()
    category_slug = request.GET.get('category', '').strip()

    posts = models.BlogPost.objects.filter(status='published').select_related('category')

    if search:
        posts = posts.filter(
            Q(title__icontains=search)
            | Q(excerpt__icontains=search)
            | Q(content__icontains=search)
            | Q(author_name__icontains=search)
            | Q(category__name__icontains=search)
        )

    if category_slug:
        posts = posts.filter(category__slug=category_slug)

    categories = models.BlogCategory.objects.annotate(
        post_count=Count('posts', filter=Q(posts__status='published'))
    )
    featured_posts = models.BlogPost.objects.filter(
        status='published', is_featured=True
    ).select_related('category')[:3]
    total_published = models.BlogPost.objects.filter(status='published').count()

    context = {
        'posts': posts,
        'categories': categories,
        'featured_posts': featured_posts,
        'search_query': search,
        'current_category': category_slug,
        'total_posts_count': total_published,
        'results_count': posts.count(),
    }
    return render(request, 'public/blog.html', context)


def public_blog_detail(request, slug):
    """Blog post detail page"""
    post = get_object_or_404(models.BlogPost, slug=slug, status='published')
    # Increment views
    models.BlogPost.objects.filter(pk=post.pk).update(views_count=F('views_count') + 1)
    post.refresh_from_db()

    related_posts = models.BlogPost.objects.filter(
        status='published', category=post.category
    ).exclude(id=post.id)[:3] if post.category else models.BlogPost.objects.filter(
        status='published'
    ).exclude(id=post.id)[:3]

    context = {
        'post': post,
        'related_posts': related_posts,
    }
    return render(request, 'public/blog_detail.html', context)


# ====================   Admin Blog Management  ===========================
def blog_list(request):
    """Admin blog post list"""
    if "name" not in request.session:
        return redirect('/manager/login/')

    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')

    posts = models.BlogPost.objects.select_related('category').all()

    if search:
        posts = posts.filter(Q(title__icontains=search) | Q(content__icontains=search))
    if status_filter:
        posts = posts.filter(status=status_filter)

    categories = models.BlogCategory.objects.all()

    context = {
        'posts': posts,
        'categories': categories,
        'search_query': search,
        'status_filter': status_filter,
    }
    return render(request, 'blog/blog_list.html', context)


def add_blog_post(request):
    """Admin add blog post"""
    if "name" not in request.session:
        return redirect('/manager/login/')

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        content = request.POST.get('content', '').strip()
        excerpt = request.POST.get('excerpt', '').strip()
        category_id = request.POST.get('category')
        author_name = request.POST.get('author_name', 'Admin').strip()
        status = request.POST.get('status', 'draft')
        is_featured = request.POST.get('is_featured') == 'on'

        if not title or not content:
            messages.error(request, '标题和内容不能为空')
            return redirect('/manager/add_blog/')

        # Generate unique slug
        base_slug = slugify(title, allow_unicode=True)
        if not base_slug:
            base_slug = hashlib.md5(title.encode()).hexdigest()[:12]
        slug = base_slug
        counter = 1
        while models.BlogPost.objects.filter(slug=slug).exists():
            slug = f"{base_slug}-{counter}"
            counter += 1

        post = models.BlogPost(
            slug=slug,
            author_name=author_name,
            status=status,
            is_featured=is_featured,
        )
        # title/content/excerpt are django-modeltranslation fields —
        # constructor kwargs silently drop them, so assign as plain
        # attributes instead.
        post.title = title
        post.content = content
        post.excerpt = excerpt

        if category_id:
            try:
                post.category = models.BlogCategory.objects.get(id=category_id)
            except models.BlogCategory.DoesNotExist:
                pass

        if 'cover_image' in request.FILES:
            post.cover_image = request.FILES['cover_image']

        if status == 'published':
            post.published_at = timezone.now()

        post.save()
        messages.success(request, f'文章 "{title}" 创建成功！')
        return redirect('/manager/blog_list/')

    categories = models.BlogCategory.objects.all()
    return render(request, 'blog/add_blog.html', {'categories': categories})


def edit_blog_post(request):
    """Admin edit blog post"""
    if "name" not in request.session:
        return redirect('/manager/login/')

    if request.method == 'GET':
        post_id = request.GET.get('id')
        post = get_object_or_404(models.BlogPost, id=post_id)
        categories = models.BlogCategory.objects.all()
        return render(request, 'blog/edit_blog.html', {'post': post, 'categories': categories})

    if request.method == 'POST':
        post_id = request.POST.get('id')
        post = get_object_or_404(models.BlogPost, id=post_id)

        post.title = request.POST.get('title', '').strip()
        post.content = request.POST.get('content', '').strip()
        post.excerpt = request.POST.get('excerpt', '').strip()
        post.author_name = request.POST.get('author_name', 'Admin').strip()
        post.is_featured = request.POST.get('is_featured') == 'on'

        new_status = request.POST.get('status', 'draft')
        if new_status == 'published' and post.status != 'published':
            post.published_at = timezone.now()
        post.status = new_status

        category_id = request.POST.get('category')
        if category_id:
            try:
                post.category = models.BlogCategory.objects.get(id=category_id)
            except models.BlogCategory.DoesNotExist:
                post.category = None
        else:
            post.category = None

        if 'cover_image' in request.FILES:
            post.cover_image = request.FILES['cover_image']

        post.save()
        messages.success(request, f'文章 "{post.title}" 更新成功！')
        return redirect('/manager/blog_list/')


def delete_blog_post(request):
    """Admin delete blog post"""
    if "name" not in request.session:
        return redirect('/manager/login/')

    post_id = request.GET.get('id')
    post = get_object_or_404(models.BlogPost, id=post_id)
    title = post.title
    post.delete()
    messages.success(request, f'文章 "{title}" 已删除')
    return redirect('/manager/blog_list/')


def manage_blog_categories(request):
    """Admin blog category management"""
    if "name" not in request.session:
        return redirect('/manager/login/')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add':
            name = request.POST.get('name', '').strip()
            description = request.POST.get('description', '').strip()
            icon = request.POST.get('icon', 'fas fa-folder').strip()
            if name:
                base_slug = slugify(name, allow_unicode=True)
                if not base_slug:
                    base_slug = hashlib.md5(name.encode()).hexdigest()[:12]
                slug = base_slug
                counter = 1
                while models.BlogCategory.objects.filter(slug=slug).exists():
                    slug = f"{base_slug}-{counter}"
                    counter += 1
                models.BlogCategory.objects.create(
                    name=name, slug=slug, description=description, icon=icon
                )
                messages.success(request, f'分类 "{name}" 创建成功！')

        elif action == 'delete':
            cat_id = request.POST.get('category_id')
            try:
                cat = models.BlogCategory.objects.get(id=cat_id)
                name = cat.name
                cat.delete()
                messages.success(request, f'分类 "{name}" 已删除')
            except models.BlogCategory.DoesNotExist:
                messages.error(request, '分类不存在')

        return redirect('/manager/blog_categories/')

    categories = models.BlogCategory.objects.annotate(
        post_count=Count('posts')
    )
    return render(request, 'blog/blog_categories.html', {'categories': categories})


def manage_book_categories(request):
    """Admin book category management used by public mobile and desktop filters."""
    if "name" not in request.session:
        return redirect('/manager/login/')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add':
            name = request.POST.get('name', '').strip()
            name_en = request.POST.get('name_en', '').strip()
            name_fr = request.POST.get('name_fr', '').strip()
            description = request.POST.get('description', '').strip()
            icon = request.POST.get('icon', 'fas fa-book').strip() or 'fas fa-book'
            color = request.POST.get('color', '#667eea').strip() or '#667eea'
            display_order = request.POST.get('display_order') or 0
            if name:
                base_slug = slugify(request.POST.get('slug', '').strip() or name, allow_unicode=True)
                if not base_slug:
                    base_slug = hashlib.md5(name.encode()).hexdigest()[:12]
                slug = base_slug
                counter = 1
                while models.BookCategory.objects.filter(slug=slug).exists():
                    slug = f"{base_slug}-{counter}"
                    counter += 1
                models.BookCategory.objects.create(
                    name=name,
                    name_en=name_en,
                    name_fr=name_fr,
                    slug=slug,
                    description=description,
                    icon=icon,
                    color=color,
                    display_order=display_order,
                    is_active=bool(request.POST.get('is_active', 'on')),
                )
                messages.success(request, f'图书分类 "{name}" 创建成功！')
            else:
                messages.error(request, '分类名称不能为空')

        elif action == 'edit':
            cat_id = request.POST.get('category_id')
            name = request.POST.get('name', '').strip()
            try:
                cat = models.BookCategory.objects.get(id=cat_id, vendor__isnull=True)
            except models.BookCategory.DoesNotExist:
                messages.error(request, '分类不存在')
                return redirect('/manager/book_categories/')
            if not name:
                messages.error(request, '分类名称不能为空')
                return redirect('/manager/book_categories/')
            cat.name = name
            cat.name_en = request.POST.get('name_en', '').strip()
            cat.name_fr = request.POST.get('name_fr', '').strip()
            cat.description = request.POST.get('description', '').strip()
            cat.icon = request.POST.get('icon', '').strip() or 'fas fa-book'
            cat.color = request.POST.get('color', '').strip() or '#667eea'
            cat.display_order = request.POST.get('display_order') or 0
            new_slug = slugify(request.POST.get('slug', '').strip() or name, allow_unicode=True)
            if not new_slug:
                new_slug = hashlib.md5(name.encode()).hexdigest()[:12]
            if new_slug != cat.slug:
                slug = new_slug
                counter = 1
                while models.BookCategory.objects.filter(slug=slug).exclude(pk=cat.pk).exists():
                    slug = f"{new_slug}-{counter}"
                    counter += 1
                cat.slug = slug
            cat.save()
            messages.success(request, f'图书分类 "{cat.name}" 更新成功！')

        elif action == 'delete':
            cat_id = request.POST.get('category_id')
            try:
                cat = models.BookCategory.objects.get(id=cat_id, vendor__isnull=True)
                name = cat.name
                cat.delete()
                messages.success(request, f'图书分类 "{name}" 已删除')
            except models.BookCategory.DoesNotExist:
                messages.error(request, '分类不存在')

        elif action == 'toggle':
            cat_id = request.POST.get('category_id')
            try:
                cat = models.BookCategory.objects.get(id=cat_id, vendor__isnull=True)
                cat.is_active = not cat.is_active
                cat.save(update_fields=['is_active'])
                messages.success(request, f'图书分类 "{cat.name}" 状态已更新')
            except models.BookCategory.DoesNotExist:
                messages.error(request, '分类不存在')

        return redirect('/manager/book_categories/')

    # Admin manages only admin/global categories (vendor=NULL) — vendor-owned
    # ones are managed independently from the vendor panel.
    categories = models.BookCategory.objects.filter(parent__isnull=True, vendor__isnull=True).annotate(
        book_count=Count('books')
    )
    return render(request, 'book/book_categories.html', {'categories': categories, 'name': request.session["name"]})


# ====================   Admin Reviews (platform-wide, grouped by vendor)  ===========================

def admin_reviews(request):
    """
    Platform-wide buyer reviews, grouped/filterable by vendor. Moved here
    from marketplace's own admin sub-panel (was marketplace:admin_post_reviews)
    so it's reachable from the main admin panel rather than buried inside
    the marketplace-only area — reviews span books too, not just marketplace
    items, so it doesn't really belong scoped under "Marketplace".
    """
    if "name" not in request.session:
        return redirect('/manager/login/')

    # Bulk vendor resolution — one query per listing kind instead of a
    # query per review. Products/Courses/SupermarketItem each belong to
    # exactly one vendor; a Book can be cross-listed by more than one
    # vendor via VendorBook, so it maps to a list (same ambiguity already
    # accepted by vendor_post_reviews' reverse lookup — a book review is
    # visible to every vendor who lists that book).
    product_vendor = dict(Product.objects.values_list('id', 'vendor_id'))
    course_vendor = dict(Course.objects.values_list('id', 'vendor_id'))
    supermarket_vendor = dict(SupermarketItem.objects.values_list('id', 'vendor_id'))
    book_vendors = {}
    for book_id, vendor_id in models.VendorBook.objects.filter(is_active=True).values_list('book_id', 'vendor_id'):
        book_vendors.setdefault(book_id, []).append(vendor_id)

    def vendor_ids_for(kind, listing_id):
        if kind == 'product':
            vid = product_vendor.get(listing_id)
            return [vid] if vid else []
        if kind == 'course':
            vid = course_vendor.get(listing_id)
            return [vid] if vid else []
        if kind == 'supermarket':
            vid = supermarket_vendor.get(listing_id)
            return [vid] if vid else []
        if kind == 'book':
            return book_vendors.get(listing_id, [])
        return []

    all_reviews = list(
        PostDeliveryReview.objects.select_related('site_user').order_by('-created_at')[:500]
    )
    vendor_ids_needed = set()
    vendor_review_counts = {}
    for review in all_reviews:
        ids = vendor_ids_for(review.listing_kind, review.listing_id)
        review._vendor_ids = ids
        vendor_ids_needed.update(ids)
        for vid in ids:
            vendor_review_counts[vid] = vendor_review_counts.get(vid, 0) + 1

    vendors_by_id = {
        v.id: v for v in models.Vendor.objects.filter(id__in=vendor_ids_needed).only('id', 'company_name')
    }
    for review in all_reviews:
        review.vendors = [vendors_by_id[vid] for vid in review._vendor_ids if vid in vendors_by_id]

    # Vendor filter dropdown — only vendors that actually have at least
    # one review right now, with a per-vendor count for context.
    vendor_options = sorted(
        (
            {'id': vid, 'name': vendors_by_id[vid].company_name if vid in vendors_by_id else f'#{vid}', 'count': count}
            for vid, count in vendor_review_counts.items()
        ),
        key=lambda v: v['name'],
    )

    vendor_filter = request.GET.get('vendor_id', '').strip()
    reviews = all_reviews
    if vendor_filter:
        try:
            vendor_filter_id = int(vendor_filter)
        except ValueError:
            vendor_filter_id = None
        if vendor_filter_id:
            reviews = [r for r in all_reviews if vendor_filter_id in r._vendor_ids]

    return render(request, 'admin/reviews.html', {
        'reviews': reviews,
        'name': request.session["name"],
        'vendor_options': vendor_options,
        'vendor_filter': vendor_filter,
    })


# ====================   Admin Contact Messages  ===========================

def admin_messages(request):
    """Redirect to unified mail management"""
    return redirect('/manager/email/')


def admin_message_detail(request, msg_id):
    """Admin: view single message detail — auto marks as read"""
    if "name" not in request.session:
        return redirect('/manager/login/')

    from manager.email_utils import ensure_platform_email_account
    ensure_platform_email_account()

    msg = get_object_or_404(models.ContactMessage, id=msg_id)
    if not msg.is_read:
        msg.is_read = True
        msg.save(update_fields=['is_read'])

    unread_count = models.ContactMessage.objects.filter(is_read=False).count()
    accounts = models.EmailAccount.objects.filter(is_active=True)
    labels = models.EmailLabel.objects.all()
    name = request.session["name"]
    return render(request, 'admin/admin_message_detail.html', {
        'msg': msg,
        'unread_count': unread_count,
        'accounts': accounts,
        'labels': labels,
        'name': name,
    })


def admin_message_toggle_read(request):
    """AJAX: toggle is_read status for a message"""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    msg_id = request.POST.get('id')
    try:
        msg = models.ContactMessage.objects.get(id=msg_id)
        msg.is_read = not msg.is_read
        msg.save(update_fields=['is_read'])
        unread_count = models.ContactMessage.objects.filter(is_read=False).count()
        return JsonResponse({'success': True, 'is_read': msg.is_read, 'unread_count': unread_count})
    except models.ContactMessage.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Not found'}, status=404)


def admin_message_delete(request):
    """Admin: delete a contact message"""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    msg_id = request.POST.get('id')
    try:
        msg = models.ContactMessage.objects.get(id=msg_id)
        msg.delete()
        unread_count = models.ContactMessage.objects.filter(is_read=False).count()
        return JsonResponse({'success': True, 'unread_count': unread_count})
    except models.ContactMessage.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Not found'}, status=404)


def admin_message_bulk_action(request):
    """Admin: bulk mark read or bulk delete messages"""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    action = request.POST.get('action')
    ids_raw = request.POST.get('ids', '')
    try:
        ids = [int(i) for i in ids_raw.split(',') if i.strip().isdigit()]
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Invalid IDs'}, status=400)

    if not ids:
        return JsonResponse({'success': False, 'error': 'No IDs provided'}, status=400)

    if action == 'mark_read':
        models.ContactMessage.objects.filter(id__in=ids).update(is_read=True)
    elif action == 'mark_unread':
        models.ContactMessage.objects.filter(id__in=ids).update(is_read=False)
    elif action == 'delete':
        models.ContactMessage.objects.filter(id__in=ids).delete()
    else:
        return JsonResponse({'success': False, 'error': 'Unknown action'}, status=400)

    unread_count = models.ContactMessage.objects.filter(is_read=False).count()
    return JsonResponse({'success': True, 'unread_count': unread_count})


def reply_to_contact(request, msg_id):
    """Reply to a contact message using a configured email account"""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    msg = get_object_or_404(models.ContactMessage, id=msg_id)
    account_id = request.POST.get('account_id')
    reply_body = request.POST.get('body', '').strip()

    if not reply_body:
        return JsonResponse({'success': False, 'error': '回复内容不能为空'})
    if not account_id:
        return JsonResponse({'success': False, 'error': '请选择发送账户'})

    account = get_object_or_404(models.EmailAccount, id=account_id, is_active=True)
    subject = f'Re: {msg.subject or "(无主题)"}'

    # Build reply with original message quoted
    full_body = (
        f'{reply_body}\n\n'
        f'--- 原始消息 ---\n'
        f'发件人: {msg.name} <{msg.email}>\n'
        f'时间: {msg.created_at.strftime("%Y-%m-%d %H:%M")}\n'
        f'主题: {msg.subject or "(无主题)"}\n\n'
        f'{msg.message}'
    )

    try:
        _send_email(
            account=account,
            to=msg.email,
            subject=subject,
            body=full_body,
        )

        # Save as sent EmailMessage for record
        models.EmailMessage.objects.create(
            account=account,
            sender_name=account.name,
            sender_email=account.email_address,
            recipients=msg.email,
            subject=subject,
            body_text=full_body,
            folder='sent',
            is_read=True,
            sent_at=timezone.now(),
        )

        msg.replied = True
        msg.replied_at = timezone.now()
        msg.admin_reply = reply_body
        msg.save(update_fields=['replied', 'replied_at', 'admin_reply'])

        return JsonResponse({'success': True, 'message': f'回复已发送至 {msg.email}'})
    except Exception as e:
        logger.error(f'Reply to contact #{msg_id} failed: {e}')
        return JsonResponse({'success': False, 'error': f'发送失败: {str(e)}'})


def contact_label_action(request):
    """Add/remove labels on contact messages"""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    action = request.POST.get('action', '')
    ids_raw = request.POST.get('ids', '')
    label_id = request.POST.get('label_id')

    try:
        ids = [int(i) for i in ids_raw.split(',') if i.strip().isdigit()]
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Invalid IDs'}, status=400)

    if not ids or not label_id:
        return JsonResponse({'success': False, 'error': 'Missing parameters'}, status=400)

    label = get_object_or_404(models.EmailLabel, id=label_id)
    msgs = models.ContactMessage.objects.filter(id__in=ids)

    if action == 'add_label':
        for m in msgs:
            m.labels.add(label)
    elif action == 'remove_label':
        for m in msgs:
            m.labels.remove(label)
    else:
        return JsonResponse({'success': False, 'error': 'Unknown action'}, status=400)

    return JsonResponse({'success': True})


# ====================   Email Management System  ===========================

import imaplib
import smtplib
import email as email_lib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.utils import parseaddr, formataddr, parsedate_to_datetime
from email.header import decode_header


def _decode_header_value(value):
    """Decode email header value (handles encoded words)"""
    if not value:
        return ''
    decoded_parts = []
    for part, charset in decode_header(value):
        if isinstance(part, bytes):
            decoded_parts.append(part.decode(charset or 'utf-8', errors='replace'))
        else:
            decoded_parts.append(part)
    return ''.join(decoded_parts)


def _apply_auto_rules(email_obj):
    """Apply auto-rules to a newly received email"""
    rules = models.EmailAutoRule.objects.filter(is_active=True).order_by('-priority')
    if email_obj.account_id:
        rules = rules.filter(
            Q(apply_to_account__isnull=True) | Q(apply_to_account=email_obj.account)
        )

    for rule in rules:
        matched = False
        pattern = rule.match_pattern.lower()
        fields_to_check = []

        if rule.match_field == 'from':
            fields_to_check = [email_obj.sender_email.lower(), email_obj.sender_name.lower()]
        elif rule.match_field == 'to':
            fields_to_check = [email_obj.recipients.lower()]
        elif rule.match_field == 'subject':
            fields_to_check = [email_obj.subject.lower()]
        elif rule.match_field == 'body':
            fields_to_check = [email_obj.body_text.lower()]
        else:  # 'any'
            fields_to_check = [
                email_obj.sender_email.lower(), email_obj.sender_name.lower(),
                email_obj.recipients.lower(), email_obj.subject.lower(),
                email_obj.body_text.lower()
            ]

        for field_val in fields_to_check:
            if pattern in field_val:
                matched = True
                break

        if matched:
            if rule.action == 'label' and rule.action_label:
                email_obj.labels.add(rule.action_label)
            elif rule.action == 'star':
                email_obj.is_starred = True
                email_obj.save(update_fields=['is_starred'])
            elif rule.action == 'archive':
                email_obj.folder = 'archive'
                email_obj.save(update_fields=['folder'])
            elif rule.action == 'delete':
                email_obj.folder = 'trash'
                email_obj.save(update_fields=['folder'])
            elif rule.action == 'mark_read':
                email_obj.is_read = True
                email_obj.save(update_fields=['is_read'])
            elif rule.action == 'auto_reply':
                try:
                    _send_email(
                        account=email_obj.account,
                        to=email_obj.sender_email,
                        subject=rule.auto_reply_subject or f'Re: {email_obj.subject}',
                        body=rule.auto_reply_body,
                        in_reply_to=email_obj.message_uid,
                    )
                except Exception as e:
                    logger.error(f'Auto-reply failed for rule {rule.name}: {e}')
            elif rule.action == 'forward' and rule.action_forward_to:
                try:
                    _send_email(
                        account=email_obj.account,
                        to=rule.action_forward_to,
                        subject=f'Fwd: {email_obj.subject}',
                        body=f'---------- Forwarded message ----------\nFrom: {email_obj.sender_name} <{email_obj.sender_email}>\nDate: {email_obj.received_at}\nSubject: {email_obj.subject}\n\n{email_obj.body_text}',
                    )
                except Exception as e:
                    logger.error(f'Forward failed for rule {rule.name}: {e}')


def _send_email(account, to, subject, body, cc='', bcc='', html_body='', in_reply_to='', attachments=None):
    """Send email through an account's SMTP settings"""
    msg = MIMEMultipart('alternative')
    msg['From'] = formataddr((account.name, account.email_address))
    msg['To'] = to
    msg['Subject'] = subject
    if cc:
        msg['Cc'] = cc
    if in_reply_to:
        msg['In-Reply-To'] = in_reply_to

    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    if html_body:
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))

    if attachments:
        for att in attachments:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(att.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{att.name}"')
            msg.attach(part)

    all_recipients = [r.strip() for r in to.split(',') if r.strip()]
    if cc:
        all_recipients += [r.strip() for r in cc.split(',') if r.strip()]
    if bcc:
        all_recipients += [r.strip() for r in bcc.split(',') if r.strip()]

    server = None
    try:
        if account.smtp_use_tls:
            server = smtplib.SMTP(account.smtp_host, account.smtp_port, timeout=30)
            server.starttls()
        else:
            server = smtplib.SMTP_SSL(account.smtp_host, account.smtp_port, timeout=30)

        server.login(account.username, account.password)
        server.send_message(msg, to_addrs=all_recipients)
        return msg['Message-ID'] or ''
    finally:
        if server:
            try:
                server.quit()
            except Exception:
                pass


def email_dashboard(request):
    """Unified mail/message management dashboard"""
    if "name" not in request.session:
        return redirect('/manager/login/')

    from manager.email_utils import ensure_platform_email_account
    ensure_platform_email_account()

    folder = request.GET.get('folder', 'inbox')
    account_id = request.GET.get('account', '')
    label_id = request.GET.get('label', '')
    search = request.GET.get('search', '').strip()
    starred_only = request.GET.get('starred', '') == '1'

    accounts = models.EmailAccount.objects.filter(is_active=True)
    labels = models.EmailLabel.objects.all()

    contact_messages = None

    if folder == 'contact':
        # Show contact form messages
        cqs = models.ContactMessage.objects.prefetch_related('labels').order_by('-created_at')
        if search:
            cqs = cqs.filter(
                Q(name__icontains=search) | Q(email__icontains=search) |
                Q(subject__icontains=search) | Q(message__icontains=search)
            )
        if label_id:
            cqs = cqs.filter(labels__id=label_id)
        contact_messages = cqs
        emails = models.EmailMessage.objects.none()
    else:
        qs = models.EmailMessage.objects.select_related('account')

        if account_id:
            qs = qs.filter(account_id=account_id)
        if folder:
            qs = qs.filter(folder=folder)
        if label_id:
            qs = qs.filter(labels__id=label_id)
        if starred_only:
            qs = qs.filter(is_starred=True)
        if search:
            qs = qs.filter(
                Q(subject__icontains=search) |
                Q(sender_name__icontains=search) |
                Q(sender_email__icontains=search) |
                Q(body_text__icontains=search) |
                Q(recipients__icontains=search)
            )

        emails = qs.order_by('-received_at', '-created_at')

    # Counts per folder
    base_qs = models.EmailMessage.objects.all()
    if account_id:
        base_qs = base_qs.filter(account_id=account_id)

    contact_unread = models.ContactMessage.objects.filter(is_read=False).count()
    contact_total = models.ContactMessage.objects.count()

    folder_counts = {
        'inbox': base_qs.filter(folder='inbox').count(),
        'sent': base_qs.filter(folder='sent').count(),
        'draft': base_qs.filter(folder='draft').count(),
        'trash': base_qs.filter(folder='trash').count(),
        'archive': base_qs.filter(folder='archive').count(),
        'starred': base_qs.filter(is_starred=True).count(),
        'contact': contact_total,
    }
    unread_count = base_qs.filter(folder='inbox', is_read=False).count() + contact_unread

    context = {
        'emails': emails,
        'contact_messages': contact_messages,
        'accounts': accounts,
        'labels': labels,
        'current_folder': folder,
        'current_account': account_id,
        'current_label': label_id,
        'search': search,
        'starred_only': starred_only,
        'folder_counts': folder_counts,
        'unread_count': unread_count,
        'contact_unread': contact_unread,
        'name': request.session.get('name', ''),
    }
    return render(request, 'admin/email_management.html', context)


def email_detail(request, email_id):
    """View a single email"""
    if "name" not in request.session:
        return redirect('/manager/login/')

    email_obj = get_object_or_404(models.EmailMessage, id=email_id)
    if not email_obj.is_read:
        email_obj.is_read = True
        email_obj.save(update_fields=['is_read'])

    # Get thread
    thread = models.EmailMessage.objects.filter(
        account=email_obj.account,
        subject__in=[email_obj.subject, f'Re: {email_obj.subject}', email_obj.subject.replace('Re: ', '', 1)]
    ).exclude(id=email_obj.id).order_by('received_at', 'created_at')[:10]

    # JSON response for reading pane
    if request.GET.get('format') == 'json':
        date_str = ''
        if email_obj.received_at:
            date_str = email_obj.received_at.strftime('%Y-%m-%d %H:%M')
        elif email_obj.sent_at:
            date_str = email_obj.sent_at.strftime('%Y-%m-%d %H:%M')
        elif email_obj.created_at:
            date_str = email_obj.created_at.strftime('%Y-%m-%d %H:%M')

        labels_data = [{'name': l.name, 'color': l.color} for l in email_obj.labels.all()]
        attachments_data = [{'filename': a.filename, 'url': a.file.url} for a in email_obj.attachments.all() if a.file]

        return JsonResponse({
            'id': email_obj.id,
            'subject': email_obj.subject,
            'sender_name': email_obj.sender_name,
            'sender_email': email_obj.sender_email,
            'recipients': email_obj.recipients,
            'cc': email_obj.cc,
            'body_text': email_obj.body_text,
            'body_html': email_obj.body_html,
            'date': date_str,
            'message_uid': email_obj.message_uid,
            'is_starred': email_obj.is_starred,
            'folder': email_obj.folder,
            'labels': labels_data,
            'attachments': attachments_data,
        })

    context = {
        'email': email_obj,
        'thread': thread,
        'accounts': models.EmailAccount.objects.filter(is_active=True),
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'admin/email_detail_partial.html', context)
    return render(request, 'admin/email_detail.html', context)


def email_compose(request):
    """Compose and send a new email"""
    if "name" not in request.session:
        return redirect('/manager/login/')

    if request.method == 'POST':
        account_id = request.POST.get('account_id')
        to = request.POST.get('to', '').strip()
        cc = request.POST.get('cc', '').strip()
        bcc = request.POST.get('bcc', '').strip()
        subject = request.POST.get('subject', '').strip()
        body = request.POST.get('body', '').strip()
        html_body = request.POST.get('html_body', '').strip()
        is_draft = request.POST.get('save_draft') == '1'
        in_reply_to = request.POST.get('in_reply_to', '').strip()

        if not account_id:
            return JsonResponse({'success': False, 'error': '请选择发送账户'})

        account = get_object_or_404(models.EmailAccount, id=account_id)

        email_obj = models.EmailMessage.objects.create(
            account=account,
            sender_name=account.name,
            sender_email=account.email_address,
            recipients=to,
            cc=cc,
            bcc=bcc,
            subject=subject or '(无主题)',
            body_text=body,
            body_html=html_body,
            folder='draft' if is_draft else 'sent',
            is_read=True,
            in_reply_to=in_reply_to,
            sent_at=None if is_draft else timezone.now(),
        )

        if not is_draft:
            try:
                attachments = request.FILES.getlist('attachments')
                _send_email(
                    account=account,
                    to=to,
                    subject=subject,
                    body=body,
                    cc=cc,
                    bcc=bcc,
                    html_body=html_body,
                    in_reply_to=in_reply_to,
                    attachments=attachments if attachments else None,
                )

                for att_file in attachments:
                    models.EmailAttachment.objects.create(
                        email=email_obj,
                        filename=att_file.name,
                        content_type=att_file.content_type or '',
                        file=att_file,
                        size=att_file.size,
                    )

                return JsonResponse({'success': True, 'message': '邮件发送成功'})
            except Exception as e:
                logger.error(f'Email send failed: {e}')
                email_obj.folder = 'draft'
                email_obj.sent_at = None
                email_obj.save()
                return JsonResponse({'success': False, 'error': f'发送失败: {str(e)}'})
        else:
            return JsonResponse({'success': True, 'message': '草稿已保存'})

    # GET: return compose data
    accounts = models.EmailAccount.objects.filter(is_active=True)
    reply_to = request.GET.get('reply_to', '')
    reply_email = None
    if reply_to:
        try:
            reply_email = models.EmailMessage.objects.get(id=reply_to)
        except models.EmailMessage.DoesNotExist:
            pass

    context = {
        'accounts': accounts,
        'reply_email': reply_email,
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'admin/email_compose_partial.html', context)
    return JsonResponse({'accounts': list(accounts.values('id', 'name', 'email_address'))})


def email_sync(request):
    """Sync emails from IMAP server(s)"""
    if "name" not in request.session:
        return redirect('/manager/login/')

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    account_id = request.POST.get('account_id', '')
    if account_id:
        accounts = models.EmailAccount.objects.filter(id=account_id, is_active=True)
    else:
        accounts = models.EmailAccount.objects.filter(is_active=True)

    total_new = 0
    errors = []

    for account in accounts:
        try:
            new_count = _fetch_emails_imap(account)
            total_new += new_count
            account.last_sync = timezone.now()
            account.save(update_fields=['last_sync'])
        except Exception as e:
            logger.error(f'IMAP sync failed for {account.email_address}: {e}')
            errors.append(f'{account.email_address}: {str(e)}')

    return JsonResponse({
        'success': True,
        'new_emails': total_new,
        'errors': errors,
        'message': f'同步完成，共 {total_new} 封新邮件' + (f'，{len(errors)} 个错误' if errors else ''),
    })


def _fetch_emails_imap(account, folder='INBOX', max_emails=50):
    """Fetch emails from IMAP for a single account with incremental sync"""
    if account.imap_use_ssl:
        mail = imaplib.IMAP4_SSL(account.imap_host, account.imap_port, timeout=30)
    else:
        mail = imaplib.IMAP4(account.imap_host, account.imap_port, timeout=30)

    mail.login(account.username, account.password)
    mail.select(folder)

    existing_uids = set(
        models.EmailMessage.objects.filter(account=account)
        .exclude(message_uid='')
        .values_list('message_uid', flat=True)
    )

    # Incremental sync: only fetch emails since last sync date
    search_criteria = 'ALL'
    if account.last_sync:
        since_date = account.last_sync.strftime('%d-%b-%Y')
        search_criteria = f'(SINCE {since_date})'

    status, data = mail.search(None, search_criteria)
    if status != 'OK':
        mail.logout()
        return 0

    email_ids = data[0].split()
    if not email_ids:
        mail.logout()
        return 0

    recent_ids = email_ids[-max_emails:]
    new_count = 0

    for eid in recent_ids:
        status, msg_data = mail.fetch(eid, '(RFC822 FLAGS)')
        if status != 'OK':
            continue

        raw_email = msg_data[0][1]
        msg = email_lib.message_from_bytes(raw_email)

        msg_uid = msg.get('Message-ID', '')
        if msg_uid in existing_uids:
            continue

        subject = _decode_header_value(msg.get('Subject', ''))
        from_raw = msg.get('From', '')
        from_name, from_email_addr = parseaddr(from_raw)
        from_name = _decode_header_value(from_name) or from_email_addr
        to_raw = _decode_header_value(msg.get('To', ''))
        cc_raw = _decode_header_value(msg.get('Cc', ''))

        date_str = msg.get('Date', '')
        received_at = None
        if date_str:
            try:
                received_at = parsedate_to_datetime(date_str)
            except Exception:
                received_at = timezone.now()
        else:
            received_at = timezone.now()

        body_text = ''
        body_html = ''
        if msg.is_multipart():
            for part in msg.walk():
                content_type = part.get_content_type()
                content_disp = str(part.get('Content-Disposition', ''))
                if 'attachment' in content_disp:
                    continue
                payload = part.get_payload(decode=True)
                if payload:
                    charset = part.get_content_charset() or 'utf-8'
                    text = payload.decode(charset, errors='replace')
                    if content_type == 'text/plain':
                        body_text = text
                    elif content_type == 'text/html':
                        body_html = text
        else:
            payload = msg.get_payload(decode=True)
            if payload:
                charset = msg.get_content_charset() or 'utf-8'
                body_text = payload.decode(charset, errors='replace')

        flags_raw = msg_data[0][0] if isinstance(msg_data[0], tuple) else b''
        is_read = b'\\Seen' in flags_raw
        in_reply_to = msg.get('In-Reply-To', '')

        email_obj = models.EmailMessage.objects.create(
            account=account,
            message_uid=msg_uid,
            sender_name=from_name,
            sender_email=from_email_addr,
            recipients=to_raw,
            cc=cc_raw,
            subject=subject or '(无主题)',
            body_text=body_text,
            body_html=body_html,
            folder='inbox',
            is_read=is_read,
            received_at=received_at,
            in_reply_to=in_reply_to,
        )

        if msg.is_multipart():
            for part in msg.walk():
                content_disp = str(part.get('Content-Disposition', ''))
                if 'attachment' in content_disp:
                    filename = _decode_header_value(part.get_filename() or 'attachment')
                    payload = part.get_payload(decode=True)
                    if payload:
                        from django.core.files.base import ContentFile
                        att = models.EmailAttachment(
                            email=email_obj,
                            filename=filename,
                            content_type=part.get_content_type() or '',
                            size=len(payload),
                        )
                        att.file.save(filename, ContentFile(payload), save=True)

        _apply_auto_rules(email_obj)
        new_count += 1

    mail.logout()
    return new_count


def email_action(request):
    """AJAX: perform actions on emails"""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    action = request.POST.get('action', '')
    email_ids_raw = request.POST.get('ids', '')

    try:
        email_ids = [int(i) for i in email_ids_raw.split(',') if i.strip().isdigit()]
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Invalid IDs'}, status=400)

    if not email_ids:
        return JsonResponse({'success': False, 'error': 'No emails selected'}, status=400)

    qs = models.EmailMessage.objects.filter(id__in=email_ids)

    if action == 'mark_read':
        qs.update(is_read=True)
    elif action == 'mark_unread':
        qs.update(is_read=False)
    elif action == 'star':
        qs.update(is_starred=True)
    elif action == 'unstar':
        qs.update(is_starred=False)
    elif action == 'toggle_star':
        for e in qs:
            e.is_starred = not e.is_starred
            e.save(update_fields=['is_starred'])
    elif action == 'trash':
        qs.update(folder='trash')
    elif action == 'archive':
        qs.update(folder='archive')
    elif action == 'move_inbox':
        qs.update(folder='inbox')
    elif action == 'delete':
        qs.delete()
    elif action == 'add_label':
        label_id = request.POST.get('label_id')
        if label_id:
            label = get_object_or_404(models.EmailLabel, id=label_id)
            for e in qs:
                e.labels.add(label)
    elif action == 'remove_label':
        label_id = request.POST.get('label_id')
        if label_id:
            label = get_object_or_404(models.EmailLabel, id=label_id)
            for e in qs:
                e.labels.remove(label)
    else:
        return JsonResponse({'success': False, 'error': 'Unknown action'}, status=400)

    unread = models.EmailMessage.objects.filter(folder='inbox', is_read=False).count()
    return JsonResponse({'success': True, 'unread_count': unread})


# ---- Email Account Management ----

def email_accounts(request):
    """Manage email accounts"""
    if "name" not in request.session:
        return redirect('/manager/login/')

    from manager.email_utils import ensure_platform_email_account
    ensure_platform_email_account()

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'add':
            account = models.EmailAccount(
                name=request.POST.get('name', '').strip(),
                email_address=request.POST.get('email_address', '').strip(),
                imap_host=request.POST.get('imap_host', 'imap.gmail.com').strip(),
                imap_port=int(request.POST.get('imap_port', 993)),
                imap_use_ssl=request.POST.get('imap_use_ssl') == 'on',
                smtp_host=request.POST.get('smtp_host', 'smtp.gmail.com').strip(),
                smtp_port=int(request.POST.get('smtp_port', 587)),
                smtp_use_tls=request.POST.get('smtp_use_tls') == 'on',
                username=request.POST.get('username', '').strip(),
                password=request.POST.get('password', '').strip(),
                is_active=True,
                is_default=request.POST.get('is_default') == 'on',
            )
            account.save()
            if account.is_default:
                models.EmailAccount.objects.exclude(id=account.id).update(is_default=False)
            return JsonResponse({'success': True, 'message': '账户添加成功', 'id': account.id})

        elif action == 'edit':
            acc_id = request.POST.get('id')
            account = get_object_or_404(models.EmailAccount, id=acc_id)
            account.name = request.POST.get('name', account.name).strip()
            account.email_address = request.POST.get('email_address', account.email_address).strip()
            account.imap_host = request.POST.get('imap_host', account.imap_host).strip()
            account.imap_port = int(request.POST.get('imap_port', account.imap_port))
            account.imap_use_ssl = request.POST.get('imap_use_ssl') == 'on'
            account.smtp_host = request.POST.get('smtp_host', account.smtp_host).strip()
            account.smtp_port = int(request.POST.get('smtp_port', account.smtp_port))
            account.smtp_use_tls = request.POST.get('smtp_use_tls') == 'on'
            account.username = request.POST.get('username', account.username).strip()
            pwd = request.POST.get('password', '').strip()
            if pwd:
                account.password = pwd
            account.is_default = request.POST.get('is_default') == 'on'
            account.is_active = request.POST.get('is_active', 'on') == 'on'
            account.save()
            if account.is_default:
                models.EmailAccount.objects.exclude(id=account.id).update(is_default=False)
            return JsonResponse({'success': True, 'message': '账户更新成功'})

        elif action == 'delete':
            acc_id = request.POST.get('id')
            models.EmailAccount.objects.filter(id=acc_id).delete()
            return JsonResponse({'success': True, 'message': '账户已删除'})

        elif action == 'test':
            acc_id = request.POST.get('id')
            account = get_object_or_404(models.EmailAccount, id=acc_id)
            errors = []
            try:
                if account.imap_use_ssl:
                    m = imaplib.IMAP4_SSL(account.imap_host, account.imap_port, timeout=15)
                else:
                    m = imaplib.IMAP4(account.imap_host, account.imap_port, timeout=15)
                m.login(account.username, account.password)
                m.logout()
            except Exception as e:
                errors.append(f'IMAP: {str(e)}')
            try:
                if account.smtp_use_tls:
                    s = smtplib.SMTP(account.smtp_host, account.smtp_port, timeout=15)
                    s.starttls()
                else:
                    s = smtplib.SMTP_SSL(account.smtp_host, account.smtp_port, timeout=15)
                s.login(account.username, account.password)
                s.quit()
            except Exception as e:
                errors.append(f'SMTP: {str(e)}')

            if errors:
                return JsonResponse({'success': False, 'error': '; '.join(errors)})
            return JsonResponse({'success': True, 'message': '连接测试成功'})

    accounts = models.EmailAccount.objects.all()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'accounts': list(accounts.values(
                'id', 'name', 'email_address', 'imap_host', 'imap_port',
                'smtp_host', 'smtp_port', 'is_active', 'is_default', 'last_sync'
            ))
        })
    return render(request, 'admin/email_accounts.html', {'accounts': accounts})


# ---- Email Labels Management ----

def email_labels(request):
    """Manage email labels"""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)

    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'add':
            label = models.EmailLabel.objects.create(
                name=request.POST.get('name', '').strip(),
                color=request.POST.get('color', '#667eea').strip(),
                icon=request.POST.get('icon', 'fa-tag').strip(),
            )
            return JsonResponse({'success': True, 'id': label.id, 'message': '标签创建成功'})
        elif action == 'edit':
            label = get_object_or_404(models.EmailLabel, id=request.POST.get('id'))
            label.name = request.POST.get('name', label.name).strip()
            label.color = request.POST.get('color', label.color).strip()
            label.icon = request.POST.get('icon', label.icon).strip()
            label.save()
            return JsonResponse({'success': True, 'message': '标签更新成功'})
        elif action == 'delete':
            models.EmailLabel.objects.filter(id=request.POST.get('id')).delete()
            return JsonResponse({'success': True, 'message': '标签已删除'})

    labels = models.EmailLabel.objects.all()
    return JsonResponse({'labels': list(labels.values('id', 'name', 'color', 'icon'))})


# ---- Email Auto-Rules Management ----

def email_rules(request):
    """Manage email auto-rules"""
    if "name" not in request.session:
        return redirect('/manager/login/')

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'add':
            label_id = request.POST.get('action_label')
            account_id = request.POST.get('apply_to_account')
            rule = models.EmailAutoRule.objects.create(
                name=request.POST.get('name', '').strip(),
                is_active=request.POST.get('is_active', 'on') == 'on',
                match_field=request.POST.get('match_field', 'any'),
                match_pattern=request.POST.get('match_pattern', '').strip(),
                action=request.POST.get('rule_action', 'label'),
                action_label_id=int(label_id) if label_id else None,
                action_forward_to=request.POST.get('action_forward_to', '').strip(),
                auto_reply_subject=request.POST.get('auto_reply_subject', '').strip(),
                auto_reply_body=request.POST.get('auto_reply_body', '').strip(),
                apply_to_account_id=int(account_id) if account_id else None,
                priority=int(request.POST.get('priority', 0)),
            )
            return JsonResponse({'success': True, 'id': rule.id, 'message': '规则创建成功'})

        elif action == 'edit':
            rule = get_object_or_404(models.EmailAutoRule, id=request.POST.get('id'))
            rule.name = request.POST.get('name', rule.name).strip()
            rule.is_active = request.POST.get('is_active', 'on') == 'on'
            rule.match_field = request.POST.get('match_field', rule.match_field)
            rule.match_pattern = request.POST.get('match_pattern', rule.match_pattern).strip()
            rule.action = request.POST.get('rule_action', rule.action)
            label_id = request.POST.get('action_label')
            rule.action_label_id = int(label_id) if label_id else None
            rule.action_forward_to = request.POST.get('action_forward_to', '').strip()
            rule.auto_reply_subject = request.POST.get('auto_reply_subject', '').strip()
            rule.auto_reply_body = request.POST.get('auto_reply_body', '').strip()
            account_id = request.POST.get('apply_to_account')
            rule.apply_to_account_id = int(account_id) if account_id else None
            rule.priority = int(request.POST.get('priority', rule.priority))
            rule.save()
            return JsonResponse({'success': True, 'message': '规则更新成功'})

        elif action == 'delete':
            models.EmailAutoRule.objects.filter(id=request.POST.get('id')).delete()
            return JsonResponse({'success': True, 'message': '规则已删除'})

        elif action == 'toggle':
            rule = get_object_or_404(models.EmailAutoRule, id=request.POST.get('id'))
            rule.is_active = not rule.is_active
            rule.save(update_fields=['is_active'])
            return JsonResponse({'success': True, 'is_active': rule.is_active})

    rules = models.EmailAutoRule.objects.select_related('action_label', 'apply_to_account').all()
    labels = models.EmailLabel.objects.all()
    accounts = models.EmailAccount.objects.filter(is_active=True)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'rules': list(rules.values(
                'id', 'name', 'is_active', 'match_field', 'match_pattern',
                'action', 'action_label__name', 'action_forward_to',
                'auto_reply_subject', 'priority'
            ))
        })
    return render(request, 'admin/email_rules.html', {
        'rules': rules, 'labels': labels, 'accounts': accounts
    })


# ==========================================
# Site User Authentication Views
# ==========================================

from manager.auth_password import (
    hash_password as _hash_password,
    set_unified_password as _set_unified_password,
    sync_password_by_email as _sync_password_by_email,
    check_email_password as _check_email_password,
    link_dual_accounts_by_email as _link_dual_accounts_by_email,
    get_linked_site_user_and_vendor as _get_linked_site_user_and_vendor,
)


import random
import string


def _generate_pin():
    """Generate a 6-digit PIN code"""
    return ''.join(random.choices(string.digits, k=6))


def _send_verification_email(email, pin_code, name):
    """Send verification PIN code via Django's email backend (English + French)."""
    subject = 'DUNO 360 - Email Verification / Vérification e-mail'
    html_body = f'''
    <div style="font-family:'Segoe UI',Arial,sans-serif;max-width:520px;margin:0 auto;background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);">
        <div style="background:linear-gradient(135deg,#14245f 0%,#1d4ed8 100%);padding:32px 28px;text-align:center;">
            <h1 style="color:#fff;margin:0;font-size:1.5rem;">DUNO 360</h1>
            <p style="color:rgba(255,255,255,0.85);margin:8px 0 0;font-size:0.95rem;">Email &amp; phone verification / Vérification e-mail et téléphone</p>
        </div>
        <div style="padding:32px 28px;">
            <p style="color:#333;font-size:1rem;margin:0 0 8px;">Hello <strong>{name}</strong>! / Bonjour <strong>{name}</strong>&nbsp;!</p>
            <p style="color:#666;font-size:0.93rem;line-height:1.7;margin:0 0 12px;">
                Thank you for registering on DUNO 360. Please use the verification code below to complete your account setup:
            </p>
            <p style="color:#666;font-size:0.93rem;line-height:1.7;margin:0 0 24px;">
                Merci de vous être inscrit sur DUNO 360. Utilisez le code ci-dessous pour finaliser la création de votre compte&nbsp;:
            </p>
            <div style="background:linear-gradient(135deg,rgba(20,36,95,0.08),rgba(29,78,216,0.08));border:2px dashed #1d4ed8;border-radius:14px;padding:24px;text-align:center;margin:0 0 24px;">
                <span style="font-size:2.5rem;font-weight:800;letter-spacing:12px;color:#1d4ed8;">{pin_code}</span>
            </div>
            <p style="color:#999;font-size:0.85rem;text-align:center;margin:0;">
                ⏰ This code expires in <strong>15 minutes</strong>. / Ce code expire dans <strong>15 minutes</strong>.
            </p>
        </div>
        <div style="background:#f8f9ff;padding:16px 28px;text-align:center;border-top:1px solid #eee;">
            <p style="color:#aaa;font-size:0.8rem;margin:0;">If you did not register, please ignore this email.<br>Si vous ne vous êtes pas inscrit, ignorez cet e-mail.</p>
        </div>
    </div>
    '''
    plain_body = (
        f'Hello {name}, your verification code is: {pin_code}. Valid for 15 minutes.\n\n'
        f'Bonjour {name}, votre code de vérification est : {pin_code}. Valide 15 minutes.'
    )

    try:
        from django.core.mail import EmailMultiAlternatives
        msg = EmailMultiAlternatives(subject, plain_body, django_settings.DEFAULT_FROM_EMAIL, [email])
        msg.attach_alternative(html_body, 'text/html')
        msg.send(fail_silently=False)
        return True
    except Exception as e:
        logger.error(f'Failed to send verification email to {email}: {e}')
        return False


def _send_registration_phone_otp(phone: str):
    """Send Twilio Verify SMS if configured. Returns (ok, error_message)."""
    from manager.twilio_verify import (
        is_twilio_verify_enabled,
        normalize_phone_e164,
        send_verification_sms,
        validate_phone_e164,
    )
    phone_e164 = normalize_phone_e164(phone)
    valid, err = validate_phone_e164(phone_e164)
    if not valid:
        return False, err
    if not is_twilio_verify_enabled():
        return True, ''
    return send_verification_sms(phone_e164)


def _check_registration_phone_otp(phone: str, code: str):
    """Validate phone OTP via Twilio. Returns (ok, error_message)."""
    from manager.twilio_verify import (
        check_verification_sms,
        is_twilio_verify_enabled,
        normalize_phone_e164,
    )
    if not is_twilio_verify_enabled():
        return True, ''
    phone_e164 = normalize_phone_e164(phone)
    return check_verification_sms(phone_e164, code)


def _verify_phone_otp_if_required(verification, phone_pin: str):
    """Check SMS OTP only when this signup session requires it."""
    from django.utils.translation import gettext as _
    if not getattr(verification, 'require_sms_verification', False):
        return True, ''
    if not (phone_pin or '').strip():
        return False, _('Enter the code received by SMS.')
    return _check_registration_phone_otp(verification.phone, phone_pin)


def _dispatch_signup_verification(verification, redirect_url: str, lang_code: str | None = None) -> dict:
    """Start phone-first verification asynchronously and return redirect immediately."""
    from manager.signup_verification import start_signup_verification_async
    return start_signup_verification_async(verification, redirect_url, lang_code)


def signup_verification_status(request):
    """Poll async signup verification dispatch (SMS + email)."""
    from django.utils.translation import gettext as _
    from manager.signup_verification import get_signup_verification_status
    from manager.twilio_verify import is_twilio_verify_enabled

    email = (request.GET.get('email') or '').strip()
    verification_type = (request.GET.get('type') or 'user').strip() or 'user'
    if not email:
        return JsonResponse({'pending': True})

    status = get_signup_verification_status(email, verification_type)
    if status:
        return JsonResponse(status)

    verification = models.EmailVerification.objects.filter(
        email=email, is_verified=False, verification_type=verification_type,
    ).first()
    if not verification:
        return JsonResponse({'pending': True})

    require_sms = bool(verification.require_sms_verification)
    sms_failed = is_twilio_verify_enabled() and not require_sms
    return JsonResponse({
        'pending': False,
        'email_sent': True,
        'require_sms': require_sms,
        'sms_failed': sms_failed,
        'sms_error': '',
        'fallback_message': _(
            'Email verification has been started automatically. '
            'Enter the 6-digit code from your inbox.'
        ) if sms_failed else '',
    })


def _verification_page_context(email, verification_type='user', request=None):
    from manager.twilio_verify import is_twilio_verify_enabled, mask_phone, normalize_phone_e164
    verification = models.EmailVerification.objects.filter(
        email=email, is_verified=False, verification_type=verification_type,
    ).first()
    phone_e164 = normalize_phone_e164(verification.phone) if verification else ''
    require_sms = bool(verification and verification.require_sms_verification)
    sms_fallback = False
    if request:
        sms_fallback = request.GET.get('sms_fallback') == '1'
    if verification and is_twilio_verify_enabled() and not require_sms:
        sms_fallback = True
    return {
        'email': email,
        'phone_masked': mask_phone(phone_e164) if phone_e164 else '',
        'twilio_sms_enabled': require_sms,
        'require_sms_verification': require_sms,
        'sms_fallback': sms_fallback,
        'verification_type': verification_type,
    }


def _parse_signup_location(request):
    """Validate country + city (Congo also requires a department) from POST
    data. Returns (country, location, city, error) — location is the Congo
    department code, blank for every other country."""
    from django.utils.translation import gettext as _
    from manager.congo_locations import (
        normalize_congo_city,
        normalize_congo_location,
        normalize_country,
        normalize_country_city,
    )
    country = normalize_country(request.POST.get('country', '') or 'Congo')
    if not country:
        return None, None, None, _('Please select your country.')

    if country == 'Congo':
        dept = normalize_congo_location(request.POST.get('location', ''))
        if not dept:
            return None, None, None, _('Please select your department.')
        city = normalize_congo_city(dept, request.POST.get('city', ''))
        if not city:
            return None, None, None, _('Please select your city.')
        return country, dept, city, None

    city = normalize_country_city(country, request.POST.get('city', ''))
    if not city:
        return None, None, None, _('Please select your city.')
    return country, '', city, None


def _signup_page_context():
    from manager.twilio_verify import is_twilio_verify_enabled
    return {'twilio_sms_enabled': is_twilio_verify_enabled()}


def user_register(request):
    """Public user registration - Step 1: collect info & send PIN"""
    from django.utils import translation
    from django.utils.translation import gettext as _

    if request.method == 'POST':
        ip = _get_client_ip(request)
        rl_key = f'signup_fail:{ip}'
        if _is_rate_limited_key(rl_key, 5):
            return JsonResponse({
                'success': False,
                'message': _('Too many attempts. Please try again in a few minutes.'),
            }, status=429)
        _record_attempt_key(rl_key, 300)

        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        password2 = request.POST.get('password2', '').strip()
        phone = request.POST.get('phone', '').strip()
        country, location, city, loc_err = _parse_signup_location(request)
        if loc_err:
            return JsonResponse({'success': False, 'message': loc_err})

        if not all([name, email, phone, password]):
            return JsonResponse({
                'success': False,
                'message': _('Please fill in all required fields, including phone.'),
            })
        if password != password2:
            return JsonResponse({'success': False, 'message': _('Passwords do not match.')})
        if len(password) < 6:
            return JsonResponse({
                'success': False,
                'message': _('Password must be at least 6 characters.'),
            })
        if models.SiteUser.objects.filter(email=email).exists():
            return JsonResponse({'success': False, 'message': _('This email is already registered.')})

        from manager.twilio_verify import normalize_phone_e164, validate_phone_e164
        phone_e164 = normalize_phone_e164(phone)
        phone_valid, phone_err = validate_phone_e164(phone_e164)
        if not phone_valid:
            return JsonResponse({'success': False, 'message': phone_err})

        # Generate PIN and store pending registration
        pin_code = _generate_pin()
        expires_at = timezone.now() + timedelta(minutes=15)

        # Delete any previous pending verifications for this email
        models.EmailVerification.objects.filter(email=email, is_verified=False).delete()

        models.EmailVerification.objects.create(
            email=email,
            pin_code=pin_code,
            name=name,
            password=_hash_password(password),
            phone=phone_e164 or phone,
            country=country,
            location=location,
            city=city,
            expires_at=expires_at,
        )

        verification = models.EmailVerification.objects.get(
            email=email, is_verified=False, verification_type='user',
        )
        result = _dispatch_signup_verification(
            verification,
            f'/manager/public/user/verify-email/?email={email}',
            translation.get_language(),
        )
        return JsonResponse(result)

    return render(request, 'public/user_register.html', _signup_page_context())


def verify_email_pin(request):
    """Step 2: User enters PIN to complete registration"""
    from django.utils.translation import gettext as _

    email = request.GET.get('email', '') or request.POST.get('email', '')

    if request.method == 'POST':
        pin = request.POST.get('pin', '').strip()
        phone_pin = request.POST.get('phone_pin', '').strip()
        email = request.POST.get('email', '').strip()

        if not pin or not email:
            return JsonResponse({'success': False, 'message': _('Please enter the verification code.')})

        try:
            verification = models.EmailVerification.objects.get(
                email=email, is_verified=False
            )
        except models.EmailVerification.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': _('Verification record not found. Please register again.'),
            })

        if verification.is_expired():
            return JsonResponse({
                'success': False,
                'message': _('Code expired. Please request a new one.'),
            })

        if verification.pin_code != pin:
            return JsonResponse({'success': False, 'message': _('Incorrect verification code.')})

        phone_ok, phone_err = _verify_phone_otp_if_required(verification, phone_pin)
        if not phone_ok:
            return JsonResponse({'success': False, 'message': phone_err})

        if models.SiteUser.objects.filter(email=email).exists():
            return JsonResponse({'success': False, 'message': _('This email is already registered.')})

        from manager.congo_locations import DEFAULT_COUNTRY, DEFAULT_CONGO_CITY
        user = models.SiteUser.objects.create(
            name=verification.name,
            email=verification.email,
            password=verification.password,
            phone=verification.phone,
            country=verification.country or DEFAULT_COUNTRY,
            location=verification.location,
            city=verification.city or DEFAULT_CONGO_CITY,
        )
        _sync_password_by_email(verification.email, verification.password)
        _link_dual_accounts_by_email(verification.email)
        verification.is_verified = True
        verification.phone_verified = True
        verification.save(update_fields=['is_verified', 'phone_verified'])

        # Create notification for new user
        create_notification(
            'new_user',
            f'新用户注册: {user.name}',
            f'{user.name} ({user.email}) 完成了注册',
            icon='fas fa-user-plus',
            color='#10b981',
            link='/manager/admin/users/',
            related_id=user.id,
        )

        # Auto-login
        request.session['site_user_id'] = user.id
        request.session['site_user_name'] = user.name

        return JsonResponse({
            'success': True,
            'message': _('Verification successful. Registration complete!'),
            'redirect': '/manager/public/user/profile/',
        })

    ctx = _verification_page_context(email, verification_type='user', request=request)
    return render(request, 'public/verify_email.html', ctx)


def resend_verification_pin(request):
    """Resend a new PIN code for pending registration"""
    from django.utils.translation import gettext as _

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': _('Invalid request.')})

    email = request.POST.get('email', '').strip()
    verification_type = request.POST.get('verification_type', 'user').strip() or 'user'
    if not email:
        return JsonResponse({'success': False, 'message': _('Email is required.')})

    try:
        verification = models.EmailVerification.objects.get(
            email=email, is_verified=False, verification_type=verification_type,
        )
    except models.EmailVerification.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': _('No pending verification found. Please register again.'),
        })

    new_pin = _generate_pin()
    verification.pin_code = new_pin
    verification.expires_at = timezone.now() + timedelta(minutes=15)
    verification.save(update_fields=['pin_code', 'expires_at'])

    sent = _send_verification_email(email, new_pin, verification.name)
    if not sent:
        return JsonResponse({
            'success': False,
            'message': _('Failed to send email. Please try again later.'),
        })

    return JsonResponse({'success': True, 'message': _('A new verification code has been sent.')})


def resend_phone_verification(request):
    """Resend Twilio Verify SMS for pending registration."""
    from django.utils.translation import gettext as _

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': _('Invalid request.')})

    email = request.POST.get('email', '').strip()
    verification_type = request.POST.get('verification_type', 'user').strip() or 'user'
    if not email:
        return JsonResponse({'success': False, 'message': _('Email is required.')})

    try:
        verification = models.EmailVerification.objects.get(
            email=email, is_verified=False, verification_type=verification_type,
        )
    except models.EmailVerification.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': _('No pending verification found. Please register again.'),
        })

    if verification.is_expired():
        return JsonResponse({
            'success': False,
            'message': _('Verification expired. Please register again.'),
        })

    sms_ok, sms_err = _send_registration_phone_otp(verification.phone)
    if not sms_ok:
        verification.require_sms_verification = False
        verification.save(update_fields=['require_sms_verification'])
        return JsonResponse({
            'success': False,
            'message': sms_err,
            'sms_disabled': True,
            'fallback_message': _(
                'SMS could not be sent. Use the email code or try again later.'
            ),
        })

    return JsonResponse({'success': True, 'message': _('New SMS code sent.')})


def _is_likely_desktop_browser(request):
    ua = (request.META.get('HTTP_USER_AGENT') or '').lower()
    if not ua:
        return True
    mobile_markers = (
        'mobile', 'android', 'iphone', 'ipad', 'ipod', 'blackberry',
        'windows phone', 'webos',
    )
    return not any(m in ua for m in mobile_markers)


def user_login(request):
    """Public user login"""
    from django.utils.translation import gettext as _

    if request.method == 'POST':
        ip = _get_client_ip(request)
        if _is_rate_limited(ip):
            return JsonResponse({
                'success': False,
                'message': _('Too many login attempts. Please try again in 5 minutes.'),
            }, status=429)

        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()

        if not all([email, password]):
            return JsonResponse({
                'success': False,
                'message': _('Please enter email and password.'),
            })

        try:
            user = models.SiteUser.objects.get(email__iexact=email, is_active=True)
        except models.SiteUser.DoesNotExist:
            _record_login_failure(ip)
            return JsonResponse({'success': False, 'message': _('Incorrect email or password.')})

        if not _check_email_password(email, password):
            _record_login_failure(ip)
            return JsonResponse({'success': False, 'message': _('Incorrect email or password.')})

        _reset_login_failures(ip)
        _link_dual_accounts_by_email(email)

        request.session.cycle_key()
        request.session['site_user_id'] = user.id
        request.session['site_user_name'] = user.name
        request.session['auth_scope'] = 'user'
        linked_vendor = models.Vendor.objects.filter(
            user_id=user.id, is_active=True, status='approved',
        ).first()
        request.session['has_vendor_access'] = bool(linked_vendor)
        if linked_vendor:
            request.session['vendor_id'] = linked_vendor.id
            request.session['vendor_name'] = linked_vendor.company_name
        next_url = (request.POST.get('next') or '').strip()
        if not next_url or not next_url.startswith('/') or next_url.startswith('//'):
            from django.urls import reverse
            next_url = (
                reverse('manager:user_profile')
                if _is_likely_desktop_browser(request)
                else '/manager/public/'
            )
        return JsonResponse({'success': True, 'message': _('Login successful.'), 'redirect': next_url})

    return render(request, 'public/user_login.html')


def user_logout(request):
    """Public user logout — flush the whole session to prevent fixation."""
    # Preserve vendor session if this is a dual-role account logging out of user scope only
    vendor_id = request.session.get('vendor_id')
    vendor_name = request.session.get('vendor_name')
    request.session.flush()
    if vendor_id:
        request.session['vendor_id'] = vendor_id
        request.session['vendor_name'] = vendor_name
        request.session['auth_scope'] = 'vendor'
    return redirect('/manager/public/')


def user_profile(request):
    """User profile page with orders, favorites, and course progress"""
    user_id = request.session.get('site_user_id')
    if not user_id:
        return redirect('/manager/public/user/login/')

    user = get_object_or_404(models.SiteUser, id=user_id, is_active=True)

    if request.method == 'POST':
        user.name = request.POST.get('name', user.name).strip()
        user.phone = request.POST.get('phone', '').strip()
        country, dept, city, loc_err = _parse_signup_location(request)
        if not loc_err:
            user.country = country
            user.location = dept
            user.city = city
        if 'avatar' in request.FILES:
            user.avatar = request.FILES['avatar']
        user.save()
        request.session['site_user_name'] = user.name
        return JsonResponse({'success': True, 'message': '资料已更新'})

    # Get book orders
    book_orders = models.Order.objects.filter(
        customer_email=user.email
    ).prefetch_related('orderitem_set__book').order_by('-created_at')[:10]

    from django.db.models import Q

    # Get marketplace orders (email or linked user id)
    mkt_orders = list(
        MarketplaceOrder.objects.filter(
            Q(user_email=user.email) | Q(user_id=user.id)
        ).order_by('-created_at')[:10]
    )

    # Merge and sort all orders
    all_orders = []
    for o in book_orders:
        all_orders.append({
            'type': 'book',
            'order': o,
            'order_number': o.order_number,
            'created_at': o.created_at,
            'total_amount': o.total_amount,
            'status_display': o.get_status_display(),
            'items': list(o.orderitem_set.all()),
        })
    for o in mkt_orders:
        all_orders.append({
            'type': 'marketplace',
            'order': o,
            'order_number': o.order_number,
            'created_at': o.created_at,
            'total_amount': o.total_amount,
            'status_display': o.get_status_display(),
            'items': list(o.items.all()),
        })
    all_orders.sort(key=lambda x: x['created_at'], reverse=True)

    # Wishlists - all types
    wishlists = models.Wishlist.objects.filter(user=user).select_related('book')
    wishlist_data = []
    for w in wishlists:
        if w.item_type == 'book' and w.book:
            wishlist_data.append({
                'id': w.id,
                'item_type': 'book',
                'item_id': w.book.id,
                'name': w.book.name,
                'price': w.book.price,
                'image_url': w.book.get_cover_url() if hasattr(w.book, 'get_cover_url') else '',
                'detail_url': f'/manager/public/books/{w.book.id}/',
            })
        else:
            item = w.get_item()
            if item:
                if w.item_type == 'course':
                    detail_url = f'/marketplace/courses/{item.slug}/'
                elif w.item_type == 'product':
                    detail_url = f'/marketplace/products/{item.slug}/'
                elif w.item_type == 'supermarket':
                    detail_url = f'/marketplace/supermarket/{item.slug}/'
                else:
                    detail_url = '#'
                wishlist_data.append({
                    'id': w.id,
                    'item_type': w.item_type,
                    'item_id': w.item_id,
                    'name': w.get_item_name(),
                    'price': w.get_item_price(),
                    'image_url': w.get_item_image_url(),
                    'detail_url': detail_url,
                })

    # Course progress
    session_key = get_session_key(request)
    courses_progress = []
    try:
        # Get all courses the user has progress in
        progress_records = CourseProgress.objects.filter(
            session_key=session_key
        ).select_related('course', 'lesson')

        course_ids = set(p.course_id for p in progress_records)
        for course_id in course_ids:
            course = Course.objects.filter(pk=course_id).first()
            if not course:
                continue
            total_lessons = CourseLesson.objects.filter(
                section__course=course
            ).count()
            completed_lessons = progress_records.filter(
                course_id=course_id, completed=True
            ).count()
            progress_pct = int((completed_lessons / total_lessons * 100)) if total_lessons > 0 else 0
            courses_progress.append({
                'course': course,
                'total_lessons': total_lessons,
                'completed_lessons': completed_lessons,
                'progress_pct': progress_pct,
                'detail_url': f'/marketplace/courses/{course.slug}/',
            })
    except Exception:
        pass

    followed_vendors = list(
        models.UserFollowedVendor.objects.filter(user=user).select_related('vendor').order_by('-followed_at')
    )
    followed_publishers = list(
        models.UserFollowedShop.objects.filter(user=user).select_related('publisher').order_by('-followed_at')
    )

    context = {
        'site_user': user,
        'orders': book_orders,  # backward compat
        'all_orders': all_orders,
        'wishlists': wishlists,  # backward compat
        'wishlist_data': wishlist_data,
        'courses_progress': courses_progress,
        'followed_vendors': followed_vendors,
        'followed_publishers': followed_publishers,
        'followed_total_count': len(followed_vendors) + len(followed_publishers),
    }
    from . import views_review

    context['pending_review_items'] = views_review.collect_pending_reviews_for_user(user)
    # Loyalty / gamification
    try:
        loyalty = models.LoyaltyPoints.objects.filter(user=user).first()
        point_transactions = models.PointTransaction.objects.filter(
            user=user
        ).order_by('-created_at')[:15] if loyalty else []
        context['loyalty'] = loyalty
        context['point_transactions'] = list(point_transactions)
    except Exception:
        context['loyalty'] = None
        context['point_transactions'] = []
    return render(request, 'public/user_profile.html', context)


def user_toggle_wishlist(request):
    """Toggle item in wishlist via AJAX - supports books and marketplace items"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request'})

    user_id = request.session.get('site_user_id')
    if not user_id:
        return JsonResponse({'success': False, 'message': '请先登录', 'login_required': True})

    item_type = request.POST.get('item_type', 'book')
    item_id = request.POST.get('item_id') or request.POST.get('book_id')
    if not item_id:
        return JsonResponse({'success': False, 'message': '缺少参数'})

    item_id = int(item_id)

    if item_type == 'book':
        book = get_object_or_404(models.Book, id=item_id)
        wish, created = models.Wishlist.objects.get_or_create(
            user_id=user_id, book=book, item_type='book',
            defaults={'item_id': item_id}
        )
    else:
        # Marketplace item - verify it exists
        item = _get_marketplace_item(item_type, item_id)
        if not item:
            return JsonResponse({'success': False, 'message': '商品不存在'})
        wish, created = models.Wishlist.objects.get_or_create(
            user_id=user_id, item_type=item_type, item_id=item_id,
            defaults={'book': None}
        )

    if not created:
        wish.delete()
        return JsonResponse({'success': True, 'wishlisted': False, 'message': '已取消收藏'})

    user = models.SiteUser.objects.filter(pk=user_id).first()
    from manager.escrow_service import resolve_vendor_for_item
    if item_type == 'book':
        item_name = book.name
        vendor = resolve_vendor_for_item('book', item_id)
    else:
        item_name = getattr(item, 'name', None) or getattr(item, 'title', 'Item')
        vendor = resolve_vendor_for_item(item_type, item_id)
    if vendor:
        create_vendor_notification(
            vendor.id,
            'wishlist_add',
            f'{user.name if user else "A customer"} added an item to favorites',
            f'{item_name} was added to a customer wishlist.',
            icon='fas fa-heart',
            color='#ec4899',
            link='/manager/vendor/dashboard/',
            related_id=wish.id,
        )
    return JsonResponse({'success': True, 'wishlisted': True, 'message': '已添加到收藏'})


def user_check_wishlist(request):
    """Check if an item is in the user's wishlist"""
    user_id = request.session.get('site_user_id')
    if not user_id:
        return JsonResponse({'wishlisted': False})

    item_type = request.GET.get('item_type', 'book')
    item_id = request.GET.get('item_id') or request.GET.get('book_id')

    if item_type == 'book':
        wishlisted = models.Wishlist.objects.filter(user_id=user_id, book_id=item_id, item_type='book').exists()
    else:
        wishlisted = models.Wishlist.objects.filter(user_id=user_id, item_type=item_type, item_id=int(item_id)).exists()

    return JsonResponse({'wishlisted': wishlisted})


# Admin CRUD for site users
def admin_site_users(request):
    """Admin view to manage site users"""
    if "name" not in request.session:
        return redirect('/manager/login/')

    users = models.SiteUser.objects.all()
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    if search:
        users = users.filter(Q(name__icontains=search) | Q(email__icontains=search) | Q(phone__icontains=search))
    if status_filter == 'active':
        users = users.filter(is_active=True)
    elif status_filter == 'inactive':
        users = users.filter(is_active=False)

    context = {
        'users': users,
        'search_query': search,
        'status_filter': status_filter,
        'total_users': models.SiteUser.objects.count(),
        'name': request.session.get('name', ''),
    }
    return render(request, 'admin/site_users.html', context)


def admin_toggle_user(request):
    """Toggle user active status"""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '未授权'})
    if request.method == 'POST':
        user = get_object_or_404(models.SiteUser, id=request.POST.get('id'))
        user.is_active = not user.is_active
        user.save(update_fields=['is_active'])
        return JsonResponse({'success': True, 'is_active': user.is_active})
    return JsonResponse({'success': False})


def _detach_vendor_protected_relations(vendor):
    """Detach historical records that must remain after a vendor account is removed."""
    if not vendor:
        return 0
    return models.PlatformEscrowTransaction.objects.filter(vendor=vendor).update(vendor=None)


def admin_delete_user(request):
    """Delete a site user"""
    if not request.session.get('is_admin') or 'name' not in request.session:
        return JsonResponse({'success': False, 'message': '未授权'})
    if request.method == 'POST':
        user_id = request.POST.get('id')
        user = models.SiteUser.objects.filter(id=user_id).first()
        if not user:
            return JsonResponse({'success': False, 'message': '用户不存在'})
        # Prevent deleting the last active user
        if models.SiteUser.objects.filter(is_active=True).count() <= 1:
            return JsonResponse({'success': False, 'message': '无法删除最后一个用户'})
        deleted_vendor_count = 0
        vendor = getattr(user, 'vendor_profile', None)
        if vendor:
            _detach_vendor_protected_relations(vendor)
            vendor.delete()
            deleted_vendor_count = 1
        user.delete()
        return JsonResponse({'success': True, 'message': '用户已删除，关联卖家与店铺也已删除' if deleted_vendor_count else '用户已删除'})
    return JsonResponse({'success': False})


# ==========================================
# Vendor / Seller Views
# ==========================================

def vendor_register(request):
    """Vendor registration page - Step 1: collect info & send PIN"""
    from django.utils import translation
    from django.utils.translation import gettext as _

    if request.method == 'POST':
        company_name = request.POST.get('company_name', '').strip()
        contact_name = request.POST.get('contact_name', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        password2 = request.POST.get('password2', '').strip()

        if not all([company_name, contact_name, email, password]):
            return JsonResponse({
                'success': False,
                'message': _('Please fill in all required fields.'),
            })
        if password != password2:
            return JsonResponse({'success': False, 'message': _('Passwords do not match.')})
        if len(password) < 6:
            return JsonResponse({
                'success': False,
                'message': _('Password must be at least 6 characters.'),
            })
        if models.Vendor.objects.filter(email=email).exists():
            return JsonResponse({
                'success': False,
                'message': _('This email is already registered as a vendor.'),
            })

        phone = request.POST.get('phone', '').strip()
        if not phone:
            return JsonResponse({
                'success': False,
                'message': _('Phone number is required.'),
            })

        country, location, city, loc_err = _parse_signup_location(request)
        if loc_err:
            return JsonResponse({'success': False, 'message': loc_err})

        from manager.twilio_verify import normalize_phone_e164, validate_phone_e164
        phone_e164 = normalize_phone_e164(phone)
        phone_valid, phone_err = validate_phone_e164(phone_e164)
        if not phone_valid:
            return JsonResponse({'success': False, 'message': phone_err})

        pin_code = _generate_pin()
        expires_at = timezone.now() + timedelta(minutes=15)

        # Clean old unverified records for this email
        models.EmailVerification.objects.filter(email=email, is_verified=False).delete()

        models.EmailVerification.objects.create(
            email=email,
            pin_code=pin_code,
            name=contact_name,
            password=_hash_password(password),
            phone=phone_e164 or phone,
            country=country,
            location=location,
            city=city,
            verification_type='vendor',
            company_name=company_name,
            description=request.POST.get('description', '').strip(),
            expires_at=expires_at,
        )

        # Handle logo: store temporarily in session
        if 'logo' in request.FILES:
            logo_file = request.FILES['logo']
            import os
            from django.conf import settings as django_conf_settings
            tmp_dir = os.path.join(django_conf_settings.MEDIA_ROOT, 'tmp_vendor_logos')
            os.makedirs(tmp_dir, exist_ok=True)
            tmp_path = os.path.join(tmp_dir, f'{email}_{logo_file.name}')
            with open(tmp_path, 'wb') as f:
                for chunk in logo_file.chunks():
                    f.write(chunk)
            request.session['vendor_tmp_logo'] = tmp_path
            request.session['vendor_tmp_logo_name'] = logo_file.name

        verification = models.EmailVerification.objects.get(
            email=email, is_verified=False, verification_type='vendor',
        )
        result = _dispatch_signup_verification(
            verification,
            f'/manager/vendor/verify-email/?email={email}',
            translation.get_language(),
        )
        return JsonResponse(result)

    return render(request, 'public/vendor_register.html', _signup_page_context())


def verify_vendor_pin(request):
    """Step 2: Vendor enters PIN to complete registration"""
    from django.utils.translation import gettext as _

    email = request.GET.get('email', '') or request.POST.get('email', '')

    if request.method == 'POST':
        pin = request.POST.get('pin', '').strip()
        phone_pin = request.POST.get('phone_pin', '').strip()
        email = request.POST.get('email', '').strip()

        if not pin or not email:
            return JsonResponse({'success': False, 'message': _('Please enter the verification code.')})

        try:
            verification = models.EmailVerification.objects.get(
                email=email, is_verified=False, verification_type='vendor'
            )
        except models.EmailVerification.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': _('Verification record not found. Please register again.'),
            })

        if verification.is_expired():
            return JsonResponse({
                'success': False,
                'message': _('Code expired. Please request a new one.'),
            })

        if verification.pin_code != pin:
            return JsonResponse({'success': False, 'message': _('Incorrect verification code.')})

        phone_ok, phone_err = _verify_phone_otp_if_required(verification, phone_pin)
        if not phone_ok:
            return JsonResponse({'success': False, 'message': phone_err})

        if models.Vendor.objects.filter(email=email).exists():
            return JsonResponse({
                'success': False,
                'message': _('This email is already registered as a vendor.'),
            })

        from manager.congo_locations import DEFAULT_COUNTRY, DEFAULT_CONGO_CITY
        vendor = models.Vendor.objects.create(
            company_name=verification.company_name,
            contact_name=verification.name,
            email=verification.email,
            password=verification.password,
            phone=verification.phone,
            country=verification.country or DEFAULT_COUNTRY,
            location=verification.location,
            city=verification.city or DEFAULT_CONGO_CITY,
            description=verification.description,
        )
        site_user = models.SiteUser.objects.filter(email__iexact=email, is_active=True).first()
        if site_user:
            vendor.user = site_user
            vendor.save(update_fields=['user', 'updated_at'])
            site_user.promote_to_seller()
            if verification.city:
                site_user.country = verification.country or DEFAULT_COUNTRY
                site_user.location = verification.location
                site_user.city = verification.city
                site_user.save(update_fields=['country', 'location', 'city', 'updated_at'])
        _sync_password_by_email(email, verification.password)
        _link_dual_accounts_by_email(email)

        # Handle temporary logo file
        tmp_logo = request.session.pop('vendor_tmp_logo', None)
        tmp_logo_name = request.session.pop('vendor_tmp_logo_name', None)
        if tmp_logo and tmp_logo_name:
            import os
            if os.path.exists(tmp_logo):
                from django.core.files import File
                with open(tmp_logo, 'rb') as f:
                    vendor.logo.save(tmp_logo_name, File(f), save=True)
                os.remove(tmp_logo)

        verification.is_verified = True
        verification.phone_verified = True
        verification.save(update_fields=['is_verified', 'phone_verified'])

        create_notification(
            'vendor_registered',
            f'新卖家注册: {vendor.company_name}',
            f'{vendor.company_name} ({vendor.email}) 完成邮箱验证并注册了卖家账户，等待审核',
            icon='fas fa-store',
            color='#667eea',
            link='/manager/admin/vendors/',
            related_id=vendor.id,
        )

        request.session['vendor_id'] = vendor.id
        request.session['vendor_name'] = vendor.company_name

        return JsonResponse({
            'success': True,
            'message': _('Verification successful. Your vendor account is pending approval.'),
            'redirect': '/manager/vendor/dashboard/',
        })

    return render(request, 'public/verify_vendor_email.html', _verification_page_context(email, verification_type='vendor', request=request))


def vendor_login(request):
    """Vendor login"""
    from django.utils.translation import gettext as _

    if request.method == 'POST':
        ip = _get_client_ip(request)
        if _is_rate_limited(ip):
            return JsonResponse({
                'success': False,
                'message': _('Too many login attempts. Please try again in 5 minutes.'),
            }, status=429)

        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()

        if not all([email, password]):
            return JsonResponse({
                'success': False,
                'message': _('Please enter email and password.'),
            })

        try:
            vendor = models.Vendor.objects.get(email__iexact=email, is_active=True)
        except models.Vendor.DoesNotExist:
            _record_login_failure(ip)
            return JsonResponse({'success': False, 'message': _('Incorrect email or password.')})

        if not _check_email_password(email, password):
            _record_login_failure(ip)
            return JsonResponse({'success': False, 'message': _('Incorrect email or password.')})

        _reset_login_failures(ip)
        vendor.refresh_from_db()
        _link_dual_accounts_by_email(email)

        request.session.cycle_key()
        request.session['vendor_id'] = vendor.id
        request.session['vendor_name'] = vendor.company_name
        request.session['auth_scope'] = 'vendor'
        if vendor.user_id and vendor.user and vendor.user.is_active:
            request.session['site_user_id'] = vendor.user_id
            request.session['site_user_name'] = vendor.user.name
            request.session['has_vendor_access'] = True
        return JsonResponse({
            'success': True,
            'message': _('Login successful.'),
            'redirect': '/manager/vendor/dashboard/',
        })

    return render(request, 'public/vendor_login.html')


def vendor_logout(request):
    """Vendor logout — keeps user session alive for dual-role accounts."""
    site_user_id = request.session.get('site_user_id')
    site_user_name = request.session.get('site_user_name')
    request.session.flush()
    if site_user_id:
        request.session['site_user_id'] = site_user_id
        request.session['site_user_name'] = site_user_name
        request.session['auth_scope'] = 'user'
    return redirect('/manager/public/')


# ==========================================
# Forgot Password / Password Reset
# ==========================================

def _send_reset_email(email, pin_code, name):
    """Send password reset PIN code (English + French)."""
    subject = 'DUNO 360 - Password Reset / Réinitialisation du mot de passe'
    html_body = f'''
    <div style="font-family:'Segoe UI',Arial,sans-serif;max-width:520px;margin:0 auto;background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);">
        <div style="background:linear-gradient(135deg,#14245f 0%,#1d4ed8 100%);padding:32px 28px;text-align:center;">
            <h1 style="color:#fff;margin:0;font-size:1.5rem;">🔐 DUNO 360</h1>
            <p style="color:rgba(255,255,255,0.85);margin:8px 0 0;font-size:0.95rem;">Password Reset / Réinitialisation du mot de passe</p>
        </div>
        <div style="padding:32px 28px;">
            <p style="color:#333;font-size:1rem;margin:0 0 8px;">Hello <strong>{name}</strong>! / Bonjour <strong>{name}</strong>&nbsp;!</p>
            <p style="color:#666;font-size:0.93rem;line-height:1.7;margin:0 0 12px;">
                You requested a password reset. Use the code below:
            </p>
            <p style="color:#666;font-size:0.93rem;line-height:1.7;margin:0 0 24px;">
                Vous avez demandé une réinitialisation de mot de passe. Utilisez le code ci-dessous&nbsp;:
            </p>
            <div style="background:linear-gradient(135deg,rgba(20,36,95,0.08),rgba(29,78,216,0.08));border:2px dashed #1d4ed8;border-radius:14px;padding:24px;text-align:center;margin:0 0 24px;">
                <span style="font-size:2.5rem;font-weight:800;letter-spacing:12px;color:#1d4ed8;">{pin_code}</span>
            </div>
            <p style="color:#999;font-size:0.85rem;text-align:center;margin:0;">
                ⏰ This code expires in <strong>15 minutes</strong>. / Ce code expire dans <strong>15 minutes</strong>.
            </p>
        </div>
        <div style="background:#f8f9ff;padding:16px 28px;text-align:center;border-top:1px solid #eee;">
            <p style="color:#aaa;font-size:0.8rem;margin:0;">If you did not request a password reset, please ignore this email.<br>Si vous n'avez pas demandé de réinitialisation, ignorez cet e-mail.</p>
        </div>
    </div>
    '''
    plain_body = (
        f'Hello {name}, your password reset code is: {pin_code}. Valid for 15 minutes.\n\n'
        f'Bonjour {name}, votre code de réinitialisation est : {pin_code}. Valide 15 minutes.'
    )

    try:
        from django.core.mail import EmailMultiAlternatives
        msg = EmailMultiAlternatives(subject, plain_body, django_settings.DEFAULT_FROM_EMAIL, [email])
        msg.attach_alternative(html_body, 'text/html')
        msg.send(fail_silently=False)
        return True
    except Exception as e:
        logger.error(f'Failed to send reset email to {email}: {e}')
        return False


def forgot_password(request):
    """Step 1: User/Vendor enters email to receive reset PIN"""
    from django.utils.translation import gettext as _

    account_type = request.GET.get('type', 'user')

    if request.method == 'POST':
        ip = _get_client_ip(request)
        rl_key = f'pwreset_fail:{ip}'
        if _is_rate_limited_key(rl_key, 5):
            return JsonResponse({
                'success': False,
                'message': _('Too many attempts. Please try again in a few minutes.'),
            }, status=429)
        _record_attempt_key(rl_key, 300)

        email = request.POST.get('email', '').strip()
        account_type = request.POST.get('account_type', 'user')

        if not email:
            return JsonResponse({'success': False, 'message': _('Please enter your email.')})

        name = ''
        if account_type == 'vendor':
            try:
                vendor = models.Vendor.objects.get(email=email, is_active=True)
                name = vendor.contact_name
            except models.Vendor.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': _('No vendor account found for this email.'),
                })
        else:
            try:
                user = models.SiteUser.objects.get(email=email, is_active=True)
                name = user.name
            except models.SiteUser.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': _('No user account found for this email.'),
                })

        models.EmailVerification.objects.filter(
            email=email, verification_type='password_reset', is_verified=False
        ).delete()

        pin_code = _generate_pin()
        models.EmailVerification.objects.create(
            email=email,
            pin_code=pin_code,
            name=name,
            password='',
            verification_type='password_reset',
            company_name=account_type,
            expires_at=timezone.now() + timezone.timedelta(minutes=15),
        )

        sent = _send_reset_email(email, pin_code, name)
        if sent:
            return JsonResponse({
                'success': True,
                'message': _('Verification code sent to your email.'),
                'email': email,
                'account_type': account_type,
            })
        if django_settings.DEBUG:
            return JsonResponse({
                'success': True,
                'message': _('Development mode: your code is %(code)s') % {'code': pin_code},
                'email': email,
                'account_type': account_type,
            })
        return JsonResponse({
            'success': False,
            'message': _('Failed to send email. Please try again later.'),
        })

    return render(request, 'public/forgot_password.html', {'account_type': account_type})


def reset_password_verify(request):
    """Step 2: Verify PIN and set new password"""
    from django.utils.translation import gettext as _

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        pin_code = request.POST.get('pin_code', '').strip()
        new_password = request.POST.get('new_password', '').strip()
        account_type = request.POST.get('account_type', 'user')

        if not all([email, pin_code, new_password]):
            return JsonResponse({'success': False, 'message': _('Please fill in all fields.')})

        if len(new_password) < 6:
            return JsonResponse({
                'success': False,
                'message': _('Password must be at least 6 characters.'),
            })

        try:
            verification = models.EmailVerification.objects.get(
                email=email,
                pin_code=pin_code,
                verification_type='password_reset',
                is_verified=False,
            )
        except models.EmailVerification.DoesNotExist:
            return JsonResponse({'success': False, 'message': _('Invalid verification code.')})

        if verification.is_expired():
            verification.delete()
            return JsonResponse({
                'success': False,
                'message': _('Code expired. Please request a new one.'),
            })

        hashed = _set_unified_password(email, new_password)
        user, vendor = _get_linked_site_user_and_vendor(email)
        if user or vendor:
            _link_dual_accounts_by_email(email)
            verification.is_verified = True
            verification.save()
            redirect_url = '/manager/vendor/login/' if account_type == 'vendor' else '/manager/public/user/login/'
            return JsonResponse({
                'success': True,
                'message': _('Password reset successful. Please sign in with your new password.'),
                'redirect': redirect_url,
            })
        return JsonResponse({
            'success': False,
            'message': _('Account not found or has been disabled.'),
        })

    return JsonResponse({'success': False, 'message': _('Invalid request.')})


def _get_vendor(request):
    """Get the active vendor profile without blocking ordinary user features."""
    vendor_id = request.session.get('vendor_id')
    if vendor_id:
        vendor = models.Vendor.objects.filter(id=vendor_id, is_active=True).first()
        if vendor:
            return vendor
        request.session.pop('vendor_id', None)
        request.session.pop('vendor_name', None)

    site_user_id = request.session.get('site_user_id')
    if site_user_id:
        vendor = models.Vendor.objects.filter(user_id=site_user_id, is_active=True, status='approved').first()
        if vendor:
            request.session['vendor_id'] = vendor.id
            request.session['vendor_name'] = vendor.company_name
            request.session['has_vendor_access'] = True
            return vendor
    return None


def publish_entry(request):
    """Guided publish entry with paid seller activation and item-type selection."""
    site_user_id = request.session.get('site_user_id')
    if not site_user_id:
        return redirect(f'/manager/public/user/login/?next=/manager/public/publish/')

    user = get_object_or_404(models.SiteUser, pk=site_user_id, is_active=True)
    vendor = models.Vendor.objects.filter(user=user).first()
    pending_payment = models.SellerActivationPayment.objects.filter(user=user, status__in=['pending', 'processing']).order_by('-created_at').first()

    if request.method == 'POST':
        action = request.POST.get('action', '').strip()
        if action == 'activate_vendor':
            phone = request.POST.get('payer_phone', '').strip() or user.phone
            payer_name = request.POST.get('payer_name', '').strip() or user.name
            if not phone:
                return JsonResponse({'success': False, 'message': '请输入付款手机号。'}, status=400)

            payment = pending_payment if pending_payment and pending_payment.status == 'pending' else None
            if not payment:
                payment = models.SellerActivationPayment.objects.create(
                    user=user,
                    vendor=vendor,
                    order_number=f'SVP-{timezone.now().strftime("%Y%m%d%H%M%S")}-{user.id}',
                    amount=Decimal('100.00'),
                    currency='XAF',
                    provider='pawapay',
                    status='pending',
                    payer_phone=phone,
                    payer_name=payer_name,
                )

            from manager.payments.pawapay import create_payment_page_session
            from django.utils import translation

            return_url = request.build_absolute_uri(
                reverse('manager:seller_activation_return', args=[payment.order_number])
            )
            lang = 'EN' if translation.get_language() == 'en' else 'FR'
            result = create_payment_page_session(
                amount=int(payment.amount),
                order_number=payment.order_number,
                return_url=return_url,
                phone_number=phone,
                country='Congo',
                reason='DUNO 360 - Activation vendeur',
                language=lang,
            )
            if not result.get('success'):
                payment.status = 'failed'
                payment.provider_message = (result.get('error') or "Échec de l'initialisation du paiement pawaPay.")[:255]
                payment.save(update_fields=['status', 'provider_message', 'updated_at'])
                return JsonResponse({
                    'success': False,
                    'message': result.get('error') or "Impossible d'ouvrir la page de paiement pawaPay. Veuillez réessayer.",
                }, status=400)

            payment.external_reference = result['deposit_id']
            payment.status = 'processing'
            payment.external_status = 'queued'
            payment.payer_phone = phone
            payment.payer_name = payer_name
            payment.save(update_fields=['external_reference', 'status', 'external_status', 'payer_phone', 'payer_name', 'updated_at'])
            request.session['pending_seller_activation_payment_id'] = payment.id

            return JsonResponse({
                'success': True,
                'redirect_url': result['redirect_url'],
            })

        if action == 'choose_type':
            if not vendor:
                return JsonResponse({'success': False, 'message': 'Veuillez activer votre espace vendeur d’abord.'}, status=400)
            paid_activation = models.SellerActivationPayment.objects.filter(user=user, status='paid').exists()
            if not paid_activation:
                return JsonResponse({'success': False, 'message': '请先完成 100 FCFA 卖家激活支付。'}, status=400)
            item_type = request.POST.get('item_type', '').strip()
            target_map = {
                'product': '/marketplace/vendor/products/add/',
                'supermarket': '/marketplace/vendor/supermarket/add/',
                'course': '/marketplace/vendor/courses/add/',
                'book': '/manager/vendor/add-book/',
            }
            target = target_map.get(item_type)
            if not target:
                return JsonResponse({'success': False, 'message': 'Type de publication invalide.'}, status=400)
            return JsonResponse({'success': True, 'redirect_url': target})

    seller_paid = models.SellerActivationPayment.objects.filter(user=user, status='paid').exists()
    if vendor and vendor.is_active and seller_paid:
        request.session['vendor_id'] = vendor.id
        request.session['vendor_name'] = vendor.company_name
        return render(request, 'public/publish_entry.html', {
            'site_user': user,
            'already_vendor': True,
            'vendor': vendor,
            'direct_type_choice': True,
            'seller_activation_paid': True,
            'pending_activation_payment': pending_payment,
        })

    context = {
        'site_user': user,
        'already_vendor': bool(vendor and vendor.is_active and seller_paid),
        'vendor': vendor,
        'direct_type_choice': False,
        'seller_activation_paid': seller_paid,
        'pending_activation_payment': pending_payment,
        'seller_activation_fee': Decimal('100.00'),
    }
    return render(request, 'public/publish_entry.html', context)


def _activate_seller_from_payment(payment):
    """Grant seller access after a SellerActivationPayment is verified paid via
    pawaPay. Idempotent and shared between seller_activation_return() (browser
    redirect back from the hosted Payment Page) and the async pawaPay webhook
    (manager/payments/views.py) — whichever confirms the payment first wins."""
    if payment.status == 'paid':
        return payment.vendor

    payment.status = 'paid'
    payment.external_status = 'completed'
    payment.provider_message = 'Paiement confirmé via pawaPay.'
    payment.paid_at = timezone.now()
    payment.save(update_fields=['status', 'external_status', 'provider_message', 'paid_at', 'updated_at'])

    user = payment.user
    vendor = payment.vendor or models.Vendor.objects.filter(user=user).first()
    if not vendor:
        vendor = models.Vendor.objects.create(
            user=user,
            company_name=f'{user.name} 的店铺',
            contact_name=user.name,
            email=user.email,
            phone=user.phone,
            password=user.password,
            description='Duno360 卖家',
            status='approved',
            is_active=True,
        )
    elif vendor.status == 'pending':
        vendor.status = 'approved'
        vendor.is_active = True
        vendor.save(update_fields=['status', 'is_active', 'updated_at'])

    payment.vendor = vendor
    payment.save(update_fields=['vendor', 'updated_at'])
    user.promote_to_seller()

    create_notification(
        'vendor_registered',
        f'卖家激活成功: {user.name}',
        f'{user.name} ({user.email}) 已完成 {payment.amount} {payment.currency} 的卖家激活支付并成功开通卖家身份。',
        icon='fas fa-store',
        color='#10b981',
        link='/manager/admin/vendors/',
        related_id=vendor.id,
    )
    return vendor


def seller_activation_return(request, order_number):
    """Landing view pawaPay's hosted Payment Page redirects back to once the
    seller-activation fee payment finishes — receives ?depositId=... and
    verifies the real status server-side (never trusts the redirect itself
    as proof of payment) before granting seller access."""
    payment = get_object_or_404(models.SellerActivationPayment, order_number=order_number)
    if payment.user_id != request.session.get('site_user_id'):
        messages.warning(request, 'Non autorisé ou session expirée.')
        return redirect('manager:public_home')

    deposit_id = request.GET.get('depositId') or payment.external_reference
    if payment.status not in ('paid', 'failed') and deposit_id:
        from manager.payments.pawapay import get_deposit_status_v2, normalize_pawapay_status
        result = get_deposit_status_v2(deposit_id)
        internal = normalize_pawapay_status(result.get('status', 'PENDING'))
        if internal == 'SUCCESSFUL':
            _activate_seller_from_payment(payment)
        elif internal == 'FAILED':
            payment.status = 'failed'
            payment.external_status = 'failed'
            payment.provider_message = (result.get('error') or 'Paiement échoué.')[:255]
            payment.save(update_fields=['status', 'external_status', 'provider_message', 'updated_at'])
        elif result.get('status') == 'NOT_FOUND':
            # PawaPay: "if the customer abandons the payment page, the deposit
            # will be NOT_FOUND and should be considered FAILED after 15
            # minutes" — without this, an abandoned Payment Page session left
            # the record stuck on 'processing' forever, blocking a clean retry.
            if _pawapay_deposit_expired(payment):
                payment.status = 'failed'
                payment.external_status = 'not_found'
                payment.provider_message = 'Paiement annulé ou expiré (page abandonnée).'
                payment.save(update_fields=['status', 'external_status', 'provider_message', 'updated_at'])
            else:
                messages.warning(request, 'Paiement annulé ou non terminé.')

    if payment.status == 'failed':
        messages.error(request, 'Le paiement a échoué. Veuillez réessayer.')
    elif payment.status == 'paid':
        messages.success(request, '✅ Paiement de 100 FCFA confirmé — votre espace vendeur est activé !')

    return redirect('manager:publish_entry')


def _pawapay_deposit_expired(payment, minutes=15):
    """True once a pawaPay Payment Page session's 15-minute window (per
    PawaPay docs) has elapsed since the payment attempt was created — used to
    safely convert a NOT_FOUND/abandoned deposit into 'failed' instead of
    leaving it stuck on 'processing' forever."""
    return (timezone.now() - payment.created_at).total_seconds() >= minutes * 60


@require_GET
def seller_activation_status(request):
    """AJAX polling endpoint used by publish_entry.html while a seller
    activation payment is still processing (same pattern as check_payment_status
    for orders, in manager/payments/views.py)."""
    site_user_id = request.session.get('site_user_id')
    if not site_user_id:
        return JsonResponse({'status': 'unknown'})

    payment_id = request.session.get('pending_seller_activation_payment_id')
    payment = None
    if payment_id:
        payment = models.SellerActivationPayment.objects.filter(id=payment_id, user_id=site_user_id).first()
    if not payment:
        # Session key can be lost across tabs/reloads — fall back to the
        # user's most recent in-flight attempt instead of erroring out.
        payment = models.SellerActivationPayment.objects.filter(
            user_id=site_user_id, status__in=['pending', 'processing'],
        ).order_by('-created_at').first()
    if not payment:
        return JsonResponse({'status': 'unknown'})

    if payment.status == 'processing' and payment.external_reference:
        from manager.payments.pawapay import get_deposit_status_v2, normalize_pawapay_status
        result = get_deposit_status_v2(payment.external_reference)
        internal = normalize_pawapay_status(result.get('status', 'PENDING'))
        if internal == 'SUCCESSFUL':
            _activate_seller_from_payment(payment)
            messages.success(request, '✅ Paiement de 100 FCFA confirmé — votre espace vendeur est activé !')
        elif internal == 'FAILED':
            payment.status = 'failed'
            payment.external_status = 'failed'
            payment.save(update_fields=['status', 'external_status', 'updated_at'])
        elif result.get('status') == 'NOT_FOUND' and _pawapay_deposit_expired(payment):
            payment.status = 'failed'
            payment.external_status = 'not_found'
            payment.provider_message = 'Paiement annulé ou expiré (page abandonnée).'
            payment.save(update_fields=['status', 'external_status', 'provider_message', 'updated_at'])

    return JsonResponse({'status': payment.status})


def _admin_apply_inventory_stock(item_type, item_id, action, delta, manual_value):
    """Update stock for a single admin inventory item."""
    if item_type == 'book' and item_id:
        item = models.Book.objects.get(id=item_id)
        if action == 'set' and manual_value is not None:
            item.inventory = max(0, int(manual_value))
        else:
            item.inventory = max(0, item.inventory + delta)
        item.save(update_fields=['inventory'])
    elif item_type == 'product' and item_id:
        item = Product.objects.get(id=item_id)
        if action == 'set' and manual_value is not None:
            item.stock = max(0, int(manual_value))
        else:
            item.stock = max(0, item.stock + delta)
        item.save(update_fields=['stock'])
    elif item_type == 'course' and item_id:
        item = Course.objects.get(id=item_id)
        if action == 'set' and manual_value is not None:
            item.stock = max(0, int(manual_value))
        else:
            item.stock = max(0, item.stock + delta)
        item.save(update_fields=['stock'])
    elif item_type == 'supermarket' and item_id:
        item = SupermarketItem.objects.get(id=item_id)
        if action == 'set' and manual_value is not None:
            item.stock = max(0, int(manual_value))
        else:
            item.stock = max(0, item.stock + delta)
        item.save(update_fields=['stock'])
    else:
        raise ValueError('Invalid inventory item.')


def _vendor_apply_inventory_stock(vendor, item_type, item_id, action, delta, manual_value):
    """Update stock for a single vendor inventory item."""
    if item_type == 'book' and item_id:
        vb = models.VendorBook.objects.get(id=item_id, vendor=vendor)
        if action == 'set' and manual_value is not None:
            vb.book.inventory = max(0, int(manual_value))
        else:
            vb.book.inventory = max(0, vb.book.inventory + delta)
        vb.book.save(update_fields=['inventory'])
    elif item_type == 'product' and item_id:
        p = Product.objects.get(id=item_id, vendor=vendor)
        if action == 'set' and manual_value is not None:
            p.stock = max(0, int(manual_value))
        else:
            p.stock = max(0, p.stock + delta)
        p.save(update_fields=['stock'])
    elif item_type == 'course' and item_id:
        c = Course.objects.get(id=item_id, vendor=vendor)
        if action == 'set' and manual_value is not None:
            c.stock = max(0, int(manual_value))
        else:
            c.stock = max(0, c.stock + delta)
        c.save(update_fields=['stock'])
    elif item_type == 'supermarket' and item_id:
        s = SupermarketItem.objects.get(id=item_id, vendor=vendor)
        if action == 'set' and manual_value is not None:
            s.stock = max(0, int(manual_value))
        else:
            s.stock = max(0, s.stock + delta)
        s.save(update_fields=['stock'])
    else:
        raise ValueError('Invalid inventory item.')


def _process_bulk_inventory(request, apply_fn, success_msg):
    """Apply bulk stock updates from POST data."""
    selected = request.POST.getlist('selected_items')
    action = request.POST.get('action')
    delta = int(request.POST.get('delta', 0) or 0)
    manual_value = request.POST.get('manual_value')
    if action == 'subtract':
        action = 'add'
        delta = -abs(delta)
    if not selected:
        messages.error(request, _('Select at least one item.'))
        return
    if action not in ('set', 'add'):
        messages.error(request, _('Invalid bulk action.'))
        return
    if action == 'add' and delta == 0:
        messages.error(request, _('Enter a quantity to add or subtract.'))
        return
    if action == 'set' and manual_value is None:
        messages.error(request, _('Enter a stock value to set.'))
        return
    updated = 0
    errors = []
    for key in selected:
        parts = key.split(':', 1)
        if len(parts) != 2:
            continue
        item_type, item_id = parts[0], parts[1]
        try:
            apply_fn(item_type, item_id, action, delta, manual_value)
            updated += 1
        except Exception as exc:
            errors.append(str(exc))
    if updated:
        messages.success(request, success_msg.format(count=updated))
    if errors:
        messages.error(request, '; '.join(errors[:3]))


def admin_inventory(request):
    """Admin inventory management for books, products, courses and supermarket items."""
    if 'name' not in request.session:
        return redirect('/manager/login')

    if request.method == 'POST':
        if request.POST.get('bulk') == '1':
            _process_bulk_inventory(
                request,
                lambda t, i, a, d, m: _admin_apply_inventory_stock(t, i, a, d, m),
                _('Stock updated for {count} item(s).'),
            )
            return redirect('manager:admin_inventory')

        item_type = request.POST.get('item_type')
        item_id = request.POST.get('item_id')
        action = request.POST.get('action')
        delta = int(request.POST.get('delta', 0) or 0)
        manual_value = request.POST.get('manual_value')
        try:
            _admin_apply_inventory_stock(item_type, item_id, action, delta, manual_value)
            messages.success(request, 'Stock mis à jour avec succès.')
        except Exception as exc:
            messages.error(request, str(exc))
        return redirect('manager:admin_inventory')

    books = models.Book.objects.select_related('publisher').all().order_by('-id')
    products = Product.objects.select_related('vendor').all().order_by('-id')
    courses = Course.objects.select_related('vendor').all().order_by('-id')
    supermarket_items = SupermarketItem.objects.select_related('vendor').all().order_by('-id')
    context = {
        'name': request.session['name'],
        'books': books,
        'products': products,
        'courses': courses,
        'supermarket_items': supermarket_items,
        'total_books': books.count(),
        'total_products': products.count(),
        'total_courses': courses.count(),
        'total_supermarket': supermarket_items.count(),
        'admin_mode': True,
    }
    return render(request, 'admin/inventory.html', context)


def admin_escrow_transactions(request):
    """Admin view: platform escrow holds and vendor payout releases."""
    if 'name' not in request.session:
        return redirect('/manager/login')

    from manager.escrow_service import (
        process_due_escrow_releases,
        admin_mark_escrow_delivered,
        admin_force_release_escrow,
        admin_cancel_escrow,
        admin_wallet_adjust,
        REFUND_HOLD_DAYS,
    )
    from manager.commission import commission_rates_for_display

    if request.method == 'POST':
        action = request.POST.get('action')
        tx_id = request.POST.get('tx_id')
        if action == 'release_due':
            count = process_due_escrow_releases()
            messages.success(request, _('{count} payout(s) released to vendors.').format(count=count))
        elif action == 'release_one' and tx_id:
            ok = admin_force_release_escrow(tx_id)
            messages.success(request, _('Payout released.')) if ok else messages.error(request, _('Could not release this payout.'))
        elif action == 'mark_delivered' and tx_id:
            ok = admin_mark_escrow_delivered(tx_id)
            messages.success(request, _('Marked as delivered — refund hold started.')) if ok else messages.error(request, _('Could not update this line.'))
        elif action == 'cancel' and tx_id:
            admin_cancel_escrow(tx_id, 'cancelled')
            messages.success(request, _('Escrow line cancelled.'))
        elif action == 'refund' and tx_id:
            admin_cancel_escrow(tx_id, 'refunded')
            messages.success(request, _('Escrow line marked as refunded.'))
        elif action == 'save_note' and tx_id:
            note = request.POST.get('notes', '').strip()
            models.PlatformEscrowTransaction.objects.filter(pk=tx_id).update(notes=note)
            messages.success(request, _('Note saved.'))
        elif action == 'wallet_adjust':
            vid = request.POST.get('vendor_id')
            amount = request.POST.get('amount')
            desc = request.POST.get('description', '').strip()
            try:
                ok = admin_wallet_adjust(int(vid), Decimal(amount), desc)
                if ok:
                    messages.success(request, _('Wallet updated.'))
                else:
                    messages.error(request, _('Wallet adjustment failed (check balance).'))
            except (TypeError, ValueError):
                messages.error(request, _('Invalid wallet adjustment.'))
        return redirect(request.get_full_path() or reverse('manager:admin_escrow'))

    status_filter = request.GET.get('status', '').strip()
    vendor_filter = request.GET.get('vendor_id', '').strip()
    item_type_filter = request.GET.get('item_type', '').strip()
    search_q = request.GET.get('q', '').strip()

    base_qs = models.PlatformEscrowTransaction.objects.all()
    qs = base_qs.select_related('vendor').order_by('-held_at')
    if status_filter:
        qs = qs.filter(status=status_filter)
    if vendor_filter:
        qs = qs.filter(vendor_id=vendor_filter)
    if item_type_filter:
        qs = qs.filter(item_type=item_type_filter)
    if search_q:
        qs = qs.filter(
            Q(transaction_ref__icontains=search_q)
            | Q(order_number__icontains=search_q)
            | Q(buyer_email__icontains=search_q)
            | Q(buyer_name__icontains=search_q)
            | Q(item_name__icontains=search_q)
        )

    stats = {}
    for row in base_qs.values('status').annotate(cnt=Count('id'), gross=Sum('gross_amount'), net=Sum('vendor_payout_amount')):
        stats[row['status']] = {'count': row['cnt'], 'gross': row['gross'], 'net': row['net']}

    wallet_totals = models.VendorWallet.objects.aggregate(
        total_balance=Sum('balance'),
        total_earned=Sum('total_earned'),
    )

    context = {
        'name': request.session['name'],
        'transactions': qs[:300],
        'status_filter': status_filter,
        'vendor_filter': vendor_filter,
        'item_type_filter': item_type_filter,
        'search_q': search_q,
        'status_choices': models.PlatformEscrowTransaction.STATUS_CHOICES,
        'item_type_choices': models.PlatformEscrowTransaction.ITEM_TYPE_CHOICES,
        'vendors': models.Vendor.objects.filter(is_active=True).order_by('company_name')[:200],
        'commission_rates': commission_rates_for_display(),
        'stats': stats,
        'stats_by_status': [
            {
                'code': code,
                'label': label,
                'count': stats.get(code, {}).get('count', 0),
                'net': stats.get(code, {}).get('net', 0) or 0,
            }
            for code, label in models.PlatformEscrowTransaction.STATUS_CHOICES
        ],
        'wallet_totals': wallet_totals,
        'refund_hold_days': REFUND_HOLD_DAYS,
    }
    return render(request, 'admin/escrow_transactions.html', context)


def public_vendor_shop(request, vendor_id):
    """Public vendor storefront page - Taobao style."""
    vendor = get_object_or_404(models.Vendor, pk=vendor_id, is_active=True)
    vendor_books = models.VendorBook.objects.filter(vendor=vendor, is_active=True).select_related('book', 'book__publisher')
    from marketplace.models import Product, Course, SupermarketItem
    vendor_products = Product.objects.filter(vendor=vendor, is_active=True)
    vendor_courses = Course.objects.filter(vendor=vendor, is_active=True)
    vendor_supermarket = SupermarketItem.objects.filter(vendor=vendor, is_active=True)

    is_following = False
    follower_count = models.UserFollowedVendor.objects.filter(vendor=vendor).count()
    user_id = request.session.get('site_user_id')
    if user_id:
        is_following = models.UserFollowedVendor.objects.filter(user_id=user_id, vendor=vendor).exists()

    total_sales = sum(vb.book.sale_num for vb in vendor_books)
    total_products = vendor_products.count()
    total_courses = vendor_courses.count()
    total_supermarket = vendor_supermarket.count()
    total_items = vendor_books.count() + total_products + total_courses + total_supermarket

    tab = request.GET.get('tab', 'all')
    if vendor.can_receive_certification() and not vendor.is_certified:
        vendor.is_certified = True
        if not vendor.certified_at:
            vendor.certified_at = timezone.now()
        vendor.save(update_fields=['is_certified', 'certified_at'])

    context = {
        'vendor': vendor,
        'vendor_books': vendor_books,
        'vendor_products': vendor_products,
        'vendor_courses': vendor_courses,
        'vendor_supermarket': vendor_supermarket,
        'is_following': is_following,
        'follower_count': follower_count,
        'active_tab': tab,
        'total_sales': total_sales,
        'total_products': total_products,
        'total_courses': total_courses,
        'total_supermarket': total_supermarket,
        'total_items': total_items,
        'certification_score': vendor.get_certification_score(),
        'certification_state': vendor.get_certification_state(),
    }
    return render(request, 'public/vendor_shop.html', context)


def _vendor_marketplace_insights_dict(vendor):
    """Analytics formerly on marketplace vendor dashboard ('product center'); merged into seller hub."""
    vendor_products = Product.objects.filter(vendor=vendor)
    vendor_courses = Course.objects.filter(vendor=vendor)
    vendor_supermarket = SupermarketItem.objects.filter(vendor=vendor)
    product_count = vendor_products.count()
    course_count = vendor_courses.count()
    supermarket_count = vendor_supermarket.count()
    active_products = vendor_products.filter(is_active=True).count()
    active_courses = vendor_courses.filter(is_active=True).count()
    active_supermarket = vendor_supermarket.filter(is_active=True).count()

    product_ids = list(vendor_products.values_list('id', flat=True))
    course_ids = list(vendor_courses.values_list('id', flat=True))
    supermarket_ids = list(vendor_supermarket.values_list('id', flat=True))

    paid_statuses = ['paid', 'processing', 'shipped', 'delivered']

    product_order_items = MarketplaceOrderItem.objects.filter(
        item_type='product', item_id__in=product_ids,
        order__status__in=paid_statuses,
    ) if product_ids else MarketplaceOrderItem.objects.none()

    course_order_items = MarketplaceOrderItem.objects.filter(
        item_type='course', item_id__in=course_ids,
        order__status__in=paid_statuses,
    ) if course_ids else MarketplaceOrderItem.objects.none()

    supermarket_order_items = MarketplaceOrderItem.objects.filter(
        item_type='supermarket', item_id__in=supermarket_ids,
        order__status__in=paid_statuses,
    ) if supermarket_ids else MarketplaceOrderItem.objects.none()

    total_product_sales = product_order_items.aggregate(s=Sum('quantity'))['s'] or 0
    total_enrollments = course_order_items.aggregate(s=Sum('quantity'))['s'] or 0
    total_supermarket_sales = supermarket_order_items.aggregate(s=Sum('quantity'))['s'] or 0

    product_rev_agg = product_order_items.aggregate(s=Sum('subtotal'))['s']
    course_rev_agg = course_order_items.aggregate(s=Sum('subtotal'))['s']
    sm_rev_agg = supermarket_order_items.aggregate(s=Sum('subtotal'))['s']
    product_revenue_mk = product_rev_agg or 0
    course_revenue_mk = course_rev_agg or 0
    supermarket_revenue_mk = sm_rev_agg or 0
    total_mk_revenue = product_revenue_mk + course_revenue_mk + supermarket_revenue_mk

    daily_labels = []
    daily_product_data = []
    daily_course_data = []
    daily_supermarket_data = []
    for i in range(6, -1, -1):
        day = (timezone.now() - timedelta(days=i)).date()
        daily_labels.append(day.strftime('%m/%d'))
        p_rev = product_order_items.filter(order__created_at__date=day).aggregate(s=Sum('subtotal'))['s'] or 0
        c_rev = course_order_items.filter(order__created_at__date=day).aggregate(s=Sum('subtotal'))['s'] or 0
        s_rev = supermarket_order_items.filter(order__created_at__date=day).aggregate(s=Sum('subtotal'))['s'] or 0
        daily_product_data.append(float(p_rev))
        daily_course_data.append(float(c_rev))
        daily_supermarket_data.append(float(s_rev))

    # Top products by actual sold quantity (not cached sales_count)
    product_sales_sub = MarketplaceOrderItem.objects.filter(
        item_type='product', item_id=OuterRef('pk'),
        order__status__in=paid_statuses,
    ).values('item_id').annotate(total=Sum('quantity')).values('total')[:1]
    annotated_products = vendor_products.annotate(
        actual_sales=Coalesce(Subquery(product_sales_sub, output_field=IntegerField()), Value(0))
    )
    top_products = list(annotated_products.order_by('-actual_sales')[:5])
    for p in top_products:
        p.display_sales = p.actual_sales

    recent_order_ids = []
    item_q = None
    if product_ids:
        item_q = Q(item_type='product', item_id__in=product_ids)
    if course_ids:
        cq = Q(item_type='course', item_id__in=course_ids)
        item_q = cq if item_q is None else (item_q | cq)
    if supermarket_ids:
        sq = Q(item_type='supermarket', item_id__in=supermarket_ids)
        item_q = sq if item_q is None else (item_q | sq)
    if item_q is not None:
        recent_order_ids = list(
            MarketplaceOrderItem.objects.filter(item_q).values_list('order_id', flat=True).distinct()[:20]
        )

    recent_orders = []
    if recent_order_ids:
        recent_orders = list(
            MarketplaceOrder.objects.filter(id__in=recent_order_ids).order_by('-created_at')[:10]
        )

    cat_data = list(vendor_products.values('category__name').annotate(cnt=Count('id')).order_by('-cnt')[:6])
    cat_labels = json.dumps([(c['category__name'] or '—') for c in cat_data])
    cat_counts = json.dumps([c['cnt'] for c in cat_data])

    return {
        'vm_product_count': product_count,
        'vm_course_count': course_count,
        'vm_supermarket_count': supermarket_count,
        'vm_active_products': active_products,
        'vm_active_courses': active_courses,
        'vm_active_supermarket': active_supermarket,
        'vm_inactive_products': max(0, product_count - active_products),
        'vm_inactive_courses': max(0, course_count - active_courses),
        'vm_inactive_supermarket': max(0, supermarket_count - active_supermarket),
        'vm_total_product_sales': total_product_sales,
        'vm_total_enrollments': total_enrollments,
        'vm_total_supermarket_sales': total_supermarket_sales,
        'vm_mk_product_revenue': product_revenue_mk,
        'vm_mk_course_revenue': course_revenue_mk,
        'vm_mk_supermarket_revenue': supermarket_revenue_mk,
        'vm_mk_total_revenue': total_mk_revenue,
        'vm_daily_labels': json.dumps(daily_labels),
        'vm_daily_product_data': json.dumps(daily_product_data),
        'vm_daily_course_data': json.dumps(daily_course_data),
        'vm_daily_supermarket_data': json.dumps(daily_supermarket_data),
        'vm_top_products': top_products,
        'vm_recent_orders': recent_orders,
        'vm_cat_labels': cat_labels,
        'vm_cat_counts': cat_counts,
    }


def _vendor_marketplace_order_ids_hub(vendor):
    """Distinct marketplace order IDs that touch this vendor's products, courses, or supermarket SKUs."""
    pids = list(Product.objects.filter(vendor=vendor).values_list('id', flat=True))
    cids = list(Course.objects.filter(vendor=vendor).values_list('id', flat=True))
    sids = list(SupermarketItem.objects.filter(vendor=vendor).values_list('id', flat=True))
    if not pids and not cids and not sids:
        return []
    q = Q()
    if pids:
        q |= Q(item_type='product', item_id__in=pids)
    if cids:
        q |= Q(item_type='course', item_id__in=cids)
    if sids:
        q |= Q(item_type='supermarket', item_id__in=sids)
    return list(
        MarketplaceOrderItem.objects.filter(q).values_list('order_id', flat=True).distinct()
    )


def _vendor_classify_mkt_order(order, pids_set, cids_set, sids_set=None):
    """Classify marketplace order lines relevant to this vendor: product / course / supermarket / mixed."""
    sids_set = sids_set or set()
    kinds = set()
    for it in order.items.all():
        if it.item_type == 'product' and it.item_id in pids_set:
            kinds.add('product')
        elif it.item_type == 'course' and it.item_id in cids_set:
            kinds.add('course')
        elif it.item_type == 'supermarket' and it.item_id in sids_set:
            kinds.add('supermarket')
    if not kinds:
        return 'mixed'
    if len(kinds) == 1:
        return kinds.pop()
    return 'mixed'


def vendor_dashboard(request):
    """Vendor dashboard"""
    # Allow admin to access any vendor dashboard
    admin_access = request.session.get("name")
    vendor = _get_vendor(request)

    if not vendor and not admin_access:
        return redirect('/manager/vendor/login/')

    # If admin is viewing, allow seeing any vendor via query param
    if admin_access and not vendor:
        vendor_id = request.GET.get('vendor_id')
        if vendor_id:
            vendor = get_object_or_404(models.Vendor, id=vendor_id)
        else:
            # Show vendor list for admin
            vendors = models.Vendor.objects.all()
            return render(request, 'admin/vendor_list.html', {
                'vendors': vendors,
                'total_vendors': vendors.count(),
            })

    vendor_books = models.VendorBook.objects.filter(vendor=vendor).select_related('book', 'book__publisher')
    active_books = sum(1 for vb in vendor_books if vb.is_active)
    inactive_books = vendor_books.count() - active_books
    total_inventory = sum(vb.book.inventory for vb in vendor_books)
    vendor_book_ids = list(vendor_books.values_list('book_id', flat=True))

    # Real per-vendor book stats from OrderItem (not cached global Book.sale_num)
    paid_book_statuses = ['paid', 'confirmed', 'processing', 'shipped', 'delivered']
    book_order_items = models.OrderItem.objects.filter(
        book_id__in=vendor_book_ids,
        order__status__in=paid_book_statuses,
    )
    total_sales = book_order_items.aggregate(s=Sum('quantity'))['s'] or 0
    total_revenue = book_order_items.aggregate(s=Sum('total_price'))['s'] or 0

    # --- Marketplace products, courses & supermarket ---
    from marketplace.models import Product, Course, SupermarketItem, MarketplaceOrder, MarketplaceOrderItem
    vendor_products = Product.objects.filter(vendor=vendor)
    vendor_courses = Course.objects.filter(vendor=vendor)
    vendor_supermarket = SupermarketItem.objects.filter(vendor=vendor)
    active_products = vendor_products.filter(is_active=True).count()
    active_courses = vendor_courses.filter(is_active=True).count()
    active_supermarket_items = vendor_supermarket.filter(is_active=True).count()

    # Real per-vendor marketplace stats from MarketplaceOrderItem (not cached fields)
    paid_mkt_statuses = ['paid', 'processing', 'shipped', 'delivered']
    product_ids = list(vendor_products.values_list('id', flat=True))
    course_ids = list(vendor_courses.values_list('id', flat=True))
    supermarket_ids = list(vendor_supermarket.values_list('id', flat=True))

    mkt_product_items = MarketplaceOrderItem.objects.filter(
        item_type='product', item_id__in=product_ids, order__status__in=paid_mkt_statuses,
    ) if product_ids else MarketplaceOrderItem.objects.none()
    mkt_course_items = MarketplaceOrderItem.objects.filter(
        item_type='course', item_id__in=course_ids, order__status__in=paid_mkt_statuses,
    ) if course_ids else MarketplaceOrderItem.objects.none()
    mkt_supermarket_items = MarketplaceOrderItem.objects.filter(
        item_type='supermarket', item_id__in=supermarket_ids, order__status__in=paid_mkt_statuses,
    ) if supermarket_ids else MarketplaceOrderItem.objects.none()

    product_revenue = mkt_product_items.aggregate(s=Sum('subtotal'))['s'] or 0
    course_revenue = mkt_course_items.aggregate(s=Sum('subtotal'))['s'] or 0
    supermarket_revenue = mkt_supermarket_items.aggregate(s=Sum('subtotal'))['s'] or 0
    combined_revenue = total_revenue + product_revenue + course_revenue + supermarket_revenue
    total_product_sales = mkt_product_items.aggregate(s=Sum('quantity'))['s'] or 0
    total_course_enrollments = mkt_course_items.aggregate(s=Sum('quantity'))['s'] or 0
    total_supermarket_sales_count = mkt_supermarket_items.aggregate(s=Sum('quantity'))['s'] or 0
    combined_sales = total_sales + total_product_sales + total_course_enrollments + total_supermarket_sales_count

    price_samples = []
    for vb in vendor_books.select_related('book'):
        price_samples.append(float(vb.book.price or 0))
    for p in vendor_products:
        price_samples.append(float(p.price))
    for c in vendor_courses:
        price_samples.append(float(c.price))
    for s in vendor_supermarket:
        price_samples.append(float(s.price))
    avg_price = (sum(price_samples) / len(price_samples)) if price_samples else 0

    total_listings = vendor_books.count() + vendor_products.count() + vendor_courses.count() + vendor_supermarket.count()
    active_listings = active_books + active_products + active_courses + active_supermarket_items
    listing_activation_pct = int(round(100 * active_listings / total_listings)) if total_listings else 0

    perf_top_items = []
    if vendor_books.exists():
        top_book = book_order_items.values('book_id').annotate(total_qty=Sum('quantity')).order_by('-total_qty').first()
        if top_book:
            vb_top = vendor_books.filter(book_id=top_book['book_id']).first()
            if vb_top:
                perf_top_items.append({
                    'kind': 'book',
                    'title': vb_top.book.name,
                    'image_url': vb_top.book.get_cover_url(),
                    'value': top_book['total_qty'],
                })
    if vendor_products.exists():
        p_top = mkt_product_items.values('item_id').annotate(total_qty=Sum('quantity')).order_by('-total_qty').first()
        if p_top:
            p_obj = vendor_products.filter(id=p_top['item_id']).first()
            if p_obj:
                perf_top_items.append({
                    'kind': 'product',
                    'title': p_obj.name,
                    'image_url': p_obj.get_image_url(),
                    'value': p_top['total_qty'],
                })
    if vendor_courses.exists():
        c_top = mkt_course_items.values('item_id').annotate(total_qty=Sum('quantity')).order_by('-total_qty').first()
        if c_top:
            c_obj = vendor_courses.filter(id=c_top['item_id']).first()
            if c_obj:
                perf_top_items.append({
                    'kind': 'course',
                    'title': c_obj.title,
                    'image_url': c_obj.get_image_url(),
                    'value': c_top['total_qty'],
                })
    if vendor_supermarket.exists():
        s_top = mkt_supermarket_items.values('item_id').annotate(total_qty=Sum('quantity')).order_by('-total_qty').first()
        if s_top:
            s_obj = vendor_supermarket.filter(id=s_top['item_id']).first()
            if s_obj:
                perf_top_items.append({
                    'kind': 'supermarket',
                    'title': s_obj.name,
                    'image_url': s_obj.get_image_url(),
                    'value': s_top['total_qty'],
                })

    best_seller_agg = book_order_items.values('book_id').annotate(total_qty=Sum('quantity')).order_by('-total_qty').first()
    best_seller = vendor_books.filter(book_id=best_seller_agg['book_id']).first() if best_seller_agg else None

    follower_count = models.UserFollowedVendor.objects.filter(vendor=vendor).count()

    # --- Marketplace orders for this vendor ---
    vendor_product_ids = list(vendor_products.values_list('id', flat=True))
    vendor_course_ids = list(vendor_courses.values_list('id', flat=True))
    vendor_supermarket_ids = list(vendor_supermarket.values_list('id', flat=True))
    marketplace_order_ids = set()
    if vendor_product_ids or vendor_course_ids or vendor_supermarket_ids:
        from django.db.models import Q
        mq = Q()
        if vendor_product_ids:
            mq |= Q(item_type='product', item_id__in=vendor_product_ids)
        if vendor_course_ids:
            mq |= Q(item_type='course', item_id__in=vendor_course_ids)
        if vendor_supermarket_ids:
            mq |= Q(item_type='supermarket', item_id__in=vendor_supermarket_ids)
        mkt_items_qs = MarketplaceOrderItem.objects.filter(mq).values_list('order_id', flat=True)
        marketplace_order_ids = set(mkt_items_qs)

    marketplace_orders = []
    if marketplace_order_ids:
        mkt_orders = MarketplaceOrder.objects.filter(
            id__in=marketplace_order_ids
        ).order_by('-created_at')[:30]
        for mo in mkt_orders:
            marketplace_orders.append({
                'order_number': mo.order_number,
                'user_name': mo.user_name or mo.user_email,
                'total_amount': mo.total_amount,
                'status': mo.get_status_display() if hasattr(mo, 'get_status_display') else mo.status,
                'status_color': mo.get_status_color(),
                'date': mo.created_at,
            })

    # --- Customer behavior analytics ---
    vendor_book_ids = [vb.book_id for vb in vendor_books]

    purchase_data = []
    if vendor_book_ids:
        purchased_items = models.OrderItem.objects.filter(
            book_id__in=vendor_book_ids,
            order__status__in=['paid', 'confirmed', 'processing', 'shipped', 'delivered']
        ).select_related('order', 'book').order_by('-order__created_at')[:50]
        for item in purchased_items:
            purchase_data.append({
                'book_name': item.book.name,
                'customer_name': item.order.customer_name,
                'customer_email': item.order.customer_email,
                'quantity': item.quantity,
                'total_price': item.total_price,
                'date': item.order.created_at,
                'status': item.order.get_status_display(),
            })

    from django.utils import timezone
    from datetime import timedelta

    cart_data = []
    if vendor_book_ids:
        cart_items = models.CartItem.objects.filter(
            book_id__in=vendor_book_ids
        ).select_related('book').order_by('-updated_at')[:50]
        for ci in cart_items:
            days_in_cart = (timezone.now() - ci.created_at).days
            cart_data.append({
                'book_name': ci.book.name,
                'quantity': ci.quantity,
                'book_price': ci.book.price,
                'session_key': ci.session_key[:8] + '...',
                'added_at': ci.created_at,
                'updated_at': ci.updated_at,
                'days_in_cart': days_in_cart,
            })

    wishlist_data = []
    if vendor_book_ids:
        wishlists = models.Wishlist.objects.filter(
            book_id__in=vendor_book_ids
        ).select_related('user', 'book').order_by('-created_at')[:50]
        for w in wishlists:
            wishlist_data.append({
                'book_name': w.book.name,
                'user_name': w.user.name,
                'user_email': w.user.email,
                'book_price': w.book.price,
                'added_at': w.created_at,
            })

    # Abandoned payments: orders placed but not paid for this vendor's books
    abandoned_orders = []
    if vendor_book_ids:
        abandoned_items = models.OrderItem.objects.filter(
            book_id__in=vendor_book_ids,
            order__status='payment_pending'
        ).select_related('order', 'book').order_by('-order__created_at')[:20]
        for item in abandoned_items:
            unit_price = item.total_price / item.quantity if item.quantity else item.total_price
            abandoned_orders.append({
                'book_name': item.book.name,
                'customer_name': item.order.customer_name,
                'customer_email': item.order.customer_email,
                'total_price': item.total_price,
                'unit_price': unit_price,
                'quantity': item.quantity,
                'date': item.order.created_at,
            })

    total_purchases = len(purchase_data)
    total_cart_items = len(cart_data)
    total_wishlists = len(wishlist_data)

    hub_book_ids = _vendor_book_order_ids(vendor)
    hub_book_order_count = models.Order.objects.filter(id__in=hub_book_ids).count() if hub_book_ids else 0
    hub_mkt_ids = _vendor_marketplace_order_ids_hub(vendor)
    hub_mkt_order_count = MarketplaceOrder.objects.filter(id__in=hub_mkt_ids).count() if hub_mkt_ids else 0

    from manager.commission import commission_rates_for_display
    escrow_qs = models.PlatformEscrowTransaction.objects.filter(vendor=vendor)
    escrow_stats = {
        'held': escrow_qs.filter(status='held').aggregate(s=Sum('vendor_payout_amount'))['s'] or Decimal('0'),
        'releasable': escrow_qs.filter(status='releasable').aggregate(s=Sum('vendor_payout_amount'))['s'] or Decimal('0'),
        'released': escrow_qs.filter(status='released').aggregate(s=Sum('vendor_payout_amount'))['s'] or Decimal('0'),
    }
    vendor_wallet = models.VendorWallet.objects.filter(vendor=vendor).first()
    recent_escrow = list(escrow_qs.order_by('-held_at')[:8])
    commission_rates = commission_rates_for_display()
    from manager.escrow_service import REFUND_HOLD_DAYS

    total_supermarket_stock = vendor_supermarket.aggregate(s=Sum('stock'))['s'] or 0

    context = {
        'vendor': vendor,
        'vendor_books': vendor_books,
        'total_books': vendor_books.count(),
        'total_sales': total_sales,
        'total_revenue': total_revenue,
        'active_books': active_books,
        'inactive_books': inactive_books,
        'avg_price': avg_price,
        'best_seller': best_seller,
        'total_inventory': total_inventory,
        'admin_access': admin_access,
        # Marketplace
        'vendor_products': vendor_products,
        'vendor_courses': vendor_courses,
        'vendor_supermarket': vendor_supermarket,
        'active_products': active_products,
        'active_courses': active_courses,
        'active_supermarket_items': active_supermarket_items,
        'product_revenue': product_revenue,
        'course_revenue': course_revenue,
        'supermarket_revenue': supermarket_revenue,
        'combined_revenue': combined_revenue,
        'combined_sales': combined_sales,
        'total_product_sales': total_product_sales,
        'total_course_enrollments': total_course_enrollments,
        'total_supermarket_sales_count': total_supermarket_sales_count,
        'total_supermarket_stock': total_supermarket_stock,
        'listing_activation_pct': listing_activation_pct,
        'total_listings': total_listings,
        'total_items': total_listings,
        'perf_top_items': perf_top_items,
        'follower_count': follower_count,
        'marketplace_orders': marketplace_orders,
        # Customer behavior
        'purchase_data': purchase_data,
        'cart_data': cart_data,
        'wishlist_data': wishlist_data,
        'total_purchases': total_purchases,
        'total_cart_items': total_cart_items,
        'total_wishlists': total_wishlists,
        'abandoned_orders': abandoned_orders,
        'hub_book_order_count': hub_book_order_count,
        'hub_mkt_order_count': hub_mkt_order_count,
        'hub_orders_total': hub_book_order_count + hub_mkt_order_count,
        'escrow_stats': escrow_stats,
        'vendor_wallet': vendor_wallet,
        'recent_escrow': recent_escrow,
        'commission_rates': commission_rates,
        'refund_hold_days': REFUND_HOLD_DAYS,
    }
    context.update(_vendor_marketplace_insights_dict(vendor))
    return render(request, 'public/vendor_dashboard.html', context)


# ─── Vendor: books catalogue (full CRUD entry page) ───────────────────────────

def vendor_books(request):
    """Seller hub — list all vendor books with CRUD shortcuts."""
    admin_access = request.session.get('name')
    vendor = _get_vendor(request)
    if not vendor and not admin_access:
        return redirect('/manager/vendor/login/')
    if admin_access and not vendor:
        vid = request.GET.get('vendor_id')
        if vid:
            vendor = get_object_or_404(models.Vendor, id=vid)
        else:
            return redirect('/manager/vendor/dashboard/')

    vendor_books_qs = models.VendorBook.objects.filter(vendor=vendor).select_related('book', 'book__publisher', 'book__category')
    inactive_ct = vendor_books_qs.filter(is_active=False).count()
    active_ct = vendor_books_qs.filter(is_active=True).count()
    total_sales = sum(vb.book.sale_num for vb in vendor_books_qs)
    total_stock = sum(vb.book.inventory for vb in vendor_books_qs)
    context = {
        'vendor': vendor,
        'admin_access': admin_access,
        'vendor_books': vendor_books_qs,
        'total_books': vendor_books_qs.count(),
        'inactive_books': inactive_ct,
        'active_books': active_ct,
        'total_sales': total_sales,
        'total_stock': total_stock,
    }
    return render(request, 'public/vendor_books.html', context)


def vendor_settings(request):
    """Dedicated store settings page (sidebar)."""
    admin_access = request.session.get('name')
    vendor = _get_vendor(request)
    if not vendor and not admin_access:
        return redirect('/manager/vendor/login/')
    if admin_access and not vendor:
        vendor_id = request.GET.get('vendor_id')
        if vendor_id:
            vendor = get_object_or_404(models.Vendor, id=vendor_id)
        else:
            return redirect('/manager/vendor/dashboard/')

    return render(request, 'public/vendor_settings.html', {
        'vendor': vendor,
        'admin_access': admin_access,
    })


@require_POST
def vendor_settings_save(request):
    """Save vendor store settings."""
    admin_access = request.session.get('name')
    vendor = _get_vendor(request)
    if not vendor and not admin_access:
        return JsonResponse({'success': False, 'message': _('未登录')})
    if admin_access and not vendor:
        vid = request.POST.get('vendor_id') or request.GET.get('vendor_id')
        if vid:
            vendor = get_object_or_404(models.Vendor, id=vid)
        else:
            return JsonResponse({'success': False, 'message': _('未指定卖家')})

    company_name = request.POST.get('company_name', '').strip()
    description = request.POST.get('description', '').strip()
    phone = request.POST.get('phone', '').strip()
    email = request.POST.get('email', '').strip()
    country, dept, city, loc_err = _parse_signup_location(request)

    if company_name:
        vendor.company_name = company_name
    if description:
        vendor.description = description
    if phone:
        vendor.phone = phone
    if email:
        vendor.email = email
    if not loc_err:
        vendor.country = country
        vendor.location = dept
        vendor.city = city
        if vendor.user_id:
            vendor.user.country = country
            vendor.user.location = dept
            vendor.user.city = city
            vendor.user.save(update_fields=['country', 'location', 'city', 'updated_at'])

    if 'logo' in request.FILES:
        vendor.logo = request.FILES['logo']

    delivery_min_raw = request.POST.get('default_delivery_days_min', '').strip()
    delivery_max_raw = request.POST.get('default_delivery_days_max', '').strip()
    if not delivery_min_raw and not delivery_max_raw:
        vendor.default_delivery_days_min = None
        vendor.default_delivery_days_max = None
    elif delivery_min_raw.isdigit() and delivery_max_raw.isdigit():
        d_min, d_max = int(delivery_min_raw), int(delivery_max_raw)
        if 0 < d_min <= d_max <= 90:
            vendor.default_delivery_days_min = d_min
            vendor.default_delivery_days_max = d_max
        else:
            messages.error(request, _('délai de livraison invalide (min doit être ≤ max, entre 1 et 90 jours).'))

    vendor.save()
    messages.success(request, _('店铺设置已保存'))
    vid = request.POST.get('vendor_id') or request.GET.get('vendor_id')
    redirect_url = reverse('manager:vendor_settings')
    if vid:
        redirect_url = f'{redirect_url}?vendor_id={vid}'
    return redirect(redirect_url)


def vendor_inventory(request):
    """Vendor inventory management — view and update stock for books, products, courses, supermarket."""
    admin_access = request.session.get('name')
    vendor = _get_vendor(request)
    if not vendor and not admin_access:
        return redirect('/manager/vendor/login/')
    if admin_access and not vendor:
        vid = request.GET.get('vendor_id')
        if vid:
            vendor = get_object_or_404(models.Vendor, id=vid)
        else:
            return redirect('/manager/vendor/dashboard/')

    if request.method == 'POST':
        if request.POST.get('bulk') == '1':
            _process_bulk_inventory(
                request,
                lambda t, i, a, d, m: _vendor_apply_inventory_stock(vendor, t, i, a, d, m),
                _('Stock updated for {count} item(s).'),
            )
            redirect_url = reverse('manager:vendor_inventory')
            if admin_access and vendor:
                redirect_url = f'{redirect_url}?vendor_id={vendor.id}'
            return redirect(redirect_url)

        item_type = request.POST.get('item_type')
        item_id = request.POST.get('item_id')
        action = request.POST.get('action')
        delta = int(request.POST.get('delta', 0) or 0)
        manual_value = request.POST.get('manual_value')

        try:
            _vendor_apply_inventory_stock(vendor, item_type, item_id, action, delta, manual_value)
            messages.success(request, _('库存已更新'))
        except Exception as e:
            messages.error(request, str(e))
        redirect_url = reverse('manager:vendor_inventory')
        if admin_access and vendor:
            redirect_url = f'{redirect_url}?vendor_id={vendor.id}'
        return redirect(redirect_url)

    from marketplace.models import Product, Course, SupermarketItem
    vendor_books = models.VendorBook.objects.filter(vendor=vendor).select_related('book')
    products = Product.objects.filter(vendor=vendor)
    courses = Course.objects.filter(vendor=vendor)
    supermarket = SupermarketItem.objects.filter(vendor=vendor)

    context = {
        'vendor': vendor,
        'admin_access': admin_access,
        'vendor_books': vendor_books,
        'products': products,
        'courses': courses,
        'supermarket_items': supermarket,
        'total_books': vendor_books.count(),
        'total_products': products.count(),
        'total_courses': courses.count(),
        'total_supermarket': supermarket.count(),
        'bulk_action_url': reverse('manager:vendor_inventory') + (f'?vendor_id={vendor.id}' if admin_access else ''),
    }
    return render(request, 'public/vendor_inventory.html', context)


def _vendor_book_order_ids(vendor):
    book_ids = list(models.VendorBook.objects.filter(vendor=vendor).values_list('book_id', flat=True))
    if not book_ids:
        return []
    return list(
        models.OrderItem.objects.filter(book_id__in=book_ids)
        .values_list('order_id', flat=True)
        .distinct()
    )


def _vendor_book_order_can_update_fulfillment(order):
    """Vendors may update status/payment unless the order is closed."""
    return order.status not in ('cancelled', 'refunded')


# 'shipped'/'delivered' are deliberately excluded — those now require the
# shipment-based flow (vendor_shipment_action) which enforces tracking info
# on ship and never lets the vendor self-report delivery.
VENDOR_BOOK_ORDER_ALLOWED_STATUSES = frozenset({'confirmed', 'processing'})


def _vendor_book_order_customer_editable(order):
    return order.status not in ('cancelled', 'refunded')


def vendor_book_orders(request):
    """Orders that contain at least one book sold by this vendor."""
    admin_access = request.session.get('name')
    vendor = _get_vendor(request)
    if not vendor and not admin_access:
        return redirect('/manager/vendor/login/')
    if admin_access and not vendor:
        vid = request.GET.get('vendor_id')
        if vid:
            vendor = get_object_or_404(models.Vendor, id=vid)
        else:
            return redirect('/manager/vendor/dashboard/')

    order_ids = _vendor_book_order_ids(vendor)
    orders = models.Order.objects.filter(id__in=order_ids).order_by('-created_at')

    status_filter = request.GET.get('status', '')
    payment_filter = request.GET.get('payment_status', '')
    search_q = request.GET.get('search', '').strip()
    if status_filter:
        orders = orders.filter(status=status_filter)
    if payment_filter:
        orders = orders.filter(payment_status=payment_filter)
    if search_q:
        orders = orders.filter(
            Q(order_number__icontains=search_q)
            | Q(customer_name__icontains=search_q)
            | Q(customer_email__icontains=search_q)
            | Q(customer_phone__icontains=search_q)
        )

    total_orders = models.Order.objects.filter(id__in=order_ids).count()
    pending_pay = models.Order.objects.filter(id__in=order_ids, payment_status='pending').count()
    paid_completed = models.Order.objects.filter(id__in=order_ids, payment_status='completed').count()

    paginator = Paginator(orders, 20)
    page = paginator.get_page(request.GET.get('page', 1))

    context = {
        'vendor': vendor,
        'admin_access': admin_access,
        'orders': page,
        'status_choices': models.ORDER_STATUS_CHOICES,
        'payment_status_choices': models.PAYMENT_STATUS_CHOICES,
        'current_status': status_filter,
        'current_payment_status': payment_filter,
        'current_search': search_q,
        'total_orders': total_orders,
        'pending_pay': pending_pay,
        'paid_completed': paid_completed,
        'fulfillment_statuses': [(k, v) for k, v in models.ORDER_STATUS_CHOICES if k not in ('cancelled', 'refunded')],
    }
    return render(request, 'public/vendor_book_orders.html', context)


def vendor_book_order_detail(request, order_id):
    vendor = _get_vendor(request)
    admin_access = request.session.get('name')
    if not vendor and not admin_access:
        return redirect('/manager/vendor/login/')
    if admin_access and not vendor:
        vid = request.GET.get('vendor_id')
        if vid:
            vendor = get_object_or_404(models.Vendor, id=vid)
        else:
            return redirect('/manager/vendor/dashboard/')

    allowed = set(_vendor_book_order_ids(vendor))
    order = get_object_or_404(models.Order, id=order_id)
    if order.id not in allowed:
        raise Http404('Order not found')

    book_ids = set(models.VendorBook.objects.filter(vendor=vendor).values_list('book_id', flat=True))
    order_items = list(
        models.OrderItem.objects.filter(order=order, book_id__in=book_ids).select_related('book')
    )
    vendor_lines_total = sum((item.total_price for item in order_items), Decimal('0'))
    vendor_shipment = models.Shipment.objects.filter(
        order_source='book', order_id=order.id, vendor=vendor,
    ).first()
    suggested_delivery_date = None
    if vendor_shipment and vendor_shipment.fulfillment_status in ('accepted', 'packing'):
        from manager.fulfillment_service import suggested_delivery_date as _suggest
        suggested_delivery_date = _suggest(vendor_shipment)

    context = {
        'vendor': vendor,
        'admin_access': admin_access,
        'order': order,
        'order_items': order_items,
        'vendor_lines_total': vendor_lines_total,
        'shipment': vendor_shipment,
        'suggested_delivery_date': suggested_delivery_date,
        'can_update_fulfillment': _vendor_book_order_can_update_fulfillment(order),
        'fulfillment_statuses': (
            [(k, v) for k, v in models.ORDER_STATUS_CHOICES if k not in ('cancelled', 'refunded')]
            if admin_access else
            [(k, v) for k, v in models.ORDER_STATUS_CHOICES if k in ('confirmed', 'processing')]
        ),
        'status_choices': models.ORDER_STATUS_CHOICES,
        'payment_status_choices': models.PAYMENT_STATUS_CHOICES,
        'can_edit_customer': _vendor_book_order_customer_editable(order),
    }
    return render(request, 'public/vendor_book_order_detail.html', context)


@require_POST
def vendor_book_order_update_status(request):
    vendor = _get_vendor(request)
    if not vendor:
        return JsonResponse({'success': False, 'message': '请以卖家身份登录'})

    order_id = request.POST.get('order_id')
    new_status = request.POST.get('status', '').strip()
    note = request.POST.get('vendor_note', '').strip()

    if new_status not in VENDOR_BOOK_ORDER_ALLOWED_STATUSES:
        return JsonResponse({'success': False, 'message': '不允许的状态'})

    allowed = set(_vendor_book_order_ids(vendor))
    order = get_object_or_404(models.Order, id=order_id)
    if order.id not in allowed:
        return JsonResponse({'success': False, 'message': '无权操作此订单'})

    if not _vendor_book_order_can_update_fulfillment(order):
        return JsonResponse({'success': False, 'message': '订单已关闭，无法更新状态'})

    old = order.status
    order.status = new_status
    if note:
        prefix = '[Vendor %s] ' % vendor.company_name[:40]
        order.admin_notes = (prefix + note + '\n' + (order.admin_notes or '')).strip()
    order.save(update_fields=['status', 'admin_notes', 'updated_at'])

    status_dict = dict(models.ORDER_STATUS_CHOICES)
    return JsonResponse({
        'success': True,
        'message': '已更新',
        'new_status': new_status,
        'new_status_display': status_dict.get(new_status, new_status),
        'new_status_color': order.get_status_color(),
    })


@require_POST
def vendor_shipment_action(request):
    """Vendor-facing shipment actions: accept / reject / ship. Shared by
    book and marketplace order detail pages — Shipment is a single model
    regardless of order_source. This is the only path to 'shipped'; there is
    no way to mark something shipped without a tracking number and carrier,
    and vendors can never set 'delivered' themselves (see fulfillment_service
    docstring for why)."""
    from manager import fulfillment_service as fs

    vendor = _get_vendor(request)
    admin_access = request.session.get('name')
    if not vendor and not admin_access:
        return JsonResponse({'success': False, 'message': '请以卖家身份登录'}, status=403)

    shipment_id = request.POST.get('shipment_id')
    action = request.POST.get('action', '').strip()
    shipment = get_object_or_404(models.Shipment, id=shipment_id)

    if vendor and shipment.vendor_id != vendor.id:
        return JsonResponse({'success': False, 'message': '无权操作此订单'}, status=403)

    if action == 'accept':
        ok = fs.accept_shipment(shipment)
    elif action == 'reject':
        reason = request.POST.get('reason', '').strip()
        if not reason:
            return JsonResponse({'success': False, 'message': '请填写拒绝原因'}, status=400)
        ok = fs.reject_shipment(shipment, reason)
    elif action == 'ship':
        tracking_number = request.POST.get('tracking_number', '').strip()
        carrier = request.POST.get('carrier', '').strip()
        eta_raw = request.POST.get('estimated_delivery_date', '').strip()
        eta = None
        if eta_raw:
            try:
                from datetime import datetime as _dt
                eta = _dt.strptime(eta_raw, '%Y-%m-%d').date()
            except ValueError:
                eta = None
        try:
            ok = fs.mark_shipped(shipment, tracking_number, carrier, estimated_delivery_date=eta)
        except ValueError as exc:
            return JsonResponse({'success': False, 'message': str(exc)}, status=400)
    else:
        return JsonResponse({'success': False, 'message': '无效操作'}, status=400)

    if not ok:
        return JsonResponse({'success': False, 'message': '当前状态无法执行此操作'}, status=400)

    shipment.refresh_from_db()
    return JsonResponse({
        'success': True,
        'message': '已更新',
        'fulfillment_status': shipment.fulfillment_status,
        'fulfillment_status_display': shipment.get_fulfillment_status_display(),
    })


def vendor_orders_hub(request):
    """Unified order list: books + marketplace (products/courses), with filters."""
    from django.urls import reverse

    admin_access = request.session.get('name')
    vendor = _get_vendor(request)
    if not vendor and not admin_access:
        return redirect('/manager/vendor/login/')
    if admin_access and not vendor:
        vid = request.GET.get('vendor_id')
        if vid:
            vendor = get_object_or_404(models.Vendor, id=vid)
        else:
            return redirect('/manager/vendor/dashboard/')

    book_ids = _vendor_book_order_ids(vendor)
    mkt_ids = _vendor_marketplace_order_ids_hub(vendor)
    pids_set = set(Product.objects.filter(vendor=vendor).values_list('id', flat=True))
    cids_set = set(Course.objects.filter(vendor=vendor).values_list('id', flat=True))
    sids_set = set(SupermarketItem.objects.filter(vendor=vendor).values_list('id', flat=True))

    channel = request.GET.get('channel', 'all').strip()
    status_filter = request.GET.get('status', '').strip()
    payment_filter = request.GET.get('payment_status', '').strip()
    search_q = request.GET.get('search', '').strip()

    book_qs = models.Order.objects.filter(id__in=book_ids)
    if status_filter:
        book_qs = book_qs.filter(status=status_filter)
    if payment_filter:
        book_qs = book_qs.filter(payment_status=payment_filter)
    if search_q:
        book_qs = book_qs.filter(
            Q(order_number__icontains=search_q)
            | Q(customer_name__icontains=search_q)
            | Q(customer_email__icontains=search_q)
            | Q(customer_phone__icontains=search_q)
        )

    mkt_qs = MarketplaceOrder.objects.filter(id__in=mkt_ids)
    if status_filter:
        mkt_qs = mkt_qs.filter(status=status_filter)
    if payment_filter:
        mkt_qs = mkt_qs.filter(payment_status=payment_filter)
    if search_q:
        mkt_qs = mkt_qs.filter(
            Q(order_number__icontains=search_q)
            | Q(user_name__icontains=search_q)
            | Q(user_email__icontains=search_q)
            | Q(customer_phone__icontains=search_q)
        )

    MAX_EACH = 600
    books_list = list(book_qs.order_by('-created_at')[:MAX_EACH])
    mkts_prefetched = list(
        mkt_qs.prefetch_related(
            Prefetch('items', queryset=MarketplaceOrderItem.objects.only('item_type', 'item_id'))
        ).order_by('-created_at')[:MAX_EACH]
    )

    for o in books_list:
        o.apply_ttl_rules()
    for o in mkts_prefetched:
        o.apply_ttl_rules()

    merged_rows = []
    for o in books_list:
        merged_rows.append({
            'kind': 'book',
            'created_at': o.created_at,
            'order': o,
            'detail_url': reverse('manager:vendor_book_order_detail', args=[o.id]),
            'mkt_line_kind': None,
        })
    for o in mkts_prefetched:
        line_kind = _vendor_classify_mkt_order(o, pids_set, cids_set, sids_set)
        merged_rows.append({
            'kind': 'marketplace',
            'created_at': o.created_at,
            'order': o,
            'detail_url': reverse('marketplace:vendor_marketplace_order_detail', args=[o.id]),
            'mkt_line_kind': line_kind,
        })

    merged_rows.sort(key=lambda r: r['created_at'], reverse=True)

    if channel == 'books':
        merged_rows = [r for r in merged_rows if r['kind'] == 'book']
    elif channel == 'marketplace':
        merged_rows = [r for r in merged_rows if r['kind'] == 'marketplace']
    elif channel == 'products':
        merged_rows = [
            r for r in merged_rows
            if r['kind'] == 'marketplace' and r['mkt_line_kind'] in ('product', 'mixed')
        ]
    elif channel == 'courses':
        merged_rows = [
            r for r in merged_rows
            if r['kind'] == 'marketplace' and r['mkt_line_kind'] in ('course', 'mixed')
        ]
    elif channel == 'supermarket':
        merged_rows = [
            r for r in merged_rows
            if r['kind'] == 'marketplace' and r['mkt_line_kind'] in ('supermarket', 'mixed')
        ]

    stats_book_total = models.Order.objects.filter(id__in=book_ids).count()
    stats_mkt_total = MarketplaceOrder.objects.filter(id__in=mkt_ids).count()
    stats_pending_pay_book = models.Order.objects.filter(id__in=book_ids, payment_status='pending').count()
    stats_pending_pay_mkt = MarketplaceOrder.objects.filter(id__in=mkt_ids, payment_status='pending').count()
    stats_orders_total = stats_book_total + stats_mkt_total
    stats_pending_pay_total = stats_pending_pay_book + stats_pending_pay_mkt

    paginator = Paginator(merged_rows, 20)
    page = paginator.get_page(request.GET.get('page', 1))

    context = {
        'vendor': vendor,
        'admin_access': admin_access,
        'orders_page': page,
        'channel': channel,
        'current_status': status_filter,
        'current_payment_status': payment_filter,
        'current_search': search_q,
        'stats_book_total': stats_book_total,
        'stats_mkt_total': stats_mkt_total,
        'stats_pending_pay_book': stats_pending_pay_book,
        'stats_pending_pay_mkt': stats_pending_pay_mkt,
        'stats_orders_total': stats_orders_total,
        'stats_pending_pay_total': stats_pending_pay_total,
        'book_status_choices': models.ORDER_STATUS_CHOICES,
        'mkt_status_choices': MarketplaceOrder.STATUS_CHOICES,
        'payment_status_choices': models.PAYMENT_STATUS_CHOICES,
        'mkt_payment_status_choices': MarketplaceOrder.PAYMENT_STATUS_CHOICES,
    }
    return render(request, 'public/vendor_orders_hub.html', context)


@require_POST
def vendor_book_order_update_customer(request):
    vendor = _get_vendor(request)
    if not vendor:
        return JsonResponse({'success': False, 'message': str(_('请以卖家身份登录'))}, status=403)

    order_id = request.POST.get('order_id')
    allowed = set(_vendor_book_order_ids(vendor))
    order = get_object_or_404(models.Order, id=order_id)
    if order.id not in allowed:
        return JsonResponse({'success': False, 'message': str(_('无权操作此订单'))}, status=403)
    if not _vendor_book_order_customer_editable(order):
        return JsonResponse({'success': False, 'message': str(_('订单已关闭，无法修改联系信息'))})

    name = request.POST.get('customer_name', '').strip()
    email = request.POST.get('customer_email', '').strip()
    phone = request.POST.get('customer_phone', '').strip()
    country = request.POST.get('country', '').strip() or order.country
    shipping_address = request.POST.get('shipping_address', '').strip()
    customer_notes = request.POST.get('customer_notes', '').strip()

    if not name or not email:
        return JsonResponse({'success': False, 'message': str(_('姓名与邮箱不能为空'))})

    order.customer_name = name[:100]
    order.customer_email = email[:254]
    order.customer_phone = phone[:20]
    order.country = country[:50]
    order.shipping_address = shipping_address
    order.customer_notes = customer_notes
    order.save(update_fields=[
        'customer_name', 'customer_email', 'customer_phone', 'country',
        'shipping_address', 'customer_notes', 'updated_at',
    ])
    return JsonResponse({'success': True, 'message': str(_('客户信息已保存'))})


@require_POST
def vendor_hub_order_update(request):
    """Update order status and/or payment status from vendor order hub (owned orders only)."""
    admin_access = request.session.get('name')
    vendor = _get_vendor(request)
    if admin_access and not vendor:
        vid = request.POST.get('vendor_id') or request.GET.get('vendor_id')
        if vid:
            vendor = models.Vendor.objects.filter(id=vid, is_active=True).first()
    if not vendor:
        return JsonResponse({'success': False, 'message': str(_('请以卖家身份登录'))}, status=403)

    kind = request.POST.get('kind', '').strip()
    order_id = request.POST.get('order_id')
    new_status = request.POST.get('status', '').strip()
    new_pay = request.POST.get('payment_status', '').strip()
    vendor_note = request.POST.get('vendor_note', '').strip()

    book_status_ok = dict(models.ORDER_STATUS_CHOICES)
    book_pay_ok = dict(models.PAYMENT_STATUS_CHOICES)
    mkt_status_ok = dict(MarketplaceOrder.STATUS_CHOICES)
    mkt_pay_ok = dict(MarketplaceOrder.PAYMENT_STATUS_CHOICES)

    # SECURITY: a genuine vendor (not admin-impersonation) must never be able
    # to self-report payment_status, nor set status to shipped/delivered/
    # cancelled/refunded directly — that used to let any vendor mark their
    # own unpaid order 'payment completed' and 'delivered' with zero
    # verification, releasing their own escrow payout. Those transitions now
    # only happen via real gateway webhooks (payment) or the shipment-based
    # accept/reject/ship/confirm-delivery flow in fulfillment_service.py.
    # Admin (staff) impersonating a vendor keeps the full override for
    # legitimate manual fixes.
    VENDOR_SAFE_STATUSES = {'confirmed', 'processing'}
    if not admin_access:
        new_pay = ''
        if new_status not in VENDOR_SAFE_STATUSES:
            new_status = ''

    if kind == 'book':
        allowed = set(_vendor_book_order_ids(vendor))
        order = get_object_or_404(models.Order, id=order_id)
        if order.id not in allowed:
            return JsonResponse({'success': False, 'message': str(_('无权操作此订单'))}, status=403)
        uf = []
        payment_routed_via_pipeline = False
        if new_pay == 'completed' and new_pay in book_pay_ok and order.payment_status != 'completed':
            # Route through the shared pipeline (shipment creation,
            # confirmation email, inventory deduction) instead of just
            # flipping the field.
            from manager.payments.views import _update_order_status
            _update_order_status(order, 'SUCCESSFUL')
            payment_routed_via_pipeline = True
        elif new_pay and new_pay in book_pay_ok:
            order.payment_status = new_pay
            uf.append('payment_status')
            if new_pay == 'completed' and not order.payment_completed_at:
                order.payment_completed_at = timezone.now()
                uf.append('payment_completed_at')
        if new_status and new_status in book_status_ok:
            order.status = new_status
            uf.append('status')
        if uf:
            note_fields = []
            if vendor_note:
                prefix = '[Vendor %s] ' % vendor.company_name[:40]
                order.admin_notes = (prefix + vendor_note + '\n' + (order.admin_notes or '')).strip()
                note_fields.append('admin_notes')
            order.save(update_fields=list(dict.fromkeys(uf + note_fields)) + ['updated_at'])
        elif vendor_note:
            prefix = '[Vendor %s] ' % vendor.company_name[:40]
            order.admin_notes = (prefix + vendor_note + '\n' + (order.admin_notes or '')).strip()
            order.save(update_fields=['admin_notes', 'updated_at'])
        elif not payment_routed_via_pipeline:
            return JsonResponse({'success': False, 'message': str(_('未更改'))}, status=400)
        return JsonResponse({'success': True, 'message': str(_('已更新'))})

    if kind == 'marketplace':
        allowed_ids = set(_vendor_marketplace_order_ids_hub(vendor))
        order = get_object_or_404(MarketplaceOrder, id=order_id)
        if order.id not in allowed_ids:
            return JsonResponse({'success': False, 'message': str(_('无权操作此订单'))}, status=403)
        uf = []
        payment_routed_via_pipeline = False
        if new_pay == 'completed' and new_pay in mkt_pay_ok and order.payment_status != 'completed':
            from manager.payments.views import _update_order_status
            _update_order_status(order, 'SUCCESSFUL')
            payment_routed_via_pipeline = True
        elif new_pay and new_pay in mkt_pay_ok:
            order.payment_status = new_pay
            uf.append('payment_status')
            if new_pay == 'completed' and not order.payment_completed_at:
                order.payment_completed_at = timezone.now()
                uf.append('payment_completed_at')
        if new_status and new_status in mkt_status_ok:
            order.status = new_status
            uf.append('status')
        if uf:
            note_fields = []
            if vendor_note:
                prefix = '[Vendor %s] ' % vendor.company_name[:40]
                order.admin_notes = (prefix + vendor_note + '\n' + (order.admin_notes or '')).strip()
                note_fields.append('admin_notes')
            order.save(update_fields=list(dict.fromkeys(uf + note_fields)) + ['updated_at'])
        elif vendor_note:
            prefix = '[Vendor %s] ' % vendor.company_name[:40]
            order.admin_notes = (prefix + vendor_note + '\n' + (order.admin_notes or '')).strip()
            order.save(update_fields=['admin_notes', 'updated_at'])
        elif not payment_routed_via_pipeline:
            return JsonResponse({'success': False, 'message': str(_('未更改'))}, status=400)
        return JsonResponse({'success': True, 'message': str(_('已更新'))})

    return JsonResponse({'success': False, 'message': str(_('参数无效'))}, status=400)




def admin_contact_quick_reply(request):
    """Admin quick-replies to a contact message (text stored, no email sent)"""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '未授权'})
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'})

    msg_id = request.POST.get('id')
    reply_text = request.POST.get('reply', '').strip()
    if not reply_text:
        return JsonResponse({'success': False, 'message': '回复内容不能为空'})

    msg = get_object_or_404(models.ContactMessage, id=msg_id)
    msg.admin_reply = reply_text
    msg.replied = True
    msg.replied_at = timezone.now()
    msg.is_read = True
    msg.save(update_fields=['admin_reply', 'replied', 'replied_at', 'is_read'])

    return JsonResponse({'success': True, 'message': '回复已发送'})


def _vendor_book_error(field, message):
    return {'success': False, 'message': message, 'field': field}


def vendor_add_book(request):
    """Vendor adds a book to sell"""
    vendor = _get_vendor(request)
    admin_access = request.session.get("name")
    if not vendor and not admin_access:
        return redirect('/manager/vendor/login/')

    # Handle inline creation of new publisher
    if request.method == 'POST' and request.POST.get('action') == 'create_publisher':
        pub_name = request.POST.get('publisher_name', '').strip()
        pub_address = request.POST.get('publisher_address', '').strip()
        if pub_name:
            # publisher_name/publisher_address are django-modeltranslation
            # fields — .objects.create() kwargs silently drop them, so
            # assign as plain attributes instead.
            pub = models.Publisher()
            pub.publisher_name = pub_name
            pub.publisher_address = pub_address
            pub.save()
            return JsonResponse({'success': True, 'id': pub.id, 'name': pub.publisher_name})
        return JsonResponse({'success': False, 'message': '出版社名称不能为空'})

    # Handle inline creation of new author
    if request.method == 'POST' and request.POST.get('action') == 'create_author':
        author_name = request.POST.get('author_name', '').strip()
        if author_name:
            author = models.Author()
            author.name = author_name
            author.save()
            return JsonResponse({'success': True, 'id': author.id, 'name': author.name})
        return JsonResponse({'success': False, 'message': '作者名称不能为空'})

    if request.method == 'POST':
        if admin_access and not vendor:
            vendor = get_object_or_404(models.Vendor, id=request.POST.get('vendor_id'))

        name = request.POST.get('name', '').strip()
        price = request.POST.get('price', '0')
        inventory = request.POST.get('inventory', '0')
        description = request.POST.get('description', '').strip()
        publisher_id = request.POST.get('publisher_id')
        category_id = request.POST.get('category_id') or None
        author_ids = request.POST.getlist('author_ids')
        book_file = request.FILES.get('book_file')
        download_link = request.POST.get('download_link', '').strip()
        delivery_days_min, delivery_days_max = _parse_delivery_days_override(request.POST)

        if not name:
            return JsonResponse(_vendor_book_error('name', 'Le titre du livre est obligatoire.'))
        if len(name) < 3:
            return JsonResponse(_vendor_book_error('name', 'Le titre du livre doit contenir au moins 3 caractères.'))
        if not description:
            return JsonResponse(_vendor_book_error('description', 'La description du livre est obligatoire.'))
        if len(description) < 12:
            return JsonResponse(_vendor_book_error('description', 'La description du livre doit contenir au moins 12 caractères.'))
        if models.VendorBook.objects.filter(vendor=vendor, book__name__iexact=name).exists():
            return JsonResponse(_vendor_book_error('name', 'Vous avez déjà publié un livre avec ce titre. Utilisez un titre différent.'))
        if not publisher_id:
            return JsonResponse(_vendor_book_error('publisher_id', 'La maison d’édition est obligatoire.'))
        if not author_ids:
            return JsonResponse(_vendor_book_error('author_ids', 'Veuillez sélectionner au moins un auteur.'))

        book = models.Book(
            price=Decimal(price),
            inventory=int(inventory),
            sale_num=0,
            publisher_id=int(publisher_id) if publisher_id else None,
            category_id=category_id,
            delivery_days_min=delivery_days_min,
            delivery_days_max=delivery_days_max,
        )
        # name/description are django-modeltranslation fields —
        # .objects.create() kwargs silently drop them, so assign as plain
        # attributes instead. Saved immediately so book.id exists below
        # (generate_cover_image needs a real PK).
        book.name = name
        book.description = description
        book.save()
        if 'cover_image' in request.FILES:
            book.cover_image = request.FILES['cover_image']
        else:
            # Auto-generate a stylish cover
            try:
                from manager.cover_generator import generate_cover_image
                book.cover_image = generate_cover_image(book.name, book.id)
            except Exception:
                pass

        if book_file:
            book.book_file = book_file
        if download_link:
            book.download_link = download_link

        book.save()

        # Associate authors with the book
        if author_ids:
            authors = models.Author.objects.filter(id__in=author_ids)
            for author in authors:
                author.book.add(book)

        models.VendorBook.objects.create(
            vendor=vendor,
            book=book,
        )
        return JsonResponse({'success': True, 'message': '图书已上架', 'redirect_url': reverse('manager:vendor_books')})

    publishers = models.Publisher.objects.all()
    authors = models.Author.objects.all()
    categories = models.BookCategory.objects.filter(Q(vendor__isnull=True) | Q(vendor=vendor), is_active=True)
    context = {
        'vendor': vendor,
        'publishers': publishers,
        'authors': authors,
        'categories': categories,
        'admin_access': admin_access,
    }
    return render(request, 'public/vendor_add_book.html', context)


def vendor_toggle_book(request):
    """Toggle a vendor book active status"""
    vendor = _get_vendor(request)
    admin_access = request.session.get("name")
    if not vendor and not admin_access:
        return JsonResponse({'success': False, 'message': '未授权'})

    if request.method == 'POST':
        vb = get_object_or_404(models.VendorBook, id=request.POST.get('id'))
        if not admin_access and vb.vendor != vendor:
            return JsonResponse({'success': False, 'message': '无权限'})
        vb.is_active = not vb.is_active
        vb.save(update_fields=['is_active'])
        return JsonResponse({'success': True, 'is_active': vb.is_active})
    return JsonResponse({'success': False})


def vendor_edit_book(request):
    """Vendor edits an existing book"""
    vendor = _get_vendor(request)
    admin_access = request.session.get("name")
    if not vendor and not admin_access:
        return redirect('/manager/vendor/login/')

    if request.method == 'GET':
        vb_id = request.GET.get('id')
        vb = get_object_or_404(models.VendorBook, id=vb_id)
        if not admin_access and vb.vendor != vendor:
            return JsonResponse({'success': False, 'message': '无权限'})
        book = vb.book
        authors = models.Author.objects.all()
        book_author_ids = list(book.author_set.values_list('id', flat=True))
        publishers = models.Publisher.objects.all()
        categories = models.BookCategory.objects.filter(Q(vendor__isnull=True) | Q(vendor=vb.vendor), is_active=True)
        return render(request, 'public/vendor_edit_book.html', {
            'vendor': vendor or vb.vendor,
            'vb': vb,
            'book': book,
            'publishers': publishers,
            'authors': authors,
            'categories': categories,
            'book_author_ids': book_author_ids,
            'admin_access': admin_access,
        })

    if request.method == 'POST':
        # Handle inline creation of new publisher
        if request.POST.get('action') == 'create_publisher':
            pub_name = request.POST.get('publisher_name', '').strip()
            pub_address = request.POST.get('publisher_address', '').strip()
            if pub_name:
                # publisher_name/publisher_address are django-modeltranslation
                # fields — .objects.create() kwargs silently drop them, so
                # assign as plain attributes instead.
                pub = models.Publisher()
                pub.publisher_name = pub_name
                pub.publisher_address = pub_address
                pub.save()
                return JsonResponse({'success': True, 'id': pub.id, 'name': pub.publisher_name})
            return JsonResponse({'success': False, 'message': '出版社名称不能为空'})

        # Handle inline creation of new author
        if request.POST.get('action') == 'create_author':
            author_name = request.POST.get('author_name', '').strip()
            if author_name:
                author = models.Author()
                author.name = author_name
                author.save()
                return JsonResponse({'success': True, 'id': author.id, 'name': author.name})
            return JsonResponse({'success': False, 'message': '作者名称不能为空'})

        vb_id = request.POST.get('vb_id')
        vb = get_object_or_404(models.VendorBook, id=vb_id)
        if not admin_access and vb.vendor != vendor:
            return JsonResponse({'success': False, 'message': '无权限'})

        book = vb.book
        name = request.POST.get('name', '').strip()
        price = request.POST.get('price', '0')
        inventory = request.POST.get('inventory', '0')
        description = request.POST.get('description', '').strip()
        publisher_id = request.POST.get('publisher_id')
        category_id = request.POST.get('category_id') or None
        author_ids = request.POST.getlist('author_ids')

        if not name:
            return JsonResponse({'success': False, 'message': '图书标题不能为空。', 'field': 'name'})
        if len(name) < 3:
            return JsonResponse({'success': False, 'message': '图书标题至少需要 3 个字符。', 'field': 'name'})
        if not description:
            return JsonResponse({'success': False, 'message': '图书描述不能为空。', 'field': 'description'})
        if len(description) < 12:
            return JsonResponse({'success': False, 'message': '图书描述至少需要 12 个字符。', 'field': 'description'})
        if models.VendorBook.objects.filter(vendor=vb.vendor, book__name__iexact=name).exclude(pk=vb.pk).exists():
            return JsonResponse({'success': False, 'message': 'Vous avez déjà publié un livre avec ce titre. Utilisez un titre différent.', 'field': 'name'})
        if not publisher_id:
            return JsonResponse({'success': False, 'message': '出版社不能为空。', 'field': 'publisher_id'})
        if not author_ids:
            return JsonResponse({'success': False, 'message': '请至少选择一位作者。', 'field': 'author_ids'})
        try:
            if Decimal(price) <= 0:
                return JsonResponse({'success': False, 'message': '价格必须大于 0。', 'field': 'price'})
        except Exception:
            return JsonResponse({'success': False, 'message': '价格格式无效。', 'field': 'price'})
        try:
            if int(inventory) < 0:
                return JsonResponse({'success': False, 'message': '库存不能小于 0。', 'field': 'inventory'})
        except Exception:
            return JsonResponse({'success': False, 'message': '库存格式无效。', 'field': 'inventory'})
        download_link = request.POST.get('download_link', '').strip()
        if download_link:
            from django.core.validators import URLValidator
            from django.core.exceptions import ValidationError
            validator = URLValidator()
            try:
                validator(download_link)
            except ValidationError:
                return JsonResponse({'success': False, 'message': '请输入有效的下载链接 URL。', 'field': 'download_link'})

        delivery_days_min, delivery_days_max = _parse_delivery_days_override(request.POST)
        book.name = name
        book.price = Decimal(price)
        book.inventory = int(inventory)
        book.description = description
        book.delivery_days_min = delivery_days_min
        book.delivery_days_max = delivery_days_max
        if publisher_id:
            book.publisher_id = int(publisher_id)
        book.category_id = category_id
        if 'cover_image' in request.FILES:
            book.cover_image = request.FILES['cover_image']
        if 'book_file' in request.FILES:
            book.book_file = request.FILES['book_file']
        if download_link:
            book.download_link = download_link
        elif request.POST.get('clear_download_link') == '1':
            book.download_link = ''
        book.save()

        # Update author associations
        book.author_set.clear()
        if author_ids:
            authors = models.Author.objects.filter(id__in=author_ids)
            for author in authors:
                author.book.add(book)

        return JsonResponse({'success': True, 'message': '图书信息已更新', 'redirect_url': reverse('manager:vendor_books')})

    return JsonResponse({'success': False})


# ─── Vendor Book Categories (own categories, separate from admin's) ────────

def vendor_book_categories(request):
    """
    Show the vendor's own book categories (editable) plus the platform
    admin's global ones (read-only reference — managed only from the admin
    panel), each with a book count. Category names are unique per vendor
    only, not checked against admin's, so seeing both side by side avoids
    a vendor unknowingly re-creating something that already exists.
    """
    vendor = _get_vendor(request)
    if not vendor:
        return redirect('/manager/vendor/login/')
    own_categories = models.BookCategory.objects.filter(vendor=vendor).annotate(
        book_count=Count('books')
    ).order_by('display_order', 'name')
    admin_categories = models.BookCategory.objects.filter(vendor__isnull=True).annotate(
        book_count=Count('books')
    ).order_by('display_order', 'name')
    return render(request, 'public/vendor_book_categories.html', {
        'vendor': vendor,
        'categories': own_categories,
        'admin_categories': admin_categories,
    })


def _translate_book_category_names(name, source_lang):
    """
    Fills in name_en/name_fr for a vendor-created BookCategory via OpenRouter.
    BookCategory doesn't use django-modeltranslation (plain name_en/name_fr
    CharFields, no auto-translate signal), so this is called directly at
    save time instead of relying on a post_save hook like Product/Course do.
    """
    from core.services.translation_service import TranslationService
    targets = [lang for lang in ('en', 'fr', 'zh-hans') if lang != source_lang][:2]
    svc = TranslationService()
    translated = svc.translate_fields(name, source=source_lang, targets=tuple(targets), content_type='general')
    return {
        'name_en': translated.get('en', ''),
        'name_fr': translated.get('fr', ''),
    }


def vendor_book_category_add(request):
    vendor = _get_vendor(request)
    if not vendor:
        return redirect('/manager/vendor/login/')

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        if not name:
            messages.error(request, 'Le nom de la catégorie est obligatoire.')
            return render(request, 'public/vendor_book_category_form.html', {'vendor': vendor, 'category': None})
        if len(name) < 2:
            messages.error(request, 'Le nom de la catégorie doit contenir au moins 2 caractères.')
            return render(request, 'public/vendor_book_category_form.html', {'vendor': vendor, 'category': None})
        if models.BookCategory.objects.filter(vendor=vendor, name__iexact=name).exists():
            messages.error(request, 'Vous avez déjà une catégorie avec ce nom.')
            return render(request, 'public/vendor_book_category_form.html', {'vendor': vendor, 'category': None})

        base_slug = slugify(name, allow_unicode=True) or f'vcat-{uuid.uuid4().hex[:8]}'
        slug = base_slug
        counter = 1
        while models.BookCategory.objects.filter(slug=slug).exists():
            slug = f'{base_slug}-{counter}'
            counter += 1

        from django.utils import translation
        current_lang = translation.get_language()
        source_lang = current_lang if current_lang in ('en', 'fr', 'zh-hans') else 'fr'
        names = _translate_book_category_names(name, source_lang)

        category = models.BookCategory.objects.create(
            vendor=vendor,
            name=name,
            name_en=names['name_en'] if source_lang != 'en' else name,
            name_fr=names['name_fr'] if source_lang != 'fr' else name,
            slug=slug,
            description=description,
            is_active=True,
        )
        messages.success(request, f'Catégorie "{category.name}" créée.')
        return redirect('manager:vendor_book_categories')

    return render(request, 'public/vendor_book_category_form.html', {'vendor': vendor, 'category': None})


def vendor_book_category_edit(request, pk):
    vendor = _get_vendor(request)
    if not vendor:
        return redirect('/manager/vendor/login/')
    category = get_object_or_404(models.BookCategory, pk=pk, vendor=vendor)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        if not name or len(name) < 2:
            messages.error(request, 'Le nom de la catégorie doit contenir au moins 2 caractères.')
            return render(request, 'public/vendor_book_category_form.html', {'vendor': vendor, 'category': category})
        if models.BookCategory.objects.filter(vendor=vendor, name__iexact=name).exclude(pk=category.pk).exists():
            messages.error(request, 'Vous avez déjà une catégorie avec ce nom.')
            return render(request, 'public/vendor_book_category_form.html', {'vendor': vendor, 'category': category})

        category.name = name
        category.name_en = request.POST.get('name_en', '').strip()
        category.name_fr = request.POST.get('name_fr', '').strip()
        category.description = description
        category.save()
        messages.success(request, f'Catégorie "{category.name}" mise à jour.')
        return redirect('manager:vendor_book_categories')

    return render(request, 'public/vendor_book_category_form.html', {'vendor': vendor, 'category': category})


@require_POST
def vendor_book_category_delete(request, pk):
    vendor = _get_vendor(request)
    if not vendor:
        return redirect('/manager/vendor/login/')
    category = get_object_or_404(models.BookCategory, pk=pk, vendor=vendor)
    category.delete()
    messages.success(request, 'Catégorie supprimée.')
    return redirect('manager:vendor_book_categories')


def vendor_delete_book(request):
    """Vendor removes a book from their listings"""
    vendor = _get_vendor(request)
    admin_access = request.session.get("name")
    if not vendor and not admin_access:
        return JsonResponse({'success': False, 'message': '未授权'})

    if request.method == 'POST':
        vb_id = request.POST.get('id')
        vb = get_object_or_404(models.VendorBook, id=vb_id)
        if not admin_access and vb.vendor != vendor:
            return JsonResponse({'success': False, 'message': '无权限'})
        vb.delete()
        return JsonResponse({'success': True, 'message': '图书已从您的店铺中移除'})
    return JsonResponse({'success': False})


# Admin returns & shipment oversight
@ensure_csrf_cookie
def admin_returns_queue(request):
    """Admin queue for open return/dispute requests, plus a proactive alert
    list for shipments running late (past their seller-acceptance SLA or
    approaching their auto-confirm-delivery safety net), and a full
    fulfillment-funnel breakdown — the kind of visibility the platform had
    none of before this system existed."""
    if "name" not in request.session:
        return redirect('/manager/login/')

    from django.utils import timezone as _tz

    search_q = request.GET.get('search', '').strip()

    pending_qs = models.OrderReturnRequest.objects.filter(status='pending').select_related(
        'shipment', 'shipment__vendor',
    )
    if search_q:
        pending_qs = pending_qs.filter(shipment__order_number__icontains=search_q)
    pending_returns = list(pending_qs.order_by('-created_at'))

    resolved_qs = models.OrderReturnRequest.objects.exclude(status='pending').select_related(
        'shipment', 'shipment__vendor',
    )
    if search_q:
        resolved_qs = resolved_qs.filter(shipment__order_number__icontains=search_q)
    resolved_paginator = Paginator(resolved_qs.order_by('-updated_at'), 10)
    resolved_returns = resolved_paginator.get_page(request.GET.get('page', 1))

    now = _tz.now()
    late_acceptance = list(
        models.Shipment.objects.filter(fulfillment_status='awaiting_acceptance', accept_by__lt=now)
        .select_related('vendor').order_by('accept_by')
    )
    approaching_auto_confirm = list(
        models.Shipment.objects.filter(
            fulfillment_status__in=('shipped', 'in_transit', 'out_for_delivery'),
            auto_confirm_at__lt=now + timedelta(days=2),
        ).select_related('vendor').order_by('auto_confirm_at')
    )

    # Full fulfillment-funnel breakdown — every status, one query.
    status_counts = dict(
        models.Shipment.objects.values_list('fulfillment_status')
        .annotate(c=Count('id')).values_list('fulfillment_status', 'c')
    )
    total_shipments = sum(status_counts.values())
    active_statuses = ('awaiting_acceptance', 'accepted', 'packing', 'shipped', 'in_transit', 'out_for_delivery')
    active_shipments_count = sum(status_counts.get(s, 0) for s in active_statuses)
    dispute_statuses = ('return_requested', 'return_approved')
    open_dispute_count = sum(status_counts.get(s, 0) for s in dispute_statuses)

    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    returned_this_month = models.Shipment.objects.filter(
        fulfillment_status='returned', updated_at__gte=month_start,
    ).count()
    delivered_this_month = models.Shipment.objects.filter(
        fulfillment_status__in=('delivered', 'completed'), delivered_at__gte=month_start,
    ).count()
    resolved_this_week = models.OrderReturnRequest.objects.exclude(status='pending').filter(
        updated_at__gte=now - timedelta(days=7),
    ).count()

    status_funnel = [
        {'key': key, 'label': label, 'count': status_counts.get(key, 0)}
        for key, label in models.Shipment.FULFILLMENT_STATUS_CHOICES
    ]
    max_funnel_count = max((row['count'] for row in status_funnel), default=0)

    context = {
        'name': request.session.get('name'),
        'pending_returns': pending_returns,
        'resolved_returns': resolved_returns,
        'late_acceptance': late_acceptance,
        'approaching_auto_confirm': approaching_auto_confirm,
        'pending_returns_count': len(pending_returns),
        'search_query': search_q,
        'total_shipments': total_shipments,
        'active_shipments_count': active_shipments_count,
        'open_dispute_count': open_dispute_count,
        'returned_this_month': returned_this_month,
        'delivered_this_month': delivered_this_month,
        'resolved_this_week': resolved_this_week,
        'status_funnel': status_funnel,
        'max_funnel_count': max_funnel_count,
    }
    return render(request, 'admin/returns_queue.html', context)


@require_POST
def admin_resolve_return(request):
    """Admin approves/rejects a pending return request."""
    from manager import fulfillment_service as fs

    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '未授权'}, status=403)

    return_id = request.POST.get('return_id')
    decision = request.POST.get('decision', '').strip()
    note = request.POST.get('note', '').strip()
    return_request = get_object_or_404(models.OrderReturnRequest, id=return_id)

    ok = fs.resolve_return_request(return_request, decision, resolution_note=note, resolved_by='admin')
    if not ok:
        return JsonResponse({'success': False, 'message': '操作失败（状态已变更或参数无效）'}, status=400)
    return JsonResponse({'success': True, 'message': '已处理'})


@require_POST
def admin_confirm_return_received(request):
    """Admin confirms a returned item was physically received back —
    finalizes the return and issues the real refund."""
    from manager import fulfillment_service as fs

    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '未授权'}, status=403)

    return_id = request.POST.get('return_id')
    note = request.POST.get('note', '').strip()
    return_request = get_object_or_404(models.OrderReturnRequest, id=return_id)

    ok = fs.confirm_return_received(return_request, resolution_note=note)
    if not ok:
        return JsonResponse({'success': False, 'message': '操作失败（退货尚未获批准）'}, status=400)
    return JsonResponse({'success': True, 'message': '已确认收货并发起退款'})


# Admin vendor management
@ensure_csrf_cookie
def admin_vendor_list(request):
    """Admin view all vendors with marketplace metrics"""
    if "name" not in request.session:
        return redirect('/manager/login/')

    vendors = models.Vendor.objects.all()
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    if search:
        vendors = vendors.filter(Q(company_name__icontains=search) | Q(email__icontains=search) | Q(contact_name__icontains=search))
    if status_filter in dict(models.VENDOR_STATUS_CHOICES):
        vendors = vendors.filter(status=status_filter)

    vendors = vendors.order_by('-is_official', '-created_at')
    from django.db.models import Sum, Count, Avg
    vendors = vendors.annotate(
        product_count=Count('products', distinct=True),
        course_count=Count('courses', distinct=True),
        supermarket_count=Count('supermarket_items', distinct=True),
        book_count=Count('vendorbook', distinct=True),
        mp_total_sales=Sum('products__sales_count'),
        mp_total_stock=Sum('products__stock'),
        mp_total_enrollments=Sum('courses__enrollment_count'),
    )

    # Global marketplace metrics
    all_vendors = models.Vendor.objects.all()
    total_vendor_books = models.VendorBook.objects.filter(vendor__isnull=False).count()
    total_mp_products = Product.objects.filter(vendor__isnull=False).count()
    total_mp_courses = Course.objects.filter(vendor__isnull=False).count()
    total_mp_supermarket = SupermarketItem.objects.filter(vendor__isnull=False).count()
    total_mp_sales = Product.objects.filter(vendor__isnull=False).aggregate(s=Sum('sales_count'))['s'] or 0
    total_mp_enrollments = Course.objects.filter(vendor__isnull=False).aggregate(s=Sum('enrollment_count'))['s'] or 0
    total_mp_revenue = MarketplaceOrderItem.objects.filter(
        order__status__in=['paid', 'processing', 'shipped', 'delivered']
    ).aggregate(s=Sum('subtotal'))['s'] or 0

    # Top selling products across all vendors
    top_products = Product.objects.filter(
        vendor__isnull=False, sales_count__gt=0
    ).select_related('vendor', 'category').order_by('-sales_count')[:5]

    # Top vendors by product sales
    top_vendors = models.Vendor.objects.annotate(
        total_sales=Sum('products__sales_count')
    ).filter(total_sales__gt=0).order_by('-total_sales')[:5]

    # Vendor being viewed in detail (read-only)
    view_vendor_id = request.GET.get('view_vendor')
    viewed_vendor = None
    vendor_products_list = []
    vendor_courses_list = []
    vendor_books_list = []
    vendor_supermarket_list = []
    if view_vendor_id:
        viewed_vendor = models.Vendor.objects.filter(id=view_vendor_id).first()
        if viewed_vendor:
            vendor_products_list = Product.objects.filter(vendor=viewed_vendor).select_related('category').order_by('-created_at')[:50]
            vendor_courses_list = Course.objects.filter(vendor=viewed_vendor).select_related('category').order_by('-created_at')[:50]
            vendor_books_list = models.VendorBook.objects.filter(vendor=viewed_vendor).select_related('book', 'book__publisher').order_by('-created_at')[:50]
            vendor_supermarket_list = SupermarketItem.objects.filter(vendor=viewed_vendor).select_related('category').order_by('-created_at')[:50]

    from manager.official_store import get_official_vendor
    from manager.commission import commission_rates_for_display
    official_vendor = get_official_vendor(create=False)

    return render(request, 'admin/vendor_list.html', {
        'vendors': vendors,
        'official_vendor': official_vendor,
        'total_vendors': models.Vendor.objects.count(),
        'search_query': search,
        'status_filter': status_filter,
        'name': request.session.get('name', ''),
        'total_mp_products': total_mp_products,
        'total_vendor_books': total_vendor_books,
        'total_mp_courses': total_mp_courses,
        'total_mp_supermarket': total_mp_supermarket,
        'total_mp_sales': total_mp_sales,
        'total_mp_enrollments': total_mp_enrollments,
        'total_mp_revenue': total_mp_revenue,
        'top_products': top_products,
        'top_vendors': top_vendors,
        'viewed_vendor': viewed_vendor,
        'vendor_products_list': vendor_products_list,
        'vendor_courses_list': vendor_courses_list,
        'vendor_books_list': vendor_books_list,
        'vendor_supermarket_list': vendor_supermarket_list,
        'commission_rates': commission_rates_for_display(),
    })


def admin_vendor_status(request):
    """Admin approve/reject/suspend vendor"""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '未授权'})
    if request.method == 'POST':
        vendor = get_object_or_404(models.Vendor, id=request.POST.get('id'))
        new_status = request.POST.get('status')
        if new_status in dict(models.VENDOR_STATUS_CHOICES):
            vendor.status = new_status
            update_fields = ['status']
            if new_status != 'approved' and vendor.is_certified:
                vendor.is_certified = False
                vendor.certified_at = None
                update_fields.extend(['is_certified', 'certified_at'])
            vendor.save(update_fields=update_fields)
            return JsonResponse({'success': True, 'message': f'状态已更新为 {vendor.get_status_display()}'})
    return JsonResponse({'success': False})


@require_POST
def admin_vendor_certify(request):
    """Grant or revoke certified seller badge (approved vendors only)."""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '未授权'}, status=403)
    vendor = get_object_or_404(models.Vendor, id=request.POST.get('id'))
    if vendor.is_official:
        return JsonResponse({'success': False, 'message': '官方直营店无需认证'})
    want_certified = request.POST.get('certified', '1') == '1'
    if want_certified:
        if not vendor.can_receive_certification():
            return JsonResponse({
                'success': False,
                'message': '仅已批准且启用的卖家可授予认证徽章',
            })
        vendor.is_certified = True
        vendor.certified_at = timezone.now()
        vendor.save(update_fields=['is_certified', 'certified_at'])
        return JsonResponse({'success': True, 'message': f'已认证卖家「{vendor.company_name}」'})
    vendor.is_certified = False
    vendor.certified_at = None
    vendor.save(update_fields=['is_certified', 'certified_at'])
    return JsonResponse({'success': True, 'message': f'已取消「{vendor.company_name}」的认证'})


@require_POST
def admin_delete_vendor_item(request):
    """Admin remove a vendor's non-compliant listing from the platform."""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '未授权'}, status=403)
    item_type = request.POST.get('type')
    item_id = request.POST.get('id')
    try:
        with transaction.atomic():
            if item_type == 'product':
                item = get_object_or_404(Product, id=item_id)
                name = item.name
                item.delete()
                return JsonResponse({'success': True, 'message': f'商品 "{name}" 已删除'})
            if item_type == 'course':
                item = get_object_or_404(Course, id=item_id)
                name = item.title
                item.delete()
                return JsonResponse({'success': True, 'message': f'课程 "{name}" 已删除'})
            if item_type == 'book':
                item = get_object_or_404(models.VendorBook, id=item_id)
                name = item.book.name
                item.delete()
                return JsonResponse({'success': True, 'message': f'图书 "{name}" 已从卖家下架'})
            if item_type == 'supermarket':
                item = get_object_or_404(SupermarketItem, id=item_id)
                name = item.name
                item.delete()
                return JsonResponse({'success': True, 'message': f'超市商品 "{name}" 已删除'})
    except Exception as exc:
        logger.exception('Admin failed to delete vendor item type=%s id=%s', item_type, item_id)
        return JsonResponse({'success': False, 'message': f'删除失败：{exc}'}, status=400)
    return JsonResponse({'success': False, 'message': '不支持的内容类型'}, status=400)


# ==========================================
# Notification System
# ==========================================

def create_vendor_notification(vendor_id, ntype, title, message, icon='fas fa-bell', color='#10b981', link='', related_id=None):
    """Create a notification for a specific vendor."""
    if not vendor_id:
        return
    already = models.VendorNotification.objects.filter(
        vendor_id=vendor_id,
        notification_type=ntype,
        related_id=related_id,
    ).exists() if related_id else False
    if already:
        return
    models.VendorNotification.objects.create(
        vendor_id=vendor_id,
        notification_type=ntype,
        title=title,
        message=message,
        icon=icon,
        color=color,
        link=link,
        related_id=related_id,
    )


def _check_vendor_notifications(vendor_id, lookback_days=30):
    """Generate pending notifications for a vendor (scoped to vendor_id only)."""
    from datetime import timedelta
    from marketplace.models import Product, Course, SupermarketItem

    now = timezone.now()
    cutoff = now - timedelta(days=lookback_days)
    vendor = models.Vendor.objects.filter(pk=vendor_id).first()
    if not vendor:
        return

    # 1. Unread messages from buyers
    unread_convos = models.Conversation.objects.filter(
        vendor_id=vendor_id,
        direct_messages__is_read=False,
    ).exclude(
        direct_messages__sender_type='vendor',
    ).distinct()
    for convo in unread_convos[:15]:
        unread_count = convo.direct_messages.filter(is_read=False).exclude(sender_type='vendor').count()
        if unread_count > 0:
            buyer_name = convo.buyer.name if convo.buyer else 'Customer'
            create_vendor_notification(
                vendor_id, 'new_message',
                f'New message from {buyer_name}',
                f'{unread_count} unread message(s) about: {convo.subject or "General"}',
                icon='fas fa-comment-dots',
                color='#3b82f6',
                link=f'/manager/vendor/messages/?conversation={convo.id}',
                related_id=convo.id,
            )

    vendor_books = list(models.VendorBook.objects.filter(vendor=vendor, is_active=True).values_list('book_id', flat=True))
    product_ids = list(Product.objects.filter(vendor=vendor).values_list('id', flat=True))
    course_ids = list(Course.objects.filter(vendor=vendor).values_list('id', flat=True))
    supermarket_ids = list(SupermarketItem.objects.filter(vendor=vendor).values_list('id', flat=True))

    # 2. Paid book orders
    if vendor_books:
        paid_book_order_ids = models.OrderItem.objects.filter(
            book_id__in=vendor_books,
            order__created_at__gte=cutoff,
            order__payment_status='completed',
        ).values_list('order_id', flat=True).distinct()
        for oid in paid_book_order_ids[:30]:
            order = models.Order.objects.filter(pk=oid).first()
            if not order:
                continue
            create_vendor_notification(
                vendor_id, 'new_order',
                'New book purchase',
                f'Order {order.order_number} — {order.customer_name} · {order.total_amount}',
                icon='fas fa-shopping-bag',
                color='#10b981',
                link=f'/manager/vendor/orders/books/{oid}/',
                related_id=oid,
            )
        abandoned_book_ids = models.OrderItem.objects.filter(
            book_id__in=vendor_books,
            order__created_at__gte=cutoff,
            order__payment_status='pending',
            order__status__in=['pending', 'payment_pending'],
        ).values_list('order_id', flat=True).distinct()
        for oid in abandoned_book_ids[:20]:
            order = models.Order.objects.filter(pk=oid).first()
            if not order:
                continue
            create_vendor_notification(
                vendor_id, 'abandoned_checkout',
                'Checkout without payment',
                f'Order {order.order_number} — {order.customer_name} did not complete payment.',
                icon='fas fa-cart-arrow-down',
                color='#f59e0b',
                link=f'/manager/vendor/orders/books/{oid}/',
                related_id=oid,
            )

    # 3. Paid marketplace orders touching this vendor
    mkt_q = Q()
    if product_ids:
        mkt_q |= Q(item_type='product', item_id__in=product_ids)
    if course_ids:
        mkt_q |= Q(item_type='course', item_id__in=course_ids)
    if supermarket_ids:
        mkt_q |= Q(item_type='supermarket', item_id__in=supermarket_ids)
    if mkt_q:
        paid_mkt_ids = MarketplaceOrderItem.objects.filter(
            mkt_q,
            order__created_at__gte=cutoff,
            order__payment_status='completed',
        ).values_list('order_id', flat=True).distinct()
        for oid in paid_mkt_ids[:30]:
            order = MarketplaceOrder.objects.filter(pk=oid).first()
            if not order:
                continue
            create_vendor_notification(
                vendor_id, 'new_order',
                'New marketplace purchase',
                f'Order {order.order_number} — {order.user_name or order.user_email}',
                icon='fas fa-store',
                color='#10b981',
                link=reverse('marketplace:vendor_marketplace_order_detail', args=[oid]),
                related_id=oid,
            )
        abandoned_mkt_ids = MarketplaceOrderItem.objects.filter(
            mkt_q,
            order__created_at__gte=cutoff,
            order__payment_status='pending',
            order__status__in=['pending', 'payment_pending'],
        ).values_list('order_id', flat=True).distinct()
        for oid in abandoned_mkt_ids[:20]:
            order = MarketplaceOrder.objects.filter(pk=oid).first()
            if not order:
                continue
            create_vendor_notification(
                vendor_id, 'abandoned_checkout',
                'Checkout without payment',
                f'Order {order.order_number} — payment not completed.',
                icon='fas fa-cart-arrow-down',
                color='#f59e0b',
                link=reverse('marketplace:vendor_marketplace_order_detail', args=[oid]),
                related_id=oid,
            )

    # 4. New followers (last month)
    for follow in models.UserFollowedVendor.objects.filter(vendor=vendor, followed_at__gte=cutoff).select_related('user')[:30]:
        create_vendor_notification(
            vendor_id, 'new_follower',
            f'{follow.user.name} follows your shop',
            f'{follow.user.email} subscribed to {vendor.company_name}.',
            icon='fas fa-user-plus',
            color='#8b5cf6',
            link='/manager/vendor/dashboard/',
            related_id=follow.id,
        )

    # 5. Wishlist adds (last month)
    for wish in models.Wishlist.objects.filter(created_at__gte=cutoff).select_related('user', 'book')[:50]:
        if wish.item_type == 'book' and wish.book_id in vendor_books:
            create_vendor_notification(
                vendor_id, 'wishlist_add',
                f'{wish.user.name} favorited a book',
                wish.book.name,
                icon='fas fa-heart',
                color='#ec4899',
                link='/manager/vendor/dashboard/',
                related_id=wish.id,
            )
        elif wish.item_type == 'product' and wish.item_id in product_ids:
            p = Product.objects.filter(pk=wish.item_id).first()
            create_vendor_notification(
                vendor_id, 'wishlist_add',
                f'{wish.user.name} favorited a product',
                p.name if p else 'Product',
                icon='fas fa-heart',
                color='#ec4899',
                link='/manager/vendor/dashboard/',
                related_id=wish.id,
            )
        elif wish.item_type == 'course' and wish.item_id in course_ids:
            c = Course.objects.filter(pk=wish.item_id).first()
            create_vendor_notification(
                vendor_id, 'wishlist_add',
                f'{wish.user.name} favorited a course',
                c.title if c else 'Course',
                icon='fas fa-heart',
                color='#ec4899',
                link='/manager/vendor/dashboard/',
                related_id=wish.id,
            )
        elif wish.item_type == 'supermarket' and wish.item_id in supermarket_ids:
            s = SupermarketItem.objects.filter(pk=wish.item_id).first()
            create_vendor_notification(
                vendor_id, 'wishlist_add',
                f'{wish.user.name} favorited a supermarket item',
                s.name if s else 'Item',
                icon='fas fa-heart',
                color='#ec4899',
                link='/manager/vendor/dashboard/',
                related_id=wish.id,
            )


@ensure_csrf_cookie
def vendor_notifications_page(request):
    """Full notifications page for vendor — last 30 days, vendor-scoped only."""
    from datetime import timedelta

    admin_access = request.session.get('name')
    vendor = _get_vendor(request)
    if not vendor and not admin_access:
        return redirect('/manager/vendor/login/')
    if admin_access and not vendor:
        vid = request.GET.get('vendor_id')
        if vid:
            vendor = get_object_or_404(models.Vendor, id=vid)
        else:
            return redirect('/manager/vendor/dashboard/')

    cutoff = timezone.now() - timedelta(days=30)
    _check_vendor_notifications(vendor.id, lookback_days=30)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'mark_all_read':
            models.VendorNotification.objects.filter(vendor=vendor, is_read=False).update(is_read=True)
            messages.success(request, _('All notifications marked as read.'))
        elif action == 'clear_read':
            models.VendorNotification.objects.filter(vendor=vendor, is_read=True, created_at__gte=cutoff).delete()
            messages.success(request, _('Read notifications cleared.'))
        redirect_url = reverse('manager:vendor_notifications')
        if admin_access:
            redirect_url = f'{redirect_url}?vendor_id={vendor.id}'
        return redirect(redirect_url)

    notifs = models.VendorNotification.objects.filter(
        vendor=vendor,
        created_at__gte=cutoff,
    ).order_by('-created_at')
    unread_count = notifs.filter(is_read=False).count()

    context = {
        'vendor': vendor,
        'admin_access': admin_access,
        'notifications': notifs,
        'unread_count': unread_count,
        'lookback_days': 30,
    }
    return render(request, 'public/vendor_notifications.html', context)


@ensure_csrf_cookie
def vendor_payments_page(request):
    """Vendor escrow & payout tracking — pending, releasable, released."""
    admin_access = request.session.get('name')
    vendor = _get_vendor(request)
    if not vendor and not admin_access:
        return redirect('/manager/vendor/login/')
    if admin_access and not vendor:
        vid = request.GET.get('vendor_id')
        if vid:
            vendor = get_object_or_404(models.Vendor, id=vid)
        else:
            return redirect('/manager/vendor/dashboard/')

    from manager.commission import commission_rates_for_display
    from manager.escrow_service import REFUND_HOLD_DAYS

    status_filter = request.GET.get('status', '').strip()
    qs = models.PlatformEscrowTransaction.objects.filter(vendor=vendor).order_by('-held_at')
    if status_filter:
        qs = qs.filter(status=status_filter)

    escrow_stats = {
        'held': qs.filter(status='held').aggregate(c=Count('id'), s=Sum('vendor_payout_amount')),
        'releasable': qs.filter(status='releasable').aggregate(c=Count('id'), s=Sum('vendor_payout_amount')),
        'released': qs.filter(status='released').aggregate(c=Count('id'), s=Sum('vendor_payout_amount')),
        'refunded': qs.filter(status__in=('refunded', 'cancelled')).aggregate(c=Count('id'), s=Sum('vendor_payout_amount')),
    }
    vendor_wallet = models.VendorWallet.objects.filter(vendor=vendor).first()
    wallet_txns = []
    if vendor_wallet:
        wallet_txns = list(
            models.VendorWalletTransaction.objects.filter(vendor=vendor).order_by('-created_at')[:30]
        )

    context = {
        'vendor': vendor,
        'admin_access': admin_access,
        'transactions': qs[:200],
        'status_filter': status_filter,
        'status_choices': models.PlatformEscrowTransaction.STATUS_CHOICES,
        'escrow_stats': escrow_stats,
        'vendor_wallet': vendor_wallet,
        'wallet_txns': wallet_txns,
        'refund_hold_days': REFUND_HOLD_DAYS,
        'commission_rates': commission_rates_for_display(),
    }
    return render(request, 'public/vendor_payments.html', context)


@ensure_csrf_cookie
def vendor_notifications_api(request):
    """API endpoint for vendor notifications (GET list, POST actions)."""
    vendor_id = request.session.get('vendor_id')
    if not vendor_id:
        return JsonResponse({'success': False, 'message': 'Not logged in'}, status=401)

    if request.method == 'GET':
        _check_vendor_notifications(vendor_id, lookback_days=7)
        from datetime import timedelta
        cutoff = timezone.now() - timedelta(days=30)
        notifs = models.VendorNotification.objects.filter(
            vendor_id=vendor_id,
            created_at__gte=cutoff,
        ).order_by('-created_at')[:50]
        unread_count = models.VendorNotification.objects.filter(vendor_id=vendor_id, is_read=False).count()
        data = [{
            'id': n.id,
            'type': n.notification_type,
            'title': n.title,
            'message': n.message,
            'icon': n.icon,
            'color': n.color,
            'link': n.link,
            'is_read': n.is_read,
            'time': n.created_at.strftime('%Y-%m-%d %H:%M'),
        } for n in notifs]
        return JsonResponse({'success': True, 'notifications': data, 'unread_count': unread_count})

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'mark_read':
            nid = request.POST.get('id')
            if nid:
                models.VendorNotification.objects.filter(id=nid, vendor_id=vendor_id).update(is_read=True)
            return JsonResponse({'success': True})
        elif action == 'mark_all_read':
            models.VendorNotification.objects.filter(vendor_id=vendor_id, is_read=False).update(is_read=True)
            return JsonResponse({'success': True})
        elif action == 'delete':
            nid = request.POST.get('id')
            if nid:
                models.VendorNotification.objects.filter(id=nid, vendor_id=vendor_id).delete()
            return JsonResponse({'success': True})
        elif action == 'clear_all':
            models.VendorNotification.objects.filter(vendor_id=vendor_id, is_read=True).delete()
            return JsonResponse({'success': True})
    return JsonResponse({'success': False})


def create_notification(ntype, title, message, icon='fas fa-bell', color='#667eea', link='', related_id=None):
    """Helper to create an admin notification"""
    models.AdminNotification.objects.create(
        notification_type=ntype,
        title=title,
        message=message,
        icon=icon,
        color=color,
        link=link,
        related_id=related_id,
    )


def check_and_create_notifications(request):
    """Generate notifications for pending events (called periodically)"""
    now = timezone.now()

    # 1. Incomplete registrations (PIN expired, not verified)
    expired_cutoff = now - timedelta(hours=1)
    expired_regs = models.EmailVerification.objects.filter(
        is_verified=False,
        expires_at__lt=now,
        created_at__gte=expired_cutoff,
    )
    for reg in expired_regs:
        exists = models.AdminNotification.objects.filter(
            notification_type='incomplete_registration',
            related_id=reg.id,
        ).exists()
        if not exists:
            create_notification(
                'incomplete_registration',
                f'用户 {reg.name} 注册未完成',
                f'{reg.email} 在注册时未在规定时间内完成验证（验证码已过期）',
                icon='fas fa-user-clock',
                color='#f59e0b',
                related_id=reg.id,
            )

    # 2. Abandoned carts: only recent carts (30min–6h old) that haven't
    #    already been notified. Carts older than 6h are likely bots/stale.
    cart_min_age = now - timedelta(hours=6)
    cart_cutoff = now - timedelta(minutes=30)
    stale_carts = models.CartItem.objects.filter(
        updated_at__lt=cart_cutoff,
        updated_at__gte=cart_min_age,
    ).values('session_key').annotate(
        item_count=Count('id'),
        latest=Max('updated_at'),
    )
    for cart in stale_carts:
        sk_prefix = cart['session_key'][:16]
        exists = models.AdminNotification.objects.filter(
            notification_type='abandoned_cart',
            message__contains=sk_prefix,
            created_at__gte=now - timedelta(hours=24),
        ).exists()
        if not exists:
            items = models.CartItem.objects.filter(session_key=cart['session_key']).select_related('book')
            book_names = ', '.join([ci.book.name for ci in items[:3]])
            create_notification(
                'abandoned_cart',
                f'有顾客可能要下单（{cart["item_count"]}件商品）',
                f'购物车包含: {book_names}... 会话: {sk_prefix}',
                icon='fas fa-shopping-cart',
                color='#06b6d4',
                link='/manager/order_list/',
            )


def admin_notifications_api(request):
    """API endpoint for notifications"""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '未授权'})

    if request.method == 'GET':
        # Check for new events first
        check_and_create_notifications(request)

        notifications = models.AdminNotification.objects.filter(is_dismissed=False)[:50]
        unread_count = models.AdminNotification.objects.filter(is_read=False, is_dismissed=False).count()

        data = [{
            'id': n.id,
            'type': n.notification_type,
            'type_display': n.get_notification_type_display(),
            'title': n.title,
            'message': n.message,
            'icon': n.icon,
            'color': n.color,
            'link': n.link,
            'is_read': n.is_read,
            'created_at': n.created_at.strftime('%Y-%m-%d %H:%M'),
            'time_ago': _time_ago(n.created_at),
        } for n in notifications]

        return JsonResponse({'success': True, 'notifications': data, 'unread_count': unread_count})

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'mark_read':
            nid = request.POST.get('id')
            if nid:
                models.AdminNotification.objects.filter(id=nid).update(is_read=True)
            return JsonResponse({'success': True})
        elif action == 'mark_all_read':
            models.AdminNotification.objects.filter(is_read=False).update(is_read=True)
            return JsonResponse({'success': True})
        elif action == 'delete':
            nid = request.POST.get('id')
            if nid:
                models.AdminNotification.objects.filter(id=nid).delete()
            return JsonResponse({'success': True})
        elif action == 'clear_all':
            models.AdminNotification.objects.filter(is_read=True).update(is_dismissed=True)
            return JsonResponse({'success': True})

    return JsonResponse({'success': False})


def _time_ago(dt):
    """Return human-readable time ago string"""
    now = timezone.now()
    diff = now - dt
    seconds = diff.total_seconds()
    if seconds < 60:
        return '刚刚'
    elif seconds < 3600:
        return f'{int(seconds // 60)} 分钟前'
    elif seconds < 86400:
        return f'{int(seconds // 3600)} 小时前'
    elif seconds < 604800:
        return f'{int(seconds // 86400)} 天前'
    else:
        return dt.strftime('%m-%d %H:%M')


def admin_notifications_page(request):
    """Full notification management page"""
    if "name" not in request.session:
        return redirect('/manager/login/')

    # Generate any pending notifications
    check_and_create_notifications(request)

    notifications = models.AdminNotification.objects.all()

    # Filters
    type_filter = request.GET.get('type', '')
    read_filter = request.GET.get('read', '')
    if type_filter:
        notifications = notifications.filter(notification_type=type_filter)
    if read_filter == 'unread':
        notifications = notifications.filter(is_read=False)
    elif read_filter == 'read':
        notifications = notifications.filter(is_read=True)

    total = models.AdminNotification.objects.count()
    unread = models.AdminNotification.objects.filter(is_read=False).count()

    notif_list = []
    for n in notifications[:200]:
        notif_list.append({
            'id': n.id,
            'type': n.notification_type,
            'type_display': n.get_notification_type_display(),
            'title': n.title,
            'message': n.message,
            'icon': n.icon,
            'color': n.color,
            'link': n.link,
            'is_read': n.is_read,
            'related_id': n.related_id,
            'created_at': n.created_at,
            'time_ago': _time_ago(n.created_at),
        })

    context = {
        'notifications': notif_list,
        'total': total,
        'unread': unread,
        'read_count': total - unread,
        'type_filter': type_filter,
        'read_filter': read_filter,
        'notification_types': models.NOTIFICATION_TYPE_CHOICES,
        'name': request.session.get('name', ''),
    }
    return render(request, 'admin/notifications.html', context)


# ==========================================
# Enhanced Admin User Management (CRUD)
# ==========================================

def admin_edit_user(request):
    """Admin edit user"""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '未授权'})

    if request.method == 'GET':
        uid = request.GET.get('id')
        user = get_object_or_404(models.SiteUser, id=uid)
        return JsonResponse({
            'success': True,
            'user': {
                'id': user.id,
                'name': user.name,
                'email': user.email,
                'phone': user.phone,
                'is_active': user.is_active,
            }
        })

    if request.method == 'POST':
        uid = request.POST.get('id')
        user = get_object_or_404(models.SiteUser, id=uid)
        user.name = request.POST.get('name', user.name).strip()
        user.email = request.POST.get('email', user.email).strip()
        user.phone = request.POST.get('phone', user.phone).strip()
        is_active = request.POST.get('is_active')
        if is_active is not None:
            user.is_active = is_active == 'true'
        user.save()
        return JsonResponse({'success': True, 'message': '用户信息已更新'})

    return JsonResponse({'success': False})


def admin_add_user(request):
    """Admin add user manually"""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '未授权'})

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        password = request.POST.get('password', '').strip()

        if not all([name, email, password]):
            return JsonResponse({'success': False, 'message': '姓名、邮箱和密码为必填项'})

        if models.SiteUser.objects.filter(email=email).exists():
            return JsonResponse({'success': False, 'message': '该邮箱已被注册'})

        user = models.SiteUser.objects.create(
            name=name,
            email=email,
            phone=phone,
            password=_set_unified_password(email, password),
            is_active=True,
        )
        create_notification(
            'new_user', f'管理员添加了用户 {name}',
            f'管理员手动添加了用户 {name} ({email})',
            icon='fas fa-user-plus', color='#10b981',
            link='/manager/admin/users/', related_id=user.id,
        )
        return JsonResponse({'success': True, 'message': f'用户 {name} 已创建'})

    return JsonResponse({'success': False})


# ==========================================
# Enhanced Admin Vendor Management (CRUD)
# ==========================================

def admin_edit_vendor(request):
    """Admin edit vendor"""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '未授权'})

    if request.method == 'GET':
        vid = request.GET.get('id')
        vendor = get_object_or_404(models.Vendor, id=vid)
        return JsonResponse({
            'success': True,
            'vendor': {
                'id': vendor.id,
                'company_name': vendor.company_name,
                'contact_name': vendor.contact_name,
                'email': vendor.email,
                'phone': vendor.phone,
                'description': vendor.description,
                'status': vendor.status,
                'is_active': vendor.is_active,
                'is_certified': vendor.is_certified,
                'is_official': vendor.is_official,
                'default_delivery_days_min': vendor.default_delivery_days_min,
                'default_delivery_days_max': vendor.default_delivery_days_max,
            }
        })

    if request.method == 'POST':
        vid = request.POST.get('id')
        vendor = get_object_or_404(models.Vendor, id=vid)
        vendor.company_name = request.POST.get('company_name', vendor.company_name).strip()
        vendor.contact_name = request.POST.get('contact_name', vendor.contact_name).strip()
        vendor.email = request.POST.get('email', vendor.email).strip()
        vendor.phone = request.POST.get('phone', vendor.phone).strip()
        vendor.description = request.POST.get('description', vendor.description).strip()
        status = request.POST.get('status')
        if status and status in dict(models.VENDOR_STATUS_CHOICES):
            vendor.status = status
        is_active = request.POST.get('is_active')
        if is_active is not None:
            vendor.is_active = is_active == 'true'
        if not vendor.is_official:
            certified_raw = request.POST.get('is_certified')
            if certified_raw is not None:
                wants_certified = certified_raw in ('1', 'true', 'on', 'yes')
                vendor.is_certified = wants_certified and vendor.status == 'approved'
                vendor.certified_at = timezone.now() if vendor.is_certified else None
        if 'logo' in request.FILES:
            vendor.logo = request.FILES['logo']
        delivery_min, delivery_max = _parse_delivery_days_override(request.POST, prefix='default_')
        vendor.default_delivery_days_min = delivery_min
        vendor.default_delivery_days_max = delivery_max
        vendor.save()
        return JsonResponse({'success': True, 'message': '卖家信息已更新'})

    return JsonResponse({'success': False})


def admin_add_vendor(request):
    """Admin add vendor manually"""
    if "name" not in request.session:
        return JsonResponse({'success': False, 'message': '未授权'})

    if request.method == 'POST':
        company_name = request.POST.get('company_name', '').strip()
        contact_name = request.POST.get('contact_name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        password = request.POST.get('password', '').strip()
        description = request.POST.get('description', '').strip()
        status = request.POST.get('status', 'approved')

        if not all([company_name, contact_name, email, password]):
            return JsonResponse({'success': False, 'message': '店铺名、联系人、邮箱和密码为必填项'})

        if models.Vendor.objects.filter(email=email).exists():
            return JsonResponse({'success': False, 'message': '该邮箱已被注册'})

        hashed = _set_unified_password(email, password)
        site_user = models.SiteUser.objects.filter(email__iexact=email, is_active=True).first()
        delivery_min, delivery_max = _parse_delivery_days_override(request.POST, prefix='default_')
        vendor = models.Vendor.objects.create(
            company_name=company_name,
            contact_name=contact_name,
            email=email,
            phone=phone,
            password=hashed,
            description=description,
            status=status,
            user=site_user,
            default_delivery_days_min=delivery_min,
            default_delivery_days_max=delivery_max,
        )
        if site_user:
            site_user.promote_to_seller()
        _link_dual_accounts_by_email(email)
        if 'logo' in request.FILES:
            vendor.logo = request.FILES['logo']
            vendor.save()

        create_notification(
            'vendor_registered', f'管理员添加了卖家 {company_name}',
            f'管理员手动添加了卖家 {company_name} ({email})',
            icon='fas fa-store', color='#667eea',
            link='/manager/admin/vendors/', related_id=vendor.id,
        )
        return JsonResponse({'success': True, 'message': f'卖家 {company_name} 已创建'})

    return JsonResponse({'success': False})


def admin_delete_vendor(request):
    """Admin delete vendor"""
    if not request.session.get('is_admin') or 'name' not in request.session:
        return JsonResponse({'success': False, 'message': '未授权'})
    if request.method == 'POST':
        vid = request.POST.get('id')
        vendor = models.Vendor.objects.filter(id=vid).first()
        if vendor:
            if vendor.is_official:
                return JsonResponse({'success': False, 'message': '官方直营店无法删除'})
            name = vendor.company_name
            detached_escrow_count = _detach_vendor_protected_relations(vendor)
            vendor.delete()
            message = f'卖家 {name} 及其店铺内容已删除'
            if detached_escrow_count:
                message += f'，{detached_escrow_count} 条历史托管记录已保留并解除关联'
            return JsonResponse({'success': True, 'message': message})
        return JsonResponse({'success': False, 'message': '卖家不存在'})
    return JsonResponse({'success': False})


# ==========================================
# Wallet / Credit System Views
# ==========================================

def _get_or_create_wallet(user):
    """Get or create wallet for a SiteUser."""
    wallet, _ = models.UserWallet.objects.get_or_create(user=user, defaults={
        'balance': Decimal('0.00'),
        'total_deposited': Decimal('0.00'),
        'total_spent': Decimal('0.00'),
    })
    return wallet


def public_wallet(request):
    """Public wallet dashboard: balance, transactions, top-up button."""
    user_id = request.session.get('site_user_id')
    if not user_id:
        return redirect('manager:user_login')
    user = get_object_or_404(models.SiteUser, pk=user_id)
    wallet = _get_or_create_wallet(user)
    transactions = user.wallet_transactions.all()[:20]
    return render(request, 'public/wallet.html', {
        'user': user,
        'wallet': wallet,
        'transactions': transactions,
    })


def public_wallet_topup(request):
    """Top-up page: amount selection + payment method."""
    user_id = request.session.get('site_user_id')
    if not user_id:
        return redirect('manager:user_login')
    user = get_object_or_404(models.SiteUser, pk=user_id)

    if request.method == 'POST':
        amount = request.POST.get('amount', '').strip()
        payment_method = request.POST.get('payment_method', '').strip()
        try:
            amount = Decimal(amount)
            if amount <= 0:
                raise ValueError
        except Exception:
            messages.error(request, 'Invalid amount')
            return redirect('manager:public_wallet_topup')

        topup = models.TopUpOrder(
            user=user,
            order_number=f'TOP{timezone.now().strftime("%Y%m%d")}{uuid.uuid4().hex[:8].upper()}',
            amount=amount,
            payment_method=payment_method,
            status='pending',
        )
        topup.save()
        messages.success(request, 'Top-up order created. Please complete payment.')
        return redirect('manager:public_wallet')

    return render(request, 'public/wallet_topup.html', {
        'user': user,
        'payment_methods': [
            ('pawapay', 'Mobile Money (PawaPay)'),
            ('kkiapay', 'Mobile Money (KKiaPay)'),
            ('orange_money', 'Orange Money'),
            ('paypal', 'PayPal'),
            ('credit_card', 'Credit Card'),
            ('bank_transfer', 'Bank Transfer'),
        ],
    })


@require_POST
def public_wallet_topup_webhook(request):
    """Webhook callback from payment gateway to confirm top-up."""
    order_number = request.POST.get('order_number', '').strip()
    transaction_id = request.POST.get('transaction_id', '').strip()
    signature = request.POST.get('signature', '').strip()

    # TODO: verify signature against gateway secret
    topup = models.TopUpOrder.objects.filter(order_number=order_number, status='pending').first()
    if not topup:
        return JsonResponse({'success': False, 'message': 'Order not found or already processed'})

    wallet = _get_or_create_wallet(topup.user)
    wallet.credit(
        topup.amount,
        source='deposit',
        description=f'Top-up via {topup.payment_method}',
        source_id=str(topup.id),
    )
    topup.status = 'completed'
    topup.transaction_id = transaction_id
    topup.completed_at = timezone.now()
    topup.save(update_fields=['status', 'transaction_id', 'completed_at'])
    return JsonResponse({'success': True, 'message': 'Top-up completed'})
